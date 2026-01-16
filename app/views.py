def home(request):
    return render(request, 'index.html')

from django.shortcuts import render
import requests
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from dotenv import load_dotenv
import os
import asyncio
import logging

from ampapi.dataclass import APIParams
from ampapi.bridge import Bridge
from ampapi.controller import AMPControllerInstance
import json
from pathlib import Path
from sqlalchemy import or_, and_
import time
from django_ratelimit.decorators import ratelimit
from casualsite.settings import get_client_ip
from .decorators import validate_json_schema, require_subscription_tier, bot_owner_required, discovery_approvers_required, server_owner_required

load_dotenv()

# Configure logger for views
logger = logging.getLogger(__name__)

# Discord bot API configuration
DISCORD_BOT_API_URL = os.getenv('DISCORD_BOT_API_URL', 'http://localhost:8001')
DISCORD_BOT_API_TOKEN = os.getenv('DISCORD_BOT_API_TOKEN', 'development-token')

# Cache for guilds_with_bot to prevent excessive API calls
_guilds_cache = {'data': None, 'timestamp': 0}
GUILDS_CACHE_TTL = 300  # Cache for 5 minutes (300 seconds) to prevent Discord API rate limiting

# Security Helper Functions
def is_safe_redirect(url):
    """
    Validate that a redirect URL is safe (same-origin).
    Prevents open redirect vulnerabilities (SEC-008).

    Args:
        url: The URL to validate

    Returns:
        bool: True if URL is safe, False otherwise
    """
    from urllib.parse import urlparse
    from django.conf import settings

    if not url:
        return False

    url = url.strip()

    # Block dangerous schemes
    if url.startswith(('javascript:', 'data:', 'vbscript:', 'file:')):
        logger.warning(f"[SECURITY] Blocked dangerous scheme in redirect: {url}")
        return False

    try:
        parsed = urlparse(url)
    except Exception:
        logger.warning(f"[SECURITY] Failed to parse redirect URL: {url}")
        return False

    # Relative URLs (no netloc) are safe if they start with /
    if not parsed.netloc:
        is_safe = url.startswith('/')
        if not is_safe:
            logger.warning(f"[SECURITY] Blocked non-absolute relative URL: {url}")
        return is_safe

    # Absolute URLs must match allowed hosts
    allowed_hosts = settings.ALLOWED_HOSTS
    # Also allow localhost for development
    if settings.DEBUG:
        allowed_hosts = list(allowed_hosts) + ['localhost', '127.0.0.1']

    is_safe = parsed.netloc in allowed_hosts
    if not is_safe:
        logger.warning(f"[SECURITY] Blocked redirect to external host: {parsed.netloc}")

    return is_safe


# LFG Game Limits by Tier
def get_lfg_game_limit(guild):
    """Get LFG game limit based on guild subscription tier."""
    if guild.is_vip or guild.subscription_tier in ['premium', 'Premium']:
        return None  # Unlimited
    elif guild.subscription_tier in ['pro', 'Pro']:
        return 10
    else:  # free tier
        return 5

def get_guilds_with_bot():
    """Get set of guild IDs where the bot is actually installed (queried from bot API with caching)."""
    global _guilds_cache

    # Check if cache is still valid
    current_time = time.time()
    if _guilds_cache['data'] is not None and (current_time - _guilds_cache['timestamp']) < GUILDS_CACHE_TTL:
        return _guilds_cache['data']

    # Cache expired or empty, fetch fresh data
    guilds_with_bot = set()
    try:
        import requests
        response = requests.get(
            f'{DISCORD_BOT_API_URL}/api/guilds',
            headers={'Authorization': f'Bearer {DISCORD_BOT_API_TOKEN}'},
            timeout=5
        )
        if response.status_code == 200:
            data = response.json()
            guilds_with_bot = set(data.get('guild_ids', []))
        else:
            logger.warning(f"Failed to get guild list from bot API: {response.status_code}")
            # Fallback to database check if API fails
            from .db import get_db_session
            from .models import Guild as GuildModel
            with get_db_session() as db:
                installed_guilds = db.query(GuildModel.guild_id).all()
                guilds_with_bot = {str(g.guild_id) for g in installed_guilds}
    except Exception as e:
        logger.error(f"Error checking installed guilds from API: {e}", exc_info=True)
        # Fallback to database check
        try:
            from .db import get_db_session
            from .models import Guild as GuildModel
            with get_db_session() as db:
                installed_guilds = db.query(GuildModel.guild_id).all()
                guilds_with_bot = {str(g.guild_id) for g in installed_guilds}
        except Exception as fallback_error:
            logger.error(f"Database fallback also failed: {fallback_error}", exc_info=True)

    # Update cache
    _guilds_cache['data'] = guilds_with_bot
    _guilds_cache['timestamp'] = current_time

    return guilds_with_bot

def get_member_guilds(request):
    """Calculate member guilds (guilds where user is member but NOT admin) that have the bot installed."""
    admin_guilds = request.session.get('discord_admin_guilds', [])
    all_guilds = request.session.get('discord_all_guilds', [])
    admin_guild_ids = {str(g['id']) for g in admin_guilds}
    member_guilds = [g for g in all_guilds if str(g['id']) not in admin_guild_ids]

    # Get guilds with bot installed
    guilds_with_bot = get_guilds_with_bot()

    # Only return member guilds where bot is installed
    return [g for g in member_guilds if str(g['id']) in guilds_with_bot]

def add_has_bot_flag(guilds):
    """Add 'has_bot' flag to each guild dict."""
    guilds_with_bot = get_guilds_with_bot()
    for guild in guilds:
        guild_id_str = str(guild['id'])
        guild['has_bot'] = guild_id_str in guilds_with_bot
    return guilds

def get_guild_with_permissions(guild_id, admin_guilds, all_guilds):
    """
    Get guild dict with full permission data (owner, permissions fields).
    Checks admin_guilds first (has owner/permissions), falls back to all_guilds.

    This ensures templates can check guild.owner status correctly.
    """
    # Try to find in admin_guilds first (has owner/permissions data)
    guild = next((g for g in admin_guilds if str(g['id']) == str(guild_id)), None)

    # Fall back to all_guilds if not found (member-only access)
    if not guild:
        guild = next((g for g in all_guilds if str(g['id']) == str(guild_id)), None)
        # Add owner=False for member-only guilds
        if guild:
            guild = dict(guild)  # Make a copy to avoid modifying session data
            guild['owner'] = False

    return guild

def check_lfg_game_limit(db, guild_id, guild):
    """Check if guild has reached their LFG game limit. Returns (can_add, current_count, limit)."""
    from .models import LFGGame

    limit = get_lfg_game_limit(guild)
    if limit is None:
        return (True, None, None)  # Unlimited

    current_count = db.query(LFGGame).filter_by(guild_id=int(guild_id), enabled=True).count()
    can_add = current_count < limit

    return (can_add, current_count, limit)


def get_discovery_lfg_post_limit(guild):
    """Get Discovery Network LFG monthly posting limit based on guild subscription tier."""
    from .module_utils import has_module_access

    # Complete Suite or LFG Module = Unlimited
    if guild.subscription_tier in ['premium', 'Premium'] or guild.is_vip:
        return None  # Unlimited

    # Check if they have LFG module
    if has_module_access(guild.guild_id, 'lfg'):
        return None  # Unlimited

    # Free tier = 5 posts per month
    return 5


def check_discovery_lfg_post_limit(db, guild_id, guild):
    """
    Check if guild has reached their Discovery Network LFG monthly posting limit.
    Returns (can_post, current_count, limit, reset_date).
    """
    from .models import LFGGroup
    from datetime import datetime, timezone
    from dateutil.relativedelta import relativedelta

    limit = get_discovery_lfg_post_limit(guild)
    if limit is None:
        return (True, None, None, None)  # Unlimited

    # Calculate start of current month (UTC)
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_start_timestamp = int(month_start.timestamp())

    # Calculate next month for reset date
    next_month = month_start + relativedelta(months=1)
    reset_date = next_month.strftime('%B %d, %Y')

    # Count LFG posts created this month in Discovery Network
    # (LFG groups that were shared to Discovery Network)
    current_count = db.query(LFGGroup).filter(
        LFGGroup.guild_id == int(guild_id),
        LFGGroup.created_at >= month_start_timestamp,
        LFGGroup.shared_to_network == True  # Discovery Network posts are flagged
    ).count()

    can_post = current_count < limit

    return (can_post, current_count, limit, reset_date)


def get_discovery_game_share_limit(guild):
    """Get Discovery Network game sharing monthly limit based on guild subscription tier."""
    from .module_utils import has_module_access

    # Complete Suite or Discovery Module = Unlimited
    if guild.subscription_tier in ['premium', 'Premium'] or guild.is_vip:
        return None  # Unlimited

    # Check if they have Discovery module
    if has_module_access(guild.guild_id, 'discovery'):
        return None  # Unlimited

    # Free tier = 3 shares per month (matches their 3 search config limit)
    return 3


def check_discovery_game_share_limit(db, guild_id, guild):
    """
    Check if guild has reached their Discovery Network game sharing monthly limit.
    Returns (can_share, current_count, limit, reset_date).
    """
    from .models import AnnouncedGame
    from datetime import datetime, timezone
    from dateutil.relativedelta import relativedelta

    limit = get_discovery_game_share_limit(guild)
    if limit is None:
        return (True, None, None, None)  # Unlimited

    # Calculate start of current month (UTC)
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_start_timestamp = int(month_start.timestamp())

    # Calculate next month for reset date
    next_month = month_start + relativedelta(months=1)
    reset_date = next_month.strftime('%B %d, %Y')

    # Count game shares created this month
    # (We'll add a 'shared_to_network' flag to track Discovery Network shares)
    current_count = db.query(AnnouncedGame).filter(
        AnnouncedGame.guild_id == int(guild_id),
        AnnouncedGame.announced_at >= month_start_timestamp
    ).count()

    can_share = current_count < limit

    return (can_share, current_count, limit, reset_date)


class DiscordBotSession:
    """Wrapper for making Discord API calls with bot token"""

    def __init__(self, base_url='https://discord.com/api/v10'):
        self.base_url = base_url
        bot_token = os.getenv('DISCORD_BOT_TOKEN')
        if not bot_token or bot_token == 'your_bot_token_here':
            self.headers = None
        else:
            self.headers = {'Authorization': f'Bot {bot_token}'}

    def get(self, path, **kwargs):
        """Make GET request to Discord API"""
        if not self.headers:
            # Return a mock response with 401 status if no token
            class MockResponse:
                status_code = 401
                text = "Bot token not configured"
                def json(self):
                    return {}
            return MockResponse()

        url = f"{self.base_url}{path}"
        return requests.get(url, headers=self.headers, **kwargs)

    def post(self, path, **kwargs):
        """Make POST request to Discord API"""
        if not self.headers:
            # Return a mock response with 401 status if no token
            class MockResponse:
                status_code = 401
                text = "Bot token not configured"
                def json(self):
                    return {}
            return MockResponse()

        url = f"{self.base_url}{path}"
        return requests.post(url, headers=self.headers, **kwargs)

    def put(self, path, **kwargs):
        """Make PUT request to Discord API"""
        if not self.headers:
            class MockResponse:
                status_code = 401
                text = "Bot token not configured"
                def json(self):
                    return {}
            return MockResponse()

        url = f"{self.base_url}{path}"
        return requests.put(url, headers=self.headers, **kwargs)

    def patch(self, path, **kwargs):
        """Make PATCH request to Discord API"""
        if not self.headers:
            class MockResponse:
                status_code = 401
                text = "Bot token not configured"
                def json(self):
                    return {}
            return MockResponse()

        url = f"{self.base_url}{path}"
        return requests.patch(url, headers=self.headers, **kwargs)

    def delete(self, path, **kwargs):
        """Make DELETE request to Discord API"""
        if not self.headers:
            class MockResponse:
                status_code = 401
                text = "Bot token not configured"
                def json(self):
                    return {}
            return MockResponse()

        url = f"{self.base_url}{path}"
        return requests.delete(url, headers=self.headers, **kwargs)


def get_bot_session(guild_id):
    """
    Get a Discord bot API session for making requests.
    Returns None if bot token is not configured.
    """
    bot_token = os.getenv('DISCORD_BOT_TOKEN')
    if not bot_token or bot_token == 'your_bot_token_here':
        logger.warning(f"DISCORD_BOT_TOKEN not configured for guild {guild_id}")
        return None

    return DiscordBotSession()


def check_lfg_manager_role(guild_id, user_id):
    """
    Check if a user has the "LFG Manager" role with required permissions.

    The LFG Manager role must have both:
    - CREATE_EVENTS permission (bit 44)
    - MANAGE_EVENTS permission (bit 33)

    Args:
        guild_id: Discord guild ID
        user_id: Discord user ID

    Returns:
        bool: True if user has valid LFG Manager role, False otherwise
    """
    if not user_id:
        return False

    try:
        import requests

        bot_token = os.getenv('DISCORD_BOT_TOKEN')
        if not bot_token:
            return False

        # Get guild roles from cache (no Discord API call!)
        from .discord_resources import get_guild_roles, get_guild_member

        roles = get_guild_roles(str(guild_id))

        lfg_manager_role_id = None
        for role in roles:
            if role.get('name') == 'LFG Manager':
                # Verify role has both CREATE_EVENTS and MANAGE_EVENTS permissions
                role_permissions = int(role.get('permissions', 0))
                # CREATE_EVENTS = 1 << 44 (17592186044416), MANAGE_EVENTS = 1 << 33 (8589934592)
                has_create = (role_permissions & (1 << 44)) != 0
                has_manage = (role_permissions & (1 << 33)) != 0
                if has_create and has_manage:
                    lfg_manager_role_id = role.get('id')
                break

        # Check if user has the LFG Manager role (from cache, no API call!)
        if lfg_manager_role_id:
            member_data = get_guild_member(str(guild_id), str(user_id))
            if member_data:
                member_roles = member_data.get('roles', [])
                if str(lfg_manager_role_id) in member_roles:
                    return True

    except Exception as e:
        logger.error(f"Error checking LFG Manager role: {e}", exc_info=True)

    return False


def log_lfg_audit(db, guild_id, group_id, action, actor_id, actor_name,
                  field_changed=None, old_value=None, new_value=None,
                  group_name=None, game_name=None):
    """
    Create an audit log entry for LFG group changes.

    Args:
        db: Database session
        guild_id: Discord guild ID
        group_id: LFG group ID (can be None for deletions after commit)
        action: Action performed ('create', 'update', 'delete')
        actor_id: Discord user ID who performed the action
        actor_name: Discord username
        field_changed: Which field was changed (for updates)
        old_value: Previous value (will be JSON encoded if complex type)
        new_value: New value (will be JSON encoded if complex type)
        group_name: Thread/group name for context
        game_name: Game name for context
    """
    try:
        from .models import LFGGroupAuditLog

        # Convert complex values to JSON strings
        if old_value is not None and not isinstance(old_value, str):
            old_value = json.dumps(old_value)
        if new_value is not None and not isinstance(new_value, str):
            new_value = json.dumps(new_value)

        audit_entry = LFGGroupAuditLog(
            guild_id=int(guild_id),
            group_id=group_id,
            action=action,
            actor_id=int(actor_id),
            actor_name=actor_name,
            field_changed=field_changed,
            old_value=old_value,
            new_value=new_value,
            group_name=group_name,
            game_name=game_name
        )
        db.add(audit_entry)
        # Note: Caller is responsible for committing the transaction

    except Exception as e:
        logger.error(f"Error creating audit log entry: {e}", exc_info=True)


def send_lfg_browser_notification(guild_id, lfg_config, notification_type, embed_data):
    """
    Send LFG Browser notification to configured channel and/or webhook.

    Args:
        guild_id: Discord guild ID
        lfg_config: LFGConfig instance with notification settings
        notification_type: Type of notification ('create', 'update', 'delete', 'join', 'leave')
        embed_data: Dict with embed fields (title, description, color, fields, etc.)
    """
    try:
        import requests

        # Check if this notification type is enabled
        notification_enabled = {
            'create': lfg_config.notify_on_group_create,
            'update': lfg_config.notify_on_group_update,
            'delete': lfg_config.notify_on_group_delete,
            'join': lfg_config.notify_on_member_join,
            'leave': lfg_config.notify_on_member_leave,
        }.get(notification_type, False)

        if not notification_enabled:
            return

        # Build embed
        embed = {
            'title': embed_data.get('title', 'LFG Group Update'),
            'description': embed_data.get('description', ''),
            'color': embed_data.get('color', 0x5865F2),
            'fields': embed_data.get('fields', []),
            'timestamp': embed_data.get('timestamp'),
            'footer': embed_data.get('footer'),
        }

        # Send to channel if configured
        if lfg_config.browser_notify_channel_id:
            bot_session = get_bot_session(guild_id)
            if bot_session:
                try:
                    # For group creation, queue bot action to create thread with interactive view
                    if notification_type == 'create' and embed_data.get('group_id'):
                        from .models import PendingAction, ActionType, ActionStatus
                        from .db import get_db_session as get_db
                        import time as time_lib

                        # Queue action for bot to create thread with interactive components
                        with get_db() as db_session:
                            action = PendingAction(
                                guild_id=int(guild_id),
                                action_type=ActionType.LFG_THREAD_CREATE,
                                payload=json.dumps({
                                    'group_id': embed_data.get('group_id'),
                                    'channel_id': str(lfg_config.browser_notify_channel_id)
                                }),
                                status=ActionStatus.PENDING,
                                priority=1,  # High priority
                                created_at=int(time_lib.time())
                            )
                            db_session.add(action)
                            db_session.flush()
                            logger.info(f"Queued LFG thread creation action for group {embed_data.get('group_id')}")
                    else:
                        # For other notifications, just post to channel
                        response = bot_session.post(
                            f'/channels/{lfg_config.browser_notify_channel_id}/messages',
                            json={'embeds': [embed]}
                        )
                        if response.status_code != 200:
                            logger.warning(f"Failed to send LFG notification to channel: {response.status_code}")
                except Exception as e:
                    logger.error(f"Error sending LFG channel notification: {e}")

        # Send to webhook if configured
        if lfg_config.webhook_url:
            try:
                response = requests.post(
                    lfg_config.webhook_url,
                    json={'embeds': [embed]},
                    timeout=5
                )
                if response.status_code not in [200, 204]:
                    logger.warning(f"Failed to send LFG webhook notification: {response.status_code}")
            except Exception as e:
                logger.error(f"Error sending LFG webhook notification: {e}")

    except Exception as e:
        logger.error(f"Error in send_lfg_browser_notification: {e}", exc_info=True)


def send_lfg_webhook_notification(webhook_url, embed_data):
    """
    Send LFG notification to a webhook.

    Args:
        webhook_url: Discord webhook URL
        embed_data: Dict with embed fields
    """
    try:
        import requests

        # Build Discord embed
        embed = {
            'title': embed_data.get('title', 'LFG Group'),
            'description': embed_data.get('description', ''),
            'color': embed_data.get('color', 0x5865F2),
            'fields': embed_data.get('fields', []),
        }

        if embed_data.get('footer'):
            embed['footer'] = embed_data['footer']

        response = requests.post(
            webhook_url,
            json={'embeds': [embed]},
            timeout=5
        )
        if response.status_code not in [200, 204]:
            logger.warning(f"Failed to send LFG webhook notification: {response.status_code}")
    except Exception as e:
        logger.error(f"Error sending LFG webhook notification: {e}")


def send_lfg_browser_dm(guild_id, user_id, embed_data):
    """
    Send a DM to a user about an LFG Browser event.

    Args:
        guild_id: Discord guild ID (for bot session)
        user_id: Discord user ID to DM
        embed_data: Dict with embed fields

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        bot_session = get_bot_session(guild_id)
        if not bot_session:
            return False

        # Create DM channel
        dm_response = bot_session.post('/users/@me/channels', json={'recipient_id': str(user_id)})
        if dm_response.status_code != 200:
            logger.warning(f"Failed to create DM channel for user {user_id}: {dm_response.status_code}")
            return False

        dm_channel = dm_response.json()
        dm_channel_id = dm_channel.get('id')

        if not dm_channel_id:
            return False

        # Build embed
        embed = {
            'title': embed_data.get('title', 'LFG Group Update'),
            'description': embed_data.get('description', ''),
            'color': embed_data.get('color', 0x5865F2),
            'fields': embed_data.get('fields', []),
            'timestamp': embed_data.get('timestamp'),
            'footer': embed_data.get('footer', {'text': 'LFG Browser'}),
        }

        # Send DM
        msg_response = bot_session.post(f'/channels/{dm_channel_id}/messages', json={'embeds': [embed]})
        if msg_response.status_code != 200:
            logger.warning(f"Failed to send DM to user {user_id}: {msg_response.status_code}")
            return False

        return True

    except Exception as e:
        logger.error(f"Error sending LFG Browser DM to {user_id}: {e}")
        return False


def send_lfg_browser_dms(guild_id, user_ids, embed_data):
    """
    Send DMs to multiple users about an LFG Browser event.

    Args:
        guild_id: Discord guild ID
        user_ids: List of Discord user IDs
        embed_data: Dict with embed fields

    Returns:
        int: Number of successful DMs sent
    """
    success_count = 0
    for user_id in user_ids:
        if send_lfg_browser_dm(guild_id, user_id, embed_data):
            success_count += 1
    return success_count


# ============================================================================
# Bulk Operation Rate Limiting Helper Functions
# ============================================================================

def get_tier_limits(guild_record):
    """
    Get daily and per-operation limits based on guild subscription tier.
    Returns: (max_items_per_operation, max_items_per_day)

    - Free: 6 items per operation, 6 total per day
    - Pro: 10 items per operation, 10 total per day
    - Premium/VIP/Engagement Module: Unlimited
    """
    from .models import SubscriptionTier
    from .module_utils import has_module_access

    # Check if guild has Engagement Module (grants unlimited bulk operations)
    if has_module_access(guild_record.guild_id, 'engagement'):
        return (None, None)  # Unlimited

    if guild_record.is_vip or guild_record.subscription_tier == SubscriptionTier.PREMIUM.value:
        return (None, None)  # Unlimited
    elif guild_record.subscription_tier == SubscriptionTier.PRO.value:
        return (10, 10)
    else:  # Free tier
        return (6, 6)


def check_daily_bulk_limit(guild_id, operation_category, items_count, guild_record):
    """
    Check if guild has exceeded daily bulk operation limit.
    Returns: (allowed: bool, error_message: str or None, usage_info: dict)
    """
    from .db import get_db_session
    from .models import DailyBulkUsage
    from datetime import datetime

    max_per_op, max_per_day = get_tier_limits(guild_record)

    # Premium/VIP has no limits
    if max_per_op is None:
        return (True, None, {'tier': 'premium', 'unlimited': True})

    # Check per-operation limit first
    if items_count > max_per_op:
        tier_name = 'Pro' if max_per_op == 10 else 'Free'
        return (False, f'{tier_name} tier limited to {max_per_op} items per operation. You have {items_count} items.', None)

    # Check daily limit
    today = int(datetime.now().strftime('%Y%m%d'))

    with get_db_session() as db:
        usage = db.query(DailyBulkUsage).filter_by(
            guild_id=int(guild_id),
            date=today,
            operation_category=operation_category
        ).first()

        current_usage = usage.items_processed if usage else 0

        if current_usage + items_count > max_per_day:
            tier_name = 'Pro' if max_per_day == 10 else 'Free'
            return (
                False,
                f'{tier_name} tier limited to {max_per_day} items per day. You have used {current_usage}/{max_per_day} today. Upgrade to Premium for unlimited operations.',
                {'current_usage': current_usage, 'max_per_day': max_per_day, 'requested': items_count}
            )

        return (True, None, {'current_usage': current_usage, 'max_per_day': max_per_day, 'remaining': max_per_day - current_usage - items_count})


def record_bulk_usage(guild_id, operation_category, items_count):
    """
    Record bulk operation usage for daily tracking.
    """
    from .db import get_db_session
    from .models import DailyBulkUsage
    from datetime import datetime
    from sqlalchemy import func

    today = int(datetime.now().strftime('%Y%m%d'))
    current_time = int(time.time())

    with get_db_session() as db:
        usage = db.query(DailyBulkUsage).filter_by(
            guild_id=int(guild_id),
            date=today,
            operation_category=operation_category
        ).first()

        if usage:
            usage.items_processed += items_count
            usage.operations_count += 1
            usage.last_operation_at = current_time
        else:
            usage = DailyBulkUsage(
                guild_id=int(guild_id),
                date=today,
                operation_category=operation_category,
                items_processed=items_count,
                operations_count=1,
                first_operation_at=current_time,
                last_operation_at=current_time
            )
            db.add(usage)


DISCORD_ACTIVITY_FILE = Path("/srv/ch-webserver/gamingactivity/activity_data.json")

def get_discord_activity():
    if DISCORD_ACTIVITY_FILE.exists():
        with open(DISCORD_ACTIVITY_FILE, "r") as f:
            return json.load(f)
    return {}


# Games tracked through AMP
STATIC_GAME_INFO = {
    "CH-7DTD01": {
        "display_name": "Dynamic Horde Protocol (PvE)",
        "description": "A custom survival world where biomes bite back. Bots spawn threats, buffs twist the rules, nothing is predictable, and that’s the point.",
        "discord_invite": "https://discord.gg/ECwJWppSjQ",
        "steam_link": "https://store.steampowered.com/app/251570/7_Days_to_Die/",
        "steam_appid": "251570",
        "connect_pw": "Join our Discord to gain access!"
    },
    "CH-Icarus01": {
        "display_name": "Prospectors Lounge",
        "description": "Custom server, No Lifing is Optional",
        "discord_invite": "https://discord.gg/ECwJWppSjQ",
        "steam_link": "https://store.steampowered.com/app/1149460/ICARUS/",
        "steam_appid": "1149460",
        "connect_pw": "Join our Discord to gain access!"
    },
    "CH-Palworld01": {
        "display_name": "Pal Sanctuary",
        "description": "Where we catch Pals and ignore responsibilities together!",
        "discord_invite": "https://discord.gg/ECwJWppSjQ",
        "steam_link": "https://store.steampowered.com/app/1623730/Palworld/",
        "steam_appid": "1623730",
        "connect_pw": "Join our Discord to gain access!"
    },

    # "CasualHeroes-Ascended01": {
    #     "display_name": "Dragonwilds",
    #     "description": "As soon as dedicated servers drop, we’re self-hosting, building custom content, and launching a one-of-a-kind Dragonwilds adventure.",
    #     "discord_invite": "https://discord.gg/WZzTppBgBz",
    #     "steam_link": "https://store.steampowered.com/app/1374490/RuneScape_Dragonwilds/",
    #     "custom_amp_img": "/static/img/games/Dragonwilds/dw_static.jpg",
    #     # "steam_appid": "1374490",
    #     "connect_pw": "N/A"
    # },

    # "Enshrouded01": {
    #     "display_name": "Enshrouded",
    #     "description": "Soulslike survival with tuned combat, custom altars, and strange terrain. Built for players who like challenge, discovery, and a bit of chaos.",
    #     "discord_invite": "https://discord.gg/CHHS",
    #     "steam_link": "https://store.steampowered.com/app/1203620/Enshrouded/",
    #     "steam_appid": "1203620",
    #     "connect_pw": "Join the Discord"
    # },
    # "CasualHeroes-Vrising01": {
    #     "display_name": "V Rising",
    #     "description": "Modded gothic survival with PvP and random preset days, castle building, and a world that rewards planning over panic. Vardoran's waiting, Rise. Bite. Build.",
    #     "discord_invite": "https://discord.gg/CHHS",
    #     "steam_link": "https://store.steampowered.com/app/1604030/V_Rising/",
    #     "steam_appid": "1604030",
    #     "connect_pw": "No Password"
    # }
}

# ============================================
# Discord-Only Games (Static Info)
# ============================================
# These provide Steam links, Discord invites, and custom images
# for games tracked via Discord activity (not AMP servers)
# NOTE: Do NOT add these to STATIC_GAME_INFO above - that's for AMP servers only!
DISCORD_GAME_STATIC_INFO = {
    "WoW": {
        "steam_link": "https://worldofwarcraft.blizzard.com/en-us/",
        "discord_invite": "https://discord.gg/ECwJWppSjQ",
        "custom_img": "/static/img/games/wow/dwarf.webp",
        "link_label": "View Site"
    },
    "ESO": {
        "steam_link": "https://store.steampowered.com/app/306130/The_Elder_Scrolls_Online/",
        "discord_invite": "https://discord.gg/ECwJWppSjQ",
        "steam_appid": "306130",
        "link_label": "View on Steam"
    },
}

# Games tracked through Discord only
DISCORD_GAMES = [
    #     {
    #     "id": "Dune",
    #     "name": "Dune: Awakening",
    #     "description": "The premier casual guild in the Dune uninverse. Chill dungeon runs, late-night banter, and a crew that’s always online. Whether you're new or a raiding vet, Casual Heroes has a spot for you.",
    #     "guild_page": "https://casual-heroes.com/dune/",
    #     "steam_link": "https://store.steampowered.com/app/1172710/Dune_Awakening/",
    #     "discord_invite": "https://discord.gg/jAJvykZvej",
    #     "steam_appid": "1172710",
    #     "online": "-",
    #     "max": "-",
    #     "link_label": "View on Steam"
    # },
    # {
    #     "id": "Dragonwilds",
    #     "name": "Dragonwilds",
    #     "description": "As soon as dedicated servers drop, we’re self-hosting, building custom content, and launching a one-of-a-kind Dragonwilds adventure.",
    #     "steam_link": "https://store.steampowered.com/app/1374490/RuneScape_Dragonwilds/",
    #     "discord_invite": "https://discord.gg/WZzTppBgBz",
    #     "steam_appid": "1374490",
    #     "custom_img": "/static/img/games/Dragonwilds/dw_static.jpg",
    #     "online": "-",
    #     "max": "-",
    #     "link_label": "View on Steam"
    # },
    # {
    #     "id": "MHW",
    #     "name": "Monster Hunter Wilds",
    #     "description": "From fashion shows to chaotic wild hunts, our Monster Hunter community is growing fast. Whether you're min-maxing DPS or just showing off your best drip, there's a spot at the campfire for you.",
    #     "steam_link": "https://store.steampowered.com/app/2246340/Monster_Hunter_Wilds/",
    #     "discord_invite": "https://discord.gg/3rKQptH7Fd",
    #     "steam_appid": "2246340",
    #     "online": "-",
    #     "max": "-",
    #     "link_label": "View on Steam"
    # },
    # {
    #     "id": "Pantheon",
    #     "name": "Pantheon: Rise of the Fallen",
    #     "description": "The premier casual guild in the Pantheon world. Chill dungeon runs, late-night banter, and a crew that’s always online. Whether you're new or a raiding vet, Casual Heroes has a spot for you.",
    #     "steam_link": "https://store.steampowered.com/app/3107230/Pantheon_Rise_of_the_Fallen/",
    #     "discord_invite": "https://discord.gg/REHJrygu64",
    #     "steam_appid": "3107230",
    #     "online": "-",
    #     "max": "-",
    #     "link_label": "View on Steam"
    # },
    # {
    #     "id": "PoE2",
    #     "name": "Path of Exile 2",
    #     "description": "You’ll always find someone theorycrafting their next crazy build here. Casual Heroes are farming, testing, and helping each other every step of the way.",
    #     "steam_link": "https://store.steampowered.com/app/2694490/Path_of_Exile_2/",
    #     "discord_invite": "https://discord.gg/fs9qAkVkxH",
    #     "steam_appid": "2694490",
    #     "online": "-",
    #     "max": "-",
    #     "link_label": "View on Steam"
    # },
    {
        "id": "WoW",
        "name": "World of Warcraft",
        "description": "Teaming up with longtime friend Eldronox and his legendary community 'Eternal Legends', we're building a World of Warcraft guild called <Casual Legends>. A chill, zero-drama space for adventurers who play at their own pace..",
        "steam_link": "https://worldofwarcraft.blizzard.com/en-us/",
        "discord_invite": "https://discord.gg/ECwJWppSjQ",
        "custom_img": "/static/img/games/wow/dwarf.webp",
        "online": "-",
        "max": "-",
        "link_label": "View Site"
    },
    {
        "id": "ESO",
        "name": "Elder Scrolls Online",
        "description": "Casual Legends is building a PC-NA ESO guild for adults who want chill runs and real progress—without the drama or sweaty expectations. New and returning players welcome. We learn together, gear up together, and push harder content when we’re ready.",
        "steam_link": "https://store.steampowered.com/app/306130/The_Elder_Scrolls_Online/",
        "discord_invite": "https://discord.gg/ECwJWppSjQ",
        "steam_appid": "306130",
        "online": "-",
        "max": "-",
        "link_label": "View Site"
    }
]

# ============================================
# AMP Instance Data Cache (PERSISTENT - Industry Standard)
# ============================================
# Uses persistent file cache system (survives restarts)
# Same approach as Discord caching for consistency
#
# ⚙️ CACHE CONFIGURATION - Adjust this to control how often AMP API is called
# Lower value = more real-time updates, but more API calls
# Higher value = fewer API calls, but slower to show status changes
#
AMP_CACHE_TTL = 60  # 60 seconds (1 minute) - Good balance for game server status
#
# Common values:
#   60   = 1 minute  (very responsive, good for game servers)
#   300  = 5 minutes (good balance)
#   600  = 10 minutes (slower updates, lower API usage)
#   1800 = 30 minutes (slow updates, minimal API usage)
#   3600 = 60 minutes (very slow updates, very low API usage)

def get_cached_instance_data(instance_name):
    """
    Get cached AMP instance data using persistent cache.

    Returns cached data if valid, None if expired or not found.
    """
    from .discord_cache import get_cache

    cache = get_cache()
    cache_key = f"amp_instance:{instance_name}"

    cached = cache.get(cache_key)
    if cached is not None:
        logger.info(f"AMP cache HIT for {instance_name}")
        return cached

    logger.debug(f"AMP cache MISS for {instance_name}")
    return None

def set_cached_instance_data(instance_name, data):
    """
    Store AMP instance data in persistent cache with TTL.

    Cache survives Django restarts.
    """
    from .discord_cache import get_cache

    cache = get_cache()
    cache_key = f"amp_instance:{instance_name}"

    cache.set(cache_key, data, ttl=AMP_CACHE_TTL)
    logger.info(f"Cached AMP data for {instance_name} (TTL: {AMP_CACHE_TTL}s)")

def clear_amp_cache(instance_name=None):
    """
    Clear AMP cache for a specific instance or all instances.

    Args:
        instance_name: Specific instance to clear, or None to clear all

    Returns:
        True if cache was cleared
    """
    from .discord_cache import get_cache

    cache = get_cache()

    if instance_name:
        cache_key = f"amp_instance:{instance_name}"
        cache.delete(cache_key)
        logger.info(f"Cleared AMP cache for {instance_name}")
        return True
    else:
        # Clear all AMP instance caches
        # Note: This requires iterating through known instances
        from .views import STATIC_GAME_INFO
        for name in STATIC_GAME_INFO.keys():
            cache.delete(f"amp_instance:{name}")
        logger.info("Cleared all AMP instance caches")
        return True

async def fetch_instance_data(instance_name):
    # Check cache first
    cached_data = get_cached_instance_data(instance_name)
    if cached_data:
        return cached_data
    _params = APIParams(
        url=os.getenv("AMP_URL"),
        user=os.getenv("AMP_USER"),
        password=os.getenv("AMP_PASSWORD")
    )
    Bridge(api_params=_params)

    controller = AMPControllerInstance()
    try:
        await controller.get_instances()
    except Exception as e:
        logger.error(f"Could not fetch AMP instances: {e}")
        return safe_amp_fallback(instance_name)

    for instance in controller.instances:
        if instance.instance_name == instance_name:
            try:
                status = await instance.get_status(format_data=False)
                ports = await instance.get_port_summaries(format_data=False)

                # Filter out internal ports and non-game ports (SFTP, etc.)
                valid_ports = [
                    p for p in ports
                    if not p.get("internalonly", False)
                    and p.get("port") is not None
                    and "sftp" not in p.get("name", "").lower()  # Exclude SFTP ports
                ]

                # Preferred port names (in order of preference)
                # Different games use different naming conventions in AMP
                preferred_order = [
                    "server and steam port",  # 7 Days to Die
                    "game port",              # Most games
                    "game and mods port",     # Some games
                    "query port",             # Fallback for some games
                ]

                game_port = next(
                    (p for name in preferred_order for p in valid_ports if name.lower() in p.get("name", "").lower()),
                    None
                )

                # If no preferred port found, use the first valid port
                if not game_port and valid_ports:
                    game_port = valid_ports[0]

                # SECURITY HARDENING: Use timeout and error handling for external IP lookup
                # This is intentional for AMP game servers - players need the public IP
                fallback_ip = "Unknown"
                if not game_port.get("ip") and not game_port.get("hostname"):
                    try:
                        # Use short timeout to avoid blocking if service is slow
                        fallback_ip = requests.get("https://ifconfig.me/ip", timeout=2).text.strip()
                    except (requests.RequestException, Exception) as e:
                        logger.warning(f"Failed to fetch public IP from ifconfig.me: {e}")
                        fallback_ip = "Unknown"

                ip = (
                    game_port.get("ip")
                    or game_port.get("hostname")
                    or fallback_ip
                )
                port = str(game_port.get("port")) if game_port else "Unknown"

                static_info = STATIC_GAME_INFO.get(instance_name, {})

                # ✅ Check if AMP reports the server as Running
                is_running = status.get("running", True)

                # Build the data object
                data = {
                    "id": instance_name,
                    "name": static_info.get("display_name", instance_name),
                    "title": static_info.get("display_name", instance_name),
                    "description": static_info.get("description", ""),
                    "discord_invite": static_info.get("discord_invite", "#"),
                    "guild_page": static_info.get("guild_page", ""),
                    "steam_link": static_info.get("steam_link", "#"),
                    "steam_appid": static_info.get("steam_appid"),
                    "custom_img": static_info.get("custom_img"),
                    "custom_amp_img": static_info.get("custom_amp_img"),
                    "online": status["metrics"]["active_users"]["raw_value"],
                    "max": status["metrics"]["active_users"]["max_value"],
                    "ip": f"{ip}:{port}",
                    "pw": static_info.get("connect_pw", "Unknown"),
                    "source": "amp",
                    "status_label": "🟢 Online" if is_running else "🔴 Offline"
                }

                # Cache the successful result
                set_cached_instance_data(instance_name, data)
                return data

            except Exception as e:
                logger.warning(f"AMP instance {instance_name} error: {e}")
                fallback = safe_amp_fallback(instance_name)
                # Cache fallback data too to prevent repeated failures
                set_cached_instance_data(instance_name, fallback)
                return fallback

    logger.info(f"AMP instance {instance_name} not found — using fallback.")
    fallback = safe_amp_fallback(instance_name)
    # Cache fallback data to prevent repeated lookups
    set_cached_instance_data(instance_name, fallback)
    return fallback


    # fallback
def safe_amp_fallback(instance_name):
    static_info = STATIC_GAME_INFO.get(instance_name, {})
    return {
        "id": instance_name,
        "name": static_info.get("display_name", instance_name),
        "title": static_info.get("display_name", instance_name),
        "description": static_info.get("description", ""),
        "discord_invite": static_info.get("discord_invite", "#"),
        "guild_page": static_info.get("guild_page", ""),
        "steam_link": static_info.get("steam_link", "#"),
        "steam_appid": static_info.get("steam_appid"),
        "custom_img": static_info.get("custom_img"),
        "custom_amp_img": static_info.get("custom_amp_img"),
        "online": "-",
        "max": "-",
        "ip": "Unavailable",
        "pw": static_info.get("connect_pw", "Unknown"),
        "source": "amp",
        "status_label": "🔴 Offline" 
    }


# Merge and render
def games_we_play(request):
    """
    Load ALL games from database (Activity Tracker).
    Supports both AMP servers and Discord-only games.
    """
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    all_games = []
    discord_activity = get_discord_activity()

    try:
        from .db import get_db_session
        from .models import SiteActivityGame

        with get_db_session() as db:
            # Load all active games from database
            db_games = db.query(SiteActivityGame).filter_by(is_active=True).order_by(SiteActivityGame.sort_order).all()

            # Collect AMP instance names to fetch in parallel
            amp_instance_names = []
            for db_game in db_games:
                if db_game.game_type in ['amp', 'both'] and db_game.amp_instance_id:
                    amp_instance_names.append(db_game.amp_instance_id)

            # Fetch AMP data in parallel
            amp_data_map = {}
            if amp_instance_names:
                amp_results = loop.run_until_complete(asyncio.gather(
                    *(fetch_instance_data(name) for name in amp_instance_names),
                    return_exceptions=True
                ))
                amp_data_map = {game.get("id"): game for game in amp_results if isinstance(game, dict)}

            # Build game cards from database
            for db_game in db_games:
                game_dict = {
                    "id": db_game.game_key,
                    "name": db_game.display_name,
                    "title": db_game.display_name,
                    "description": db_game.description or "",
                    "steam_appid": db_game.steam_appid,
                    "custom_img": db_game.custom_img,
                    "steam_link": db_game.steam_link,
                    "discord_invite": db_game.discord_invite,
                    "link_label": db_game.link_label or "View Site",
                    "online": "-",
                    "max": "-",
                    "live_now": False,
                }

                # For AMP servers: inject live server data
                if db_game.game_type in ['amp', 'both'] and db_game.amp_instance_id:
                    amp_data = amp_data_map.get(db_game.amp_instance_id)
                    if amp_data:
                        game_dict["source"] = "amp"
                        game_dict["online"] = amp_data.get("online", "-")
                        game_dict["max"] = amp_data.get("max", "-")
                        game_dict["live_now"] = amp_data.get("live_now", False)
                        game_dict["ip"] = amp_data.get("ip", "Unavailable")
                        game_dict["connect_pw"] = amp_data.get("connect_pw", "Unknown")
                        game_dict["status_label"] = amp_data.get("status_label", "Unknown")

                # For Discord games: inject Discord activity data
                if db_game.game_type in ['discord', 'both']:
                    stats = discord_activity.get(db_game.game_key)
                    if stats:
                        game_dict["source"] = "discord"
                        game_dict["online"] = stats.get("active", "-")
                        game_dict["max"] = stats.get("total", "-")
                        game_dict["live_now"] = stats.get("active", 0) > 0

                all_games.append(game_dict)

            logger.info(f"Loaded {len(all_games)} active games from database ({len(amp_data_map)} AMP, {len(all_games) - len(amp_data_map)} Discord)")
    except Exception as e:
        logger.error(f"Failed to load games from database: {e}")
        # Empty list if database fails
        all_games = []

    return render(request, 'gamesweplay.html', { 'games': all_games })


# ============================================================================
# SITE ACTIVITY TRACKER - BOT OWNER ADMIN PANEL
# ============================================================================
# Configuration interface for Discord game activity tracking
# SECURITY: BOT OWNER ONLY - These views are protected by @bot_owner_required

@bot_owner_required
def site_activity_tracker_admin(request, guild_id):
    """
    Admin panel for configuring site activity tracker.

    BOT OWNER ONLY - Configure which games to track and their Discord role mappings.
    Integrated into QuestLog guild dashboard.
    """
    from .db import get_db_session
    from .models import SiteActivityGame, SiteActivityGuildRole, Guild as GuildModel

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Get guild context (for sidebar and base_guild.html)
    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    guild_record = None

    with get_db_session() as db:
        # Get guild record for premium features
        guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()

        # Get all configured games with their role mappings
        games = db.query(SiteActivityGame).order_by(SiteActivityGame.sort_order, SiteActivityGame.display_name).all()

        # Get role mappings for each game
        games_data = []
        for game in games:
            roles = db.query(SiteActivityGuildRole).filter_by(game_id=game.id, is_active=True).all()

            games_data.append({
                'id': game.id,
                'game_key': game.game_key,
                'display_name': game.display_name,
                'description': game.description,
                'activity_keywords': json.loads(game.activity_keywords),
                'is_active': game.is_active,
                'sort_order': game.sort_order,
                'roles': [{
                    'id': r.id,
                    'guild_id': r.guild_id,
                    'role_id': r.role_id,
                    'guild_name': r.guild_name,
                    'role_name': r.role_name,
                } for r in roles]
            })

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'games': games_data,
        'data_file': str(DISCORD_ACTIVITY_FILE),
        'active_page': 'site_activity_tracker',
    }

    return render(request, 'questlog/site_activity_tracker.html', context)


@bot_owner_required
@require_http_methods(["GET", "POST", "PUT", "DELETE"])
def api_site_activity_games(request, guild_id, game_id=None):
    """
    API endpoint for managing site activity games.

    BOT OWNER ONLY

    GET: List all games or get specific game
    POST: Create new game
    PUT: Update existing game
    DELETE: Delete game
    """
    from .db import get_db_session
    from .models import SiteActivityGame, SiteActivityGuildRole

    with get_db_session() as db:
        if request.method == "GET":
            if game_id:
                # Get specific game
                game = db.query(SiteActivityGame).filter_by(id=game_id).first()
                if not game:
                    return JsonResponse({'error': 'Game not found'}, status=404)

                roles = db.query(SiteActivityGuildRole).filter_by(game_id=game.id).all()

                return JsonResponse({
                    'id': game.id,
                    'game_key': game.game_key,
                    'display_name': game.display_name,
                    'description': game.description,
                    'game_type': game.game_type,
                    'amp_instance_id': game.amp_instance_id,
                    'steam_appid': game.steam_appid,
                    'steam_link': game.steam_link,
                    'discord_invite': game.discord_invite,
                    'custom_img': game.custom_img,
                    'link_label': game.link_label,
                    'activity_keywords': json.loads(game.activity_keywords),
                    'is_active': game.is_active,
                    'sort_order': game.sort_order,
                    'roles': [{
                        'id': r.id,
                        'guild_id': str(r.guild_id),  # Return as string to preserve full Discord ID in JavaScript
                        'role_id': str(r.role_id),    # Return as string to preserve full Discord ID in JavaScript
                        'guild_name': r.guild_name,
                        'role_name': r.role_name,
                        'is_active': r.is_active,
                    } for r in roles]
                })
            else:
                # List all games
                games = db.query(SiteActivityGame).order_by(SiteActivityGame.sort_order).all()
                return JsonResponse({
                    'games': [{
                        'id': g.id,
                        'game_key': g.game_key,
                        'display_name': g.display_name,
                        'description': g.description,
                        'game_type': g.game_type,
                        'amp_instance_id': g.amp_instance_id,
                        'steam_appid': g.steam_appid,
                        'steam_link': g.steam_link,
                        'discord_invite': g.discord_invite,
                        'custom_img': g.custom_img,
                        'link_label': g.link_label,
                        'activity_keywords': json.loads(g.activity_keywords),
                        'is_active': g.is_active,
                        'sort_order': g.sort_order,
                    } for g in games]
                })

        elif request.method == "POST":
            # Create new game
            data = json.loads(request.body)

            # Validation
            if not data.get('game_key') or not data.get('display_name'):
                return JsonResponse({'error': 'game_key and display_name are required'}, status=400)

            # Check for duplicate game_key
            existing = db.query(SiteActivityGame).filter_by(game_key=data['game_key']).first()
            if existing:
                return JsonResponse({'error': 'Game with this key already exists'}, status=400)

            game = SiteActivityGame(
                game_key=data['game_key'],
                display_name=data['display_name'],
                description=data.get('description', ''),
                game_type=data.get('game_type', 'discord'),
                amp_instance_id=data.get('amp_instance_id'),
                steam_appid=data.get('steam_appid'),
                steam_link=data.get('steam_link'),
                discord_invite=data.get('discord_invite'),
                custom_img=data.get('custom_img'),
                link_label=data.get('link_label', 'View Site'),
                activity_keywords=json.dumps(data.get('activity_keywords', [])),
                is_active=data.get('is_active', True),
                sort_order=data.get('sort_order', 0),
                created_at=int(time.time()),
                updated_at=int(time.time())
            )

            db.add(game)
            db.commit()
            db.refresh(game)

            # Handle role mappings if provided
            role_mappings = data.get('role_mappings', [])
            if role_mappings:
                from .models import SiteActivityGuildRole
                for mapping in role_mappings:
                    if mapping.get('guild_id') and mapping.get('role_id'):
                        role = SiteActivityGuildRole(
                            game_id=game.id,
                            guild_id=int(str(mapping['guild_id'])),  # Convert to string first to avoid JS number precision loss
                            role_id=int(str(mapping['role_id'])),    # Convert to string first to avoid JS number precision loss
                            guild_name=mapping.get('guild_name'),
                            role_name=mapping.get('role_name'),
                            created_at=int(time.time())
                        )
                        db.add(role)
                db.commit()

            return JsonResponse({
                'success': True,
                'game': {
                    'id': game.id,
                    'game_key': game.game_key,
                    'display_name': game.display_name
                }
            })

        elif request.method == "PUT":
            # Update existing game
            if not game_id:
                return JsonResponse({'error': 'game_id required for update'}, status=400)

            game = db.query(SiteActivityGame).filter_by(id=game_id).first()
            if not game:
                return JsonResponse({'error': 'Game not found'}, status=404)

            data = json.loads(request.body)

            # Update fields
            if 'display_name' in data:
                game.display_name = data['display_name']
            if 'description' in data:
                game.description = data['description']
            if 'game_type' in data:
                game.game_type = data['game_type']
            if 'amp_instance_id' in data:
                game.amp_instance_id = data['amp_instance_id']
            if 'steam_appid' in data:
                game.steam_appid = data['steam_appid']
            if 'steam_link' in data:
                game.steam_link = data['steam_link']
            if 'discord_invite' in data:
                game.discord_invite = data['discord_invite']
            if 'custom_img' in data:
                game.custom_img = data['custom_img']
            if 'link_label' in data:
                game.link_label = data['link_label']
            if 'activity_keywords' in data:
                game.activity_keywords = json.dumps(data['activity_keywords'])
            if 'is_active' in data:
                game.is_active = data['is_active']
            if 'sort_order' in data:
                game.sort_order = data['sort_order']

            game.updated_at = int(time.time())

            # Handle role mappings if provided
            if 'role_mappings' in data:
                from .models import SiteActivityGuildRole

                # Get existing role mappings
                existing_roles = db.query(SiteActivityGuildRole).filter_by(game_id=game_id).all()
                existing_ids = {role.id for role in existing_roles}

                # Process new role mappings
                new_role_ids = set()
                for mapping in data['role_mappings']:
                    if mapping.get('guild_id') and mapping.get('role_id'):
                        # If mapping has an ID, it's an existing one (update)
                        if mapping.get('id'):
                            role = db.query(SiteActivityGuildRole).filter_by(id=mapping['id']).first()
                            if role:
                                role.guild_id = int(str(mapping['guild_id']))  # Convert to string first to avoid JS number precision loss
                                role.role_id = int(str(mapping['role_id']))    # Convert to string first to avoid JS number precision loss
                                role.guild_name = mapping.get('guild_name')
                                role.role_name = mapping.get('role_name')
                                new_role_ids.add(role.id)
                        else:
                            # New mapping
                            role = SiteActivityGuildRole(
                                game_id=game_id,
                                guild_id=int(str(mapping['guild_id'])),  # Convert to string first to avoid JS number precision loss
                                role_id=int(str(mapping['role_id'])),    # Convert to string first to avoid JS number precision loss
                                guild_name=mapping.get('guild_name'),
                                role_name=mapping.get('role_name'),
                                created_at=int(time.time())
                            )
                            db.add(role)
                            db.flush()
                            new_role_ids.add(role.id)

                # Delete role mappings that were removed
                roles_to_delete = existing_ids - new_role_ids
                if roles_to_delete:
                    db.query(SiteActivityGuildRole).filter(
                        SiteActivityGuildRole.id.in_(roles_to_delete)
                    ).delete(synchronize_session=False)

            db.commit()

            return JsonResponse({'success': True})

        elif request.method == "DELETE":
            # Delete game
            if not game_id:
                return JsonResponse({'error': 'game_id required for delete'}, status=400)

            game = db.query(SiteActivityGame).filter_by(id=game_id).first()
            if not game:
                return JsonResponse({'error': 'Game not found'}, status=404)

            db.delete(game)
            db.commit()

            return JsonResponse({'success': True})


@bot_owner_required
@require_http_methods(["POST", "DELETE"])
def api_site_activity_roles(request, guild_id, role_mapping_id=None):
    """
    API endpoint for managing game role mappings.

    BOT OWNER ONLY

    POST: Add guild/role mapping to a game
    DELETE: Remove role mapping
    """
    from .db import get_db_session
    from .models import SiteActivityGame, SiteActivityGuildRole

    with get_db_session() as db:
        if request.method == "POST":
            # Add role mapping
            data = json.loads(request.body)

            # Validation
            required_fields = ['game_id', 'guild_id', 'role_id']
            if not all(data.get(f) for f in required_fields):
                return JsonResponse({'error': 'game_id, guild_id, and role_id are required'}, status=400)

            # Check game exists
            game = db.query(SiteActivityGame).filter_by(id=data['game_id']).first()
            if not game:
                return JsonResponse({'error': 'Game not found'}, status=404)

            # Check for duplicate
            existing = db.query(SiteActivityGuildRole).filter_by(
                game_id=data['game_id'],
                guild_id=data['guild_id'],
                role_id=data['role_id']
            ).first()
            if existing:
                return JsonResponse({'error': 'This role mapping already exists'}, status=400)

            role_mapping = SiteActivityGuildRole(
                game_id=data['game_id'],
                guild_id=data['guild_id'],
                role_id=data['role_id'],
                guild_name=data.get('guild_name', ''),
                role_name=data.get('role_name', ''),
                is_active=data.get('is_active', True),
                created_at=int(time.time()),
                updated_at=int(time.time())
            )

            db.add(role_mapping)
            db.commit()
            db.refresh(role_mapping)

            return JsonResponse({
                'success': True,
                'role_mapping': {
                    'id': role_mapping.id,
                    'game_id': role_mapping.game_id,
                    'guild_id': role_mapping.guild_id,
                    'role_id': role_mapping.role_id
                }
            })

        elif request.method == "DELETE":
            # Delete role mapping
            if not role_mapping_id:
                return JsonResponse({'error': 'role_mapping_id required'}, status=400)

            role_mapping = db.query(SiteActivityGuildRole).filter_by(id=role_mapping_id).first()
            if not role_mapping:
                return JsonResponse({'error': 'Role mapping not found'}, status=404)

            db.delete(role_mapping)
            db.commit()

            return JsonResponse({'success': True})


# Leave this here
def home(request):
    return render(request, 'index.html')


def dune_page(request):
    dune_data = {
        "total": "-",
        "online": "-",
        "active": "-"
    }

    if DISCORD_ACTIVITY_FILE.exists():
        try:
            with DISCORD_ACTIVITY_FILE.open("r") as f:
                all_activity = json.load(f)
                raw = all_activity.get("Dune", {})
                logger.debug(f"Raw Dune Activity: {raw}")

                # Safely cast integers
                dune_data["total"] = int(raw["total"]) if str(raw.get("total", "")).isdigit() else "-"
                dune_data["online"] = int(raw["online"]) if str(raw.get("online", "")).isdigit() else "-"
                dune_data["active"] = int(raw["active"]) if str(raw.get("active", "")).isdigit() else "-"
        except Exception as e:
            logger.error(f"Dune page failed to load activity data: {e}")

    logger.debug(f"Final dune_data: {dune_data}")
    return render(request, "dune.html", {
        "dune_activity": dune_data
    })


def pantheon_page(request):
    pantheon_data = {
        "total": "-",
        "online": "-",
        "active": "-"
    }

    if DISCORD_ACTIVITY_FILE.exists():
        try:
            with DISCORD_ACTIVITY_FILE.open("r") as f:
                all_activity = json.load(f)
                raw = all_activity.get("Pantheon", {})
                logger.debug(f"Raw Pantheon Activity: {raw}")

                # Safely cast integers
                pantheon_data["total"] = int(raw["total"]) if str(raw.get("total", "")).isdigit() else "-"
                pantheon_data["online"] = int(raw["online"]) if str(raw.get("online", "")).isdigit() else "-"
                pantheon_data["active"] = int(raw["active"]) if str(raw.get("active", "")).isdigit() else "-"
        except Exception as e:
            logger.error(f"Pantheon page failed to load activity data: {e}")

    logger.debug(f"Final pantheon_data: {pantheon_data}")
    return render(request, "pantheon.html", {
        "pantheon_activity": pantheon_data
    })

def wow_page(request):
    wow_data = {
        "total": "-",
        "online": "-",
        "active": "-"
    }

    if DISCORD_ACTIVITY_FILE.exists():
        try:
            with DISCORD_ACTIVITY_FILE.open("r") as f:
                all_activity = json.load(f)
                raw = all_activity.get("WoW", {})
                logger.debug("[WoW PAGE] Raw WoW Activity: %s", raw)

                # Safely cast integers
                wow_data["total"] = int(raw["total"]) if str(raw.get("total", "")).isdigit() else "-"
                wow_data["online"] = int(raw["online"]) if str(raw.get("online", "")).isdigit() else "-"
                wow_data["active"] = int(raw["active"]) if str(raw.get("active", "")).isdigit() else "-"
        except Exception as e:
            logger.error("[WoW PAGE] Failed to load activity data", exc_info=True)

    logger.debug("[WoW PAGE] Final wow_data: %s", wow_data)
    return render(request, "wow.html", {
        "wow_activity": wow_data
    })



def eso_page(request):
    eso_data = {
        "total": "-",
        "online": "-",
        "active": "-"
    }

    if DISCORD_ACTIVITY_FILE.exists():
        try:
            with DISCORD_ACTIVITY_FILE.open("r") as f:
                all_activity = json.load(f)
                raw = all_activity.get("ESO", {})
                logger.debug("[ESO PAGE] Raw ESO Activity: %s", raw)

                # Safely cast integers
                eso_data["total"] = int(raw["total"]) if str(raw.get("total", "")).isdigit() else "-"
                eso_data["online"] = int(raw["online"]) if str(raw.get("online", "")).isdigit() else "-"
                eso_data["active"] = int(raw["active"]) if str(raw.get("active", "")).isdigit() else "-"
        except Exception as e:
            logger.error("[ESO PAGE] Failed to load activity data", exc_info=True)

    logger.debug("[ESO PAGE] Final eso_data: %s", eso_data)
    return render(request, "eso.html", {
        "eso_activity": eso_data
    })




def get_discord_activity_counts():
    if not DISCORD_ACTIVITY_FILE.exists():
        return {}

    with DISCORD_ACTIVITY_FILE.open("r") as f:
        return json.load(f)

articles = [
    {
        "slug": "survival-games-2025",
        "title": "Top Survival Games in 2025",
        "author": "FullData",
        "games": [
            {
                "title": "Dune: Awakening",
                "summary": """Set on the unforgiving planet of Arrakis, Dune: Awakening blends survival mechanics with MMO elements. 
                Players must navigate sandstorms, harvest spice, and avoid colossal sandworms. The game emphasizes base-building, resource management, and PvP combat.
                While the world-building and atmosphere have been praised, some players have noted that combat mechanics feel clunky and could use refinement.
                The game's success will likely hinge on how well it balances its ambitious features.""",
                "image": "img/games/survivalgames/Dune1.jpg"
            },
            {
                "title": "Subnautica 2",
                "summary": """Diving back into the depths, Subnautica 2 offers a new alien ocean world to explore. 
                The sequel introduces co-op gameplay, allowing up to four players to explore together. Players can expect new biomes, creatures, and crafting options. 
                Early impressions highlight the game's immersive environment and improved mechanics. 
                However, some fans express concerns about the game's shorter story length, aiming for around 15 hours, and the introduction of microtransactions.""",
                "image": "img/games/survivalgames/sub2.jpg"
            },
            {
                "title": "The Alters",
                "summary": """The Alters presents a unique survival experience where players create alternate versions of themselves to survive on a hostile planet. 
                Each "alter" possesses different skills, aiding in tasks like base-building and exploration. 
                The game's narrative-driven approach has been lauded for its depth and originality. 
                However, some players feel that the gameplay leans heavily on dialogue and could benefit from more interactive elements.""",
                "image": "img/games/survivalgames/Alters.jpg"
            },
            {
                "title": "RuneScape: Dragonwilds",
                "summary": """A spin-off from the classic MMO, RuneScape: Dragonwilds ventures into survival territory. 
                Set in the continent of Ashenfall, players engage in base-building, crafting, and combat against dragons. 
                The game has seen a strong start, with over 600,000 copies sold and positive reviews highlighting its engaging mechanics. 
                Nonetheless, some players feel that the game lacks depth in its current state and hope for more content in future updates.""",
                "image": "img/games/survivalgames/dragonwilds_static.jpg"
            },
            {
                "title": "V Rising: Invaders of Oakveil",
                "summary": """V Rising: Invaders of Oakveil is out now and it’s a massive step forward for the game. 
                It builds smartly on what V Rising already did well — from world design to progression — and adds meaningful features like the cursed forest biome, PvP duel arenas, and deeper character customization.
                What really stands out in this update is how it pushes both PvE and PvP players forward. The cursed forest introduces new tactical layers with poison-based enemies and new gear, while the duel arenas finally give PvP-focused players a structured way to test their builds.
                If you were already a fan of V Rising, this update makes the game feel more complete. And if you're new? There’s never been a better time to jump in.""",
                "image": "img/games/survivalgames/V-Rising-Invaders-of-Oakveil.jpg"
            },
            {
                "title": "The Forever Winter",
                "summary": """Set in a post-apocalyptic world, The Forever Winter combines survival horror with extraction shooter mechanics. 
                Players scavenge resources while avoiding massive war machines. The game's unique art style, inspired by anime and dystopian themes, has garnered attention. 
                While the dynamic encounter system keeps gameplay fresh, some players have criticized certain mechanics, like the water system, though recent updates have addressed these concerns.""",
                "image": "img/games/survivalgames/TheForeverWinter_SteamImage.jpg"
            },
            {
                "title": "Oppidum",
                "summary": """Aimed at a broader audience, Oppidum offers a more accessible survival experience. 
                With its colorful visuals and simplified mechanics, it's reminiscent of titles like The Legend of Zelda. Players can engage in crafting, farming, and exploration. 
                While the game has been praised for its charm and cooperative gameplay, some have pointed out issues like limited inventory space and a cumbersome travel system.""",
                "image": "img/games/survivalgames/Oppidum.png"
            },
            {
                "title": "Autonomica",
                "summary": """Autonomica stands out with its solarpunk aesthetic and ambitious blend of farming, automation, and time-travel elements. 
                Players can build automated farms, engage in mech battles, and even pursue romantic relationships with NPCs. 
                The game's Kickstarter success indicates strong interest, but some are cautious about how all these features will integrate seamlessly.""",
                "image": "img/games/survivalgames/autonomica.jpg"
            },
            {
                "title": "Terminator: Survivors",
                "summary": """Set between Judgment Day and the rise of John Connor's resistance, Terminator: Survivors is an open-world survival game where players scavenge resources while evading Skynet's machines. 
                The game emphasizes co-op gameplay and base-building. 
                While the premise is intriguing, some players are concerned about the game's delayed release and hope that it delivers a polished experience upon launch.""",
                "image": "img/games/survivalgames/Terminiator.webp"
            },
            {
                "title": "Outward 2",
                "summary": """Building upon its predecessor, Outward 2 offers a challenging action RPG experience with survival elements. 
                Players can expect improved combat mechanics, a richer world, and the ability to drop backpacks to manage weight. 
                While early impressions are positive, some have noted issues like unpredictable enemy movements and occasional bugs. 
                The developers' shift from Unity to Unreal Engine 5 suggests a commitment to enhancing the game's quality.""",
                "image": "img/games/survivalgames/outward2.jpg"
            }
        ]
    }
]

def features(request):
    articles = [
        {
            "slug": "survival-games-2025",
            "title": "Top Survival Games in 2025",
            "summary": "What's next after V Rising and Enshrouded?",
            "image_url": "img/games/survivalgames/survival2025.png"
        },
    ]
    return render(request, "features/features.html", {"articles": articles})


def features_detail(request, slug):
    article = next((a for a in articles if a["slug"] == slug), None)
    if not article:
        return render(request, "404.html", status=404)
    return render(request, "features/article_details.html", {"article": article})


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            messages.error(request, "Invalid username or password.")
    return render(request, "auth/login.html")

from django.contrib.auth.decorators import login_required

@login_required
def dashboard(request):
    return render(request, "dashboard.html")

def hosting(request):
    return render(request, 'hosting.html')

def sevendtd(request):
    sevendtd_instance = asyncio.run(fetch_instance_data("CasualHeroes-7DTD01"))

    return render(request, '7dtd.html', {
        "sevendtd_instance": sevendtd_instance
    })


def dragonwilds(request):
    return render(request, 'dragonwilds.html')

def gameshype(request):
    return render(request, 'gameshype.html')

def gamesuggest(request):
    return render(request, 'gamesuggest.html')

def enshrouded(request):
    # enshrouded_instance = asyncio.run(fetch_instance_data("Enshrouded01"))

    return render(request, 'enshrouded.html', {
        # "enshrouded_instance": enshrouded_instance
    })

def vrising(request):
    # vrising_instance = asyncio.run(fetch_instance_data("CasualHeroes-Vrising01"))

    return render(request, 'vrising.html', {
        # "vrising_instance": vrising_instance
    })

def palworld(request):
    palworld_instance = asyncio.run(fetch_instance_data("CH-Palworld01"))

    return render(request, 'palworld.html', {
        "palworld_instance": palworld_instance
    })

def icarus(request):
    icarus_instance = asyncio.run(fetch_instance_data("CH-Icarus01"))

    return render(request, 'icarus.html', {
        "icarus_instance": icarus_instance
    })
def guides(request):
    return render(request, 'guides.html')

def content(request):
    return render(request, 'content.html')

def aboutus(request):
    return render(request, 'aboutus.html')

def privacy(request):
    return render(request, 'privacy.html')

def terms(request):
    return render(request, 'terms.html')

def contactus(request):
    return render(request, 'contactus.html')

def faq(request):
    return render(request, 'faq.html')

def questlog_overview(request):
    return render(request, 'questlog_overview.html')

def questlog_login(request):
    """Redirect to QuestLog dashboard with Discord OAuth"""
    import secrets
    from .discord_auth import get_discord_login_url

    # Check if already logged in
    discord_user = request.session.get('discord_user')
    if discord_user:
        # Validate token is still valid before redirecting to dashboard
        access_token = discord_user.get('access_token')
        if access_token:
            try:
                validate_response = requests.get(
                    'https://discord.com/api/users/@me',
                    headers={'Authorization': f'Bearer {access_token}'},
                    timeout=5
                )

                # If token is valid, redirect to dashboard
                if validate_response.status_code == 200:
                    return redirect('https://dashboard.casual-heroes.com/questlog/')

                # If token expired (401), clear session and continue to re-auth below
                if validate_response.status_code == 401:
                    logger.warning(f"Discord token expired during login check, clearing session")
                    request.session.flush()
            except requests.RequestException:
                # Network error - fail open and redirect to dashboard
                # Dashboard will handle auth validation
                return redirect('https://dashboard.casual-heroes.com/questlog/')

    # Generate a state token for CSRF protection
    state = secrets.token_urlsafe(32)
    request.session['discord_oauth_state'] = state

    # Set the next URL to the dashboard subdomain
    request.session['discord_login_next'] = 'https://dashboard.casual-heroes.com/questlog/'

    # Force session save before OAuth redirect
    request.session.modified = True
    request.session.save()

    # Use the existing Discord OAuth login URL generator
    login_url = get_discord_login_url(state=state)
    return redirect(login_url)

def creator_of_the_month_page(request):
    """Public page showing all Creator of the Month winners across all guilds."""
    from .db import get_db_session
    from .models import CreatorOfTheMonth, Guild
    import datetime

    with get_db_session() as db:
        # Get all COTM records, newest first
        cotm_records = db.query(CreatorOfTheMonth).order_by(
            CreatorOfTheMonth.year.desc(),
            CreatorOfTheMonth.month.desc()
        ).all()

        # Enrich with guild info
        for record in cotm_records:
            guild = db.query(Guild).filter_by(guild_id=record.guild_id).first()
            record.guild_name = guild.name if guild else f"Guild {record.guild_id}"
            record.guild_icon = guild.icon_url if guild else None

            # Format month name
            record.month_name = datetime.datetime(record.year, record.month, 1).strftime("%B")

        return render(request, 'questlog/creator_of_the_month.html', {
            'cotm_records': cotm_records,
            'page_title': 'Creator of the Month - Hall of Fame'
        })

def creator_of_the_week_page(request):
    """Public page showing all Creator of the Week winners across all guilds."""
    from .db import get_db_session
    from .models import CreatorOfTheWeek, Guild

    with get_db_session() as db:
        # Get all COTW records, newest first
        cotw_records = db.query(CreatorOfTheWeek).order_by(
            CreatorOfTheWeek.year.desc(),
            CreatorOfTheWeek.week.desc()
        ).all()

        # Enrich with guild info
        for record in cotw_records:
            guild = db.query(Guild).filter_by(guild_id=record.guild_id).first()
            record.guild_name = guild.name if guild else f"Guild {record.guild_id}"
            record.guild_icon = guild.icon_url if guild else None

        return render(request, 'questlog/creator_of_the_week.html', {
            'cotw_records': cotw_records,
            'page_title': 'Creator of the Week - Hall of Fame'
        })

@staff_member_required
def analytics_view(request):
    return render(request, 'admin/analytics.html')


# Discord OAuth2 Authentication

from .discord_auth import (
    get_discord_login_url,
    exchange_code_for_token,
    get_discord_user,
    get_discord_guilds,
    get_discord_avatar_url,
    revoke_token,
)
import secrets
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.models import User


def discord_login(request):
    """Initiate Discord OAuth2 login flow"""
    # Generate a state token for CSRF protection
    state = secrets.token_urlsafe(32)
    request.session['discord_oauth_state'] = state

    # Store the 'next' URL if provided (SEC-008 fix: validate redirect URL)
    next_url = request.GET.get('next', '/dashboard/')
    if not is_safe_redirect(next_url):
        logger.warning(f"[SECURITY] SEC-008: Blocked unsafe redirect URL from OAuth login: {next_url}")
        next_url = '/dashboard/'
    request.session['discord_login_next'] = next_url

    login_url = get_discord_login_url(state=state)
    return redirect(login_url)


@ratelimit(key=get_client_ip, rate='10/m', method='GET')
def discord_callback(request):
    """Handle Discord OAuth2 callback"""
    error = request.GET.get('error')
    if error:
        messages.error(request, f"Discord login failed: {error}")
        return redirect('home')

    code = request.GET.get('code')
    state = request.GET.get('state')

    # Verify state token
    stored_state = request.session.get('discord_oauth_state')
    if not state or state != stored_state:
        messages.error(request, "Invalid state token. Please try again.")
        return redirect('home')

    # Clear the state from session
    del request.session['discord_oauth_state']

    try:
        # Exchange code for token
        token_data = exchange_code_for_token(code)
        access_token = token_data['access_token']
        refresh_token = token_data.get('refresh_token')

        # Get Discord user info
        discord_user = get_discord_user(access_token)

        # Get user's guilds
        discord_guilds = get_discord_guilds(access_token)

        # Store Discord info in session
        request.session['discord_user'] = {
            'id': discord_user['id'],
            'username': discord_user['username'],
            'global_name': discord_user.get('global_name') or discord_user['username'],
            'email': discord_user.get('email'),
            'avatar': discord_user.get('avatar'),
            'avatar_url': get_discord_avatar_url(discord_user['id'], discord_user.get('avatar')),
            'discriminator': discord_user.get('discriminator', '0'),
            'access_token': access_token,
            'refresh_token': refresh_token,
        }

        # Store all guilds first
        all_guilds = [
            {'id': g['id'], 'name': g['name'], 'icon': g.get('icon')}
            for g in discord_guilds
        ]
        request.session['discord_all_guilds'] = all_guilds

        # Store guilds where user has Discord admin permissions
        # 0x8 = Administrator, 0x20 = Manage Server (for trusted mods)
        admin_guilds = [
            {
                'id': g['id'],
                'name': g['name'],
                'icon': g.get('icon'),
                'owner': g.get('owner', False),
                'permissions': g.get('permissions', 0),
            }
            for g in discord_guilds
            if g.get('owner') or (int(g.get('permissions', 0)) & 0x8) or (int(g.get('permissions', 0)) & 0x20)
        ]

        # HYBRID ADMIN: Also check custom admin roles and add those guilds to admin list
        user_id = discord_user.get('id')
        if user_id:
            custom_admin_guilds = get_guilds_with_custom_admin_access(user_id, all_guilds)
            # Add custom admin guilds that aren't already in admin_guilds
            admin_guild_ids = {g['id'] for g in admin_guilds}
            for custom_guild in custom_admin_guilds:
                if custom_guild['id'] not in admin_guild_ids:
                    admin_guilds.append(custom_guild)
                    logger.info(f"[HYBRID_ADMIN] User {user_id} granted admin access to guild {custom_guild['id']} via custom roles")

        request.session['discord_admin_guilds'] = admin_guilds

        # Optionally link to Django user (create if doesn't exist)
        try:
            user, created = User.objects.get_or_create(
                username=f"discord_{discord_user['id']}",
                defaults={
                    'email': discord_user.get('email', ''),
                    'first_name': discord_user.get('global_name', discord_user['username'])[:30],
                }
            )
            if not created and discord_user.get('email'):
                user.email = discord_user['email']
                user.save()

            # Log in the Django user
            from django.contrib.auth import login as auth_login
            # Specify backend to avoid "multiple backends" error
            auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        except Exception as e:
            # Even if Django user creation fails, session-based auth works
            logger.warning(f"Could not create Django user: {e}")

        messages.success(request, f"Welcome, {discord_user.get('global_name', discord_user['username'])}!")

        # Explicitly save session before redirect to ensure data persists
        # This is critical in multi-worker environments like Gunicorn
        request.session.modified = True
        request.session.save()

        # Redirect to stored 'next' URL or dashboard (SEC-008 fix: validate again)
        next_url = request.session.pop('discord_login_next', '/dashboard/')
        if not is_safe_redirect(next_url):
            logger.warning(f"[SECURITY] SEC-008: Blocked unsafe redirect URL from OAuth callback: {next_url}")
            next_url = '/dashboard/'

        # Save again after popping discord_login_next
        request.session.modified = True
        request.session.save()

        return redirect(next_url)

    except Exception as e:
        logger.error(f"Discord OAuth callback failed: {e}")
        messages.error(request, "Failed to authenticate with Discord. Please try again.")
        return redirect('home')


def discord_logout(request):
    """Log out user and revoke Discord token"""
    discord_user = request.session.get('discord_user')

    if discord_user and discord_user.get('access_token'):
        # Attempt to revoke the Discord token
        revoke_token(discord_user['access_token'])

    # Clear Discord session data
    for key in ['discord_user', 'discord_admin_guilds', 'discord_all_guilds']:
        if key in request.session:
            del request.session[key]

    # Log out Django user if logged in
    auth_logout(request)

    messages.info(request, "You have been logged out.")
    return redirect('home')


@require_http_methods(["POST"])
def discord_refresh_guilds(request):
    """Refresh the user's guild list from Discord without logging out"""
    discord_user = request.session.get('discord_user')

    if not discord_user:
        return JsonResponse({'error': 'Not authenticated with Discord'}, status=401)

    access_token = discord_user.get('access_token')
    if not access_token:
        return JsonResponse({'error': 'No access token found'}, status=401)

    try:
        # Fetch fresh guild list from Discord
        guilds_response = requests.get(
            'https://discord.com/api/users/@me/guilds',
            headers={'Authorization': f'Bearer {access_token}'}
        )

        if guilds_response.status_code != 200:
            logger.error(f"Failed to fetch guilds: {guilds_response.status_code} {guilds_response.text}")

            # If token is expired (401), clear session and trigger re-auth
            if guilds_response.status_code == 401:
                request.session.flush()  # Clear the expired session
                return JsonResponse({
                    'error': 'Session expired',
                    'requires_reauth': True,
                    'redirect_url': '/auth/discord/login/'
                }, status=401)

            return JsonResponse({'error': 'Failed to fetch guilds from Discord'}, status=guilds_response.status_code)

        discord_guilds = guilds_response.json()

        # Store all guilds first
        all_guilds = [
            {'id': g['id'], 'name': g['name'], 'icon': g.get('icon')}
            for g in discord_guilds
        ]
        request.session['discord_all_guilds'] = all_guilds

        # Update admin guilds (same logic as in callback)
        admin_guilds = [
            {
                'id': g['id'],
                'name': g['name'],
                'icon': g.get('icon'),
                'owner': g.get('owner', False),
                'permissions': g.get('permissions', 0),
            }
            for g in discord_guilds
            if g.get('owner') or (int(g.get('permissions', 0)) & 0x8) or (int(g.get('permissions', 0)) & 0x20)
        ]

        # HYBRID ADMIN: Also check custom admin roles and add those guilds to admin list
        user_id = request.session.get('discord_user', {}).get('id')
        if user_id:
            custom_admin_guilds = get_guilds_with_custom_admin_access(user_id, all_guilds)
            # Add custom admin guilds that aren't already in admin_guilds
            admin_guild_ids = {g['id'] for g in admin_guilds}
            for custom_guild in custom_admin_guilds:
                if custom_guild['id'] not in admin_guild_ids:
                    admin_guilds.append(custom_guild)
                    logger.info(f"[HYBRID_ADMIN] User {user_id} granted admin access to guild {custom_guild['id']} via custom roles (refresh)")

        # Update session
        request.session['discord_admin_guilds'] = admin_guilds

        request.session.modified = True
        request.session.save()

        logger.info(f"Refreshed guilds for user {discord_user.get('username')}: found {len(admin_guilds)} admin guilds")

        return JsonResponse({
            'success': True,
            'admin_guilds_count': len(admin_guilds),
            'total_guilds_count': len(discord_guilds)
        })

    except Exception as e:
        logger.error(f"Error refreshing guilds: {e}", exc_info=True)
        return JsonResponse({'error': 'An error occurred while refreshing guilds'}, status=500)


def discord_required(view_func):
    """
    Decorator to require Discord authentication - NO CRAWLER BYPASS.

    Use this for:
    - All state-changing views (POST/PUT/DELETE)
    - Admin pages
    - Any sensitive endpoint

    For public pages that need Open Graph support, use @discord_required_read_only instead.

    Security: SEC-001 fix - removed crawler bypass to prevent authentication bypass attacks.
    """
    def wrapper(request, *args, **kwargs):
        discord_user = request.session.get('discord_user')

        if not discord_user:
            # For API endpoints, return JSON error
            if request.path.startswith('/api/'):
                from django.http import JsonResponse
                return JsonResponse({'error': 'Not authenticated'}, status=401)
            # For page views, redirect to login
            messages.warning(request, "Please log in with Discord to access this page.")
            return redirect(f"/auth/discord/login/?next={request.path}")

        # Validate token is still valid by making a lightweight Discord API call
        access_token = discord_user.get('access_token')
        if access_token:
            try:
                # Quick validation call to check if token is still valid
                validate_response = requests.get(
                    'https://discord.com/api/users/@me',
                    headers={'Authorization': f'Bearer {access_token}'},
                    timeout=5
                )

                # If token is expired (401), clear session and redirect to login
                if validate_response.status_code == 401:
                    logger.warning(f"Discord token expired for user {discord_user.get('username')}, forcing re-auth")
                    request.session.flush()

                    # For API endpoints, return JSON error
                    if request.path.startswith('/api/'):
                        from django.http import JsonResponse
                        return JsonResponse({
                            'error': 'Session expired',
                            'requires_reauth': True,
                            'redirect_url': '/auth/discord/login/'
                        }, status=401)

                    # For page views, redirect to login with message
                    messages.warning(request, "Your Discord session has expired. Please log in again.")
                    return redirect(f"/auth/discord/login/?next={request.path}")
            except requests.RequestException:
                # If validation fails due to network error, allow through (fail open for availability)
                # The actual API calls will fail and handle appropriately
                pass

        return view_func(request, *args, **kwargs)
    return wrapper


def discord_required_read_only(view_func):
    """
    Decorator for read-only pages that need Open Graph support for social media previews.

    Allows trusted social media crawlers to access pages for GET requests ONLY.
    All POST/PUT/DELETE requests require full authentication.

    Use this for:
    - Public profile pages
    - Event pages that should show previews
    - Guild pages with public leaderboards

    Security: SEC-001 fix - restricts crawler bypass to GET requests only with specific crawlers.
    """
    def wrapper(request, *args, **kwargs):
        # Only allow crawler bypass for GET requests
        if request.method == 'GET':
            user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
            # Specific trusted crawlers only - NO generic 'bot' pattern
            is_trusted_crawler = any(crawler in user_agent for crawler in [
                'facebookexternalhit',  # Facebook
                'twitterbot',            # Twitter/X
                'linkedinbot',           # LinkedIn
                'discordbot',            # Discord
                'slackbot',              # Slack
                'telegrambot',           # Telegram
                'whatsapp',              # WhatsApp
                'pinterest',             # Pinterest
            ])

            if is_trusted_crawler:
                # Log crawler access for monitoring
                logger.info(f"Crawler access: {user_agent[:100]} -> {request.path}")
                return view_func(request, *args, **kwargs)

        # All non-GET requests and non-crawler GET requests require auth
        if not request.session.get('discord_user'):
            # For API endpoints, return JSON error
            if request.path.startswith('/api/'):
                from django.http import JsonResponse
                return JsonResponse({'error': 'Not authenticated'}, status=401)
            # For page views, redirect to login
            messages.warning(request, "Please log in with Discord to access this page.")
            return redirect(f"/auth/discord/login/?next={request.path}")
        return view_func(request, *args, **kwargs)
    return wrapper


@discord_required
def user_profile(request):
    """User profile page showing Discord account info and connected guilds"""
    import json as json_lib

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])
    all_guilds = request.session.get('discord_all_guilds', [])

    context = {
        'discord_user': discord_user,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'all_guilds': all_guilds,
        'all_guilds_json': json_lib.dumps(all_guilds),  # JSON-encoded for JavaScript
        'guild_count': len(all_guilds),
        'admin_guild_count': len(admin_guilds),
    }
    return render(request, 'auth/profile.html', context)


@discord_required
def questlog_dashboard(request):
    """Main Warden bot dashboard - select a guild to manage"""
    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Add 'has_bot' flag to admin guilds
    admin_guilds = add_has_bot_flag(admin_guilds)

    # Get member guilds (already filtered to only show guilds with bot)
    member_guilds = get_member_guilds(request)

    context = {
        'discord_user': discord_user,
        'admin_guilds': admin_guilds,
        'member_guilds': member_guilds,
        'total_guilds': len(admin_guilds) + len(member_guilds),
        'bot_client_id': os.getenv('DISCORD_CLIENT_ID', ''),
    }
    return render(request, 'questlog/dashboard.html', context)


@discord_required
def guild_dashboard(request, guild_id):
    """Dashboard for a specific guild - shows admin config or member portal based on permissions"""
    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])
    admin_guilds = add_has_bot_flag(admin_guilds)
    all_guilds = request.session.get('discord_all_guilds', [])

    # Check if user is admin OR member of this guild
    is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)
    guild = get_guild_with_permissions(guild_id, admin_guilds, all_guilds)

    if not guild:
        messages.error(request, "You are not a member of this server.")
        return redirect('questlog_dashboard')

    # If not admin, show member landing page instead
    if not is_admin:
        return render(request, 'questlog/member_portal.html', {
            'discord_user': discord_user,
            'guild': guild,
            'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
            'member_guilds': get_member_guilds(request),
            'is_admin': False,
        })

    # Fetch feature statuses
    feature_status = {
        'welcome_enabled': False,
        'discovery_enabled': False,
        'creator_discovery_enabled': False,
        'game_discovery_enabled': False,
        'discovery_feature_interval': 3,
        'verification_enabled': False,
    }

    # Initialize metrics
    metrics = {
        'total_members': 0,
        'active_today': 0,
        'engagement_rate': 0,
        'warnings_this_week': 0,
    }

    guild_record = None

    try:
        from .db import get_db_session
        from .models import (
            WelcomeConfig, DiscoveryConfig, VerificationConfig, VerificationType,
            GuildMember, Warning, LevelRole, ReactRole, ChannelStatTracker
        )
        from datetime import datetime, timedelta

        logger.info(f"[METRICS DEBUG] Loading dashboard for guild {guild_id}")

        with get_db_session() as db:
            # Check welcome messages
            welcome_config = db.query(WelcomeConfig).filter_by(guild_id=int(guild_id)).first()
            if welcome_config:
                feature_status['welcome_enabled'] = welcome_config.enabled
                logger.info(f"[METRICS DEBUG] Welcome: {welcome_config.enabled}")

            # Check discovery
            discovery_config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if discovery_config:
                feature_status['discovery_enabled'] = discovery_config.enabled or discovery_config.game_discovery_enabled
                feature_status['creator_discovery_enabled'] = discovery_config.enabled
                feature_status['game_discovery_enabled'] = discovery_config.game_discovery_enabled
                feature_status['discovery_feature_interval'] = discovery_config.feature_interval_hours
                logger.info(f"[METRICS DEBUG] Discovery - enabled: {discovery_config.enabled}, game_discovery: {discovery_config.game_discovery_enabled}, combined: {feature_status['discovery_enabled']}")
            else:
                logger.info(f"[METRICS DEBUG] No DiscoveryConfig found for guild {guild_id}")

            # Check verification
            verification_config = db.query(VerificationConfig).filter_by(guild_id=int(guild_id)).first()
            if verification_config:
                feature_status['verification_enabled'] = verification_config.verification_type != VerificationType.NONE
                logger.info(f"[METRICS DEBUG] Verification: {feature_status['verification_enabled']}")

            # Calculate metrics from Discord data (cached by bot)
            # Get guild record to access cached Discord stats
            from .models import Guild as GuildModel
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()

            # Total Members (from Discord, cached by bot)
            metrics['total_members'] = guild_record.member_count if guild_record and guild_record.member_count else 0
            logger.info(f"[METRICS DEBUG] Total members (from Discord cache): {metrics['total_members']}")

            # Active Today = Currently Online (from Discord presence, cached by bot)
            metrics['active_today'] = guild_record.online_count if guild_record and guild_record.online_count else 0
            logger.info(f"[METRICS DEBUG] Currently online: {metrics['active_today']}")

            # Engagement Rate - Based on REAL activity (messages, voice, reactions, media)
            # Count members who were active in last 7 days (check timestamp columns)
            week_start = int(time.time()) - (7 * 86400)  # 7 days ago
            engaged_members = db.query(GuildMember).filter(
                GuildMember.guild_id == int(guild_id),
                GuildMember.is_bot == False,
                or_(
                    GuildMember.last_message_ts >= week_start,
                    GuildMember.last_media_ts >= week_start,
                    GuildMember.last_voice_join_ts >= week_start,
                    GuildMember.last_react_ts >= week_start,
                    GuildMember.last_command_ts >= week_start,
                    GuildMember.last_gaming_ts >= week_start
                )
            ).count()

            if metrics['total_members'] > 0:
                metrics['engagement_rate'] = round((engaged_members / metrics['total_members']) * 100, 1)
            else:
                metrics['engagement_rate'] = 0
            logger.info(f"[METRICS DEBUG] Engagement rate (7-day activity): {metrics['engagement_rate']}% ({engaged_members}/{metrics['total_members']})")

            # Get active members from DB for XP leaderboards (last 24h)
            today_start = int(time.time()) - 86400  # 24 hours ago
            active_members = db.query(GuildMember).filter(
                GuildMember.guild_id == int(guild_id),
                GuildMember.is_bot == False,
                GuildMember.last_active >= today_start
            ).all()

            # Exclude current viewer from leaderboards
            current_user_id = int(discord_user.get('id', 0)) if discord_user.get('id') else None
            eligible_members = [m for m in active_members if m.user_id != current_user_id] if current_user_id else active_members

            # Top XP Earner Today (among active members, excluding viewer)
            if eligible_members:
                top_xp_member = max(eligible_members, key=lambda m: m.xp)
                metrics['top_xp_earner'] = {
                    'name': top_xp_member.display_name or top_xp_member.username or f"User#{top_xp_member.user_id}",
                    'xp': round(top_xp_member.xp, 1)
                }
            else:
                metrics['top_xp_earner'] = None
            logger.info(f"[METRICS DEBUG] Top XP earner: {metrics['top_xp_earner']}")

            # Top Token Holder Today (among active members, excluding viewer)
            if eligible_members:
                top_token_member = max(eligible_members, key=lambda m: m.hero_tokens)
                metrics['top_token_holder'] = {
                    'name': top_token_member.display_name or top_token_member.username or f"User#{top_token_member.user_id}",
                    'tokens': top_token_member.hero_tokens
                }
            else:
                metrics['top_token_holder'] = None
            logger.info(f"[METRICS DEBUG] Top token holder: {metrics['top_token_holder']}")

            # Warnings This Week
            week_start = int(time.time()) - (7 * 86400)  # 7 days ago
            warnings_count = db.query(Warning).filter(
                Warning.guild_id == int(guild_id),
                Warning.issued_at >= week_start
            ).count()
            metrics['warnings_this_week'] = warnings_count
            logger.info(f"[METRICS DEBUG] Warnings this week: {metrics['warnings_this_week']}")

            # Module stats
            metrics['xp_tracked_members'] = metrics['total_members']  # All members tracked
            metrics['level_roles'] = db.query(LevelRole).filter_by(guild_id=int(guild_id)).count()
            metrics['reaction_role_menus'] = db.query(ReactRole).filter_by(guild_id=int(guild_id)).count()
            metrics['member_trackers'] = db.query(ChannelStatTracker).filter_by(guild_id=int(guild_id)).count()

            logger.info(f"[METRICS DEBUG] Module stats - XP: {metrics['xp_tracked_members']}, LevelRoles: {metrics['level_roles']}, ReactionRoles: {metrics['reaction_role_menus']}, Trackers: {metrics['member_trackers']}")
            logger.info(f"[METRICS DEBUG] Final metrics: {metrics}")

    except Exception as e:
        logger.error(f"Could not fetch feature statuses or metrics for guild {guild_id}: {e}", exc_info=True)

    # Get subscription tier info
    is_vip = guild_record.is_vip if guild_record else False
    subscription_tier = guild_record.subscription_tier if guild_record else 'free'
    billing_cycle = guild_record.billing_cycle if guild_record else None

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'member_guilds': get_member_guilds(request),
        'is_admin': is_admin,
        'feature_status': feature_status,
        'metrics': metrics,
        'guild_record': guild_record,
        'is_vip': is_vip,
        'subscription_tier': subscription_tier,
        'billing_cycle': billing_cycle,
        'active_page': 'dashboard',
    }
    return render(request, 'questlog/guild_dashboard.html', context)


@discord_required
def force_sync_guild(request, guild_id):
    """Force an immediate sync of guild stats from Discord via bot API."""
    from django.http import JsonResponse

    try:
        # Call bot API to trigger sync
        bot_api_url = os.getenv('WARDEN_BOT_API_URL', 'http://localhost:8001')
        bot_api_token = os.getenv('DISCORD_BOT_API_TOKEN')
        response = requests.post(
            f'{bot_api_url}/api/sync/{guild_id}',
            headers={'Authorization': f'Bearer {bot_api_token}'},
            timeout=10
        )

        if response.status_code == 200:
            # Invalidate cache so new data loads immediately
            from .discord_resources import invalidate_guild_cache
            invalidate_guild_cache(guild_id)
            return JsonResponse({'success': True, 'message': 'Guild synced successfully'})
        else:
            return JsonResponse({
                'success': False,
                'error': response.json().get('error', 'Unknown error')
            }, status=response.status_code)

    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to sync guild {guild_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to connect to bot API'
        }, status=503)


def invalidate_cache(request, guild_id):
    """Invalidate Django cache for a guild (called by bot after auto-sync)."""
    from django.http import JsonResponse
    from .discord_resources import invalidate_guild_cache

    try:
        invalidate_guild_cache(guild_id)
        logger.info(f"Cache invalidated for guild {guild_id} via API")
        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Failed to invalidate cache for guild {guild_id}: {e}")
        return JsonResponse({'success': False, 'error': 'An internal error occurred. Please try again later.'}, status=500)


# Flair Store

@discord_required
def flair_store(request, guild_id):
    """Flair store for members to purchase and equip flairs."""
    from django.http import JsonResponse
    from .db import get_db_session
    from .models import GuildMember, Guild as GuildModel, GuildFlair
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    all_guilds = request.session.get('discord_all_guilds', [])
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Verify user is a member of this guild
    guild = next((g for g in all_guilds if str(g['id']) == str(guild_id)), None)
    if not guild:
        messages.error(request, "You are not a member of this server.")
        return redirect('questlog_dashboard')

    # Handle POST (purchase flair)
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

        # Handle flair removal (no replacement)
        if data.get('remove_flair'):
            try:
                with get_db_session() as db:
                    member = db.query(GuildMember).filter_by(
                        guild_id=int(guild_id),
                        user_id=int(discord_user['id'])
                    ).first()

                    if not member:
                        return JsonResponse({
                            'success': False,
                            'error': 'You are not a member of this server yet. Please interact in Discord first.'
                        }, status=404)

                    # Clear current flair
                    member.flair = None

                    # Add pending action to remove flair roles in Discord
                    from .models import PendingAction, ActionType, ActionStatus
                    action = PendingAction(
                        guild_id=int(guild_id),
                        action_type=ActionType.FLAIR_ASSIGN.value,
                        status=ActionStatus.PENDING.value,
                        payload=json.dumps({
                            'target_user_id': int(discord_user['id']),
                            'flair_name': None  # signals removal
                        }),
                        triggered_by=int(discord_user['id']),
                        triggered_by_name=discord_user.get('username', 'Unknown')
                    )
                    db.add(action)

                    logger.info(f"User {discord_user.get('username')} removed flair in guild {guild_id}")

                    return JsonResponse({
                        'success': True,
                        'message': 'Flair removed. Your Discord roles will update automatically.'
                    })
            except Exception as e:
                logger.error(f"Error removing flair: {e}", exc_info=True)
                return JsonResponse({'success': False, 'error': 'Failed to remove flair'}, status=500)

        flair_name = data.get('flair_name')

        # Look up flair in database
        try:
            with get_db_session() as db:
                flair = db.query(GuildFlair).filter_by(
                    guild_id=int(guild_id),
                    flair_name=flair_name,
                    enabled=True
                ).first()

                if not flair:
                    return JsonResponse({
                        'success': False,
                        'error': 'Flair not found or is disabled'
                    }, status=404)

                # Check if custom flair requires premium
                if flair.flair_type == 'custom':
                    guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
                    is_premium = (guild_record and (
                        guild_record.subscription_tier == 'complete' or guild_record.is_vip
                    ))
                    if not is_premium:
                        return JsonResponse({
                            'success': False,
                            'error': 'Custom flairs require the Complete Suite or Engagement module!'
                        }, status=403)

                cost = flair.cost

        except Exception as e:
            logger.error(f"Error looking up flair: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Failed to look up flair'}, status=500)

        # Purchase flair
        try:
            with get_db_session() as db:
                member = db.query(GuildMember).filter_by(
                    guild_id=int(guild_id),
                    user_id=int(discord_user['id'])
                ).first()

                if not member:
                    return JsonResponse({
                        'success': False,
                        'error': 'You are not a member of this server yet. Please interact in Discord first.'
                    }, status=404)

                # Check if user already owns this flair
                owned_flairs = []
                if member.owned_flairs:
                    try:
                        owned_flairs = json.loads(member.owned_flairs)
                    except:
                        owned_flairs = []

                already_owned = flair_name in owned_flairs

                # Only charge tokens if not already owned
                if not already_owned:
                    if member.hero_tokens < cost:
                        return JsonResponse({
                            'success': False,
                            'error': f'Insufficient tokens. You need {cost} but only have {member.hero_tokens}.'
                        }, status=400)

                    # Deduct tokens and add to owned flairs
                    if cost > 0:
                        member.hero_tokens -= cost
                    owned_flairs.append(flair_name)
                    member.owned_flairs = json.dumps(owned_flairs)

                # Equip the flair
                member.flair = flair_name

                # Create pending action for bot to assign Discord role
                from .models import PendingAction, ActionType, ActionStatus
                action = PendingAction(
                    guild_id=int(guild_id),
                    action_type=ActionType.FLAIR_ASSIGN.value,
                    status=ActionStatus.PENDING.value,
                    payload=json.dumps({
                        'target_user_id': int(discord_user['id']),
                        'flair_name': flair_name
                    }),
                    triggered_by=int(discord_user['id']),
                    triggered_by_name=discord_user.get('username', 'Unknown')
                )
                db.add(action)

                # Log appropriate action
                action_verb = "equipped" if already_owned else "purchased"
                logger.info(f"User {discord_user['username']} {action_verb} flair {flair_name} in guild {guild_id}")

                # Return appropriate message
                if already_owned:
                    message = f'Switched to {flair_name}! Your Discord role will update automatically.'
                else:
                    message = f'Successfully purchased {flair_name} for {cost} tokens! Your Discord role will update automatically.'

                return JsonResponse({
                    'success': True,
                    'flair': flair_name,
                    'tokens_remaining': member.hero_tokens,
                    'message': message,
                    'already_owned': already_owned
                })

        except Exception as e:
            logger.error(f"Error purchasing flair: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'An error occurred while processing your purchase. Please try again or contact support.'
            }, status=500)

    # GET request - show flair store
    normal_flairs = {}
    seasonal_flairs = {}
    custom_flairs = {}
    current_tokens = 0
    current_flair = None
    owned_flairs = []
    subscription_tier = 'free'
    is_vip = False
    is_premium = False
    is_pro = False
    token_name = "Hero Tokens"  # Default
    token_emoji = ":coin:"      # Default

    try:
        with get_db_session() as db:
            # Get member info
            member = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=int(discord_user['id'])
            ).first()

            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()

            current_tokens = member.hero_tokens if member else 0
            current_flair = member.flair if member else None

            # Get list of owned flairs
            if member and member.owned_flairs:
                try:
                    import json
                    owned_flairs = json.loads(member.owned_flairs)
                except:
                    owned_flairs = []

            subscription_tier = guild_record.subscription_tier if guild_record else 'free'
            is_vip = guild_record.is_vip if guild_record else False
            billing_cycle = guild_record.billing_cycle if guild_record else None
            token_name = guild_record.token_name if guild_record and guild_record.token_name else "Hero Tokens"
            token_emoji = guild_record.token_emoji if guild_record and guild_record.token_emoji else ":coin:"

            # Premium access = Complete tier OR VIP
            is_premium = subscription_tier == 'complete' or is_vip
            # All premium features available in Complete tier
            is_pro = is_premium

            # Load flairs from database
            flairs = db.query(GuildFlair).filter_by(
                guild_id=int(guild_id),
                enabled=True
            ).order_by(
                GuildFlair.flair_type,
                GuildFlair.display_order,
                GuildFlair.id
            ).all()

            # Organize flairs by type
            for flair in flairs:
                if flair.flair_type == 'normal':
                    normal_flairs[flair.flair_name] = flair.cost
                elif flair.flair_type == 'seasonal':
                    seasonal_flairs[flair.flair_name] = flair.cost
                elif flair.flair_type == 'custom':
                    custom_flairs[flair.flair_name] = flair.cost

    except Exception as e:
        logger.error(f"Error loading flair store: {e}", exc_info=True)

    # Check if user has admin permissions for this guild
    is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)

    # Check if guild has engagement module access
    has_engagement_module = has_module_access(guild_id, 'engagement')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,  # Add for sidebar navigation
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': is_admin,
        'normal_flairs': normal_flairs,
        'seasonal_flairs': seasonal_flairs,
        'custom_flairs': custom_flairs,
        'current_tokens': current_tokens,
        'current_flair': current_flair,
        'owned_flairs': owned_flairs,
        'subscription_tier': subscription_tier,
        'is_vip': is_vip,
        'billing_cycle': billing_cycle,
        'is_premium': is_premium,
        'is_pro': is_pro,
        'token_name': token_name,
        'token_emoji': token_emoji,
        'has_engagement_module': has_engagement_module,
        'has_any_module': has_any_module,
        'active_page': 'flair_store',
    }
    return render(request, 'questlog/flair_store.html', context)


# Member Profile
@discord_required
def member_profile(request, guild_id):
    """View member's own profile with XP, tokens, level, and stats."""
    from .db import get_db_session
    from .models import GuildMember, Guild as GuildModel

    discord_user = request.session.get('discord_user', {})
    all_guilds = request.session.get('discord_all_guilds', [])
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Check if user is member of this guild
    guild = next((g for g in all_guilds if str(g['id']) == str(guild_id)), None)
    if not guild:
        messages.error(request, "You are not a member of this server.")
        return redirect('questlog_dashboard')

    is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)

    # Get member stats from database
    member_stats = None
    guild_record = None
    token_name = "Hero Tokens"
    token_emoji = ":coin:"

    try:
        with get_db_session() as db:
            member = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=int(discord_user['id'])
            ).first()

            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            token_name = guild_record.token_name if guild_record and guild_record.token_name else "Hero Tokens"
            token_emoji = guild_record.token_emoji if guild_record and guild_record.token_emoji else ":coin:"

            # Build avatar URL
            avatar_url = None
            if discord_user.get('avatar'):
                avatar_url = f"https://cdn.discordapp.com/avatars/{discord_user['id']}/{discord_user['avatar']}.png?size=256"
            else:
                # Default Discord avatar (based on discriminator or user ID)
                discriminator = int(discord_user.get('discriminator', '0'))
                default_num = (discriminator % 5) if discriminator > 0 else (int(discord_user['id']) >> 22) % 6
                avatar_url = f"https://cdn.discordapp.com/embed/avatars/{default_num}.png"

            if member:
                member_stats = {
                    'display_name': member.display_name or member.username,
                    'avatar_url': avatar_url,
                    'level': member.level,
                    'xp': member.xp,
                    'tokens': member.hero_tokens,
                    'flair': member.flair,
                    'message_count': member.message_count,
                    'voice_minutes': member.voice_minutes,
                    'reaction_count': member.reaction_count,
                    'media_count': member.media_count,
                }
            else:
                member_stats = {
                    'display_name': discord_user.get('username', 'Unknown'),
                    'avatar_url': avatar_url,
                    'level': 0,
                    'xp': 0,
                    'tokens': 0,
                    'flair': None,
                    'message_count': 0,
                    'voice_minutes': 0,
                    'reaction_count': 0,
                    'media_count': 0,
                }
    except Exception as e:
        logger.error(f"Error loading member profile: {e}", exc_info=True)
        member_stats = {}

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,  # Add for sidebar navigation
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': is_admin,
        'member_stats': member_stats,
        'token_name': token_name,
        'token_emoji': token_emoji,
        'active_page': 'profile',
    }
    return render(request, 'questlog/member_profile.html', context)


# Leaderboards
@discord_required
def guild_leaderboards(request, guild_id):
    """View guild leaderboards for XP, tokens, and activity."""
    from .db import get_db_session
    from .models import GuildMember, Guild as GuildModel

    discord_user = request.session.get('discord_user', {})
    all_guilds = request.session.get('discord_all_guilds', [])
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Check if user is member of this guild
    guild = next((g for g in all_guilds if str(g['id']) == str(guild_id)), None)
    if not guild:
        messages.error(request, "You are not a member of this server.")
        return redirect('questlog_dashboard')

    is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)

    # Get leaderboards from database
    xp_leaders = []
    token_leaders = []
    message_leaders = []
    token_name = "Hero Tokens"
    token_emoji = ":coin:"

    try:
        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            token_name = guild_record.token_name if guild_record and guild_record.token_name else "Hero Tokens"
            token_emoji = guild_record.token_emoji if guild_record and guild_record.token_emoji else ":coin:"

            # Top 10 by XP
            xp_leaders = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                is_bot=False
            ).order_by(GuildMember.xp.desc()).limit(10).all()

            # Top 10 by Tokens
            token_leaders = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                is_bot=False
            ).order_by(GuildMember.hero_tokens.desc()).limit(10).all()

            # Top 10 by Messages
            message_leaders = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                is_bot=False
            ).order_by(GuildMember.message_count.desc()).limit(10).all()

    except Exception as e:
        logger.error(f"Error loading leaderboards: {e}", exc_info=True)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,  # Add for sidebar navigation
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': is_admin,
        'xp_leaders': xp_leaders,
        'token_leaders': token_leaders,
        'message_leaders': message_leaders,
        'token_name': token_name,
        'token_emoji': token_emoji,
        'active_page': 'leaderboards',
    }
    return render(request, 'questlog/leaderboards.html', context)


# Flair Management Dashboard (Premium Feature)

@discord_required
def flair_management(request, guild_id):
    """Flair management page - bulk editor for customizing flairs and prices."""
    from .db import get_db_session
    from .models import Guild as GuildModel
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Check premium status
    is_premium = False
    is_vip = False
    subscription_tier = 'free'

    try:
        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                subscription_tier = guild_record.subscription_tier
                is_vip = guild_record.is_vip
                billing_cycle = guild_record.billing_cycle
                is_premium = subscription_tier == 'complete' or is_vip

    except Exception as e:
        logger.warning(f"Could not fetch guild premium status: {e}")

    # Check if guild has engagement module access
    has_engagement_module = has_module_access(guild_id, 'engagement')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,  # User already validated as admin above
        'is_premium': is_premium,
        'is_vip': is_vip,
        'subscription_tier': subscription_tier,
        'billing_cycle': billing_cycle if 'billing_cycle' in locals() else None,
        'has_engagement_module': has_engagement_module,
        'has_any_module': has_any_module,
        'active_page': 'flair_management',
    }
    return render(request, 'questlog/flair_management.html', context)


# Tracker Management Dashboard Page

@discord_required
def guild_trackers(request, guild_id):
    """Manage channel stat trackers for a guild."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Fetch existing trackers from database
    trackers = []
    try:
        from .db import get_db_session
        from .models import ChannelStatTracker

        with get_db_session() as db:
            db_trackers = db.query(ChannelStatTracker).filter_by(
                guild_id=int(guild_id)
            ).all()

            trackers = [
                {
                    'id': t.id,
                    'channel_id': str(t.channel_id),
                    'role_id': str(t.role_id),
                    'label': t.label,
                    'emoji': t.emoji,
                    'game_name': t.game_name,
                    'show_playing_count': t.show_playing_count,
                    'enabled': t.enabled,
                    'last_topic': t.last_topic,
                }
                for t in db_trackers
            ]
    except Exception as e:
        logger.warning(f"Could not fetch trackers: {e}")

    # Check if guild has LFG module access
    has_lfg_module = has_module_access(guild_id, 'lfg')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'trackers': trackers,
        'has_lfg_module': has_lfg_module,
        'has_any_module': has_any_module,
        'active_page': 'trackers',
    }
    return render(request, 'questlog/trackers.html', context)


# Tracker API Endpoints (REST API for AJAX calls)

from django.http import JsonResponse, HttpResponseForbidden, HttpResponseNotFound, HttpResponse
import json as json_lib


# PERSISTENT CACHE for Discord guild permissions (industry standard: persistent cache)
_PERMISSION_CACHE_TTL = 30  # 30 seconds cache for permission checks (fast updates with minimal DB load)

# Rate limit tracking constants (per-user tracking)
_DISCORD_API_RATE_LIMIT_PER_USER = 1  # Max 1 call per user per window (Discord's actual limit)
_RATE_LIMIT_WINDOW = 3  # 3 second window per user to be VERY safe

# Emergency circuit breaker - completely stop API calls if we get too many 429s
_CIRCUIT_BREAKER_THRESHOLD = 5  # After 5 429s in a row
_CIRCUIT_BREAKER_DURATION = 300  # Stop all calls for 5 minutes


def check_custom_admin_roles(guild_id, user_id):
    """
    Check if user has any of the guild's custom admin roles.

    Uses cached_members from Guild table (synced by bot) to check role membership.
    NO Discord API calls - uses database cache only.

    Returns:
        bool: True if user has admin role, False otherwise
    """
    try:
        from .db import get_db_session
        from .models import Guild
        import json

        with get_db_session() as db:
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()

            if not guild:
                return False

            # Check if guild has custom admin roles configured
            admin_roles_json = guild.admin_roles if hasattr(guild, 'admin_roles') else None
            if not admin_roles_json:
                return False  # No custom roles configured

            # Parse admin role IDs
            try:
                admin_role_ids = json.loads(admin_roles_json) if isinstance(admin_roles_json, str) else admin_roles_json
                if not admin_role_ids or not isinstance(admin_role_ids, list):
                    return False
            except (json.JSONDecodeError, TypeError):
                return False

            # Get cached members data
            if not guild.cached_members:
                return False

            try:
                cached_members = json.loads(guild.cached_members)
            except (json.JSONDecodeError, TypeError):
                return False

            # Find this user in cached members
            user_data = next((m for m in cached_members if str(m.get('id')) == str(user_id)), None)
            if not user_data:
                return False

            # Check if user has any of the admin roles
            user_roles = user_data.get('roles', [])
            has_admin_role = any(str(role_id) in [str(r) for r in user_roles] for role_id in admin_role_ids)

            if has_admin_role:
                logger.info(f"[ADMIN_ROLE_CHECK] User {user_id} has custom admin role in guild {guild_id}")

            return has_admin_role

    except Exception as e:
        logger.error(f"Error checking custom admin roles for user {user_id} in guild {guild_id}: {e}")
        return False


def get_guilds_with_custom_admin_access(user_id, all_guilds):
    """
    Get list of guilds where user has custom admin role access.

    Args:
        user_id: Discord user ID
        all_guilds: List of guild dicts from Discord OAuth (with id, name, icon)

    Returns:
        List of guild dicts where user has custom admin roles
    """
    custom_admin_guilds = []

    for guild in all_guilds:
        guild_id = guild.get('id')
        if guild_id and check_custom_admin_roles(guild_id, user_id):
            custom_admin_guilds.append({
                'id': guild_id,
                'name': guild.get('name'),
                'icon': guild.get('icon'),
                'owner': False,
                'permissions': 0,  # No Discord permissions, but has custom admin role
                'custom_admin': True,  # Flag to indicate this is from custom roles
            })

    return custom_admin_guilds


def api_auth_required(view_func):
    """
    Check Discord auth and guild admin access for API endpoints.

    HYBRID PERMISSION CHECK:
    1. Discord OAuth permissions (Administrator or Manage Server)
    2. Custom admin roles configured in guild settings

    ZERO DISCORD API CALLS - Uses session data + database cache only.
    """
    def wrapper(request, guild_id, *args, **kwargs):
        from .discord_cache import get_cache

        cache = get_cache()

        discord_user = request.session.get('discord_user')
        if not discord_user:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        user_id = discord_user.get('id')
        cache_key = f"discord_permission:{user_id}:{guild_id}"

        # Check persistent cache first (30 minute TTL)
        cached = cache.get(cache_key)
        if cached is not None:
            logger.debug(f"Permission cache HIT for user {user_id} guild {guild_id}")
            if not cached['is_admin']:
                return JsonResponse({'error': 'No admin access to this guild'}, status=403)
            return view_func(request, guild_id, *args, **kwargs)

        # Cache MISS: Check both Discord permissions AND custom admin roles
        is_admin = False

        # Check #1: Discord OAuth permissions (session data)
        admin_guilds = request.session.get('discord_admin_guilds', [])
        guild = next((g for g in admin_guilds if str(g['id']) == str(guild_id)), None)

        if guild:
            is_admin = True
            logger.debug(f"Permission check: User {user_id} has Discord admin permissions in guild {guild_id}")

        # Check #2: Custom admin roles (database lookup - fast, no API call)
        if not is_admin:
            has_custom_role = check_custom_admin_roles(guild_id, user_id)
            if has_custom_role:
                is_admin = True
                logger.debug(f"Permission check: User {user_id} has custom admin role in guild {guild_id}")

        # Cache the result
        cache.set(cache_key, {'is_admin': is_admin}, ttl=_PERMISSION_CACHE_TTL)

        if not is_admin:
            logger.debug(f"Permission cache MISS: User {user_id} has no admin access to guild {guild_id}")
            return JsonResponse({'error': 'No admin access to this guild'}, status=403)

        logger.debug(f"Permission cache MISS: Cached admin access for user {user_id} guild {guild_id}")
        return view_func(request, guild_id, *args, **kwargs)
    return wrapper


def api_member_auth_required(view_func):
    """Check Discord auth and guild member access for API endpoints."""
    def wrapper(request, guild_id, *args, **kwargs):
        discord_user = request.session.get('discord_user')
        if not discord_user:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        all_guilds = request.session.get('discord_all_guilds', [])
        guild = next((g for g in all_guilds if str(g['id']) == str(guild_id)), None)
        if not guild:
            return JsonResponse({'error': 'No access to this guild'}, status=403)

        return view_func(request, guild_id, *args, **kwargs)
    return wrapper


def api_owner_required(view_func):
    """
    Check Discord auth and guild OWNER access for API endpoints.

    OWNER-ONLY CHECK:
    - Only the Discord server owner can access this endpoint
    - More restrictive than @api_auth_required (which allows admins)
    - Used for sensitive operations like Discovery Network apply/leave/rejoin

    ZERO DISCORD API CALLS - Uses session data + database only.
    """
    def wrapper(request, guild_id, *args, **kwargs):
        from .db import get_db_session
        from .models import Guild

        discord_user = request.session.get('discord_user')
        if not discord_user:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        user_id = int(discord_user.get('id'))

        # Check database for guild owner
        with get_db_session() as db:
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()

            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            if not guild.owner_id or int(guild.owner_id) != user_id:
                return JsonResponse({
                    'error': 'Owner access required',
                    'detail': 'Only the server owner can perform this action'
                }, status=403)

        return view_func(request, guild_id, *args, **kwargs)
    return wrapper


# Flair Management API Endpoints

@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_flair_list(request, guild_id):
    """GET /api/guild/<id>/flairs/ - Get all flairs for a guild."""
    try:
        from .models import GuildFlair
        from .db import get_db_session

        with get_db_session() as db:
            flairs = db.query(GuildFlair).filter_by(
                guild_id=int(guild_id)
            ).order_by(
                GuildFlair.flair_type,
                GuildFlair.display_order,
                GuildFlair.id
            ).all()

            flair_data = [
                {
                    'id': f.id,
                    'flair_name': f.flair_name,
                    'flair_type': f.flair_type,
                    'cost': f.cost,
                    'enabled': f.enabled,
                    'created_by': str(f.created_by) if f.created_by else None,
                    'display_order': f.display_order,
                }
                for f in flairs
            ]

            return JsonResponse({
                'success': True,
                'flairs': flair_data
            })

    except Exception as e:
        logger.error(f"Error fetching flairs: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user', rate='20/h', method='POST', block=True)
def api_flair_bulk_update(request, guild_id):
    """POST /api/guild/<id>/flairs/bulk-update/ - Update multiple flairs at once."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST allowed'}, status=405)

    try:
        from .models import GuildFlair
        from .db import get_db_session
        import json

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        flairs_to_update = data.get('flairs', [])

        if not flairs_to_update:
            return JsonResponse({'error': 'No flairs provided'}, status=400)

        rename_results = []

        with get_db_session() as db:
            updated = 0
            for flair_data in flairs_to_update:
                flair_id = int(flair_data['id'])
                flair = db.query(GuildFlair).filter_by(
                    id=flair_id,
                    guild_id=int(guild_id)
                ).first()

                if flair:
                    old_name = flair.flair_name
                    old_display_order = flair.display_order
                    new_name = flair_data.get('flair_name', flair.flair_name)
                    new_display_order = int(flair_data.get('display_order', flair.display_order))

                    # Update flair properties
                    flair.flair_name = new_name
                    flair.cost = int(flair_data.get('cost', flair.cost))
                    flair.enabled = flair_data.get('enabled', flair.enabled)
                    flair.display_order = new_display_order
                    flair.updated_at = int(time.time())
                    updated += 1

                    # Rename Discord role if flair name changed (for all flair types)
                    if new_name != old_name:
                        try:
                            bot_session = get_bot_session(int(guild_id))
                            if bot_session:
                                # Get all roles in the guild
                                roles_resp = bot_session.get(f'/guilds/{guild_id}/roles')
                                if roles_resp.status_code == 200:
                                    roles = roles_resp.json()
                                    # Find the role with the old name
                                    matching_role = next((r for r in roles if r['name'] == old_name), None)

                                    if matching_role:
                                        # Rename the role
                                        patch_resp = bot_session.patch(
                                            f'/guilds/{guild_id}/roles/{matching_role["id"]}',
                                            json={'name': new_name}
                                        )
                                        if patch_resp.status_code == 200:
                                            logger.info(f"Renamed Discord role '{old_name}' to '{new_name}' in guild {guild_id}")
                                            rename_results.append({'old': old_name, 'new': new_name, 'success': True})
                                        else:
                                            logger.warning(f"Failed to rename Discord role '{old_name}' in guild {guild_id}: {patch_resp.status_code}")
                                            rename_results.append({'old': old_name, 'new': new_name, 'success': False})
                                    else:
                                        logger.warning(f"Discord role '{old_name}' not found in guild {guild_id}")
                                else:
                                    logger.warning(f"Failed to fetch roles for guild {guild_id}: {roles_resp.status_code}")
                            else:
                                logger.warning(f"Bot session not available for guild {guild_id}")
                        except Exception as e:
                            logger.error(f"Error renaming Discord role '{old_name}' to '{new_name}' in guild {guild_id}: {e}", exc_info=True)

                    # Update Discord role position if display_order changed
                    if new_display_order != old_display_order:
                        try:
                            bot_session = get_bot_session(int(guild_id))
                            if bot_session:
                                # Get all roles in the guild to find the role ID
                                roles_resp = bot_session.get(f'/guilds/{guild_id}/roles')
                                if roles_resp.status_code == 200:
                                    roles = roles_resp.json()
                                    # Find the role with the current flair name (use new_name in case it was renamed)
                                    matching_role = next((r for r in roles if r['name'] == new_name), None)

                                    if matching_role:
                                        # Update the role position
                                        # Discord API expects an array of role position objects
                                        patch_resp = bot_session.patch(
                                            f'/guilds/{guild_id}/roles',
                                            json=[{'id': matching_role['id'], 'position': new_display_order}]
                                        )
                                        if patch_resp.status_code == 200:
                                            logger.info(f"Updated Discord role '{new_name}' position from {old_display_order} to {new_display_order} in guild {guild_id}")
                                        else:
                                            logger.warning(f"Failed to update Discord role '{new_name}' position in guild {guild_id}: {patch_resp.status_code}")
                                    else:
                                        logger.warning(f"Discord role '{new_name}' not found in guild {guild_id} for position update")
                                else:
                                    logger.warning(f"Failed to fetch roles for guild {guild_id} for position update: {roles_resp.status_code}")
                            else:
                                logger.warning(f"Bot session not available for guild {guild_id} for position update")
                        except Exception as e:
                            logger.error(f"Error updating Discord role '{new_name}' position in guild {guild_id}: {e}", exc_info=True)

            db.commit()

            return JsonResponse({
                'success': True,
                'updated': updated,
                'roles_renamed': len(rename_results),
                'rename_details': rename_results
            })

    except Exception as e:
        logger.error(f"Error bulk updating flairs: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_flair_create(request, guild_id):
    """POST /api/guild/<id>/flairs/create/ - Create a new custom flair."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST allowed'}, status=405)

    try:
        from .models import GuildFlair
        from .db import get_db_session
        import json

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        flair_name = data.get('flair_name', '').strip()
        cost = int(data.get('cost', 100))
        role_position = data.get('role_position')  # Optional position for the Discord role

        if not flair_name:
            return JsonResponse({'error': 'Flair name is required'}, status=400)

        if len(flair_name) > 100:
            return JsonResponse({'error': 'Flair name must be 100 characters or less'}, status=400)

        # Check if flair already exists
        with get_db_session() as db:
            existing = db.query(GuildFlair).filter_by(
                guild_id=int(guild_id),
                flair_name=flair_name
            ).first()

            if existing:
                return JsonResponse({'error': 'A flair with this name already exists'}, status=400)

            # Create new flair
            new_flair = GuildFlair(
                guild_id=int(guild_id),
                flair_name=flair_name,
                flair_type='custom',
                cost=cost,
                enabled=True,
                created_by=int(request.session.get('discord_user', {}).get('id', 0)),
                display_order=999,  # Put custom flairs at the end
                created_at=int(time.time()),
                updated_at=int(time.time())
            )

            db.add(new_flair)
            db.commit()

            # Create the Discord role for this flair
            role_created = False
            role_error = None
            try:
                bot_session = get_bot_session(int(guild_id))
                if bot_session:
                    # Create role with default settings (no special permissions, not hoisted by default)
                    role_data = {
                        'name': flair_name,
                        'permissions': '0',
                        'color': 0,
                        'hoist': False,  # Don't hoist by default
                        'mentionable': False,
                    }

                    resp = bot_session.post(f'/guilds/{guild_id}/roles', json=role_data)
                    if resp.status_code == 200:
                        role = resp.json()
                        role_created = True
                        logger.info(f"Created Discord role '{flair_name}' (ID: {role['id']}) for flair in guild {guild_id}")

                        # Set role position if specified
                        if role_position is not None:
                            try:
                                position = int(role_position)
                                patch_resp = bot_session.patch(
                                    f'/guilds/{guild_id}/roles',
                                    json=[{'id': role['id'], 'position': position}]
                                )
                                if patch_resp.status_code != 200:
                                    logger.warning(f"Failed to set role position for '{flair_name}' in guild {guild_id}: {patch_resp.status_code}")
                            except Exception as e:
                                logger.warning(f"Error setting role position for '{flair_name}' in guild {guild_id}: {e}")
                    else:
                        role_error = f"Discord API returned status {resp.status_code}"
                        logger.warning(f"Failed to create Discord role for flair '{flair_name}' in guild {guild_id}: {role_error}")
                else:
                    role_error = "Bot session not available"
                    logger.warning(f"Could not create Discord role for flair '{flair_name}': Bot not connected to guild {guild_id}")
            except Exception as e:
                role_error = str(e)
                logger.error(f"Error creating Discord role for flair '{flair_name}' in guild {guild_id}: {e}", exc_info=True)

            return JsonResponse({
                'success': True,
                'flair': {
                    'id': new_flair.id,
                    'flair_name': new_flair.flair_name,
                    'flair_type': new_flair.flair_type,
                    'cost': new_flair.cost,
                    'enabled': new_flair.enabled,
                    'created_by': str(new_flair.created_by) if new_flair.created_by else None,
                    'display_order': new_flair.display_order,
                },
                'role_created': role_created,
                'role_info': 'Discord role created successfully' if role_created else f'Flair created but Discord role creation failed: {role_error}'
            })

    except Exception as e:
        logger.error(f"Error creating flair: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_flair_delete(request, guild_id, flair_id):
    """DELETE /api/guild/<id>/flairs/<flair_id>/ - Delete a custom flair."""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Only DELETE allowed'}, status=405)

    try:
        from .models import GuildFlair
        from .db import get_db_session

        with get_db_session() as db:
            flair = db.query(GuildFlair).filter_by(
                id=int(flair_id),
                guild_id=int(guild_id)
            ).first()

            if not flair:
                return JsonResponse({'error': 'Flair not found'}, status=404)

            # Only allow deleting custom flairs
            if flair.flair_type != 'custom':
                return JsonResponse({'error': 'Cannot delete default flairs'}, status=403)

            db.delete(flair)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': f'Flair "{flair.flair_name}" deleted'
            })

    except Exception as e:
        logger.error(f"Error deleting flair: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_flair_create_default_roles(request, guild_id):
    """POST /api/guild/<id>/flairs/create-default-roles/ - Queue creation of default flair roles."""
    try:
        from .models import PendingAction, ActionType, ActionStatus
        from .db import get_db_session

        discord_user = request.session.get('discord_user', {})

        with get_db_session() as db:
            # Avoid duplicate pending actions
            existing = db.query(PendingAction).filter_by(
                guild_id=int(guild_id),
                action_type=ActionType.FLAIR_SEED_ROLES.value,
                status=ActionStatus.PENDING.value
            ).first()

            if existing:
                return JsonResponse({
                    'success': True,
                    'message': 'A flair role creation task is already queued.'
                })

            action = PendingAction(
                guild_id=int(guild_id),
                action_type=ActionType.FLAIR_SEED_ROLES.value,
                status=ActionStatus.PENDING.value,
                payload=json.dumps({}),
                triggered_by=int(discord_user.get('id', 0)),
                triggered_by_name=discord_user.get('username', 'Unknown')
            )
            db.add(action)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Default flair roles will be created and hoisted automatically.'
            })

    except Exception as e:
        logger.error(f"Error queueing default flair role creation for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to queue flair role creation'}, status=500)


# Tracker API Endpoints

@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_trackers_list(request, guild_id):
    """GET /api/guild/<id>/trackers/ - List all trackers for a guild."""
    try:
        from .db import get_db_session
        from .models import ChannelStatTracker

        with get_db_session() as db:
            trackers = db.query(ChannelStatTracker).filter_by(
                guild_id=int(guild_id)
            ).all()

            return JsonResponse({
                'success': True,
                'trackers': [
                    {
                        'id': t.id,
                        'channel_id': str(t.channel_id),
                        'role_id': str(t.role_id),
                        'label': t.label,
                        'emoji': t.emoji,
                        'game_name': t.game_name,
                        'show_playing_count': t.show_playing_count,
                        'enabled': t.enabled,
                        'last_topic': t.last_topic,
                        'last_updated': t.last_updated,
                    }
                    for t in trackers
                ]
            })
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_tracker_create(request, guild_id):
    """POST /api/guild/<id>/trackers/ - Create a new tracker."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    required = ['channel_id', 'role_id', 'label']
    for field in required:
        if field not in data:
            return JsonResponse({'error': f'Missing required field: {field}'}, status=400)

    try:
        from .db import get_db_session
        from .models import ChannelStatTracker

        discord_user = request.session.get('discord_user', {})

        with get_db_session() as db:
            # Check if tracker already exists for this channel
            existing = db.query(ChannelStatTracker).filter_by(
                guild_id=int(guild_id),
                channel_id=int(data['channel_id'])
            ).first()

            if existing:
                return JsonResponse({
                    'error': 'A tracker already exists for this channel'
                }, status=400)

            # Create new tracker
            tracker = ChannelStatTracker(
                guild_id=int(guild_id),
                channel_id=int(data['channel_id']),
                role_id=int(data['role_id']),
                label=data['label'],
                emoji=data.get('emoji'),
                game_name=data.get('game_name'),
                show_playing_count=bool(data.get('game_name')),
                enabled=True,
                created_by=int(discord_user.get('id', 0))
            )

            db.add(tracker)
            db.flush()  # Get the ID

            return JsonResponse({
                'success': True,
                'tracker': {
                    'id': tracker.id,
                    'channel_id': str(tracker.channel_id),
                    'role_id': str(tracker.role_id),
                    'label': tracker.label,
                    'emoji': tracker.emoji,
                    'game_name': tracker.game_name,
                    'show_playing_count': tracker.show_playing_count,
                    'enabled': tracker.enabled,
                }
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["PATCH", "PUT"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['PATCH', 'PUT'], block=True)
def api_tracker_update(request, guild_id, tracker_id):
    """PATCH /api/guild/<id>/trackers/<tracker_id>/ - Update a tracker."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import ChannelStatTracker

        with get_db_session() as db:
            tracker = db.query(ChannelStatTracker).filter_by(
                id=int(tracker_id),
                guild_id=int(guild_id)
            ).first()

            if not tracker:
                return JsonResponse({'error': 'Tracker not found'}, status=404)

            # Update allowed fields
            if 'label' in data:
                tracker.label = data['label']
            if 'emoji' in data:
                tracker.emoji = data['emoji'] or None
            if 'role_id' in data:
                tracker.role_id = int(data['role_id'])
            if 'game_name' in data:
                tracker.game_name = data['game_name'] or None
                tracker.show_playing_count = bool(data['game_name'])
            if 'enabled' in data:
                tracker.enabled = bool(data['enabled'])

            # Reset last_topic to force update on next bot cycle
            tracker.last_topic = None

            return JsonResponse({
                'success': True,
                'tracker': {
                    'id': tracker.id,
                    'channel_id': str(tracker.channel_id),
                    'role_id': str(tracker.role_id),
                    'label': tracker.label,
                    'emoji': tracker.emoji,
                    'game_name': tracker.game_name,
                    'show_playing_count': tracker.show_playing_count,
                    'enabled': tracker.enabled,
                }
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_tracker_delete(request, guild_id, tracker_id):
    """DELETE /api/guild/<id>/trackers/<tracker_id>/ - Delete a tracker."""
    try:
        from .db import get_db_session
        from .models import ChannelStatTracker

        with get_db_session() as db:
            tracker = db.query(ChannelStatTracker).filter_by(
                id=int(tracker_id),
                guild_id=int(guild_id)
            ).first()

            if not tracker:
                return JsonResponse({'error': 'Tracker not found'}, status=404)

            db.delete(tracker)

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_guild_resources(request, guild_id):
    """GET /api/guild/<id>/resources/ - Fetch guild channels, roles, and cached members."""
    try:
        from .db import get_db_session
        from .models import GuildMember as MemberModel

        # Get Discord bot token from session or settings
        admin_guilds = request.session.get('discord_admin_guilds', [])
        guild = next((g for g in admin_guilds if g['id'] == guild_id), None)

        if not guild:
            return JsonResponse({'error': 'Guild not found'}, status=404)

        response_data = {
            'channels': [],
            'roles': [],
            'members': []
        }

        # Get channels from Discord guild data (if available in session)
        if 'channels' in guild:
            response_data['channels'] = [
                {'id': str(ch.get('id')), 'name': ch.get('name'), 'type': ch.get('type')}
                for ch in guild.get('channels', [])
            ]

        # Get roles from Discord guild data (if available in session)
        if 'roles' in guild:
            response_data['roles'] = [
                {'id': str(role.get('id')), 'name': role.get('name'), 'color': role.get('color', 0)}
                for role in guild.get('roles', [])
            ]

        # Get cached members from database
        try:
            with get_db_session() as db:
                members = db.query(MemberModel).filter_by(
                    guild_id=int(guild_id)
                ).limit(1000).all()  # Limit to 1000 for performance

                response_data['members'] = [
                    {'id': str(m.user_id), 'name': m.display_name or f'User {m.user_id}'}
                    for m in members
                ]
        except Exception as e:
            logger.warning(f"Could not fetch members from DB: {e}")

        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Error fetching guild resources: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_guild_members(request, guild_id):
    """GET /api/guild/<id>/members/ - Fetch cached members (id + name)."""
    try:
        from .db import get_db_session
        from .models import GuildMember as MemberModel

        with get_db_session() as db:
            members = db.query(MemberModel).filter_by(guild_id=int(guild_id)).limit(1000).all()
            members_list = [
                {
                    'id': str(m.user_id),
                    'username': m.display_name or f'User {m.user_id}',
                    'display_name': m.display_name or f'User {m.user_id}'
                }
                for m in members
            ]
        return JsonResponse({'success': True, 'members': members_list})
    except Exception as e:
        logger.error(f"Error fetching guild members for {guild_id}: {e}")
        return JsonResponse({'success': False, 'error': 'An internal error occurred. Please try again later.'}, status=500)


# REMOVED DUPLICATE - See the correct api_guild_emojis implementation at line ~6252


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_message_action(request, guild_id):
    """POST /api/guild/<id>/messages/ - Queue message/embed send/edit/broadcast."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    mode = data.get('mode')
    if mode not in ('send', 'send_embed', 'edit', 'edit_embed', 'broadcast'):
        return JsonResponse({'error': 'Invalid mode'}, status=400)

    try:
        from .actions import queue_action, ActionType

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        payload = data
        payload['mode'] = mode
        payload['guild_id'] = int(guild_id)

        action_id = queue_action(
            guild_id=int(guild_id),
            action_type=ActionType.MESSAGE_SEND,
            payload=payload,
            triggered_by=triggered_by,
            triggered_by_name=triggered_by_name,
            source='website'
        )

        return JsonResponse({'success': True, 'action_id': action_id, 'message': 'Queued'})

    except Exception as e:
        logger.error(f"Error queueing message action for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to queue message action'}, status=500)


# Scheduled Messages Endpoints

@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_scheduled_messages_create(request, guild_id):
    """POST /api/guild/<id>/scheduled-messages/ - Create a scheduled message."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    # Validate required fields
    message_type = data.get('message_type')
    if message_type not in ('message', 'embed', 'broadcast'):
        return JsonResponse({'error': 'Invalid message_type. Must be message, embed, or broadcast'}, status=400)

    scheduled_time = data.get('scheduled_time')
    if not scheduled_time:
        return JsonResponse({'error': 'scheduled_time is required'}, status=400)

    content_data = data.get('content_data')
    if not content_data:
        return JsonResponse({'error': 'content_data is required'}, status=400)

    # Validate destination
    if message_type in ('message', 'embed') and not data.get('channel_id'):
        return JsonResponse({'error': 'channel_id is required for message and embed types'}, status=400)

    if message_type == 'broadcast' and not data.get('category_id'):
        return JsonResponse({'error': 'category_id is required for broadcast type'}, status=400)

    try:
        from .db import get_db_session
        from .models import ScheduledMessage

        discord_user = request.session.get('discord_user', {})
        user_id = int(discord_user.get('id', 0))

        with get_db_session() as db:
            scheduled_msg = ScheduledMessage(
                guild_id=int(guild_id),
                message_type=message_type,
                channel_id=int(data.get('channel_id')) if data.get('channel_id') else None,
                category_id=int(data.get('category_id')) if data.get('category_id') else None,
                scheduled_time=int(scheduled_time),
                timezone=data.get('timezone', 'UTC'),
                content_data=json_lib.dumps(content_data),
                status='pending',
                created_by=user_id,
                created_at=int(time.time()),
                updated_at=int(time.time())
            )
            db.add(scheduled_msg)
            db.commit()
            db.refresh(scheduled_msg)

            return JsonResponse({
                'success': True,
                'message': 'Scheduled message created successfully',
                'id': scheduled_msg.id
            })

    except Exception as e:
        logger.error(f"Error creating scheduled message for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to create scheduled message'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_scheduled_messages_list(request, guild_id):
    """GET /api/guild/<id>/scheduled-messages/ - List all scheduled messages for a guild."""
    try:
        from .db import get_db_session
        from .models import ScheduledMessage

        # Optional status filter
        status_filter = request.GET.get('status', None)

        with get_db_session() as db:
            from .models import Guild

            # Get guild to access cached channels/roles
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Parse cached resources
            cached_channels = json_lib.loads(guild.cached_channels) if guild.cached_channels else []
            channel_map = {str(ch['id']): ch['name'] for ch in cached_channels}

            query = db.query(ScheduledMessage).filter_by(guild_id=int(guild_id))

            if status_filter:
                query = query.filter_by(status=status_filter)

            # Order by scheduled_time ascending (soonest first)
            scheduled_messages = query.order_by(ScheduledMessage.scheduled_time.asc()).all()

            messages_data = []
            for msg in scheduled_messages:
                # Get channel/category name
                channel_name = None
                category_name = None

                if msg.channel_id:
                    channel_name = channel_map.get(str(msg.channel_id), f"Unknown ({msg.channel_id})")
                if msg.category_id:
                    category_name = channel_map.get(str(msg.category_id), f"Unknown ({msg.category_id})")

                messages_data.append({
                    'id': msg.id,
                    'message_type': msg.message_type,
                    'channel_id': str(msg.channel_id) if msg.channel_id else None,
                    'channel_name': channel_name,
                    'category_id': str(msg.category_id) if msg.category_id else None,
                    'category_name': category_name,
                    'scheduled_time': msg.scheduled_time,
                    'timezone': msg.timezone,
                    'content_data': json_lib.loads(msg.content_data),
                    'status': msg.status,
                    'sent_at': msg.sent_at,
                    'error_message': msg.error_message,
                    'created_by': str(msg.created_by),
                    'created_at': msg.created_at,
                    'updated_at': msg.updated_at
                })

            return JsonResponse({
                'success': True,
                'messages': messages_data,
                'count': len(messages_data)
            })

    except Exception as e:
        logger.error(f"Error listing scheduled messages for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to list scheduled messages'}, status=500)


@require_http_methods(["PUT"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='PUT', block=True)
def api_scheduled_messages_update(request, guild_id, message_id):
    """PUT /api/guild/<id>/scheduled-messages/<msg_id>/ - Update a scheduled message."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import ScheduledMessage

        with get_db_session() as db:
            scheduled_msg = db.query(ScheduledMessage).filter_by(
                id=int(message_id),
                guild_id=int(guild_id)
            ).first()

            if not scheduled_msg:
                return JsonResponse({'error': 'Scheduled message not found'}, status=404)

            # Only allow editing pending messages
            if scheduled_msg.status != 'pending':
                return JsonResponse({
                    'error': f'Cannot edit {scheduled_msg.status} messages. Only pending messages can be edited.'
                }, status=400)

            # Update fields if provided
            if 'scheduled_time' in data:
                scheduled_msg.scheduled_time = int(data['scheduled_time'])

            if 'timezone' in data:
                scheduled_msg.timezone = data['timezone']

            if 'content_data' in data:
                scheduled_msg.content_data = json_lib.dumps(data['content_data'])

            if 'channel_id' in data:
                scheduled_msg.channel_id = int(data['channel_id']) if data['channel_id'] else None

            if 'category_id' in data:
                scheduled_msg.category_id = int(data['category_id']) if data['category_id'] else None

            scheduled_msg.updated_at = int(time.time())

            db.commit()
            db.refresh(scheduled_msg)

            return JsonResponse({
                'success': True,
                'message': 'Scheduled message updated successfully'
            })

    except Exception as e:
        logger.error(f"Error updating scheduled message {message_id} for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to update scheduled message'}, status=500)


@require_http_methods(["DELETE"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_scheduled_messages_cancel(request, guild_id, message_id):
    """DELETE /api/guild/<id>/scheduled-messages/<msg_id>/ - Cancel a scheduled message."""
    try:
        from .db import get_db_session
        from .models import ScheduledMessage

        with get_db_session() as db:
            scheduled_msg = db.query(ScheduledMessage).filter_by(
                id=int(message_id),
                guild_id=int(guild_id)
            ).first()

            if not scheduled_msg:
                return JsonResponse({'error': 'Scheduled message not found'}, status=404)

            # Only allow cancelling pending messages
            if scheduled_msg.status != 'pending':
                return JsonResponse({
                    'error': f'Cannot cancel {scheduled_msg.status} messages. Only pending messages can be cancelled.'
                }, status=400)

            scheduled_msg.status = 'cancelled'
            scheduled_msg.updated_at = int(time.time())

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Scheduled message cancelled successfully'
            })

    except Exception as e:
        logger.error(f"Error cancelling scheduled message {message_id} for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to cancel scheduled message'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user', rate='5/h', method='POST', block=True)
def api_guild_leave(request, guild_id):
    """POST /api/guild/<id>/leave/ - Make the bot leave the specified guild."""
    try:
        # Verify user has admin access to this guild
        admin_guilds = request.session.get('discord_admin_guilds', [])
        guild = next((g for g in admin_guilds if g['id'] == guild_id), None)

        if not guild:
            return JsonResponse({'error': 'You do not have admin access to this server'}, status=403)

        # Get bot token from environment
        bot_token = os.getenv('DISCORD_BOT_TOKEN')

        if not bot_token or bot_token == 'your_bot_token_here':
            logger.error("DISCORD_BOT_TOKEN not configured")
            return JsonResponse({'error': 'Bot token not configured'}, status=500)

        # Make Discord API call to leave the guild
        response = requests.delete(
            f'https://discord.com/api/v10/users/@me/guilds/{guild_id}',
            headers={'Authorization': f'Bot {bot_token}'}
        )

        if response.status_code == 204:
            logger.info(f"Bot successfully left guild {guild_id} ({guild.get('name', 'Unknown')})")
            return JsonResponse({
                'success': True,
                'message': f'Successfully removed bot from {guild.get("name", "server")}'
            })
        elif response.status_code == 404:
            return JsonResponse({'error': 'Bot is not in this server or server not found'}, status=404)
        else:
            logger.error(f"Failed to leave guild {guild_id}: {response.status_code} - {response.text}")
            return JsonResponse({
                'error': f'Failed to leave server (Discord API error: {response.status_code})'
            }, status=500)

    except Exception as e:
        logger.error(f"Error leaving guild {guild_id}: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# XP & Leveling Dashboard

@discord_required
def guild_xp(request, guild_id):
    """XP and leveling management page for a guild."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Fetch XP config and leaderboard from database
    xp_config = None
    leaderboard = []
    level_roles = []
    total_members = 0
    guild_record = None

    try:
        from .db import get_db_session
        from .models import XPConfig, GuildMember, LevelRole, Guild as GuildModel, DailyBulkUsage
        from datetime import datetime

        # Get today's date as integer (YYYYMMDD)
        today_date = int(datetime.now().strftime('%Y%m%d'))

        with get_db_session() as db:
            # Get guild record for tier checking
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()

            # Get or create XP config
            config = db.query(XPConfig).filter_by(guild_id=int(guild_id)).first()
            if config:
                xp_config = {
                    'xp_enabled': config.xp_enabled,
                    'message_xp': config.message_xp,
                    'media_multiplier': config.media_multiplier,
                    'reaction_xp': config.reaction_xp,
                    'voice_xp': config.voice_xp_per_interval,
                    'command_xp': config.command_xp,
                    'gaming_xp': config.gaming_xp_per_interval,
                    'invite_xp': config.invite_xp,
                    'message_cooldown': config.message_cooldown,
                    'voice_interval': config.voice_interval,
                    'max_level': config.max_level,
                    'tokens_active': config.tokens_per_100_xp_active,
                    'tokens_passive': config.tokens_per_100_xp_passive,
                    'track_messages': config.track_messages,
                    'track_media': config.track_media,
                    'track_reactions': config.track_reactions,
                    'track_voice': config.track_voice,
                    'track_gaming': config.track_gaming,
                    'track_game_launch': config.track_game_launch,
                }

            # Get top 50 leaderboard
            members = db.query(GuildMember).filter_by(
                guild_id=int(guild_id)
            ).order_by(GuildMember.xp.desc()).limit(50).all()

            leaderboard = [
                {
                    'user_id': str(m.user_id),
                    'display_name': m.display_name or f'User {m.user_id}',
                    'username': m.username,
                    'xp': round(m.xp, 1),
                    'level': m.level,
                    'hero_tokens': m.hero_tokens,
                    'message_count': m.message_count,
                    'voice_minutes': m.voice_minutes,
                }
                for m in members
            ]

            total_members = db.query(GuildMember).filter_by(
                guild_id=int(guild_id)
            ).count()

            # Get level roles
            roles = db.query(LevelRole).filter_by(
                guild_id=int(guild_id)
            ).order_by(LevelRole.level).all()

            level_roles = [
                {
                    'id': r.id,
                    'level': r.level,
                    'role_id': str(r.role_id),
                    'role_name': r.role_name,
                    'remove_previous': r.remove_previous,
                }
                for r in roles
            ]

            # Get daily bulk operation usage for today
            bulk_edit_usage = db.query(DailyBulkUsage).filter_by(
                guild_id=int(guild_id),
                date=today_date,
                operation_category='xp_bulk_edit'
            ).first()

            import_usage = db.query(DailyBulkUsage).filter_by(
                guild_id=int(guild_id),
                date=today_date,
                operation_category='xp_import'
            ).first()

            bulk_edit_items_today = bulk_edit_usage.items_processed if bulk_edit_usage else 0
            import_items_today = import_usage.items_processed if import_usage else 0

    except Exception as e:
        logger.warning(f"Could not fetch XP data: {e}")
        bulk_edit_items_today = 0
        import_items_today = 0

    # Check if guild has engagement module access
    has_engagement_module = has_module_access(guild_id, 'engagement')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'xp_config': xp_config or {},
        'leaderboard': leaderboard,
        'level_roles': level_roles,
        'total_members': total_members,
        'has_engagement_module': has_engagement_module,
        'has_any_module': has_any_module,
        'active_page': 'xp',
        'bulk_edit_items_today': bulk_edit_items_today,
        'import_items_today': import_items_today,
    }
    return render(request, 'questlog/xp.html', context)


# XP API Endpoints

@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_xp_config(request, guild_id):
    """GET /api/guild/<id>/xp/config/ - Get XP configuration."""
    try:
        from .db import get_db_session
        from .models import XPConfig
        from sqlalchemy.exc import ProgrammingError
        from sqlalchemy import text

        with get_db_session() as db:
            try:
                config = db.query(XPConfig).filter_by(guild_id=int(guild_id)).first()
            except ProgrammingError as pe:
                logger.error(f"XP config query failed (likely missing columns). Defaulting config. {pe}")
                # Attempt to add missing column if absent
                try:
                    db.execute(text("ALTER TABLE xp_configs ADD COLUMN xp_enabled TINYINT(1) DEFAULT 1"))
                    db.flush()
                    config = db.query(XPConfig).filter_by(guild_id=int(guild_id)).first()
                except Exception as alter_err:
                    logger.error(f"Could not auto-add xp_enabled column: {alter_err}")
                config = None

            if not config:
                # Return defaults
                return JsonResponse({
                    'success': True,
                    'config': {
                        'xp_enabled': False,  # XP disabled by default
                        'message_xp': 1.5,
                        'media_multiplier': 1.3,
                        'reaction_xp': 1.0,
                        'voice_xp': 1.3,
                        'command_xp': 1.0,
                        'gaming_xp': 1.2,
                        'invite_xp': 50.0,
                        'message_cooldown': 60,
                        'voice_interval': 5400,
                        'max_level': 99,
                        'tokens_active': 15,
                        'tokens_passive': 5,
                        # XP source toggles (OPT-IN: disabled by default)
                        'track_messages': False,
                        'track_media': False,
                        'track_reactions': False,
                        'track_voice': False,
                        'track_gaming': False,
                        'track_game_launch': False,
                    }
                })

            return JsonResponse({
                'success': True,
                'config': {
                    'xp_enabled': config.xp_enabled,
                    'message_xp': config.message_xp,
                    'media_multiplier': config.media_multiplier,
                    'reaction_xp': config.reaction_xp,
                    'voice_xp': config.voice_xp_per_interval,
                    'command_xp': config.command_xp,
                    'gaming_xp': config.gaming_xp_per_interval,
                    'invite_xp': config.invite_xp,
                    'message_cooldown': config.message_cooldown,
                    'voice_interval': config.voice_interval,
                    'max_level': config.max_level,
                    'tokens_active': config.tokens_per_100_xp_active,
                    'tokens_passive': config.tokens_per_100_xp_passive,
                    # XP source toggles
                    'track_messages': config.track_messages,
                    'track_media': config.track_media,
                    'track_reactions': config.track_reactions,
                    'track_voice': config.track_voice,
                    'track_gaming': config.track_gaming,
                    'track_game_launch': config.track_game_launch,
                }
            })
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST", "PATCH"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['POST', 'PATCH'], block=True)
def api_xp_config_update(request, guild_id):
    """POST /api/guild/<id>/xp/config/ - Update XP configuration."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import XPConfig, Guild as GuildModel
        from sqlalchemy.exc import ProgrammingError
        from sqlalchemy import text

        with get_db_session() as db:
            # Ensure guild exists
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                guild_record = GuildModel(guild_id=int(guild_id))
                db.add(guild_record)
                db.flush()

            try:
                config = db.query(XPConfig).filter_by(guild_id=int(guild_id)).first()
            except ProgrammingError as pe:
                logger.error(f"XP config query failed (likely missing columns). Defaulting config. {pe}")
                try:
                    db.execute(text("ALTER TABLE xp_configs ADD COLUMN xp_enabled TINYINT(1) DEFAULT 1"))
                    db.flush()
                    config = db.query(XPConfig).filter_by(guild_id=int(guild_id)).first()
                except Exception as alter_err:
                    logger.error(f"Could not auto-add xp_enabled column: {alter_err}")
                config = None
            if not config:
                config = XPConfig(guild_id=int(guild_id))
                db.add(config)

            # Update fields
            if 'message_xp' in data:
                config.message_xp = float(data['message_xp'])
            if 'media_multiplier' in data:
                config.media_multiplier = float(data['media_multiplier'])
            if 'reaction_xp' in data:
                config.reaction_xp = float(data['reaction_xp'])
            if 'voice_xp' in data:
                config.voice_xp_per_interval = float(data['voice_xp'])
            if 'command_xp' in data:
                config.command_xp = float(data['command_xp'])
            if 'gaming_xp' in data:
                config.gaming_xp_per_interval = float(data['gaming_xp'])
            if 'invite_xp' in data:
                config.invite_xp = float(data['invite_xp'])
            if 'message_cooldown' in data:
                config.message_cooldown = int(data['message_cooldown'])
            if 'voice_interval' in data:
                config.voice_interval = int(data['voice_interval'])
            if 'max_level' in data:
                config.max_level = int(data['max_level'])
            if 'tokens_active' in data:
                config.tokens_per_100_xp_active = int(data['tokens_active'])
            if 'tokens_passive' in data:
                config.tokens_per_100_xp_passive = int(data['tokens_passive'])

            # Update XP source toggles
            if 'track_messages' in data:
                config.track_messages = bool(data['track_messages'])
                logger.info(f"[XP Config] Setting track_messages to {config.track_messages} for guild {guild_id}")
            if 'track_media' in data:
                config.track_media = bool(data['track_media'])
                logger.info(f"[XP Config] Setting track_media to {config.track_media} for guild {guild_id}")
            if 'track_reactions' in data:
                config.track_reactions = bool(data['track_reactions'])
                logger.info(f"[XP Config] Setting track_reactions to {config.track_reactions} for guild {guild_id}")
            if 'track_voice' in data:
                config.track_voice = bool(data['track_voice'])
                logger.info(f"[XP Config] Setting track_voice to {config.track_voice} for guild {guild_id}")
            if 'track_gaming' in data:
                config.track_gaming = bool(data['track_gaming'])
                logger.info(f"[XP Config] Setting track_gaming to {config.track_gaming} for guild {guild_id}")
            if 'track_game_launch' in data:
                config.track_game_launch = bool(data['track_game_launch'])
                logger.info(f"[XP Config] Setting track_game_launch to {config.track_game_launch} for guild {guild_id}")

            db.flush()  # Ensure changes are flushed before commit
            logger.info(f"[XP Config] Saved XP config for guild {guild_id}")
            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_xp_toggle(request, guild_id):
    """POST /api/guild/<id>/xp/toggle/ - Toggle XP system on/off."""
    import json
    try:
        from .db import get_db_session
        from .models import XPConfig
        from sqlalchemy.exc import ProgrammingError
        from sqlalchemy import text

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        xp_enabled = data.get('xp_enabled', False)

        with get_db_session() as db:
            try:
                config = db.query(XPConfig).filter_by(guild_id=int(guild_id)).first()
            except ProgrammingError as pe:
                logger.error(f"XP toggle query failed (likely missing columns). Creating default. {pe}")
                try:
                    db.execute(text("ALTER TABLE xp_configs ADD COLUMN xp_enabled TINYINT(1) DEFAULT 1"))
                    db.flush()
                    config = db.query(XPConfig).filter_by(guild_id=int(guild_id)).first()
                except Exception as alter_err:
                    logger.error(f"Could not auto-add xp_enabled column: {alter_err}")
                config = None

            if not config:
                # Create new config with the enabled status
                config = XPConfig(guild_id=int(guild_id), xp_enabled=xp_enabled)
                db.add(config)
            else:
                # Update existing config
                config.xp_enabled = xp_enabled

            # Keep guild table flag in sync (bot checks Guild.xp_enabled)
            from .models import Guild as GuildModel
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                guild_record.xp_enabled = xp_enabled

            db.commit()

            return JsonResponse({
                'success': True,
                'xp_enabled': config.xp_enabled
            })

    except Exception as e:
        logger.error(f"Error toggling XP system: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_xp_leaderboard(request, guild_id):
    """GET /api/guild/<id>/xp/leaderboard/ - Get XP leaderboard."""
    try:
        from .db import get_db_session
        from .models import GuildMember

        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 50))
        offset = (page - 1) * per_page

        with get_db_session() as db:
            members = db.query(GuildMember).filter_by(
                guild_id=int(guild_id)
            ).order_by(GuildMember.xp.desc()).offset(offset).limit(per_page).all()

            total = db.query(GuildMember).filter_by(guild_id=int(guild_id)).count()

            return JsonResponse({
                'success': True,
                'leaderboard': [
                    {
                        'user_id': str(m.user_id),
                        'display_name': m.display_name or f'User {m.user_id}',
                        'username': m.username,
                        'xp': round(m.xp, 1),
                        'level': m.level,
                        'hero_tokens': m.hero_tokens,
                        'message_count': m.message_count,
                        'voice_minutes': m.voice_minutes,
                    }
                    for m in members
                ],
                'total': total,
                'page': page,
                'per_page': per_page,
                'total_pages': (total + per_page - 1) // per_page,
            })
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
@validate_json_schema({
    "type": "object",
    "properties": {
        "action": {"type": "string", "enum": ["add", "subtract", "set"]},
        "amount": {"type": "number", "minimum": -1000000, "maximum": 1000000},
        "field": {"type": "string", "enum": ["xp", "tokens"]}
    },
    "required": ["action", "amount", "field"]
})
@api_auth_required
def api_xp_member_update(request, guild_id, user_id):
    """POST /api/guild/<id>/xp/member/<user_id>/ - Update a member's XP/tokens."""
    data = request.validated_data

    try:
        from .actions import queue_xp_add, queue_xp_set, queue_tokens_add, queue_tokens_set

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        action_type = data.get('action', 'set')  # 'add', 'subtract', 'set'
        amount = float(data.get('amount', 0))
        field = data.get('field', 'xp')  # 'xp' or 'tokens'

        if field == 'xp':
            if action_type == 'add':
                action_id = queue_xp_add(
                    guild_id=int(guild_id),
                    user_id=int(user_id),
                    amount=amount,
                    triggered_by=triggered_by,
                    triggered_by_name=triggered_by_name
                )
            elif action_type == 'subtract':
                # Use negative amount for subtraction
                action_id = queue_xp_add(
                    guild_id=int(guild_id),
                    user_id=int(user_id),
                    amount=-abs(amount),
                    triggered_by=triggered_by,
                    triggered_by_name=triggered_by_name
                )
            elif action_type == 'set':
                action_id = queue_xp_set(
                    guild_id=int(guild_id),
                    user_id=int(user_id),
                    amount=amount,
                    triggered_by=triggered_by,
                    triggered_by_name=triggered_by_name
                )
            else:
                return JsonResponse({'error': 'Invalid action type'}, status=400)
        elif field == 'tokens':
            if action_type == 'add':
                action_id = queue_tokens_add(
                    guild_id=int(guild_id),
                    user_id=int(user_id),
                    amount=int(amount),
                    triggered_by=triggered_by,
                    triggered_by_name=triggered_by_name
                )
            elif action_type == 'subtract':
                # Use negative amount for subtraction
                action_id = queue_tokens_add(
                    guild_id=int(guild_id),
                    user_id=int(user_id),
                    amount=-abs(int(amount)),
                    triggered_by=triggered_by,
                    triggered_by_name=triggered_by_name
                )
            elif action_type == 'set':
                action_id = queue_tokens_set(
                    guild_id=int(guild_id),
                    user_id=int(user_id),
                    amount=int(amount),
                    triggered_by=triggered_by,
                    triggered_by_name=triggered_by_name
                )
            else:
                return JsonResponse({'error': 'Invalid action type'}, status=400)
        else:
            return JsonResponse({'error': 'Invalid field'}, status=400)

        return JsonResponse({
            'success': True,
            'action_id': action_id,
            'message': f'Action queued (ID: {action_id})'
        })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_xp_level_roles(request, guild_id):
    """GET /api/guild/<id>/xp/roles/ - Get level roles."""
    try:
        from .db import get_db_session
        from .models import LevelRole

        with get_db_session() as db:
            roles = db.query(LevelRole).filter_by(
                guild_id=int(guild_id)
            ).order_by(LevelRole.level).all()

            return JsonResponse({
                'success': True,
                'level_roles': [
                    {
                        'id': r.id,
                        'level': r.level,
                        'role_id': str(r.role_id),
                        'role_name': r.role_name,
                        'remove_previous': r.remove_previous,
                    }
                    for r in roles
                ]
            })
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_xp_level_role_create(request, guild_id):
    """POST /api/guild/<id>/xp/roles/ - Create a level role."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    required = ['level', 'role_id']
    for field in required:
        if field not in data:
            return JsonResponse({'error': f'Missing required field: {field}'}, status=400)

    try:
        from .db import get_db_session
        from .models import LevelRole

        with get_db_session() as db:
            # Check if level already has a role
            existing = db.query(LevelRole).filter_by(
                guild_id=int(guild_id),
                level=int(data['level'])
            ).first()

            if existing:
                return JsonResponse({'error': 'A role already exists for this level'}, status=400)

            role = LevelRole(
                guild_id=int(guild_id),
                level=int(data['level']),
                role_id=int(data['role_id']),
                role_name=data.get('role_name'),
                remove_previous=data.get('remove_previous', True)
            )
            db.add(role)
            db.flush()

            return JsonResponse({
                'success': True,
                'level_role': {
                    'id': role.id,
                    'level': role.level,
                    'role_id': str(role.role_id),
                    'role_name': role.role_name,
                    'remove_previous': role.remove_previous,
                }
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_xp_level_role_bulk_create(request, guild_id):
    """POST /api/guild/<id>/xp/roles/bulk-create/ - Create multiple level roles at once."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if 'roles' not in data or not isinstance(data['roles'], list):
        return JsonResponse({'error': 'Missing or invalid "roles" array'}, status=400)

    if len(data['roles']) == 0:
        return JsonResponse({'error': 'No roles provided'}, status=400)

    try:
        from .db import get_db_session
        from .models import LevelRole

        created_count = 0
        errors = []

        with get_db_session() as db:
            for idx, role_data in enumerate(data['roles']):
                # Validate required fields
                if 'level' not in role_data or 'role_id' not in role_data:
                    errors.append(f"Row {idx + 1}: Missing level or role_id")
                    continue

                try:
                    level = int(role_data['level'])
                    role_id = int(role_data['role_id'])
                except (ValueError, TypeError):
                    errors.append(f"Row {idx + 1}: Invalid level or role_id")
                    continue

                # Check if level already has a role
                existing = db.query(LevelRole).filter_by(
                    guild_id=int(guild_id),
                    level=level
                ).first()

                if existing:
                    errors.append(f"Row {idx + 1}: Level {level} already has a role assigned")
                    continue

                # Create new level role
                role = LevelRole(
                    guild_id=int(guild_id),
                    level=level,
                    role_id=role_id,
                    role_name=role_data.get('role_name'),
                    remove_previous=role_data.get('remove_previous', True)
                )
                db.add(role)
                created_count += 1

            db.flush()

            return JsonResponse({
                'success': True,
                'created': created_count,
                'errors': errors if errors else None
            })

    except Exception as e:
        logger.error(f"Bulk level role creation error: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_xp_level_role_delete(request, guild_id, role_id):
    """DELETE /api/guild/<id>/xp/roles/<role_id>/ - Delete a level role."""
    try:
        from .db import get_db_session
        from .models import LevelRole

        with get_db_session() as db:
            role = db.query(LevelRole).filter_by(
                id=int(role_id),
                guild_id=int(guild_id)
            ).first()

            if not role:
                return JsonResponse({'error': 'Level role not found'}, status=404)

            db.delete(role)
            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE"])
@api_auth_required
@ratelimit(key='user', rate='10/m', method='DELETE', block=True)
def api_xp_member_delete(request, guild_id, user_id):
    """DELETE /api/guild/<id>/xp/member/<user_id>/delete/ - Remove a member from XP system."""
    try:
        from .db import get_db_session
        from .models import GuildMember

        with get_db_session() as db:
            member = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=int(user_id)
            ).first()

            if not member:
                return JsonResponse({'error': 'Member not found'}, status=404)

            display_name = member.display_name or f"User {user_id}"
            db.delete(member)

            return JsonResponse({
                'success': True,
                'message': f'Removed {display_name} from XP system'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_xp_member_add(request, guild_id):
    """POST /api/guild/<id>/xp/member/add/ - Manually add a member to XP system."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import GuildMember

        user_id = int(data.get('user_id'))
        xp = float(data.get('xp', 0))
        level = int(data.get('level', 0))
        hero_tokens = int(data.get('hero_tokens', 0))
        display_name = data.get('display_name', '')

        with get_db_session() as db:
            # Check if member already exists
            existing = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=user_id
            ).first()

            if existing:
                return JsonResponse({'error': 'Member already exists in XP system'}, status=400)

            # Create new member
            member = GuildMember(
                guild_id=int(guild_id),
                user_id=user_id,
                xp=xp,
                level=level,
                hero_tokens=hero_tokens,
                display_name=display_name
            )
            db.add(member)

            return JsonResponse({
                'success': True,
                'message': f'Added {display_name or user_id} to XP system'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user', rate='10/h', method='POST', block=True)
def api_xp_import_csv(request, guild_id):
    """POST /api/guild/<id>/xp/import/ - Import XP data from CSV or XLSX file."""
    try:
        import logging
        logger = logging.getLogger(__name__)

        if 'csv_file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)

        upload_file = request.FILES['csv_file']
        filename = upload_file.name.lower()

        # Validate file size (10MB limit)
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        if upload_file.size > MAX_FILE_SIZE:
            return JsonResponse({'error': 'File too large (max 10MB)'}, status=400)

        # SECURITY: Only accept .xlsx files (more reliable than CSV, prevents encoding issues)
        if not filename.endswith('.xlsx'):
            return JsonResponse({'error': 'Invalid file type. Only .xlsx files allowed'}, status=400)

        # SECURITY: Validate content-type header to prevent extension spoofing
        content_type = upload_file.content_type.lower() if upload_file.content_type else ''
        allowed_content_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/octet-stream',  # Sometimes browsers send this for Excel files
        ]
        if content_type and content_type not in allowed_content_types:
            logger.warning(f"[XP IMPORT] Rejected file with invalid content-type: {content_type} (filename: {filename})")
            return JsonResponse({'error': f'Invalid file type. Content-type {content_type} not allowed'}, status=400)

        # LOG who is uploading XP files
        user_session = request.session.get('discord_user', {})
        logger.warning(
            f"[XP IMPORT TRACKER] User {user_session.get('username', 'UNKNOWN')} "
            f"(ID: {user_session.get('id', 'UNKNOWN')}) uploading XP file '{filename}' "
            f"for guild {guild_id}"
        )

        from .db import get_db_session
        from .models import GuildMember

        # Parse XLSX file
        from openpyxl import load_workbook
        from io import BytesIO

        wb = load_workbook(BytesIO(upload_file.read()))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Read data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = str(value) if value is not None else ''
            rows.append(row_dict)

        # Check tier limits and daily usage
        from .models import Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Check daily bulk limit
            allowed, error_msg, usage_info = check_daily_bulk_limit(
                guild_id, 'xp_import', len(rows), guild_record
            )

            if not allowed:
                return JsonResponse({
                    'error': error_msg,
                    'limit_exceeded': True,
                    'usage_info': usage_info
                }, status=403)

        created = 0
        updated = 0
        skipped = 0
        errors = []

        with get_db_session() as db:
            for row in rows:
                try:
                    # Handle scientific notation in user_id
                    user_id_str = row.get('user_id', '').strip()
                    if not user_id_str:
                        skipped += 1
                        continue

                    # Clean up any Excel artifacts
                    user_id_str = user_id_str.replace("'", "").replace('"', '').replace('=', '')

                    # Convert directly to int - DO NOT use float() as it loses precision for large numbers!
                    user_id = int(user_id_str)

                    # VALIDATE: Detect corrupted IDs from Excel (trailing zeros = precision loss)
                    user_id_check = str(user_id)
                    if user_id_check.endswith('000000000'):  # 9+ trailing zeros = corrupted
                        errors.append(f"CORRUPTED ID detected: {user_id_check} for {row.get('display_name', 'unknown')}. "
                                    f"Excel converted IDs to scientific notation! "
                                    f"FIX: Format user_id column as TEXT before pasting IDs, or use Google Sheets.")
                        skipped += 1
                        continue

                    # VALIDATE: Check if string was converted through float() (subtle corruption)
                    # If the original string doesn't match the parsed int, it was corrupted
                    if user_id_str != str(user_id):
                        errors.append(
                            f"CORRUPTED ID detected: input '{user_id_str}' became {user_id} for "
                            f"{row.get('display_name', 'unknown')}. This is float precision loss! "
                            f"FIX: Check your Excel/CSV formatting."
                        )
                        logger.error(
                            f"[XP IMPORT] REJECTED CORRUPTED ID: '{user_id_str}' -> {user_id} "
                            f"for {row.get('display_name', 'unknown')}"
                        )
                        skipped += 1
                        continue

                    # Validate ID is reasonable (Discord IDs are 17-19 digits)
                    if user_id < 100000000000000000 or user_id > 9999999999999999999:
                        errors.append(f"Invalid user_id {user_id} for {row.get('display_name', 'unknown')} - must be 17-19 digits")
                        skipped += 1
                        continue

                    xp = float(row.get('xp', 0))
                    level = int(row.get('level', 0))
                    hero_tokens = int(row.get('hero_tokens', 0))
                    display_name = row.get('display_name', '').strip()

                    # Skip members with 0 XP and level
                    if xp == 0 and level == 0:
                        skipped += 1
                        continue

                    # Check if member exists
                    member = db.query(GuildMember).filter_by(
                        guild_id=int(guild_id),
                        user_id=user_id
                    ).first()

                    if member:
                        # Update existing
                        member.xp = xp
                        member.level = level
                        member.hero_tokens = hero_tokens
                        if display_name:
                            member.display_name = display_name
                        updated += 1
                    else:
                        # Create new
                        logger.warning(
                            f"[XP IMPORT TRACKER] CREATING NEW MEMBER: guild={guild_id}, "
                            f"user_id={user_id}, display_name={display_name}, xp={xp}, "
                            f"from_file={filename}"
                        )
                        member = GuildMember(
                            guild_id=int(guild_id),
                            user_id=user_id,
                            xp=xp,
                            level=level,
                            hero_tokens=hero_tokens,
                            display_name=display_name
                        )
                        db.add(member)
                        created += 1

                except Exception as e:
                    errors.append(f"Row error: {str(e)}")
                    continue

        # Record usage for daily tracking (only count successfully processed rows)
        processed_count = created + updated
        if processed_count > 0:
            record_bulk_usage(guild_id, 'xp_import', processed_count)

        return JsonResponse({
            'success': True,
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:10]  # Return first 10 errors
        })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@require_subscription_tier('pro', 'premium')
@ratelimit(key='user_or_ip', rate='10/h', method='GET', block=True)
def api_xp_export_csv(request, guild_id):
    """GET /api/guild/<id>/xp/export/ - Export XP data as Excel (Pro/Premium only, 10 req/hour)."""
    try:
        from .db import get_db_session
        from .models import GuildMember
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from io import BytesIO

        with get_db_session() as db:
            members = db.query(GuildMember).filter_by(guild_id=int(guild_id)).order_by(GuildMember.xp.desc()).all()

            # Create XLSX workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "XP Data"

            # Write headers
            headers = ['user_id', 'display_name', 'xp', 'level', 'hero_tokens', 'message_count']
            ws.append(headers)

            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Write data
            for member in members:
                ws.append([
                    str(member.user_id),
                    member.display_name or '',
                    member.xp,
                    member.level,
                    member.hero_tokens,
                    member.message_count or 0
                ])

            # Format user_id column as TEXT (prevents Excel scientific notation)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=1).number_format = '@'  # @ = TEXT format

            # Adjust column widths
            ws.column_dimensions['A'].width = 20  # user_id
            ws.column_dimensions['B'].width = 25  # display_name
            ws.column_dimensions['C'].width = 12  # xp
            ws.column_dimensions['D'].width = 10  # level
            ws.column_dimensions['E'].width = 15  # hero_tokens
            ws.column_dimensions['F'].width = 15  # message_count

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Create response
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="guild_{guild_id}_xp_export.xlsx"'

            return response

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET", "POST"])
@api_auth_required
@require_subscription_tier('pro', 'premium')
@ratelimit(key='user_or_ip', rate='5/h', method='POST', block=True)
def api_xp_bulk_edit(request, guild_id):
    """GET/POST /api/guild/<id>/xp/bulk-edit/ - Bulk edit XP for active members (Pro/Premium only, 5 req/hour)."""
    try:
        from .db import get_db_session
        from .models import GuildMember
        import json

        if request.method == 'GET':
            # Fetch all members from the database
            with get_db_session() as db:
                members = db.query(GuildMember).filter(
                    GuildMember.guild_id == int(guild_id)
                ).order_by(GuildMember.xp.desc()).all()

                member_data = [
                    {
                        'user_id': str(member.user_id),
                        'display_name': member.display_name or f'User {member.user_id}',
                        'xp': float(member.xp),
                        'level': member.level,
                        'hero_tokens': member.hero_tokens or 0
                    }
                    for member in members
                ]

                return JsonResponse({
                    'success': True,
                    'members': member_data
                })

        elif request.method == 'POST':
            # Update members
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON'}, status=400)
            members_to_update = data.get('members', [])

            if not members_to_update:
                return JsonResponse({'error': 'No members provided'}, status=400)

            # Check tier limits and daily usage
            from .models import Guild as GuildModel

            with get_db_session() as db:
                guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
                if not guild_record:
                    return JsonResponse({'error': 'Guild not found'}, status=404)

                # Check daily bulk limit
                allowed, error_msg, usage_info = check_daily_bulk_limit(
                    guild_id, 'xp_bulk', len(members_to_update), guild_record
                )

                if not allowed:
                    return JsonResponse({
                        'error': error_msg,
                        'limit_exceeded': True,
                        'usage_info': usage_info
                    }, status=403)

            with get_db_session() as db:
                updated = 0
                for member_data in members_to_update:
                    user_id = int(member_data['user_id'])
                    member = db.query(GuildMember).filter_by(
                        guild_id=int(guild_id),
                        user_id=user_id
                    ).first()

                    if member:
                        # Update XP, Level, Hero Tokens
                        member.xp = float(member_data.get('xp', member.xp))
                        member.level = int(member_data.get('level', member.level))
                        member.hero_tokens = int(member_data.get('hero_tokens', member.hero_tokens or 0))

                        updated += 1

                db.commit()

                # Record usage for daily tracking
                record_bulk_usage(guild_id, 'xp_bulk', len(members_to_update))

                return JsonResponse({
                    'success': True,
                    'updated': updated
                })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# XP Boost Events API

@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_xp_boost_events_list(request, guild_id):
    """GET /api/guild/<id>/xp/boost-events/ - Get all boost events for a guild."""
    try:
        from .db import get_db_session
        from .models import XPBoostEvent

        with get_db_session() as db:
            events = db.query(XPBoostEvent).filter_by(
                guild_id=int(guild_id)
            ).order_by(XPBoostEvent.is_default.desc(), XPBoostEvent.name).all()

            return JsonResponse({
                'success': True,
                'events': [
                    {
                        'id': e.id,
                        'name': e.name,
                        'description': e.description,
                        'multiplier': float(e.multiplier),
                        'start_time': e.start_time,
                        'end_time': e.end_time,
                        'is_active': e.is_active,
                        'is_default': e.is_default,
                        'scope': e.scope,
                        'scope_id': str(e.scope_id) if e.scope_id else None,
                        'token_bonus': e.token_bonus,
                        'announcement_channel_id': str(e.announcement_channel_id) if e.announcement_channel_id else None,
                        'announcement_role_id': str(e.announcement_role_id) if e.announcement_role_id else None,
                        'created_at': e.created_at,
                        'updated_at': e.updated_at,
                    }
                    for e in events
                ]
            })
    except Exception as e:
        logger.error(f"Error fetching boost events for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["PATCH"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='PATCH', block=True)
def api_xp_boost_event_update(request, guild_id, event_id):
    """PATCH /api/guild/<id>/xp/boost-events/<event_id>/ - Update a boost event."""
    try:
        import json as json_lib
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import XPBoostEvent, PendingAction, ActionType

        with get_db_session() as db:
            event = db.query(XPBoostEvent).filter_by(
                id=int(event_id),
                guild_id=int(guild_id)
            ).first()

            if not event:
                return JsonResponse({'error': 'Boost event not found'}, status=404)

            # Track old is_active state for announcement triggering
            was_active = event.is_active
            should_announce = False

            # Update fields
            if 'is_active' in data:
                new_is_active = bool(data['is_active'])
                event.is_active = new_is_active
                # Queue announcement if event is being activated
                if not was_active and new_is_active:
                    should_announce = True

            if 'name' in data:
                event.name = data['name']
            if 'description' in data:
                event.description = data['description']
            if 'multiplier' in data:
                event.multiplier = float(data['multiplier'])
            if 'token_bonus' in data:
                event.token_bonus = int(data['token_bonus'])
            if 'scope' in data:
                event.scope = data['scope']
            if 'scope_id' in data:
                event.scope_id = int(data['scope_id']) if data['scope_id'] else None

            # Timing fields
            if 'start_time' in data:
                event.start_time = int(data['start_time']) if data['start_time'] else None
            if 'end_time' in data:
                event.end_time = int(data['end_time']) if data['end_time'] else None

            # Announcement fields
            if 'announcement_channel_id' in data:
                event.announcement_channel_id = int(data['announcement_channel_id']) if data['announcement_channel_id'] else None
            if 'announcement_role_id' in data:
                event.announcement_role_id = int(data['announcement_role_id']) if data['announcement_role_id'] else None

            event.updated_at = int(time.time())
            db.commit()

            # Queue announcement action if event was just activated
            if should_announce and event.announcement_channel_id:
                import json as json_module

                action_data = {
                    'event_id': event.id,
                    'event_name': event.name,
                    'multiplier': float(event.multiplier),
                    'description': event.description,
                    'channel_id': str(event.announcement_channel_id),
                    'role_id': str(event.announcement_role_id) if event.announcement_role_id else None,
                    'end_time': event.end_time,
                }

                pending_action = PendingAction(
                    guild_id=int(guild_id),
                    action_type=ActionType.BOOST_EVENT_START,
                    payload=json_module.dumps(action_data),
                    status='pending'
                )
                db.add(pending_action)
                db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Boost event updated successfully'
            })

    except Exception as e:
        logger.error(f"Error updating boost event {event_id} for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_xp_boost_event_delete(request, guild_id, event_id):
    """DELETE /api/guild/<id>/xp/boost-events/<event_id>/ - Delete a custom boost event."""
    try:
        from .db import get_db_session
        from .models import XPBoostEvent
        from .module_utils import has_module_access

        # Only allow deletion of custom events (not defaults)
        # Require Engagement Module
        if not has_module_access(guild_id, 'engagement'):
            return JsonResponse({'error': 'Engagement Module required to delete custom boost events'}, status=403)

        with get_db_session() as db:
            event = db.query(XPBoostEvent).filter_by(
                id=int(event_id),
                guild_id=int(guild_id)
            ).first()

            if not event:
                return JsonResponse({'error': 'Boost event not found'}, status=404)

            if event.is_default:
                return JsonResponse({'error': 'Cannot delete default boost events'}, status=400)

            db.delete(event)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Boost event deleted successfully'
            })

    except Exception as e:
        logger.error(f"Error deleting boost event {event_id} for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_xp_boost_event_create(request, guild_id):
    """POST /api/guild/<id>/xp/boost-events/create/ - Create a custom boost event."""
    try:
        import json as json_lib
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import XPBoostEvent
        from .module_utils import has_module_access

        # Require Engagement Module for custom events
        if not has_module_access(guild_id, 'engagement'):
            return JsonResponse({'error': 'Engagement Module required to create custom boost events'}, status=403)

        # Validate required fields
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'error': 'Event name is required'}, status=400)
        if len(name) > 255:
            return JsonResponse({'error': 'Event name must be 255 characters or less'}, status=400)

        description = data.get('description', '').strip()

        # Validate multiplier
        try:
            multiplier = float(data.get('multiplier', 2.0))
            if multiplier < 1.0 or multiplier > 10.0:
                return JsonResponse({'error': 'Multiplier must be between 1.0 and 10.0'}, status=400)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid multiplier value'}, status=400)

        # Validate scope
        scope = data.get('scope', 'server')
        if scope not in ['server', 'role', 'channel']:
            return JsonResponse({'error': 'Invalid scope. Must be server, role, or channel'}, status=400)

        # Validate scope_id if scope is role or channel
        scope_id = None
        if scope in ['role', 'channel']:
            scope_id_str = data.get('scope_id', '').strip()
            if not scope_id_str:
                return JsonResponse({'error': f'Scope ID is required when scope is {scope}'}, status=400)
            try:
                scope_id = int(scope_id_str)
            except (ValueError, TypeError):
                return JsonResponse({'error': 'Invalid scope ID'}, status=400)

        # Validate token bonus
        try:
            token_bonus = int(data.get('token_bonus', 0))
            if token_bonus < 0:
                return JsonResponse({'error': 'Token bonus cannot be negative'}, status=400)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid token bonus value'}, status=400)

        # Parse timing fields
        start_time = None
        end_time = None
        if 'start_time' in data and data['start_time']:
            try:
                start_time = int(data['start_time'])
            except (ValueError, TypeError):
                return JsonResponse({'error': 'Invalid start time'}, status=400)

        if 'end_time' in data and data['end_time']:
            try:
                end_time = int(data['end_time'])
            except (ValueError, TypeError):
                return JsonResponse({'error': 'Invalid end time'}, status=400)

        # Validate end_time is after start_time
        if start_time and end_time and end_time <= start_time:
            return JsonResponse({'error': 'End time must be after start time'}, status=400)

        # Parse announcement fields
        announcement_channel_id = None
        announcement_role_id = None
        if 'announcement_channel_id' in data and data['announcement_channel_id']:
            try:
                announcement_channel_id = int(data['announcement_channel_id'])
            except (ValueError, TypeError):
                return JsonResponse({'error': 'Invalid announcement channel ID'}, status=400)

        if 'announcement_role_id' in data and data['announcement_role_id']:
            try:
                announcement_role_id = int(data['announcement_role_id'])
            except (ValueError, TypeError):
                return JsonResponse({'error': 'Invalid announcement role ID'}, status=400)

        current_time = int(time.time())

        with get_db_session() as db:
            # Create new boost event
            new_event = XPBoostEvent(
                guild_id=int(guild_id),
                name=name,
                description=description or None,
                multiplier=multiplier,
                start_time=start_time,
                end_time=end_time,
                is_active=False,
                is_default=False,  # Custom events are not defaults
                scope=scope,
                scope_id=scope_id,
                token_bonus=token_bonus,
                announcement_channel_id=announcement_channel_id,
                announcement_role_id=announcement_role_id,
                created_at=current_time,
                updated_at=current_time
            )

            db.add(new_event)
            db.commit()
            db.refresh(new_event)

            return JsonResponse({
                'success': True,
                'message': 'Boost event created successfully',
                'event': {
                    'id': new_event.id,
                    'name': new_event.name,
                    'description': new_event.description,
                    'multiplier': float(new_event.multiplier),
                    'start_time': new_event.start_time,
                    'end_time': new_event.end_time,
                    'is_active': new_event.is_active,
                    'is_default': new_event.is_default,
                    'scope': new_event.scope,
                    'scope_id': str(new_event.scope_id) if new_event.scope_id else None,
                    'token_bonus': new_event.token_bonus,
                    'announcement_channel_id': str(new_event.announcement_channel_id) if new_event.announcement_channel_id else None,
                    'announcement_role_id': str(new_event.announcement_role_id) if new_event.announcement_role_id else None,
                }
            })

    except Exception as e:
        logger.error(f"Error creating boost event for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# Role Management Dashboard (with CSV Import - Pro/Premium)

@discord_required
def guild_roles(request, guild_id):
    """Role management page for a guild."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Fetch guild info including subscription tier
    is_premium = False
    pending_actions = []
    recent_imports = []
    guild_record = None

    try:
        from .db import get_db_session
        from .models import Guild as GuildModel, PendingAction, ActionStatus, BulkImportJob

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                is_premium = guild_record.is_premium()

            # Get pending role actions
            actions = db.query(PendingAction).filter_by(
                guild_id=int(guild_id)
            ).filter(
                PendingAction.action_type.in_(['role_add', 'role_remove', 'role_bulk_add', 'role_bulk_remove'])
            ).order_by(PendingAction.created_at.desc()).limit(10).all()

            pending_actions = [
                {
                    'id': a.id,
                    'action_type': a.action_type.value,
                    'status': a.status.value,
                    'created_at': a.created_at,
                    'triggered_by_name': a.triggered_by_name,
                }
                for a in actions
            ]

            # Get recent bulk imports
            imports = db.query(BulkImportJob).filter_by(
                guild_id=int(guild_id),
                job_type='role_assign'
            ).order_by(BulkImportJob.created_at.desc()).limit(5).all()

            recent_imports = [
                {
                    'id': i.id,
                    'filename': i.filename,
                    'status': i.status,
                    'total_records': i.total_records,
                    'success_count': i.success_count,
                    'error_count': i.error_count,
                    'created_at': i.created_at,
                }
                for i in imports
            ]

    except Exception as e:
        logger.warning(f"Could not fetch role data: {e}")

    # Check if guild has roles module access
    has_roles_module = has_module_access(guild_id, 'roles')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'is_premium': is_premium,
        'pending_actions': pending_actions,
        'recent_imports': recent_imports,
        'has_roles_module': has_roles_module,
        'has_any_module': has_any_module,
        'active_page': 'roles',
    }
    return render(request, 'questlog/roles.html', context)


def _serialize_reaction_role_menus(db, guild_id, message_id=None):
    """Build reaction role menu payloads grouped by message_id."""
    from .models import ReactRole

    query = db.query(ReactRole).filter_by(guild_id=int(guild_id))
    if message_id is not None:
        query = query.filter_by(message_id=int(message_id))

    roles = query.all()
    menus = {}

    for rr in roles:
        menu = menus.get(rr.message_id)
        if not menu:
            menu = {
                'id': rr.message_id,
                'message_id': str(rr.message_id),
                'channel_id': str(rr.channel_id),
                'title': None,
                'description': None,
                'enabled': True,
                'roles': []
            }
            menus[rr.message_id] = menu

        menu['roles'].append({
            'id': rr.id,
            'emoji': rr.emoji,
            'role_id': str(rr.role_id),
            'role_name': rr.role_name or '',
            'remove_on_unreact': bool(rr.remove_on_unreact),
            'exclusive_group': rr.exclusive_group or ''
        })

    results = []
    for menu in menus.values():
        menu['role_count'] = len(menu['roles'])
        results.append(menu)

    # Try to hydrate title/description from the Discord message embed
    bot_session = get_bot_session(int(guild_id))
    if bot_session:
        for menu in results:
            try:
                resp = bot_session.get(f"/channels/{menu['channel_id']}/messages/{menu['message_id']}")
                if resp.status_code == 200:
                    msg_json = resp.json()
                    embeds = msg_json.get('embeds', [])
                    if embeds:
                        embed = embeds[0]
                        menu['title'] = embed.get('title') or menu['title'] or f"Menu #{menu['message_id']}"
                        menu['description'] = embed.get('description') or menu['description'] or ''
                    else:
                        menu['title'] = menu['title'] or f"Menu #{menu['message_id']}"
                        menu['description'] = menu['description'] or ''
                else:
                    menu['title'] = menu['title'] or f"Menu #{menu['message_id']}"
                    menu['description'] = menu['description'] or ''
            except Exception as e:
                logger.warning(f"Could not fetch reaction role message {menu['message_id']}: {e}")
                menu['title'] = menu['title'] or f"Menu #{menu['message_id']}"
                menu['description'] = menu['description'] or ''
    else:
        for menu in results:
            menu['title'] = menu['title'] or f"Menu #{menu['message_id']}"
            menu['description'] = menu['description'] or ''

    # If a specific menu was requested, return it directly
    if message_id is not None:
        return results[0] if results else None

    return results


@discord_required
def guild_reaction_roles(request, guild_id):
    """Reaction roles management page for a guild."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Fetch reaction role menus from database
    reaction_menus = []
    guild_record = None

    try:
        from .db import get_db_session
        from .models import Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            reaction_menus = _serialize_reaction_role_menus(db, guild_id)
    except Exception as e:
        logger.warning(f"Could not fetch reaction role menus: {e}")

    # Check if guild has roles module access
    has_roles_module = has_module_access(guild_id, 'roles')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'reaction_menus': reaction_menus,
        'has_roles_module': has_roles_module,
        'has_any_module': has_any_module,
        'active_page': 'reaction_roles',
    }
    return render(request, 'questlog/reaction_roles.html', context)


@require_http_methods(["GET", "POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_reaction_roles(request, guild_id):
    """
    GET /api/guild/<id>/reaction-roles/ - List reaction role menus
    POST /api/guild/<id>/reaction-roles/ - Create and deploy a new reaction role menu
    """
    try:
        from .db import get_db_session
        from .models import ReactRole
        from urllib.parse import quote

        guild_id_int = int(guild_id)

        if request.method == "GET":
            with get_db_session() as db:
                menus = _serialize_reaction_role_menus(db, guild_id)
                return JsonResponse({'success': True, 'menus': menus})

        # POST - create menu and send message
        try:
            data = json_lib.loads(request.body)
        except json_lib.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        channel_id = data.get('channel_id')
        roles_data = data.get('roles', [])
        title = data.get('title') or 'Choose your roles'
        description = data.get('description') or ''

        # Bot session required to post message and add reactions
        bot_session = get_bot_session(guild_id_int)
        if not bot_session:
            return JsonResponse({'error': 'Bot not connected to this guild'}, status=400)

        try:
            channel_id_int = int(channel_id)
        except (TypeError, ValueError):
            return JsonResponse({'error': 'channel_id is required'}, status=400)

        if not isinstance(roles_data, list) or len(roles_data) == 0:
            return JsonResponse({'error': 'At least one role entry is required'}, status=400)

        parsed_roles = []
        seen_emojis = set()
        for item in roles_data:
            emoji = (item.get('emoji') or '').strip()
            role_id = item.get('role_id')
            if not emoji or not role_id:
                return JsonResponse({'error': 'Each entry needs emoji and role_id'}, status=400)
            if emoji in seen_emojis:
                return JsonResponse({'error': f'Duplicate emoji "{emoji}" in menu'}, status=400)
            seen_emojis.add(emoji)

            try:
                role_id_int = int(role_id)
            except (TypeError, ValueError):
                return JsonResponse({'error': 'role_id must be numeric'}, status=400)

            parsed_roles.append({
                'emoji': emoji,
                'role_id': role_id_int,
                'role_name': item.get('role_name') or '',
                'remove_on_unreact': bool(item.get('remove_on_unreact', True)),
                'exclusive_group': (item.get('exclusive_group') or '').strip() or None,
            })

        # Build embed description
        lines = []
        if description:
            lines.append(description)
            lines.append('')
        for entry in parsed_roles:
            role_display = entry['role_name'] or f"Role {entry['role_id']}"
            flags = []
            if entry['exclusive_group']:
                flags.append(f"Group: {entry['exclusive_group']}")
            if not entry['remove_on_unreact']:
                flags.append("Keeps role on unreact")
            flags_text = f" ({'; '.join(flags)})" if flags else ''
            lines.append(f"{entry['emoji']} — {role_display}{flags_text}")

        embed_description = "\n".join(lines) if lines else "Choose your roles below."

        # Send message via bot
        message_payload = {
            'content': '',
            'embeds': [{
                'title': title,
                'description': embed_description,
                'color': 0xF6C454
            }]
        }

        msg_resp = bot_session.post(f'/channels/{channel_id_int}/messages', json=message_payload)
        if msg_resp.status_code not in (200, 201):
            logger.error(f"Failed to send reaction role message: {msg_resp.status_code} {msg_resp.text}")
            return JsonResponse({'error': 'Failed to send message to channel'}, status=500)

        msg_json = msg_resp.json()
        try:
            message_id_int = int(msg_json.get('id'))
        except (TypeError, ValueError):
            logger.error(f"Reaction role message response missing ID: {msg_json}")
            return JsonResponse({'error': 'Failed to send message to channel'}, status=500)

        # Add reactions
        for entry in parsed_roles:
            try:
                emoji_encoded = quote(entry['emoji'])
                react_resp = bot_session.put(f'/channels/{channel_id_int}/messages/{message_id_int}/reactions/{emoji_encoded}/@me')
                if react_resp.status_code not in (200, 204):
                    logger.warning(f"Failed to add reaction {entry['emoji']} to message {message_id_int}: {react_resp.status_code} {react_resp.text}")
            except Exception as react_err:
                logger.warning(f"Error adding reaction {entry['emoji']} to message {message_id_int}: {react_err}")

        with get_db_session() as db:
            for entry in parsed_roles:
                db.add(ReactRole(
                    guild_id=guild_id_int,
                    message_id=message_id_int,
                    channel_id=channel_id_int,
                    emoji=entry['emoji'],
                    role_id=entry['role_id'],
                    role_name=entry['role_name'],
                    remove_on_unreact=entry['remove_on_unreact'],
                    exclusive_group=entry['exclusive_group']
                ))

            db.flush()
            menu = _serialize_reaction_role_menus(db, guild_id, message_id_int)

            return JsonResponse({'success': True, 'menu': menu})

    except Exception as e:
        logger.error(f"Error handling reaction roles: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET", "PUT", "DELETE"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method=['PUT', 'DELETE'], block=True)
def api_reaction_role_detail(request, guild_id, message_id):
    """
    GET /api/guild/<id>/reaction-roles/<message_id>/ - Get a menu
    PUT /api/guild/<id>/reaction-roles/<message_id>/ - Update menu roles and message
    DELETE /api/guild/<id>/reaction-roles/<message_id>/ - Delete menu (and message)
    """
    try:
        from .db import get_db_session
        from .models import ReactRole
        from urllib.parse import quote

        guild_id_int = int(guild_id)

        try:
            message_id_int = int(message_id)
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Invalid message_id'}, status=400)

        if request.method == "GET":
            with get_db_session() as db:
                menu = _serialize_reaction_role_menus(db, guild_id, message_id_int)
                if not menu:
                    return JsonResponse({'error': 'Menu not found'}, status=404)
                return JsonResponse({'success': True, 'menu': menu})

        if request.method == "DELETE":
            with get_db_session() as db:
                existing = db.query(ReactRole).filter_by(
                    guild_id=guild_id_int,
                    message_id=message_id_int
                ).all()
                channel_id = existing[0].channel_id if existing else None
                deleted = db.query(ReactRole).filter_by(
                    guild_id=guild_id_int,
                    message_id=message_id_int
                ).delete()

            if deleted == 0:
                return JsonResponse({'error': 'Menu not found'}, status=404)

            # Attempt to delete the Discord message too
            if channel_id:
                bot_session = get_bot_session(guild_id_int)
                if bot_session:
                    bot_session.delete(f'/channels/{channel_id}/messages/{message_id_int}')

            return JsonResponse({'success': True})

        # PUT - update menu
        try:
            data = json_lib.loads(request.body)
        except json_lib.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        channel_id = data.get('channel_id')
        roles_data = data.get('roles', [])
        title = data.get('title') or 'Choose your roles'
        description = data.get('description') or ''

        try:
            channel_id_int = int(channel_id)
        except (TypeError, ValueError):
            return JsonResponse({'error': 'channel_id is required'}, status=400)

        if not isinstance(roles_data, list) or len(roles_data) == 0:
            return JsonResponse({'error': 'At least one role entry is required'}, status=400)

        # Normalize roles: keep last occurrence per role, then last per emoji
        entries = []
        for item in roles_data:
            emoji = (item.get('emoji') or '').strip()
            role_id = item.get('role_id')
            if not emoji or not role_id:
                return JsonResponse({'error': 'Each entry needs emoji and role_id'}, status=400)
            try:
                role_id_int = int(role_id)
            except (TypeError, ValueError):
                return JsonResponse({'error': 'role_id must be numeric'}, status=400)
            entries.append({
                'emoji': emoji,
                'role_id': role_id_int,
                'role_name': item.get('role_name') or '',
                'remove_on_unreact': bool(item.get('remove_on_unreact', True)),
                'exclusive_group': (item.get('exclusive_group') or '').strip() or None,
            })

        # Last entry per role_id wins
        by_role = {}
        for entry in entries:
            by_role[entry['role_id']] = entry

        # Preserve only the final value per role
        parsed_roles = list(by_role.values())

        bot_session = get_bot_session(guild_id_int)
        if not bot_session:
            return JsonResponse({'error': 'Bot not connected to this guild'}, status=400)

        # Build embed description
        lines = []
        if description:
            lines.append(description)
            lines.append('')
        for entry in parsed_roles:
            role_display = entry['role_name'] or f"Role {entry['role_id']}"
            flags = []
            if entry['exclusive_group']:
                flags.append(f"Group: {entry['exclusive_group']}")
            if not entry['remove_on_unreact']:
                flags.append("Keeps role on unreact")
            flags_text = f" ({'; '.join(flags)})" if flags else ''
            lines.append(f"{entry['emoji']} — {role_display}{flags_text}")

        embed_description = "\n".join(lines) if lines else "Choose your roles below."

        logger.info(f"[ReactionRole PUT] Updating message {message_id_int}: {len(parsed_roles)} roles, new emojis: {[e['emoji'] for e in parsed_roles]}")

        with get_db_session() as db:
            existing = db.query(ReactRole).filter_by(
                guild_id=guild_id_int,
                message_id=message_id_int
            ).all()
            if not existing:
                return JsonResponse({'error': 'Menu not found'}, status=404)

            # Get old emojis so we can remove them from the message
            old_emojis = {record.emoji for record in existing}
            logger.info(f"[ReactionRole PUT] Old emojis: {old_emojis}")

            # Delete existing rows for this menu
            db.query(ReactRole).filter_by(
                guild_id=guild_id_int,
                message_id=message_id_int
            ).delete()
            db.commit()

        # Update the existing Discord message (PATCH instead of delete/recreate)
        message_payload = {
            'embeds': [{
                'title': title,
                'description': embed_description,
                'color': 0xF6C454
            }]
        }
        logger.info(f"[ReactionRole PUT] Sending PATCH to Discord: {message_payload}")

        try:
            patch_resp = bot_session.patch(
                f'/channels/{channel_id_int}/messages/{message_id_int}',
                json=message_payload
            )
            logger.info(f"[ReactionRole PUT] Discord PATCH response: {patch_resp.status_code}")
            if patch_resp.status_code not in (200, 201):
                logger.error(f"Failed to update message {message_id_int}: {patch_resp.status_code} {patch_resp.text}")
                return JsonResponse({'error': 'Failed to update message on Discord'}, status=500)
        except Exception as patch_err:
            logger.error(f"Error updating reaction-role message: {patch_err}", exc_info=True)
            return JsonResponse({'error': 'Failed to update message on Discord'}, status=500)

        # Remove old reactions that are no longer in the new config
        new_emojis = {entry['emoji'] for entry in parsed_roles}
        emojis_to_remove = old_emojis - new_emojis
        logger.info(f"[ReactionRole PUT] Removing old emojis: {emojis_to_remove}, keeping: {new_emojis}")

        for emoji in emojis_to_remove:
            try:
                emoji_encoded = quote(emoji)
                logger.info(f"[ReactionRole PUT] Deleting reaction {emoji} (encoded: {emoji_encoded})")
                del_resp = bot_session.delete(f'/channels/{channel_id_int}/messages/{message_id_int}/reactions/{emoji_encoded}')
                logger.info(f"[ReactionRole PUT] Delete reaction response: {del_resp.status_code}")
            except Exception as remove_err:
                logger.warning(f"Error removing old reaction {emoji} from message {message_id_int}: {remove_err}")

        # Add new reactions (Discord API is idempotent, so re-adding existing ones is safe)
        for entry in parsed_roles:
            try:
                emoji_encoded = quote(entry['emoji'])
                logger.info(f"[ReactionRole PUT] Adding reaction {entry['emoji']} (encoded: {emoji_encoded})")
                react_resp = bot_session.put(f'/channels/{channel_id_int}/messages/{message_id_int}/reactions/{emoji_encoded}/@me')
                logger.info(f"[ReactionRole PUT] Add reaction response: {react_resp.status_code}")
            except Exception as react_err:
                logger.warning(f"Error adding reaction {entry['emoji']} to message {message_id_int}: {react_err}")

        # Persist new rows with same message_id
        with get_db_session() as db:
            for entry in parsed_roles:
                db.add(ReactRole(
                    guild_id=guild_id_int,
                    message_id=message_id_int,
                    channel_id=channel_id_int,
                    emoji=entry['emoji'],
                    role_id=entry['role_id'],
                    role_name=entry['role_name'],
                    remove_on_unreact=entry['remove_on_unreact'],
                    exclusive_group=entry['exclusive_group']
                ))
            db.commit()
            menu = _serialize_reaction_role_menus(db, guild_id, message_id_int)

        return JsonResponse({'success': True, 'menu': menu})

    except Exception as e:
        logger.error(f"Error handling reaction role menu {message_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='20/m', method='POST', block=True)
def api_role_action(request, guild_id):
    """POST /api/guild/<id>/roles/action/ - Queue a role action."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    action = data.get('action')  # 'add' or 'remove'
    user_id = data.get('user_id')
    role_id = data.get('role_id')
    reason = data.get('reason', f'Role {action} via web dashboard')

    if not all([action, user_id, role_id]):
        return JsonResponse({'error': 'Missing required fields'}, status=400)

    try:
        from .actions import queue_action, ActionType

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        # Prepare payload
        payload = {
            'user_id': int(user_id),
            'role_id': int(role_id),
            'reason': reason
        }

        if action == 'add':
            action_id = queue_action(
                guild_id=int(guild_id),
                action_type=ActionType.ROLE_ADD,
                payload=payload,
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name,
                source='website'
            )
        elif action == 'remove':
            action_id = queue_action(
                guild_id=int(guild_id),
                action_type=ActionType.ROLE_REMOVE,
                payload=payload,
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name,
                source='website'
            )
        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)

        return JsonResponse({
            'success': True,
            'action_id': action_id,
            'message': f'Role {action} queued (ID: {action_id})'
        })

    except Exception as e:
        logger.error(f"Error queuing role action: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@require_subscription_tier('pro', 'premium')
@ratelimit(key='user_or_ip', rate='5/h', method='POST', block=True)
def api_role_bulk_import(request, guild_id):
    """POST /api/guild/<id>/roles/import/ - Import XLSX for bulk role assignment (Pro/Premium only, 5 req/hour)."""
    # Check subscription tier and determine limits
    try:
        from .db import get_db_session
        from .models import Guild as GuildModel, SubscriptionTier

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Determine import limit based on tier
            # Free tier blocked by @require_subscription_tier decorator
            # VIP and Premium: Unlimited
            # Pro: 10 users
            if guild_record.is_vip or guild_record.subscription_tier == SubscriptionTier.PREMIUM.value:
                max_users = None  # Unlimited
            else:  # Pro tier
                max_users = 10
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)

    # Get uploaded file
    excel_file = request.FILES.get('file')
    if not excel_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    # Validate file size (5MB limit)
    MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
    if excel_file.size > MAX_FILE_SIZE:
        return JsonResponse({'error': 'File too large (max 5MB)'}, status=400)

    # Validate file extension (case-insensitive, prevent mixed-case tricks)
    if not excel_file.name.lower().endswith('.xlsx'):
        return JsonResponse({'error': 'Invalid file type. Only .xlsx files allowed'}, status=400)

    # SECURITY: Validate content-type header to prevent extension spoofing
    content_type = excel_file.content_type.lower() if excel_file.content_type else ''
    allowed_content_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/octet-stream',  # Sometimes browsers send this for Excel files
    ]
    if content_type and content_type not in allowed_content_types:
        logger.warning(f"[ROLE IMPORT] Rejected file with invalid content-type: {content_type} (filename: {excel_file.name})")
        return JsonResponse({'error': f'Invalid file type. Content-type {content_type} not allowed'}, status=400)

    role_id = request.POST.get('role_id')
    action = request.POST.get('action', 'add')  # 'add' or 'remove'

    if not role_id:
        return JsonResponse({'error': 'Missing role_id'}, status=400)

    try:
        from openpyxl import load_workbook
        from io import BytesIO
        from .actions import (
            queue_bulk_role_add, create_bulk_import_job, update_bulk_import_progress
        )

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        # Read XLSX file
        wb = load_workbook(BytesIO(excel_file.read()), read_only=True, data_only=True)
        ws = wb.active

        # Get header row
        headers = [cell.value for cell in ws[1]]

        # Find user_id column (flexible column names)
        user_id_col = None
        for idx, header in enumerate(headers):
            if header and str(header).lower() in ['user_id', 'discord_id', 'id']:
                user_id_col = idx
                break

        if user_id_col is None:
            return JsonResponse({
                'error': 'No user_id column found. Expected columns: user_id, discord_id, or id'
            }, status=400)

        user_ids = []
        errors = []

        # Read data rows (skip header)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row_num > 10001:  # Limit to 10,000 rows
                break

            user_id = row[user_id_col] if user_id_col < len(row) else None

            if user_id:
                # Convert to string first, then to int (handles TEXT formatted cells)
                user_id_str = str(user_id).strip()
                try:
                    user_ids.append(int(user_id_str))
                except ValueError:
                    errors.append({'row': row_num, 'error': f'Invalid user_id: {user_id_str}'})
            else:
                errors.append({'row': row_num, 'error': 'Missing user_id'})

        wb.close()

        if not user_ids:
            return JsonResponse({
                'error': 'No valid user IDs found in XLSX',
                'parse_errors': errors
            }, status=400)

        # Check daily bulk limit (replaces simple per-operation check)
        allowed, error_msg, usage_info = check_daily_bulk_limit(
            guild_id, 'role_bulk', len(user_ids), guild_record
        )

        if not allowed:
            return JsonResponse({
                'error': error_msg,
                'limit_exceeded': True,
                'usage_info': usage_info
            }, status=403)

        # Create bulk import job for tracking
        job_id = create_bulk_import_job(
            guild_id=int(guild_id),
            job_type='role_assign',
            filename=excel_file.name,
            total_records=len(user_ids),
            triggered_by=triggered_by,
            triggered_by_name=triggered_by_name
        )

        # Queue the bulk action
        if action == 'add':
            action_id = queue_bulk_role_add(
                guild_id=int(guild_id),
                role_id=int(role_id),
                user_ids=user_ids,
                reason=f'Bulk import from {excel_file.name}',
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name
            )
        else:
            from .actions import queue_action, ActionType
            action_id = queue_action(
                guild_id=int(guild_id),
                action_type=ActionType.ROLE_BULK_REMOVE,
                payload={'role_id': int(role_id), 'user_ids': user_ids},
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name,
                source='xlsx_import'
            )

        # Record usage for daily tracking
        record_bulk_usage(guild_id, 'role_bulk', len(user_ids))

        return JsonResponse({
            'success': True,
            'job_id': job_id,
            'action_id': action_id,
            'total_users': len(user_ids),
            'parse_errors': errors[:10] if errors else [],  # Return first 10 errors
            'message': f'Queued bulk {action} for {len(user_ids)} users'
        })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_bulk_import_status(request, guild_id, job_id):
    """GET /api/guild/<id>/roles/import/<job_id>/ - Get bulk import status."""
    try:
        from .actions import get_bulk_import_job

        job = get_bulk_import_job(int(job_id))
        if not job or job['guild_id'] != int(guild_id):
            return JsonResponse({'error': 'Job not found'}, status=404)

        return JsonResponse({'success': True, 'job': job})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_role_export_template(request, guild_id):
    """GET /api/guild/<id>/roles/export-template/ - Export blank XLSX template for bulk role import."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        from django.http import HttpResponse

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Role Import Template"

        # Add headers with formatting
        headers = ['user_id']
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

        # Add a few example rows with instructions
        ws.append(['1234567890123456'])  # Example Discord ID
        ws.append(['9876543210987654'])  # Example Discord ID

        # Format user_id column as TEXT (prevents scientific notation)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=1).number_format = '@'

        # Set column width
        ws.column_dimensions['A'].width = 20

        # Add instructions sheet
        ws_info = wb.create_sheet(title="Instructions")
        ws_info['A1'] = "Bulk Role Import Instructions"
        ws_info['A1'].font = Font(bold=True, size=14)

        instructions = [
            "",
            "How to use this template:",
            "1. Fill in the 'user_id' column with Discord user IDs",
            "2. Delete the example rows",
            "3. Save the file",
            "4. Upload it via the Bulk Import button",
            "5. Select the role to assign/remove",
            "",
            "Tips:",
            "- User IDs are 17-19 digit numbers",
            "- You can export current server members from Discord",
            "- The user_id column is pre-formatted as TEXT to prevent Excel issues",
            "- Maximum 10,000 rows per import",
        ]

        for i, instruction in enumerate(instructions, start=2):
            ws_info[f'A{i}'] = instruction

        ws_info.column_dimensions['A'].width = 70

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return as download
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="role_import_template.xlsx"'
        return response

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_role_export_current(request, guild_id):
    """GET /api/guild/<id>/roles/export-roles/ - Export all guild roles with their permissions."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from django.http import HttpResponse
        from datetime import datetime

        # Get bot session
        guild_id_int = int(guild_id)
        bot_session = get_bot_session(guild_id_int)
        if not bot_session:
            return JsonResponse({'error': 'Bot not connected to this guild'}, status=400)

        # Fetch guild roles
        roles_resp = bot_session.get(f'/guilds/{guild_id_int}/roles')
        if roles_resp.status_code != 200:
            logger.error(f"Failed to fetch roles: {roles_resp.status_code} {roles_resp.text}")
            return JsonResponse({'error': 'Failed to fetch roles from Discord'}, status=500)

        roles_data = roles_resp.json()

        # Discord permission flags (all major permissions)
        permission_flags = [
            ('create_instant_invite', 0x1),
            ('kick_members', 0x2),
            ('ban_members', 0x4),
            ('administrator', 0x8),
            ('manage_channels', 0x10),
            ('manage_guild', 0x20),
            ('add_reactions', 0x40),
            ('view_audit_log', 0x80),
            ('priority_speaker', 0x100),
            ('stream', 0x200),
            ('view_channel', 0x400),
            ('send_messages', 0x800),
            ('send_tts_messages', 0x1000),
            ('manage_messages', 0x2000),
            ('embed_links', 0x4000),
            ('attach_files', 0x8000),
            ('read_message_history', 0x10000),
            ('mention_everyone', 0x20000),
            ('use_external_emojis', 0x40000),
            ('view_guild_insights', 0x80000),
            ('connect', 0x100000),
            ('speak', 0x200000),
            ('mute_members', 0x400000),
            ('deafen_members', 0x800000),
            ('move_members', 0x1000000),
            ('use_vad', 0x2000000),
            ('change_nickname', 0x4000000),
            ('manage_nicknames', 0x8000000),
            ('manage_roles', 0x10000000),
            ('manage_webhooks', 0x20000000),
            ('manage_expressions', 0x40000000),
            ('use_application_commands', 0x80000000),
            ('request_to_speak', 0x100000000),
            ('manage_events', 0x200000000),
            ('manage_threads', 0x400000000),
            ('create_public_threads', 0x800000000),
            ('create_private_threads', 0x1000000000),
            ('use_external_stickers', 0x2000000000),
            ('send_messages_in_threads', 0x4000000000),
            ('use_embedded_activities', 0x8000000000),
            ('moderate_members', 0x10000000000),
        ]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Role Definitions"

        # Build headers: basic info + all permissions
        headers = ['role_id', 'name', 'color', 'position', 'hoist', 'mentionable', 'permissions_value']
        headers.extend([perm[0] for perm in permission_flags])
        ws.append(headers)

        # Style header row
        for idx, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add role data
        for role in sorted(roles_data, key=lambda r: r.get('position', 0), reverse=True):
            permissions_int = int(role.get('permissions', 0))

            # Build row with basic info
            row_data = [
                str(role.get('id', '')),
                str(role.get('name', '')),
                str(role.get('color', '0')),
                str(role.get('position', '0')),
                'TRUE' if role.get('hoist', False) else 'FALSE',
                'TRUE' if role.get('mentionable', False) else 'FALSE',
                str(permissions_int)
            ]

            # Add permission flags
            for perm_name, perm_value in permission_flags:
                has_permission = bool(permissions_int & perm_value)
                row_data.append('TRUE' if has_permission else 'FALSE')

            ws.append(row_data)

        # Format all cells as TEXT
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '@'

        # Set column widths
        ws.column_dimensions['A'].width = 20  # role_id
        ws.column_dimensions['B'].width = 25  # name
        ws.column_dimensions['C'].width = 12  # color
        ws.column_dimensions['D'].width = 10  # position
        ws.column_dimensions['E'].width = 10  # hoist
        ws.column_dimensions['F'].width = 12  # mentionable
        ws.column_dimensions['G'].width = 20  # permissions_value

        # Set permission column widths
        for col_idx in range(8, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 18

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return as download
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="guild_{guild_id}_role_definitions_{timestamp}.xlsx"'
        return response

    except Exception as e:
        logger.error(f"Role export error: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_role_create(request, guild_id):
    """POST /api/guild/<id>/roles/create/ - Create a new Discord role."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    required = ['name']
    for field in required:
        if field not in data:
            return JsonResponse({'error': f'Missing required field: {field}'}, status=400)

    try:
        # Get bot session
        guild_id_int = int(guild_id)
        bot_session = get_bot_session(guild_id_int)
        if not bot_session:
            return JsonResponse({'error': 'Bot not connected to this guild'}, status=400)

        # Prepare role data for Discord API
        role_data = {
            'name': data['name'],
            'permissions': str(data.get('permissions', '0')),
            'color': int(data.get('color', 0)),
            'hoist': bool(data.get('hoist', False)),
            'mentionable': bool(data.get('mentionable', False)),
        }

        # Create role via Discord API
        resp = bot_session.post(f'/guilds/{guild_id_int}/roles', json=role_data)

        if resp.status_code == 200:
            role = resp.json()

            # If a position was requested, attempt to move the role using Discord's bulk role move
            desired_position = data.get('position')
            if desired_position is not None:
                try:
                    desired_position = int(desired_position)
                    patch_resp = bot_session.patch(
                        f'/guilds/{guild_id_int}/roles',
                        json=[{'id': role['id'], 'position': desired_position}]
                    )
                    if patch_resp.status_code != 200:
                        logger.warning(
                            f"Failed to set role position for role {role['id']} in guild {guild_id_int}: "
                            f"{patch_resp.status_code} {patch_resp.text}"
                        )
                except Exception as e:
                    logger.warning(f"Error adjusting role position in guild {guild_id_int}: {e}", exc_info=True)

            return JsonResponse({
                'success': True,
                'message': f'Role "{data["name"]}" created successfully!',
                'role': {
                    'id': role['id'],
                    'name': role['name'],
                    'color': role['color'],
                }
            })
        else:
            error_data = resp.json() if resp.headers.get('content-type') == 'application/json' else {}
            return JsonResponse({
                'error': error_data.get('message', f'Discord API error: {resp.status_code}')
            }, status=500)

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


def _ensure_admin(guild_id, request):
    all_admin_guilds = request.session.get('discord_admin_guilds', [])
    return any(str(g['id']) == str(guild_id) for g in all_admin_guilds)


@login_required
def guild_raffles(request, guild_id):
    """Render raffles dashboard."""
    from .db import get_db_session
    from .module_utils import has_module_access, has_any_module_access

    admin_guilds = request.session.get('discord_admin_guilds', [])
    guild = next((g for g in admin_guilds if str(g.get('id')) == str(guild_id)), None)
    if not guild:
        messages.error(request, "You don't have permission to manage raffles.")
        return redirect('questlog_dashboard')

    # Require Admin or Manage Server
    permissions = int(guild.get('permissions', 0))
    is_admin = (permissions & 0x8) == 0x8 or (permissions & 0x20) == 0x20
    if not is_admin:
        messages.error(request, "You need Admin or Manage Server permission.")
        return redirect('questlog_dashboard')

    is_vip = False
    subscription_tier = 'free'
    guild_record = None
    try:
        with get_db_session() as db:
            from .models import GuildMember, Guild as GuildModel
            member = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=int(request.session.get('discord_user', {}).get('id', 0))
            ).first()
            current_tokens = member.hero_tokens if member else 0

            # Get subscription tier info
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                is_vip = guild_record.is_vip
                subscription_tier = guild_record.subscription_tier if guild_record.subscription_tier else 'free'
                billing_cycle = guild_record.billing_cycle
    except Exception:
        current_tokens = 0
        billing_cycle = None

    # Check if guild has engagement module access
    has_engagement_module = has_module_access(guild_id, 'engagement')
    has_any_module = has_any_module_access(guild_id)

    return render(request, 'questlog/raffles.html', {
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'discord_user': request.session.get('discord_user', {}),
        'is_admin': True,
        'current_tokens': current_tokens,
        'is_vip': is_vip,
        'subscription_tier': subscription_tier,
        'billing_cycle': billing_cycle if 'billing_cycle' in locals() else None,
        'has_engagement_module': has_engagement_module,
        'has_any_module': has_any_module,
        'active_page': 'raffles',
    })


@discord_required
def guild_raffle_browser(request, guild_id):
    """Member-facing raffle browser - view and enter raffles."""
    from .db import get_db_session

    # Get all guilds the user is in (both admin and member)
    admin_guilds = request.session.get('discord_admin_guilds', [])
    member_guilds = get_member_guilds(request)
    all_guilds_session = request.session.get('discord_all_guilds', [])
    all_guilds = admin_guilds + member_guilds

    # Check if user is in this guild (get from admin_guilds first for owner/permissions)
    guild = get_guild_with_permissions(guild_id, admin_guilds, all_guilds_session)
    if not guild:
        messages.error(request, "You are not a member of this server.")
        return redirect('questlog_dashboard')

    # Get user's token balance
    current_tokens = 0
    try:
        with get_db_session() as db:
            from .models import GuildMember
            member = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=int(request.session.get('discord_user', {}).get('id', 0))
            ).first()
            current_tokens = member.hero_tokens if member else 0
    except Exception:
        current_tokens = 0

    # Check if user is admin for conditional UI elements
    is_admin = any(str(g.get('id')) == str(guild_id) for g in admin_guilds)

    return render(request, 'questlog/raffle_browser.html', {
        'guild': guild,
        'admin_guilds': admin_guilds,
        'member_guilds': member_guilds,
        'discord_user': request.session.get('discord_user', {}),
        'is_admin': is_admin,
        'current_tokens': current_tokens,
        'active_page': 'raffle_browser',
    })


def _raffle_status(raffle):
    now = int(time.time())
    if raffle.winners:
        return 'completed'
    if not raffle.active:
        return 'closed'
    # Only mark as ended if auto_pick is enabled (bot will handle it)
    # If auto_pick is off, keep it active so admins can manually pick winners
    if raffle.end_at and now > raffle.end_at and raffle.auto_pick:
        return 'ended'
    if raffle.start_at and now < raffle.start_at:
        return 'scheduled'
    return 'active'


def _draw_winners(db, raffle):
    import random
    from .models import RaffleEntry
    entries = db.query(RaffleEntry).filter_by(raffle_id=raffle.id).all()
    if not entries:
        raffle.winners = json_lib.dumps([])
        raffle.active = False
        return []

    population = []
    for e in entries:
        population.append({'user_id': e.user_id, 'username': e.username, 'weight': max(1, e.tickets)})

    # weighted sampling without replacement
    winners = []
    remaining = population[:]
    total_winners = max(1, raffle.max_winners or 1)
    rng = random.SystemRandom()

    for _ in range(total_winners):
        if not remaining:
            break
        total_weight = sum(item['weight'] for item in remaining)
        pick = rng.uniform(0, total_weight)
        cumulative = 0
        chosen = None
        for item in remaining:
            cumulative += item['weight']
            if pick <= cumulative:
                chosen = item
                break
        if chosen:
            winners.append({'user_id': chosen['user_id'], 'username': chosen.get('username')})
            remaining = [r for r in remaining if r['user_id'] != chosen['user_id']]

    raffle.winners = json_lib.dumps(winners)
    raffle.active = False
    return winners


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user', rate='10/h', method='POST', block=True)
def api_role_bulk_create(request, guild_id):
    """POST /api/guild/<id>/roles/bulk-create/ - Create multiple roles from XLSX import."""
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    upload_file = request.FILES['file']

    # Validate file size (5MB limit for role imports)
    MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
    if upload_file.size > MAX_FILE_SIZE:
        return JsonResponse({'error': 'File too large (max 5MB)'}, status=400)

    # Validate file extension (case-insensitive, prevent mixed-case tricks)
    if not upload_file.name.lower().endswith('.xlsx'):
        return JsonResponse({'error': 'Invalid file type. Only .xlsx files allowed'}, status=400)

    # SECURITY: Validate content-type header to prevent extension spoofing
    content_type = upload_file.content_type.lower() if upload_file.content_type else ''
    allowed_content_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/octet-stream',  # Sometimes browsers send this for Excel files
    ]
    if content_type and content_type not in allowed_content_types:
        logger.warning(f"[ROLE CREATE] Rejected file with invalid content-type: {content_type} (filename: {upload_file.name})")
        return JsonResponse({'error': f'Invalid file type. Content-type {content_type} not allowed'}, status=400)

    # Check subscription tier and determine limits
    try:
        from .db import get_db_session
        from .models import Guild as GuildModel, SubscriptionTier

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Determine role creation limit based on tier
            # VIP and Premium: Unlimited (cap at 100 for safety)
            # Pro: 10 roles
            # Free: 6 roles
            if guild_record.is_vip or guild_record.subscription_tier == SubscriptionTier.PREMIUM.value:
                max_roles = 100  # Safety limit
            elif guild_record.subscription_tier == SubscriptionTier.PRO.value:
                max_roles = 10
            else:  # Free tier
                max_roles = 6
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_raffle_list(request, guild_id):
    """List raffles; auto-finalize ended raffles with auto_pick. (Members can view)"""
    from .db import get_db_session
    from .models import Raffle, RaffleEntry

    # Check if user is logged in (but don't require admin access)
    discord_user = request.session.get('discord_user')
    if not discord_user:
        return JsonResponse({'error': 'Not authenticated'}, status=401)

    now = int(time.time())

    try:
        with get_db_session() as db:
            # auto finalize ended raffles that requested auto_pick and have no winners
            auto_raffles = db.query(Raffle).filter(
                Raffle.guild_id == int(guild_id),
                Raffle.active == True,
                Raffle.auto_pick == True,
                Raffle.end_at != None,
                Raffle.end_at < now,
                Raffle.winners == None
            ).all()
            for r in auto_raffles:
                _draw_winners(db, r)
            if auto_raffles:
                db.flush()

            raffles = db.query(Raffle).filter_by(guild_id=int(guild_id)).order_by(Raffle.id.desc()).all()

            active, ended = [], []
            admin = _ensure_admin(guild_id, request)

            for r in raffles:
                entry_count = db.query(RaffleEntry).filter_by(raffle_id=r.id).count()
                winners = []
                if r.winners:
                    try:
                        winners = json_lib.loads(r.winners)
                    except Exception:
                        winners = []

                row = {
                    'id': r.id,
                    'title': r.title,
                    'description': r.description,
                    'cost_tokens': r.cost_tokens,
                    'max_winners': r.max_winners,
                    'max_entries_per_user': r.max_entries_per_user,
                    'start_at': r.start_at,
                    'end_at': r.end_at,
                    'auto_pick': r.auto_pick,
                    'active': r.active,
                    'announce_channel_id': str(r.announce_channel_id) if r.announce_channel_id else None,
                    'announce_role_id': str(r.announce_role_id) if r.announce_role_id else None,
                    'announce_message': r.announce_message,
                    'winner_message': r.winner_message,
                    'entry_emoji': r.entry_emoji,
                    'announce_message_id': str(r.announce_message_id) if r.announce_message_id else None,
                    'reminder_channel_id': str(r.reminder_channel_id) if r.reminder_channel_id else None,
                    'entry_count': entry_count,
                    'winners': winners,
                    'can_manage': admin,
                }

                if _raffle_status(r) in ['completed', 'ended', 'closed']:
                    ended.append(row)
                else:
                    active.append(row)

            return JsonResponse({'active': active, 'ended': ended})
    except Exception as e:
        logger.error(f"Error listing raffles: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to load raffles'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_raffle_create(request, guild_id):
    """Create a raffle (admin only)."""
    from .db import get_db_session
    from .models import Raffle, Guild as GuildModel

    logger.info(f"Raffle create request for guild {guild_id}")

    if not _ensure_admin(guild_id, request):
        logger.warning(f"Permission denied for guild {guild_id}")
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        data = json_lib.loads(request.body)
        logger.info(f"Raffle data received: {data}")
        logger.info(f"Channel ID raw: {data.get('announce_channel_id')} (type: {type(data.get('announce_channel_id'))})")
        logger.info(f"Role ID raw: {data.get('announce_role_id')} (type: {type(data.get('announce_role_id'))})")
    except json_lib.JSONDecodeError as e:
        logger.error(f"Invalid JSON: {e}")
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    def _to_int(val):
        try:
            return int(val) if val else None
        except Exception:
            return None

    try:
        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                logger.warning(f"Guild {guild_id} not found in database")
                # Create guild record if it doesn't exist
                guild_record = GuildModel(guild_id=int(guild_id))
                db.add(guild_record)
                db.flush()

            raffle = Raffle(
                guild_id=int(guild_id),
                title=data.get('title', '').strip(),
                description=data.get('description', ''),
                cost_tokens=max(0, int(data.get('cost_tokens', 0))),
                max_winners=max(1, int(data.get('max_winners', 1))),
                max_entries_per_user=_to_int(data.get('max_entries_per_user')),
                start_at=data.get('start_at') or None,
                end_at=data.get('end_at') or None,
                auto_pick=bool(data.get('auto_pick', False)),
                announce_channel_id=_to_int(data.get('announce_channel_id')),
                announce_role_id=_to_int(data.get('announce_role_id')),
                announce_message=(data.get('announce_message') or '').strip(),
                winner_message=(data.get('winner_message') or '').strip(),
                entry_emoji=data.get('entry_emoji') or "🎟️",
                reminder_channel_id=_to_int(data.get('reminder_channel_id')),
                active=True,
                winners=None,
                created_by=_to_int(request.session.get('discord_user', {}).get('id')),
                created_by_name=request.session.get('discord_user', {}).get('username')
            )
            db.add(raffle)
            db.flush()

            logger.info(f"Raffle created successfully with ID: {raffle.id}")

            return JsonResponse({'success': True, 'id': raffle.id})
    except Exception as e:
        logger.error(f"Error creating raffle: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to create raffle'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_raffle_update(request, guild_id, raffle_id):
    """Update an existing raffle (admin only)."""
    from .db import get_db_session
    from .models import Raffle
    if not _ensure_admin(guild_id, request):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    def _to_int(val):
        try:
            return int(val)
        except Exception:
            return None

    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        with get_db_session() as db:
            raffle = db.query(Raffle).filter_by(id=raffle_id, guild_id=int(guild_id)).first()
            if not raffle:
                return JsonResponse({'error': 'Raffle not found'}, status=404)

            raffle.title = data.get('title', raffle.title).strip()
            raffle.description = data.get('description', raffle.description)
            raffle.cost_tokens = max(0, int(data.get('cost_tokens', raffle.cost_tokens or 0)))
            raffle.max_winners = max(1, int(data.get('max_winners', raffle.max_winners or 1)))
            raffle.max_entries_per_user = _to_int(data.get('max_entries_per_user'))
            raffle.start_at = data.get('start_at') or None
            raffle.end_at = data.get('end_at') or None
            raffle.auto_pick = bool(data.get('auto_pick', raffle.auto_pick))
            raffle.announce_channel_id = _to_int(data.get('announce_channel_id'))
            raffle.announce_role_id = _to_int(data.get('announce_role_id'))
            raffle.announce_message = (data.get('announce_message') or '').strip()
            raffle.winner_message = (data.get('winner_message') or '').strip()
            raffle.entry_emoji = data.get('entry_emoji') or raffle.entry_emoji or "🎟️"
            raffle.reminder_channel_id = _to_int(data.get('reminder_channel_id'))
            # reset announce message id if channel changed so bot can re-post
            if data.get('announce_channel_id') and str(raffle.announce_channel_id) != str(data.get('announce_channel_id')):
                raffle.announce_message_id = None
            return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Error updating raffle: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to update raffle'}, status=500)


@require_http_methods(["POST"])
@ratelimit(key='user_or_ip', rate='10/m', method='POST', block=True)
def api_raffle_enter(request, guild_id, raffle_id):
    """Enter a raffle using tokens. (Members can enter)"""
    from .db import get_db_session
    from .models import Raffle, RaffleEntry, GuildMember

    # Check if user is logged in (but don't require admin access)
    discord_user = request.session.get('discord_user')
    if not discord_user:
        return JsonResponse({'error': 'Not authenticated'}, status=401)

    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    tickets = max(1, int(data.get('tickets', 1)))

    try:
        with get_db_session() as db:
            raffle = db.query(Raffle).filter_by(id=raffle_id, guild_id=int(guild_id)).first()
            if not raffle:
                return JsonResponse({'error': 'Raffle not found'}, status=404)

            # status checks
            status = _raffle_status(raffle)
            if status == 'completed':
                return JsonResponse({'error': 'Raffle already completed'}, status=400)
            if status == 'closed':
                return JsonResponse({'error': 'Raffle closed'}, status=400)

            now = int(time.time())
            if raffle.start_at and now < raffle.start_at:
                return JsonResponse({'error': 'Raffle has not started yet'}, status=400)
            if raffle.end_at and now > raffle.end_at:
                return JsonResponse({'error': 'Raffle has ended'}, status=400)

            member = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=int(discord_user.get('id', 0))
            ).first()
            if not member:
                return JsonResponse({'error': 'You must interact in the server first.'}, status=400)

            # Check max entries per user limit
            if raffle.max_entries_per_user:
                from sqlalchemy import func
                existing_entries = db.query(func.sum(RaffleEntry.tickets)).filter_by(
                    raffle_id=raffle.id,
                    user_id=int(discord_user.get('id'))
                ).scalar() or 0

                total_entries = existing_entries + tickets
                if total_entries > raffle.max_entries_per_user:
                    remaining_slots = raffle.max_entries_per_user - existing_entries
                    if remaining_slots <= 0:
                        return JsonResponse({
                            'error': f'You have already purchased the maximum of {raffle.max_entries_per_user} entries for this raffle.'
                        }, status=400)
                    else:
                        return JsonResponse({
                            'error': f'This raffle has a limit of {raffle.max_entries_per_user} entries per person. You have {existing_entries} entries and can only buy {remaining_slots} more.'
                        }, status=400)

            cost = raffle.cost_tokens * tickets
            if member.hero_tokens < cost:
                return JsonResponse({'error': f'Not enough tokens. Need {cost}, have {member.hero_tokens}.'}, status=400)

            member.hero_tokens -= cost
            entry = RaffleEntry(
                raffle_id=raffle.id,
                user_id=int(discord_user.get('id')),
                username=discord_user.get('username'),
                tickets=tickets
            )
            db.add(entry)

            return JsonResponse({
                'success': True,
                'message': 'Entry submitted!',
                'tokens_remaining': member.hero_tokens
            })
    except Exception as e:
        logger.error(f"Error entering raffle: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to enter raffle'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_raffle_pick(request, guild_id, raffle_id):
    """Pick winners (admin only)."""
    from .db import get_db_session
    from .models import Raffle
    if not _ensure_admin(guild_id, request):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        with get_db_session() as db:
            raffle = db.query(Raffle).filter_by(id=raffle_id, guild_id=int(guild_id)).first()
            if not raffle:
                return JsonResponse({'error': 'Raffle not found'}, status=404)

            winners = _draw_winners(db, raffle)
            return JsonResponse({
                'success': True,
                'message': 'Winners selected!' if winners else 'No entries to draw from.',
                'winners': winners
            })
    except Exception as e:
        logger.error(f"Error picking winners: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to pick winners'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_raffle_start_now(request, guild_id, raffle_id):
    """Manually start a raffle immediately (set start_at to now and clear announce message so bot can post)."""
    from .db import get_db_session
    from .models import Raffle
    if not _ensure_admin(guild_id, request):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    now = int(time.time())
    try:
        with get_db_session() as db:
            raffle = db.query(Raffle).filter_by(id=raffle_id, guild_id=int(guild_id)).first()
            if not raffle:
                return JsonResponse({'error': 'Raffle not found'}, status=404)
            raffle.start_at = now
            raffle.active = True
            raffle.announce_message_id = None
            # if no end supplied, leave as None and rely on manual end/pick
            return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Error starting raffle {raffle_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to start raffle'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_raffle_end_now(request, guild_id, raffle_id):
    """Manually end a raffle immediately; marks end_at and triggers draw if auto_pick."""
    from .db import get_db_session
    from .models import Raffle
    if not _ensure_admin(guild_id, request):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    now = int(time.time())
    try:
        with get_db_session() as db:
            raffle = db.query(Raffle).filter_by(id=raffle_id, guild_id=int(guild_id)).first()
            if not raffle:
                return JsonResponse({'error': 'Raffle not found'}, status=404)
            raffle.end_at = now
            raffle.active = False
            if raffle.auto_pick:
                _draw_winners(db, raffle)
            return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Error ending raffle {raffle_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to end raffle'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_role_export_create_template(request, guild_id):
    """GET /api/guild/<id>/roles/export-create-template/ - Export XLSX template for bulk role creation."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        from django.http import HttpResponse

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Role Creation Template"

        # Add headers with formatting
        headers = ['name', 'color', 'permissions', 'hoist', 'mentionable']
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

        # Add example rows
        ws.append(['Member', '#5865F2', '0', 'FALSE', 'TRUE'])
        ws.append(['Moderator', '#F23F43', '8388614', 'TRUE', 'TRUE'])
        ws.append(['VIP', '#FEE75C', '0', 'TRUE', 'FALSE'])

        # Set column widths
        ws.column_dimensions['A'].width = 20  # name
        ws.column_dimensions['B'].width = 12  # color
        ws.column_dimensions['C'].width = 15  # permissions
        ws.column_dimensions['D'].width = 12  # hoist
        ws.column_dimensions['E'].width = 15  # mentionable

        # Add instructions sheet
        ws_info = wb.create_sheet(title="Instructions")
        ws_info['A1'] = "Bulk Role Creation Instructions"
        ws_info['A1'].font = Font(bold=True, size=14)

        instructions = [
            "",
            "How to use this template:",
            "1. Fill in the role details in the 'Role Creation Template' sheet",
            "2. Delete the example rows",
            "3. Save the file",
            "4. Upload it via the Bulk Create button",
            "",
            "Column Descriptions:",
            "• name - Role name (required)",
            "• color - Hex color code with # (e.g., #FF5733)",
            "• permissions - Permission number (0 = no permissions, use Discord permission calculator)",
            "• hoist - Display separately in member list (TRUE/FALSE)",
            "• mentionable - Allow role to be mentioned (TRUE/FALSE)",
            "",
            "Common Permission Numbers:",
            "• 0 = No permissions",
            "• 8 = Administrator (DANGEROUS - full access)",
            "• 8388614 = Moderator (Kick, Ban, Manage Messages, etc.)",
            "• 104189504 = Read & Send Messages",
            "",
            "Tips:",
            "• Maximum 100 roles per import",
            "• Use TRUE/FALSE for boolean fields",
            "• Colors should be hex codes (#RRGGBB)",
        ]

        for i, instruction in enumerate(instructions, start=2):
            ws_info[f'A{i}'] = instruction

        ws_info.column_dimensions['A'].width = 70

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return as download
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="role_creation_template.xlsx"'
        return response

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# Audit Logs Dashboard

@discord_required
def guild_audit_logs(request, guild_id):
    """View audit logs for a guild."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Get filter parameters
    action_filter = request.GET.get('action', '')
    actor_filter = request.GET.get('actor', '')
    target_filter = request.GET.get('target', '')
    days = int(request.GET.get('days', 7))

    # Fetch audit logs from database
    audit_logs = []
    total_count = 0
    action_types = []
    guild_record = None

    try:
        from .db import get_db_session
        from .models import AuditLog, AuditAction, Guild as GuildModel

        with get_db_session() as db:
            # Get guild info
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            is_premium = guild_record.is_premium() if guild_record else False

            # Limit days based on subscription
            if not is_premium:
                days = min(days, 7)  # Free: 7 days max
            elif guild_record and guild_record.subscription_tier == 'pro':
                days = min(days, 90)  # Pro: 90 days
            else:
                days = min(days, 30)  # Premium: 30 days

            # Calculate time threshold
            time_threshold = int(time.time()) - (days * 24 * 60 * 60)

            # Build query
            query = db.query(AuditLog).filter(
                AuditLog.guild_id == int(guild_id),
                AuditLog.timestamp >= time_threshold
            )

            # Apply filters
            if action_filter:
                try:
                    action_enum = AuditAction(action_filter)
                    query = query.filter(AuditLog.action == action_enum)
                except ValueError:
                    pass

            if actor_filter:
                query = query.filter(AuditLog.actor_id == int(actor_filter))

            if target_filter:
                query = query.filter(AuditLog.target_id == int(target_filter))

            # Get total count
            total_count = query.count()

            # Get logs (most recent first, limit 100)
            logs = query.order_by(AuditLog.timestamp.desc()).limit(100).all()

            audit_logs = [
                {
                    'id': log.id,
                    'action': log.action.value,
                    'action_display': log.action.value.replace('_', ' ').title(),
                    'category': log.action_category or get_action_category(log.action.value),
                    'actor_id': str(log.actor_id) if log.actor_id else None,
                    'actor_name': log.actor_name or 'System',
                    'target_id': str(log.target_id) if log.target_id else None,
                    'target_name': log.target_name,
                    'target_type': log.target_type,
                    'reason': log.reason,
                    'details': log.details,
                    'timestamp': log.timestamp,
                }
                for log in logs
            ]

            # Get unique action types for filter dropdown
            action_types = [a.value for a in AuditAction]

    except Exception as e:
        logger.warning(f"Could not fetch audit logs: {e}")

    # Get audit log channel configuration
    audit_log_channel_id = ''
    if guild_record:
        audit_log_channel_id = str(guild_record.log_channel_id) if guild_record.log_channel_id else ''

    # Check if guild has moderation module access
    has_moderation_module = has_module_access(guild_id, 'moderation')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'audit_logs': audit_logs,
        'total_count': total_count,
        'action_types': action_types,
        'audit_log_channel_id': audit_log_channel_id,
        'audit_enabled': bool(guild_record.audit_logging_enabled) if guild_record else False,
        'audit_event_config': guild_record.audit_event_config if guild_record else '',
        'has_moderation_module': has_moderation_module,
        'has_any_module': has_any_module,
        'current_filters': {
            'action': action_filter,
            'actor': actor_filter,
            'target': target_filter,
            'days': days,
        },
        'active_page': 'audit',
    }
    return render(request, 'questlog/audit.html', context)


def get_action_category(action: str) -> str:
    """Get category for an audit action."""
    if action.startswith('member_'):
        return 'members'
    elif action.startswith('role_'):
        return 'roles'
    elif action.startswith('channel_'):
        return 'channels'
    elif action.startswith('message_'):
        return 'messages'
    elif action.startswith('verification_'):
        return 'verification'
    elif action in ['raid_detected', 'lockdown_activated', 'lockdown_deactivated']:
        return 'security'
    else:
        return 'other'


# Audit Logs API Endpoints

@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_audit_logs(request, guild_id):
    """GET /api/guild/<id>/audit/ - Get paginated audit logs."""
    try:
        from .db import get_db_session
        from .models import AuditLog, AuditAction, Guild as GuildModel

        # Pagination
        page = int(request.GET.get('page', 1))
        per_page = min(int(request.GET.get('per_page', 50)), 100)
        offset = (page - 1) * per_page

        # Filters
        action_filter = request.GET.get('action', '')
        actor_filter = request.GET.get('actor', '')
        target_filter = request.GET.get('target', '')
        days = int(request.GET.get('days', 7))

        with get_db_session() as db:
            # Check subscription for day limits
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            is_premium = guild_record.is_premium() if guild_record else False

            if not is_premium:
                days = min(days, 7)
            elif guild_record and guild_record.subscription_tier == 'pro':
                days = min(days, 90)
            else:
                days = min(days, 30)

            time_threshold = int(time.time()) - (days * 24 * 60 * 60)

            # Build query
            query = db.query(AuditLog).filter(
                AuditLog.guild_id == int(guild_id),
                AuditLog.timestamp >= time_threshold
            )

            if action_filter:
                try:
                    action_enum = AuditAction(action_filter)
                    query = query.filter(AuditLog.action == action_enum)
                except ValueError:
                    pass

            if actor_filter:
                query = query.filter(AuditLog.actor_id == int(actor_filter))

            if target_filter:
                query = query.filter(AuditLog.target_id == int(target_filter))

            total = query.count()
            logs = query.order_by(AuditLog.timestamp.desc()).offset(offset).limit(per_page).all()

            return JsonResponse({
                'success': True,
                'logs': [
                    {
                        'id': log.id,
                        'action': log.action.value,
                        'action_display': log.action.value.replace('_', ' ').title(),
                        'category': log.action_category or get_action_category(log.action.value),
                        'actor_id': str(log.actor_id) if log.actor_id else None,
                        'actor_name': log.actor_name or 'System',
                        'target_id': str(log.target_id) if log.target_id else None,
                        'target_name': log.target_name,
                        'target_type': log.target_type,
                        'reason': log.reason,
                        'details': log.details,
                        'timestamp': log.timestamp,
                    }
                    for log in logs
                ],
                'total': total,
                'page': page,
                'per_page': per_page,
                'total_pages': (total + per_page - 1) // per_page,
                'max_days': days,
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_audit_stats(request, guild_id):
    """GET /api/guild/<id>/audit/stats/ - Get audit log statistics."""
    try:
        from .db import get_db_session
        from .models import AuditLog, AuditAction
        from sqlalchemy import func

        days = int(request.GET.get('days', 7))
        time_threshold = int(time.time()) - (days * 24 * 60 * 60)

        with get_db_session() as db:
            # Count by action type
            action_counts = db.query(
                AuditLog.action,
                func.count(AuditLog.id).label('count')
            ).filter(
                AuditLog.guild_id == int(guild_id),
                AuditLog.timestamp >= time_threshold
            ).group_by(AuditLog.action).all()

            # Top actors
            top_actors = db.query(
                AuditLog.actor_id,
                AuditLog.actor_name,
                func.count(AuditLog.id).label('count')
            ).filter(
                AuditLog.guild_id == int(guild_id),
                AuditLog.timestamp >= time_threshold,
                AuditLog.actor_id.isnot(None)
            ).group_by(AuditLog.actor_id, AuditLog.actor_name).order_by(
                func.count(AuditLog.id).desc()
            ).limit(10).all()

            # Total count
            total = db.query(AuditLog).filter(
                AuditLog.guild_id == int(guild_id),
                AuditLog.timestamp >= time_threshold
            ).count()

            return JsonResponse({
                'success': True,
                'stats': {
                    'total': total,
                    'by_action': {
                        action.value: count for action, count in action_counts
                    },
                    'top_actors': [
                        {
                            'actor_id': str(actor_id),
                            'actor_name': actor_name or f'User {actor_id}',
                            'count': count
                        }
                        for actor_id, actor_name, count in top_actors
                    ],
                    'days': days,
                }
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET", "POST", "PATCH"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method=['POST', 'PATCH'], block=True)
def api_audit_config_update(request, guild_id):
    """GET/POST /api/guild/<id>/audit/config/ - Fetch or update audit log configuration."""
    # GET: return current config
    if request.method == "GET":
        try:
            from .db import get_db_session
            from .models import Guild as GuildModel

            with get_db_session() as db:
                guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
                return JsonResponse({
                    'success': True,
                    'log_channel_id': str(guild_record.log_channel_id) if guild_record and guild_record.log_channel_id else '',
                    'audit_enabled': bool(guild_record.audit_logging_enabled) if guild_record else False,
                    'event_config': json_lib.loads(guild_record.audit_event_config or '{}') if guild_record else {}
                })
        except Exception as e:
            logger.error(f"Error fetching audit config: {e}", exc_info=True)
            return JsonResponse({'error': 'Failed to fetch audit config'}, status=500)

    # POST/PATCH: update config
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                guild_record = GuildModel(guild_id=int(guild_id))
                db.add(guild_record)

            # Update log channel
            if 'log_channel_id' in data:
                try:
                    guild_record.log_channel_id = int(data['log_channel_id']) if data['log_channel_id'] else None
                except (TypeError, ValueError):
                    guild_record.log_channel_id = None
            if data.get('audit_enabled') is False:
                guild_record.log_channel_id = None
                guild_record.audit_logging_enabled = False
            elif data.get('audit_enabled') is True:
                guild_record.audit_logging_enabled = True
            if 'event_config' in data:
                guild_record.audit_event_config = json_lib.dumps(data.get('event_config') or {})

            logger.info(f"Updated audit log channel for guild {guild_id} to {guild_record.log_channel_id}")
            db.commit()
            return JsonResponse({
                'success': True,
                'log_channel_id': str(guild_record.log_channel_id) if guild_record.log_channel_id else '',
                'audit_enabled': guild_record.audit_logging_enabled,
                'event_config': json_lib.loads(guild_record.audit_event_config or '{}')
            })

    except Exception as e:
        logger.error(f"Error updating audit config: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@require_subscription_tier('pro', 'premium')
@ratelimit(key='user_or_ip', rate='10/h', method='GET', block=True)
def api_audit_export(request, guild_id):
    """GET /api/guild/<id>/audit/export/ - Export audit logs to CSV (Pro/Premium only, 10 req/hour)."""
    import csv
    import io
    from django.http import HttpResponse

    try:
        from .db import get_db_session
        from .models import AuditLog

        # Get filter params
        action_filter = request.GET.get('action', '')
        actor_filter = request.GET.get('actor', '')
        target_filter = request.GET.get('target', '')
        days = int(request.GET.get('days', 30))

        with get_db_session() as db:
            cutoff_time = int(time.time()) - (days * 86400)

            query = db.query(AuditLog).filter(
                AuditLog.guild_id == int(guild_id),
                AuditLog.timestamp >= cutoff_time
            )

            if action_filter:
                query = query.filter(AuditLog.action == action_filter)
            if actor_filter:
                query = query.filter(AuditLog.actor_name.ilike(f'%{actor_filter}%'))
            if target_filter:
                query = query.filter(AuditLog.target_name.ilike(f'%{target_filter}%'))

            logs = query.order_by(AuditLog.timestamp.desc()).limit(10000).all()

            # Create CSV
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(['Timestamp', 'Action', 'Actor', 'Target', 'Reason', 'Details'])

            # Write data
            for log in logs:
                from datetime import datetime
                dt = datetime.fromtimestamp(log.timestamp)
                writer.writerow([
                    dt.strftime('%Y-%m-%d %H:%M:%S'),
                    log.action.value if hasattr(log.action, 'value') else str(log.action),
                    log.actor_name or 'N/A',
                    log.target_name or 'N/A',
                    log.reason or '',
                    log.details or ''
                ])

            # Create HTTP response
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="audit_logs_{guild_id}_{int(time.time())}.csv"'
            return response

    except Exception as e:
        logger.error(f"Error exporting audit logs: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to export logs'}, status=500)


# Welcome/Goodbye Messages Dashboard

@discord_required
def guild_welcome(request, guild_id):
    """Welcome and goodbye message configuration page."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Fetch welcome config from database
    welcome_config = None
    guild_channels = []
    guild_record = None

    try:
        from .db import get_db_session
        from .models import WelcomeConfig, Guild as GuildModel

        with get_db_session() as db:
            # Ensure guild exists
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()

            # Get welcome config
            config = db.query(WelcomeConfig).filter_by(guild_id=int(guild_id)).first()

            if config:
                welcome_config = {
                    'enabled': config.enabled,
                    'channel_message_enabled': config.channel_message_enabled,
                    'channel_message': config.channel_message,
                    'channel_embed_enabled': config.channel_embed_enabled,
                    'channel_embed_title': config.channel_embed_title,
                    'channel_embed_color': config.channel_embed_color,
                    'channel_embed_thumbnail': config.channel_embed_thumbnail,
                    'channel_embed_footer': config.channel_embed_footer,
                    'dm_enabled': config.dm_enabled,
                    'dm_message': config.dm_message,
                    'goodbye_enabled': config.goodbye_enabled,
                    'goodbye_message': config.goodbye_message,
                    'auto_role_id': str(config.auto_role_id) if config.auto_role_id else '',
                    'goodbye_channel_id': ''
                }

                # Get welcome channel from guild record
                if guild_record:
                    welcome_config['welcome_channel_id'] = str(guild_record.welcome_channel_id) if guild_record.welcome_channel_id else ''
            else:
                # Defaults
                welcome_config = {
                    'enabled': True,
                    'channel_message_enabled': True,
                    'channel_message': 'Welcome to **{server}**, {user}! You are member #{member_count}.',
                    'channel_embed_enabled': True,
                    'channel_embed_title': 'Welcome!',
                    'channel_embed_color': 0x5865F2,
                    'channel_embed_thumbnail': True,
                    'channel_embed_footer': '',
                    'dm_enabled': False,
                    'dm_message': 'Welcome to **{server}**! Please read the rules and enjoy your stay.',
                    'goodbye_enabled': False,
                    'goodbye_message': '**{username}** has left the server.',
                    'auto_role_id': '',
                    'welcome_channel_id': '',
                    'goodbye_channel_id': ''
                }

    except Exception as e:
        logger.warning(f"Could not fetch welcome config: {e}")
        welcome_config = {}

    # Check if guild has engagement module access
    has_engagement_module = has_module_access(guild_id, 'engagement')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'welcome_config': welcome_config,
        'has_engagement_module': has_engagement_module,
        'has_any_module': has_any_module,
        'active_page': 'welcome',
    }
    return render(request, 'questlog/welcome.html', context)


@require_http_methods(["POST", "PATCH"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['POST', 'PATCH'], block=True)
def api_welcome_config_update(request, guild_id):
    """POST /api/guild/<id>/welcome/config/ - Update welcome configuration."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import WelcomeConfig, Guild as GuildModel

        with get_db_session() as db:
            # Ensure guild exists
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                guild_record = GuildModel(guild_id=int(guild_id))
                db.add(guild_record)
                db.flush()

            # Get or create welcome config
            config = db.query(WelcomeConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                config = WelcomeConfig(guild_id=int(guild_id))
                db.add(config)

            # Update fields
            if 'enabled' in data:
                config.enabled = bool(data['enabled'])
            if 'channel_message_enabled' in data:
                config.channel_message_enabled = bool(data['channel_message_enabled'])
            if 'channel_message' in data:
                config.channel_message = data['channel_message']
            if 'channel_embed_enabled' in data:
                config.channel_embed_enabled = bool(data['channel_embed_enabled'])
            if 'channel_embed_title' in data:
                config.channel_embed_title = data['channel_embed_title']
            if 'channel_embed_color' in data:
                config.channel_embed_color = int(data['channel_embed_color'])
            if 'channel_embed_thumbnail' in data:
                config.channel_embed_thumbnail = bool(data['channel_embed_thumbnail'])
            if 'channel_embed_footer' in data:
                config.channel_embed_footer = data['channel_embed_footer'] or None
            if 'dm_enabled' in data:
                config.dm_enabled = bool(data['dm_enabled'])
            if 'dm_message' in data:
                config.dm_message = data['dm_message']
            if 'goodbye_enabled' in data:
                config.goodbye_enabled = bool(data['goodbye_enabled'])
            if 'goodbye_message' in data:
                config.goodbye_message = data['goodbye_message']
            if 'goodbye_channel_id' in data:
                config.goodbye_channel_id = int(data['goodbye_channel_id']) if data['goodbye_channel_id'] else None
            if 'auto_role_id' in data:
                config.auto_role_id = int(data['auto_role_id']) if data['auto_role_id'] else None

            # Update welcome channel in guild record
            if 'welcome_channel_id' in data:
                guild_record.welcome_channel_id = int(data['welcome_channel_id']) if data['welcome_channel_id'] else None

            config.updated_at = int(time.time())

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_welcome_config(request, guild_id):
    """GET /api/guild/<id>/welcome/config/ - Get welcome configuration."""
    try:
        from .db import get_db_session
        from .models import WelcomeConfig, Guild as GuildModel

        with get_db_session() as db:
            config = db.query(WelcomeConfig).filter_by(guild_id=int(guild_id)).first()
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()

            if not config:
                return JsonResponse({
                    'success': True,
                    'config': {
                        'enabled': True,
                        'channel_message_enabled': True,
                        'channel_message': 'Welcome to **{server}**, {user}! You are member #{member_count}.',
                        'channel_embed_enabled': True,
                        'channel_embed_title': 'Welcome!',
                        'channel_embed_color': 0x5865F2,
                        'channel_embed_thumbnail': True,
                        'channel_embed_footer': '',
                        'dm_enabled': False,
                        'dm_message': 'Welcome to **{server}**! Please read the rules and enjoy your stay.',
                        'goodbye_enabled': False,
                        'goodbye_message': '**{username}** has left the server.',
                        'auto_role_id': '',
                        'welcome_channel_id': '',
                    }
                })

            return JsonResponse({
                'success': True,
                'config': {
                    'enabled': config.enabled,
                    'xp_enabled': config.xp_enabled,
                    'channel_message_enabled': config.channel_message_enabled,
                    'channel_message': config.channel_message,
                    'channel_embed_enabled': config.channel_embed_enabled,
                    'channel_embed_title': config.channel_embed_title,
                    'channel_embed_color': config.channel_embed_color,
                    'channel_embed_thumbnail': config.channel_embed_thumbnail,
                    'channel_embed_footer': config.channel_embed_footer or '',
                    'dm_enabled': config.dm_enabled,
                    'dm_message': config.dm_message,
                    'goodbye_enabled': config.goodbye_enabled,
                    'goodbye_message': config.goodbye_message,
                    'auto_role_id': str(config.auto_role_id) if config.auto_role_id else '',
                    'welcome_channel_id': str(guild_record.welcome_channel_id) if guild_record and guild_record.welcome_channel_id else '',
                }
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_welcome_test(request, guild_id):
    """POST /api/guild/<id>/welcome/test/ - Send a test welcome message."""
    try:
        from .actions import queue_action, ActionType

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        message_type = request.GET.get('type', 'welcome')  # 'welcome' or 'goodbye'

        action_id = queue_action(
            guild_id=int(guild_id),
            action_type=ActionType.MESSAGE_SEND,
            payload={
                'type': f'test_{message_type}',
                'target_user_id': triggered_by,
            },
            triggered_by=triggered_by,
            triggered_by_name=triggered_by_name,
            source='website'
        )

        return JsonResponse({
            'success': True,
            'action_id': action_id,
            'message': f'Test {message_type} message queued!'
        })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# Level-Up Messages Dashboard

@discord_required
def guild_levelup(request, guild_id):
    """Level-up message configuration page."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    levelup_config = None
    guild_record = None

    try:
        from .db import get_db_session
        from .models import LevelUpConfig, Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            config = db.query(LevelUpConfig).filter_by(guild_id=int(guild_id)).first()

            if config:
                levelup_config = {
                    'enabled': config.enabled,
                    'destination': config.destination,
                    'message': config.message,
                    'use_embed': config.use_embed,
                    'embed_color': config.embed_color,
                    'show_progress': config.show_progress,
                    'ping_user': config.ping_user,
                    'ping_on_role_only': config.ping_on_role_only,
                    'announce_role_reward': config.announce_role_reward,
                    'role_reward_message': config.role_reward_message,
                    'milestone_levels': config.milestone_levels or '[]',
                    'milestone_message': config.milestone_message,
                    'quiet_hours_enabled': config.quiet_hours_enabled,
                    'quiet_hours_start': config.quiet_hours_start,
                    'quiet_hours_end': config.quiet_hours_end,
                    'level_up_channel_id': str(guild_record.level_up_channel_id) if guild_record and guild_record.level_up_channel_id else '',
                }
            else:
                levelup_config = {
                    'enabled': True,
                    'destination': 'current',
                    'message': "Congrats {user}! You've reached **Level {level}**!",
                    'use_embed': True,
                    'embed_color': 0x5865F2,
                    'show_progress': True,
                    'ping_user': True,
                    'ping_on_role_only': False,
                    'announce_role_reward': True,
                    'role_reward_message': "You've also earned the **{role}** role!",
                    'milestone_levels': '[10, 25, 50, 100]',
                    'milestone_message': "Incredible! You've hit the **Level {level}** milestone!",
                    'quiet_hours_enabled': False,
                    'quiet_hours_start': 22,
                    'quiet_hours_end': 8,
                    'level_up_channel_id': '',
                }

    except Exception as e:
        logger.warning(f"Could not fetch level-up config: {e}")
        levelup_config = {}

    # Check if guild has engagement module access
    has_engagement_module = has_module_access(guild_id, 'engagement')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'levelup_config': levelup_config,
        'has_engagement_module': has_engagement_module,
        'has_any_module': has_any_module,
        'active_page': 'levelup',
    }
    return render(request, 'questlog/levelup.html', context)


# Message System
@discord_required
def guild_messages(request, guild_id):
    """Message System dashboard (send/edit messages and embeds)."""
    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])
    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    guild_record = None
    try:
        from .db import get_db_session
        from .models import Guild as GuildModel
        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
    except Exception as e:
        logger.warning(f"Could not fetch guild record: {e}")

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'active_page': 'messages',
    }
    return render(request, 'questlog/messages.html', context)


@require_http_methods(["POST", "PATCH"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['POST', 'PATCH'], block=True)
def api_levelup_config_update(request, guild_id):
    """POST /api/guild/<id>/levelup/config/ - Update level-up configuration."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import LevelUpConfig, Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                guild_record = GuildModel(guild_id=int(guild_id))
                db.add(guild_record)
                db.flush()

            config = db.query(LevelUpConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                config = LevelUpConfig(guild_id=int(guild_id))
                db.add(config)

            # Update fields
            if 'enabled' in data:
                config.enabled = bool(data['enabled'])
            if 'destination' in data:
                config.destination = data['destination']
            if 'message' in data:
                config.message = data['message']
            if 'use_embed' in data:
                config.use_embed = bool(data['use_embed'])
            if 'embed_color' in data:
                config.embed_color = int(data['embed_color'])
            if 'show_progress' in data:
                config.show_progress = bool(data['show_progress'])
            if 'ping_user' in data:
                config.ping_user = bool(data['ping_user'])
            if 'ping_on_role_only' in data:
                config.ping_on_role_only = bool(data['ping_on_role_only'])
            if 'announce_role_reward' in data:
                config.announce_role_reward = bool(data['announce_role_reward'])
            if 'role_reward_message' in data:
                config.role_reward_message = data['role_reward_message']
            if 'milestone_levels' in data:
                config.milestone_levels = data['milestone_levels']
            if 'milestone_message' in data:
                config.milestone_message = data['milestone_message']
            if 'quiet_hours_enabled' in data:
                config.quiet_hours_enabled = bool(data['quiet_hours_enabled'])
            if 'quiet_hours_start' in data:
                config.quiet_hours_start = int(data['quiet_hours_start'])
            if 'quiet_hours_end' in data:
                config.quiet_hours_end = int(data['quiet_hours_end'])

            # Update channel in guild record
            if 'level_up_channel_id' in data:
                guild_record.level_up_channel_id = int(data['level_up_channel_id']) if data['level_up_channel_id'] else None

            config.updated_at = int(time.time())

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_guild_channels(request, guild_id):
    """GET /api/guild/<id>/channels/ - Get guild channels (cached)."""
    try:
        from .discord_resources import get_guild_channels

        # Get force_refresh from query params
        force_refresh = request.GET.get('force_refresh', 'false').lower() == 'true'

        # Fetch channels with caching
        channels = get_guild_channels(guild_id, force_refresh=force_refresh)
        # Normalize IDs to strings to avoid JS precision loss
        norm_channels = []
        for c in channels:
            norm_channels.append({
                'id': str(c.get('id')),
                'name': c.get('name'),
                'type': c.get('type')
            })

        return JsonResponse({'channels': norm_channels})

    except Exception as e:
        logger.error(f"Failed to fetch channels for guild {guild_id}: {e}")
        return JsonResponse({'error': 'Failed to fetch channels', 'channels': []}, status=500)


@require_http_methods(["GET", "PATCH"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_guild_roles(request, guild_id):
    """
    GET /api/guild/<id>/roles/ - Get guild roles (cached).
    PATCH /api/guild/<id>/roles/ - Update Discord role positions.
    """
    if request.method == 'GET':
        try:
            from .discord_resources import get_guild_roles

            # Get force_refresh from query params
            force_refresh = request.GET.get('force_refresh', 'false').lower() == 'true'

            # Fetch roles with caching
            roles = get_guild_roles(guild_id, force_refresh=force_refresh)

            # Add is_everyone flag for each role
            norm_roles = []
            for role in roles:
                norm_roles.append({
                    'id': str(role.get('id')),
                    'name': role.get('name'),
                    'position': role.get('position'),
                    'is_everyone': role.get('name') == '@everyone' or str(role.get('id')) == str(guild_id)
                })

            return JsonResponse({'success': True, 'roles': norm_roles})

        except Exception as e:
            logger.error(f"Failed to fetch roles for guild {guild_id}: {e}")
            return JsonResponse({'error': 'Failed to fetch roles', 'roles': []}, status=500)

    elif request.method == 'PATCH':
        try:
            import json

            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON'}, status=400)

            # Data should be an array of role position updates
            # Format: [{"id": "role_id", "position": 5}, ...]
            if not isinstance(data, list):
                return JsonResponse({'error': 'Expected array of role updates'}, status=400)

            bot_session = get_bot_session(int(guild_id))
            if not bot_session:
                return JsonResponse({'error': 'Bot session not available'}, status=500)

            # Update role positions via Discord API
            # NOTE: Discord may auto-adjust positions to prevent conflicts or enforce hierarchy rules
            resp = bot_session.patch(f'/guilds/{guild_id}/roles', json=data)

            if resp.status_code == 200:
                # Discord returns the updated roles array
                updated_roles = resp.json()
                logger.info(f"Discord accepted position update for guild {guild_id}. Response: {updated_roles}")

                # Check if Discord actually applied the requested position
                requested_role = data[0]  # We only send one role at a time
                actual_role = next((r for r in updated_roles if r['id'] == requested_role['id']), None)

                if actual_role and actual_role['position'] != requested_role['position']:
                    logger.warning(
                        f"Discord adjusted role position for {requested_role['id']} in guild {guild_id}: "
                        f"requested {requested_role['position']}, actual {actual_role['position']}"
                    )
                    return JsonResponse({
                        'success': True,
                        'updated': len(data),
                        'warning': f'Discord adjusted position to {actual_role["position"]} due to role hierarchy rules'
                    })

                return JsonResponse({'success': True, 'updated': len(data)})
            else:
                error_text = resp.text
                logger.error(f"Failed to update Discord role positions in guild {guild_id}: {resp.status_code} - {error_text}")
                return JsonResponse({'error': 'Failed to update role positions'}, status=500)

        except Exception as e:
            logger.error(f"Error updating Discord role positions in guild {guild_id}: {e}", exc_info=True)
            return JsonResponse({'error': 'An internal error occurred'}, status=500)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_guild_emojis(request, guild_id):
    """GET /api/guild/<id>/emojis/ - Get ALL guild custom emojis from database cache + standard emojis."""
    try:
        from .discord_resources import get_guild_emojis

        # Get force_refresh from query params
        force_refresh = request.GET.get('force_refresh', 'false').lower() == 'true'

        # Fetch ALL custom emojis from database (synced by bot)
        custom = get_guild_emojis(guild_id, force_refresh=force_refresh)

        # Standard emoji list (commonly used ones)
        standard = [
            {'name': '😀', 'id': None},
            {'name': '😁', 'id': None},
            {'name': '😂', 'id': None},
            {'name': '🤣', 'id': None},
            {'name': '😎', 'id': None},
            {'name': '😍', 'id': None},
            {'name': '👍', 'id': None},
            {'name': '👎', 'id': None},
            {'name': '🎉', 'id': None},
            {'name': '✅', 'id': None},
            {'name': '❌', 'id': None},
            {'name': '🔥', 'id': None},
            {'name': '⭐', 'id': None},
            {'name': '💎', 'id': None},
            {'name': '🎮', 'id': None},
            {'name': '🎯', 'id': None},
            {'name': '🏆', 'id': None},
            {'name': '💰', 'id': None},
            {'name': '🎁', 'id': None},
            {'name': '🎲', 'id': None},
        ]

        return JsonResponse({'success': True, 'standard': standard, 'custom': custom})

    except Exception as e:
        logger.error(f"Failed to fetch emojis for guild {guild_id}: {e}")
        return JsonResponse({'success': False, 'error': 'Failed to fetch emojis', 'standard': [], 'custom': []}, status=500)


# Admin/Server Settings Dashboard

@discord_required
def guild_settings(request, guild_id):
    """Server settings configuration page."""
    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    settings = {}
    guild_record = None

    try:
        from .db import get_db_session
        from .models import Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()

            if guild_record:
                settings = {
                    'prefix': guild_record.bot_prefix or '!',
                    'language': guild_record.language or 'en',
                    'timezone': guild_record.timezone or 'UTC',
                    'token_name': guild_record.token_name or 'Hero Tokens',
                    'token_emoji': guild_record.token_emoji or ':coin:',
                    'mod_log_channel_id': str(guild_record.mod_log_channel_id) if guild_record.mod_log_channel_id else '',
                    'admin_roles': guild_record.admin_roles or '[]',
                    'subscription_tier': guild_record.subscription_tier if guild_record.subscription_tier else 'free',
                    'billing_cycle': guild_record.billing_cycle,
                    'is_vip': guild_record.is_vip,
                }
            else:
                settings = {
                    'prefix': '!',
                    'language': 'en',
                    'timezone': 'UTC',
                    'token_name': 'Hero Tokens',
                    'token_emoji': ':coin:',
                    'mod_log_channel_id': '',
                    'admin_roles': '[]',
                    'subscription_tier': 'free',
                    'billing_cycle': None,
                    'is_vip': False,
                }

    except Exception as e:
        logger.warning(f"Could not fetch settings: {e}")

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'settings': settings,
        'active_page': 'settings',
    }
    return render(request, 'questlog/settings.html', context)


@discord_required
def guild_game_servers(request, guild_id):
    """AMP Game Server Management page - Shows hosting upsell or AMP panel access."""
    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])
    all_guilds = request.session.get('discord_all_guilds', [])

    # Check if user is admin
    guild_check = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild_check:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Get guild with owner/permissions data for sidebar
    guild = get_guild_with_permissions(guild_id, admin_guilds, all_guilds)

    try:
        from .db import get_db_session
        from .models import Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()

            context = {
                'guild': guild,
                'guild_record': guild_record,
                'admin_guilds': admin_guilds,
                'member_guilds': get_member_guilds(request),
                'is_admin': True,
                'active_page': 'game_servers',
            }
            return render(request, 'questlog/game_servers.html', context)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading game servers page: {e}", exc_info=True)
        messages.error(request, "Failed to load game servers page.")
        return redirect('guild_dashboard', guild_id=guild_id)


@discord_required
@server_owner_required
def guild_billing(request, guild_id):
    """Billing and subscription management page - SERVER OWNER ONLY."""
    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])
    all_guilds = request.session.get('discord_all_guilds', [])

    # Get guild with owner/permissions data for sidebar
    guild = get_guild_with_permissions(guild_id, admin_guilds, all_guilds)

    guild_record = None
    try:
        from django.conf import settings
        from .db import get_db_session
        from .models import Guild as GuildModel, GuildModule
        from .modules_config import MODULES, BUNDLES, get_all_modules
        from .module_utils import get_guild_modules

        # Get all available modules
        all_modules = get_all_modules()

        # Get currently active modules for this guild
        active_modules = get_guild_modules(guild_id)

        # Check if guild has VIP status or Complete tier
        is_vip = False
        subscription_tier = 'free'
        has_all_modules = False
        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                is_vip = guild_record.is_vip
                subscription_tier = guild_record.subscription_tier if guild_record.subscription_tier else 'free'
                billing_cycle = guild_record.billing_cycle
                # VIP or Complete tier get all modules
                has_all_modules = is_vip or subscription_tier == 'complete' or billing_cycle == 'lifetime'
            else:
                billing_cycle = None

        # Build module data with active status
        modules_data = []
        for module_key, module_config in all_modules.items():
            modules_data.append({
                'key': module_key,
                'name': module_config['name'],
                'short_name': module_config['short_name'],
                'description': module_config['description'],
                'icon': module_config['icon'],
                'color': module_config['color'],
                'features': module_config['features'],
                'price_monthly': module_config['price_monthly'],
                'price_yearly': module_config['price_yearly'],
                'is_active': module_key in active_modules or has_all_modules,
            })

        # Calculate pricing (filter out None values for yearly since individual modules don't have yearly pricing)
        total_monthly = sum(m['price_monthly'] for m in modules_data if m['price_monthly'] is not None)
        total_yearly = sum(m['price_yearly'] for m in modules_data if m['price_yearly'] is not None)

        # Get Stripe subscription details if available
        subscription_data = None
        if guild_record and guild_record.stripe_subscription_id:
            # Skip if it's a lifetime subscription (not a real Stripe subscription ID)
            if not guild_record.stripe_subscription_id.startswith('lifetime_'):
                try:
                    import stripe
                    stripe.api_key = settings.STRIPE_SECRET_KEY
                    subscription = stripe.Subscription.retrieve(guild_record.stripe_subscription_id)

                    # Extract price info from subscription items using dict-style access
                    amount = 0
                    interval = 'month'
                    if 'items' in subscription and subscription['items'].get('data'):
                        items_data = subscription['items']['data']
                        if items_data and len(items_data) > 0:
                            price = items_data[0]['price']
                            amount = price['unit_amount'] / 100
                            interval = price.get('recurring', {}).get('interval', 'month')

                    subscription_data = {
                        'id': subscription.id,
                        'status': subscription.get('status', 'active'),
                        'current_period_end': subscription['items']['data'][0].get('current_period_end') if subscription.get('items', {}).get('data') else None,
                        'cancel_at_period_end': subscription.get('cancel_at_period_end', False),
                        'plan_name': subscription.get('metadata', {}).get('plan_name', 'Complete Suite'),
                        'amount': amount,
                        'interval': interval,
                    }
                    logger.info(f"Fetched subscription for guild {guild_id}: status={subscription_data['status']}, cancel_at_period_end={subscription_data['cancel_at_period_end']}, current_period_end={subscription_data['current_period_end']}")
                except Exception as e:
                    logger.error(f"Error fetching Stripe subscription: {e}", exc_info=True)

        context = {
            'discord_user': discord_user,
            'guild': guild,
            'guild_record': guild_record,
            'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
            'is_admin': True,
            'active_page': 'billing',
            'modules': modules_data,
            'bundles': BUNDLES,
            'active_modules': active_modules,
            'active_module_count': len(active_modules),
            'total_modules': len(all_modules),
            'total_monthly': total_monthly,
            'total_yearly': total_yearly,
            'is_vip': is_vip,
            'subscription_tier': subscription_tier,
            'billing_cycle': billing_cycle if 'billing_cycle' in locals() else None,
            'subscription': subscription_data,
        }
        return render(request, 'questlog/billing.html', context)

    except Exception as e:
        logger.error(f"Error loading billing page: {e}")
        messages.error(request, "Error loading billing information.")
        return redirect('guild_dashboard', guild_id=guild_id)


@require_http_methods(["POST", "PATCH"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['POST', 'PATCH'], block=True)
def api_settings_update(request, guild_id):
    """POST /api/guild/<id>/settings/ - Update server settings."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                guild_record = GuildModel(guild_id=int(guild_id))
                db.add(guild_record)

            # Update fields
            if 'prefix' in data:
                guild_record.bot_prefix = data['prefix'][:10] if data['prefix'] else '!'
            if 'language' in data:
                guild_record.language = data['language']
            if 'timezone' in data:
                guild_record.timezone = data['timezone']
            if 'token_name' in data:
                guild_record.token_name = data['token_name'][:50] if data['token_name'] else 'Hero Tokens'
            if 'token_emoji' in data:
                guild_record.token_emoji = data['token_emoji'][:20] if data['token_emoji'] else ':coin:'
            if 'mod_log_channel_id' in data:
                guild_record.mod_log_channel_id = int(data['mod_log_channel_id']) if data['mod_log_channel_id'] else None
            if 'admin_roles' in data:
                # Store admin_roles as JSON string (already stringified from frontend)
                guild_record.admin_roles = data['admin_roles'] if data['admin_roles'] else None

            db.commit()

            # Note: Permission cache will auto-expire in 30 seconds, so changes take effect quickly
            # No need to manually clear cache - it's already short-lived

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@server_owner_required
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def api_settings_reset(request, guild_id):
    """POST /api/guild/<id>/settings/reset/ - Reset guild settings to defaults. SERVER OWNER ONLY."""
    try:
        from .db import get_db_session
        from .models import Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                # Reset to defaults
                guild_record.bot_prefix = '!'
                guild_record.language = 'en'
                guild_record.timezone = 'UTC'
                guild_record.token_name = 'Hero Tokens'
                guild_record.token_emoji = ':coin:'
                guild_record.mod_log_channel_id = None

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error(f"Error resetting settings for guild {guild_id}: {e}")
        return JsonResponse({'error': 'Failed to reset settings'}, status=500)


# ============================================================================
# Stripe Integration Views
# ============================================================================

@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='5/m', method='POST', block=True)
def stripe_create_checkout(request, guild_id):
    """
    POST /api/guild/<id>/stripe/checkout/ - Create Stripe checkout session.

    Body: {
        "items": [{"type": "module", "key": "engagement"}],
        "billing_cycle": "monthly"  # or "yearly"
    }
    """
    try:
        from .stripe_utils import create_checkout_session
        import json
        from django.conf import settings

        data = json.loads(request.body)
        items = data.get('items', [])
        billing_cycle = data.get('billing_cycle', 'monthly')

        if not items:
            return JsonResponse({'error': 'No items specified'}, status=400)

        # Create URLs for success and cancel
        base_url = f"https://{request.get_host()}"
        success_url = f"{base_url}/questlog/guild/{guild_id}/billing/?success=true"
        cancel_url = f"{base_url}/questlog/guild/{guild_id}/billing/?cancelled=true"

        # Create checkout session
        session = create_checkout_session(
            guild_id=guild_id,
            items=items,
            billing_cycle=billing_cycle,
            success_url=success_url,
            cancel_url=cancel_url
        )

        return JsonResponse({
            'success': True,
            'session_id': session.id,
            'checkout_url': session.url
        })

    except Exception as e:
        logger.error(f"Error creating checkout session: {e}")
        return JsonResponse({'error': 'Payment processing failed'}, status=500)


@require_http_methods(["POST"])
@csrf_exempt  # Stripe webhook uses signature verification instead of CSRF
def stripe_webhook(request):
    """
    POST /webhooks/stripe/ - Handle Stripe webhook events.

    This endpoint is called by Stripe to notify us of subscription events.
    CSRF exempt because Stripe webhook signature verification is more secure.
    """
    try:
        from django.conf import settings
        from .stripe_utils import (
            handle_checkout_completed,
            handle_subscription_updated,
            handle_subscription_deleted
        )
        import stripe

        payload = request.body
        sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')

        if not sig_header:
            return JsonResponse({'error': 'Missing signature'}, status=400)

        # Verify webhook signature
        try:
            event = stripe.Webhook.construct_event(
                payload, sig_header, settings.STRIPE_WEBHOOK_SECRET
            )
        except ValueError:
            return JsonResponse({'error': 'Invalid payload'}, status=400)
        except stripe.error.SignatureVerificationError:
            return JsonResponse({'error': 'Invalid signature'}, status=400)

        # Handle the event
        event_type = event['type']
        event_data = event['data']['object']

        logger.info(f"Received Stripe webhook: {event_type}")

        if event_type == 'checkout.session.completed':
            handle_checkout_completed(event_data)
        elif event_type == 'customer.subscription.updated':
            handle_subscription_updated(event_data)
        elif event_type == 'customer.subscription.deleted':
            handle_subscription_deleted(event_data)
        elif event_type == 'invoice.payment_failed':
            # Handle failed payments (optional - could send notification)
            logger.warning(f"Payment failed for subscription: {event_data.get('subscription')}")

        return JsonResponse({'success': True})

    except Exception as e:
        logger.error(f"Stripe webhook error: {e}")
        return JsonResponse({'error': 'Failed to process webhook'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
def stripe_cancel_subscription(request, guild_id):
    """
    POST /api/guild/<id>/stripe/cancel/ - Cancel guild's subscription.
    """
    try:
        from .stripe_utils import cancel_subscription

        success, message = cancel_subscription(guild_id)

        if success:
            return JsonResponse({'success': True, 'message': message})
        else:
            return JsonResponse({'error': message}, status=400)

    except Exception as e:
        logger.error(f"Error cancelling subscription: {e}")
        return JsonResponse({'error': 'Payment processing failed'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
def stripe_subscription_status(request, guild_id):
    """
    GET /api/guild/<id>/stripe/status/ - Get subscription status.
    """
    try:
        from .stripe_utils import get_subscription_status

        status = get_subscription_status(guild_id)

        if status:
            return JsonResponse({'success': True, 'subscription': status})
        else:
            return JsonResponse({'success': True, 'subscription': None})

    except Exception as e:
        logger.error(f"Error fetching subscription status: {e}")
        return JsonResponse({'error': 'Payment processing failed'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
def stripe_billing_portal(request, guild_id):
    """
    POST /api/guild/<id>/stripe/portal/ - Create Stripe billing portal session.
    """
    try:
        from django.conf import settings
        from .db import get_db_session
        from .models import Guild as GuildModel
        import stripe
        stripe.api_key = settings.STRIPE_SECRET_KEY

        with get_db_session() as db:
            guild = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild or not guild.stripe_customer_id:
                return JsonResponse({'error': 'No active subscription found'}, status=404)

            # Create portal session
            # Construct return URL from request
            return_url = request.build_absolute_uri(f'/questlog/guild/{guild_id}/billing/?portal=success')
            portal_session = stripe.billing_portal.Session.create(
                customer=guild.stripe_customer_id,
                return_url=return_url
            )

            return JsonResponse({'success': True, 'portal_url': portal_session.url})

    except Exception as e:
        logger.error(f"Error creating billing portal: {e}")
        return JsonResponse({'error': 'Payment processing failed'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
def stripe_transfer_subscription(request, guild_id):
    """
    POST /api/guild/<id>/stripe/transfer/ - Transfer subscription to another guild.
    """
    try:
        import json
        from .db import get_db_session
        from .models import Guild as GuildModel, GuildModule

        data = json.loads(request.body)
        new_guild_id = data.get('new_guild_id')

        if not new_guild_id:
            return JsonResponse({'error': 'new_guild_id is required'}, status=400)

        with get_db_session() as db:
            # Get source guild
            source_guild = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not source_guild:
                return JsonResponse({'error': 'Source guild not found'}, status=404)

            # Get or create target guild
            target_guild = db.query(GuildModel).filter_by(guild_id=int(new_guild_id)).first()
            if not target_guild:
                target_guild = GuildModel(
                    guild_id=int(new_guild_id),
                    is_vip=False,
                    subscription_tier='free',
                )
                db.add(target_guild)
                db.flush()

            # Transfer subscription data
            target_guild.stripe_customer_id = source_guild.stripe_customer_id
            target_guild.stripe_subscription_id = source_guild.stripe_subscription_id
            target_guild.subscription_tier = source_guild.subscription_tier

            # Transfer all modules
            source_modules = db.query(GuildModule).filter_by(guild_id=int(guild_id)).all()
            for src_module in source_modules:
                # Check if target already has this module
                target_module = db.query(GuildModule).filter_by(
                    guild_id=int(new_guild_id),
                    module_name=src_module.module_name
                ).first()

                if target_module:
                    # Update existing
                    target_module.enabled = src_module.enabled
                    target_module.expires_at = src_module.expires_at
                    target_module.stripe_subscription_id = src_module.stripe_subscription_id
                else:
                    # Create new
                    new_module = GuildModule(
                        guild_id=int(new_guild_id),
                        module_name=src_module.module_name,
                        enabled=src_module.enabled,
                        expires_at=src_module.expires_at,
                        stripe_subscription_id=src_module.stripe_subscription_id,
                    )
                    db.add(new_module)

            # Clear source guild subscription
            source_guild.stripe_customer_id = None
            source_guild.stripe_subscription_id = None
            source_guild.subscription_tier = 'free'

            # Disable all source modules
            for src_module in source_modules:
                src_module.enabled = False

            db.commit()

            logger.info(f"Transferred subscription from guild {guild_id} to {new_guild_id}")
            return JsonResponse({'success': True, 'message': 'Subscription transferred successfully'})

    except Exception as e:
        logger.error(f"Error transferring subscription: {e}")
        return JsonResponse({'error': 'Payment processing failed'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@server_owner_required
@ratelimit(key='user_or_ip', rate='2/h', method='POST', block=True)
def api_settings_remove_data(request, guild_id):
    """POST /api/guild/<id>/settings/remove-data/ - Remove all guild data. SERVER OWNER ONLY."""
    try:
        from .db import get_db_session
        from .models import Guild as GuildModel

        with get_db_session() as db:
            # Delete guild record (cascade will handle related data)
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                db.delete(guild_record)
                logger.info(f"Removed all data for guild {guild_id}")

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error(f"Error removing data for guild {guild_id}: {e}")
        return JsonResponse({'error': 'Failed to remove data'}, status=500)


# Verification Configuration Dashboard

@discord_required
def guild_verification(request, guild_id):
    """Verification configuration page."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    verification_config = {}
    guild_record = None

    try:
        from .db import get_db_session
        from .models import VerificationConfig, Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            config = db.query(VerificationConfig).filter_by(guild_id=int(guild_id)).first()

            if config:
                verification_config = {
                    'verification_type': config.verification_type.value,
                    'require_account_age': config.require_account_age,
                    'min_account_age_days': config.min_account_age_days,
                    'button_text': config.button_text,
                    'captcha_length': config.captcha_length,
                    'captcha_timeout_seconds': config.captcha_timeout_seconds,
                    'require_rules_read': config.require_rules_read,
                    'require_intro_message': config.require_intro_message,
                    'intro_channel_id': str(config.intro_channel_id) if config.intro_channel_id else '',
                    'verification_instructions': config.verification_instructions or '',
                    'verified_message': config.verified_message or 'You have been verified!',
                    'verification_timeout_hours': config.verification_timeout_hours,
                    'kick_on_timeout': config.kick_on_timeout,
                    'verification_channel_id': str(guild_record.verification_channel_id) if guild_record and guild_record.verification_channel_id else '',
                    'verified_role_id': str(guild_record.verified_role_id) if guild_record and guild_record.verified_role_id else '',
                }
            else:
                verification_config = {
                    'verification_type': 'button',
                    'require_account_age': True,
                    'min_account_age_days': 7,
                    'button_text': 'I agree to the rules',
                    'captcha_length': 6,
                    'captcha_timeout_seconds': 300,
                    'require_rules_read': False,
                    'require_intro_message': False,
                    'intro_channel_id': '',
                    'verification_instructions': '',
                    'verified_message': 'You have been verified!',
                    'verification_timeout_hours': 24,
                    'kick_on_timeout': False,
                    'verification_channel_id': '',
                    'verified_role_id': '',
                }

    except Exception as e:
        logger.warning(f"Could not fetch verification config: {e}")

    # Check if guild has moderation module access
    has_moderation_module = has_module_access(guild_id, 'moderation')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'verification_config': verification_config,
        'has_moderation_module': has_moderation_module,
        'has_any_module': has_any_module,
        'active_page': 'verification',
    }
    return render(request, 'questlog/verification.html', context)


@require_http_methods(["POST", "PATCH"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['POST', 'PATCH'], block=True)
def api_verification_config_update(request, guild_id):
    """POST /api/guild/<id>/verification/config/ - Update verification configuration."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import VerificationConfig, VerificationType, Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                guild_record = GuildModel(guild_id=int(guild_id))
                db.add(guild_record)
                db.flush()

            config = db.query(VerificationConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                config = VerificationConfig(guild_id=int(guild_id))
                db.add(config)

            # Update fields
            if 'verification_type' in data:
                config.verification_type = VerificationType(data['verification_type'])
            if 'require_account_age' in data:
                config.require_account_age = bool(data['require_account_age'])
            if 'min_account_age_days' in data:
                config.min_account_age_days = int(data['min_account_age_days'])
            if 'button_text' in data:
                config.button_text = data['button_text'][:100]
            if 'captcha_length' in data:
                config.captcha_length = max(4, min(10, int(data['captcha_length'])))
            if 'captcha_timeout_seconds' in data:
                config.captcha_timeout_seconds = int(data['captcha_timeout_seconds'])
            if 'require_rules_read' in data:
                config.require_rules_read = bool(data['require_rules_read'])
            if 'require_intro_message' in data:
                config.require_intro_message = bool(data['require_intro_message'])
            if 'intro_channel_id' in data:
                config.intro_channel_id = int(data['intro_channel_id']) if data['intro_channel_id'] else None
            if 'verification_instructions' in data:
                config.verification_instructions = data['verification_instructions']
            if 'verified_message' in data:
                config.verified_message = data['verified_message']
            if 'verification_timeout_hours' in data:
                config.verification_timeout_hours = int(data['verification_timeout_hours'])
            if 'kick_on_timeout' in data:
                config.kick_on_timeout = bool(data['kick_on_timeout'])

            # Update guild channels/roles
            if 'verification_channel_id' in data:
                guild_record.verification_channel_id = int(data['verification_channel_id']) if data['verification_channel_id'] else None
            if 'verified_role_id' in data:
                guild_record.verified_role_id = int(data['verified_role_id']) if data['verified_role_id'] else None

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# Moderation Dashboard

@discord_required
def guild_moderation(request, guild_id):
    """Moderation dashboard - warnings, bans, timeouts."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    mod_actions = []
    stats = {'total': 0, 'active': 0, 'week': 0}
    guild_record = None

    try:
        from .db import get_db_session
        from .models import Warning, ModAction, GuildMember, Guild
        from datetime import datetime

        with get_db_session() as db:
            guild_record = db.get(Guild, int(guild_id))

            # Get recent warnings
            recent_warnings = db.query(Warning).filter(
                Warning.guild_id == int(guild_id)
            ).order_by(Warning.issued_at.desc()).limit(100).all()

            # Get recent mod actions (timeouts, kicks, bans, etc.)
            # Only include user-facing moderation actions, not admin actions
            user_mod_actions = ['timeout', 'untimeout', 'kick', 'ban', 'unban', 'mute', 'unmute', 'jail', 'unjail']
            recent_mod_actions = db.query(ModAction).filter(
                ModAction.guild_id == int(guild_id),
                ModAction.action_type.in_(user_mod_actions)
            ).order_by(ModAction.timestamp.desc()).limit(100).all()

            # Combine warnings into mod_actions format
            for w in recent_warnings:
                # Get user info
                member = db.query(GuildMember).filter_by(
                    guild_id=int(guild_id),
                    user_id=w.user_id
                ).first()

                username = member.display_name or member.username or f'User {w.user_id}' if member else f'User {w.user_id}'
                formatted_date = datetime.fromtimestamp(w.issued_at).strftime('%b %d, %Y at %I:%M %p') if w.issued_at else 'Unknown'

                mod_actions.append({
                    'id': f'warning_{w.id}',
                    'action_type': 'warning',
                    'warning_id': w.id,
                    'user_id': str(w.user_id),
                    'username': username,
                    'warning_type': w.warning_type.value,
                    'reason': w.reason,
                    'severity': w.severity,
                    'moderator_id': str(w.issued_by) if w.issued_by else 'AutoMod',
                    'moderator_name': w.issued_by_name or 'AutoMod',
                    'timestamp': w.issued_at,
                    'formatted_date': formatted_date,
                    'is_active': w.is_active,
                    'pardoned': w.pardoned,
                    'duration': None,
                })

            # Add other mod actions
            for ma in recent_mod_actions:
                # Get target user info
                if ma.target_id:
                    member = db.query(GuildMember).filter_by(
                        guild_id=int(guild_id),
                        user_id=ma.target_id
                    ).first()
                    username = member.display_name or member.username or ma.target_name or f'User {ma.target_id}' if member else ma.target_name or f'User {ma.target_id}'
                else:
                    username = ma.target_name or 'Unknown'

                formatted_date = datetime.fromtimestamp(ma.timestamp).strftime('%b %d, %Y at %I:%M %p') if ma.timestamp else 'Unknown'

                # Determine if action is still active (for timeouts)
                is_active = True
                if ma.action_type in ['timeout']:
                    is_active = True  # We'll update this based on Discord API if needed
                elif ma.action_type in ['untimeout', 'kick', 'ban', 'unban', 'unjail', 'unmute']:
                    is_active = False

                mod_actions.append({
                    'id': f'action_{ma.id}',
                    'action_type': ma.action_type,
                    'mod_action_id': ma.id,
                    'user_id': str(ma.target_id) if ma.target_id else None,
                    'username': username,
                    'reason': ma.reason,
                    'moderator_id': str(ma.mod_id),
                    'moderator_name': ma.mod_name or f'Moderator {ma.mod_id}',
                    'timestamp': ma.timestamp,
                    'formatted_date': formatted_date,
                    'is_active': is_active,
                    'pardoned': False,
                    'duration': ma.duration,
                })

            # Sort all actions by timestamp descending
            mod_actions.sort(key=lambda x: x['timestamp'], reverse=True)

            # Stats - combine both warnings and mod actions
            week_ago = int(time.time()) - (7 * 24 * 60 * 60)
            warnings_count = db.query(Warning).filter(Warning.guild_id == int(guild_id)).count()
            actions_count = db.query(ModAction).filter(
                ModAction.guild_id == int(guild_id),
                ModAction.action_type.in_(user_mod_actions)
            ).count()

            stats['total'] = warnings_count + actions_count
            stats['active'] = db.query(Warning).filter(Warning.guild_id == int(guild_id), Warning.is_active == True).count()
            stats['week'] = (
                db.query(Warning).filter(Warning.guild_id == int(guild_id), Warning.issued_at >= week_ago).count() +
                db.query(ModAction).filter(
                    ModAction.guild_id == int(guild_id),
                    ModAction.action_type.in_(user_mod_actions),
                    ModAction.timestamp >= week_ago
                ).count()
            )

    except Exception as e:
        logger.warning(f"Could not fetch moderation data: {e}")
        guild_record = None

    # Check if guild has moderation module access
    has_moderation_module = has_module_access(guild_id, 'moderation')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'guild_record': guild_record,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'mod_actions': mod_actions,
        'stats': stats,
        'mod_enabled': bool(guild_record.mod_enabled) if guild_record else False,
        'has_moderation_module': has_moderation_module,
        'has_any_module': has_any_module,
        'active_page': 'moderation',
    }
    return render(request, 'questlog/moderation.html', context)


@discord_required
def guild_moderation_settings(request, guild_id):
    """Moderation settings - configure jail, mute, mod log."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    # Get guild settings from database
    settings = {}
    channels = []
    roles = []

    try:
        from .db import get_db_session
        from .models import Guild
        import json as json_lib

        with get_db_session() as db:
            db_guild = db.get(Guild, int(guild_id))
            if db_guild:
                settings = {
                    'mod_log_channel_id': str(db_guild.mod_log_channel_id) if db_guild.mod_log_channel_id else None,
                    'jail_channel_id': str(db_guild.jail_channel_id) if db_guild.jail_channel_id else None,
                    'jail_role_id': str(db_guild.jail_role_id) if db_guild.jail_role_id else None,
                    'muted_role_id': str(db_guild.muted_role_id) if db_guild.muted_role_id else None,
                    'mod_enabled': bool(db_guild.mod_enabled),
                }

                # Parse cached channels and roles
                if db_guild.cached_channels:
                    try:
                        channels = json_lib.loads(db_guild.cached_channels)
                    except:
                        channels = []

                if db_guild.cached_roles:
                    try:
                        roles = json_lib.loads(db_guild.cached_roles)
                    except:
                        roles = []

    except Exception as e:
        logger.warning(f"Could not fetch moderation settings: {e}")

    # Check if guild has moderation module access
    has_moderation_module = has_module_access(guild_id, 'moderation')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'settings': settings,
        'channels': channels,
        'roles': roles,
        'has_moderation_module': has_moderation_module,
        'has_any_module': has_any_module,
        'active_page': 'moderation_settings',
    }
    return render(request, 'questlog/moderation_settings.html', context)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_warning_pardon(request, guild_id, warning_id):
    """POST /api/guild/<id>/warnings/<warning_id>/pardon/ - Pardon a warning."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        data = {}

    try:
        from .db import get_db_session
        from .models import Warning

        discord_user = request.session.get('discord_user', {})
        pardoned_by = int(discord_user.get('id', 0))

        with get_db_session() as db:
            warning = db.query(Warning).filter(
                Warning.id == int(warning_id),
                Warning.guild_id == int(guild_id)
            ).first()

            if not warning:
                return JsonResponse({'error': 'Warning not found'}, status=404)

            warning.is_active = False
            warning.pardoned = True
            warning.pardoned_by = pardoned_by
            warning.pardoned_at = int(time.time())
            warning.pardon_reason = data.get('reason', '')

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_mod_untimeout(request, guild_id):
    """POST /api/guild/<id>/mod/untimeout/ - Remove timeout from a user."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_id = data.get('user_id')
    reason = data.get('reason', 'Timeout removed via web dashboard')

    if not user_id:
        return JsonResponse({'error': 'user_id is required'}, status=400)

    # Get requester's Discord ID from session
    discord_user = request.session.get('discord_user')
    requester_id = discord_user.get('id') if discord_user else None

    try:
        # Call Discord bot API to remove timeout
        response = requests.post(
            f"{DISCORD_BOT_API_URL}/mod/untimeout",
            json={
                'guild_id': guild_id,
                'user_id': user_id,
                'requester_id': requester_id,
                'reason': reason,
            },
            headers={'Authorization': f'Bearer {DISCORD_BOT_API_TOKEN}'},
            timeout=10
        )

        if response.status_code == 200:
            return JsonResponse({'success': True, 'message': 'Timeout removed successfully'})
        else:
            error_msg = response.json().get('error', 'Failed to remove timeout')
            return JsonResponse({'error': error_msg}, status=response.status_code)

    except Exception as e:
        logger.error(f"Error removing timeout: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def api_mod_kick(request, guild_id):
    """POST /api/guild/<id>/mod/kick/ - Kick a user from the server."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_id = data.get('user_id')
    reason = data.get('reason', 'Kicked via web dashboard')

    if not user_id:
        return JsonResponse({'error': 'user_id is required'}, status=400)

    # Get requester's Discord ID from session
    discord_user = request.session.get('discord_user')
    requester_id = discord_user.get('id') if discord_user else None

    try:
        # Call Discord bot API to kick user
        response = requests.post(
            f"{DISCORD_BOT_API_URL}/mod/kick",
            json={
                'guild_id': guild_id,
                'user_id': user_id,
                'requester_id': requester_id,
                'reason': reason,
            },
            headers={'Authorization': f'Bearer {DISCORD_BOT_API_TOKEN}'},
            timeout=10
        )

        if response.status_code == 200:
            return JsonResponse({'success': True, 'message': 'User kicked successfully'})
        else:
            error_msg = response.json().get('error', 'Failed to kick user')
            return JsonResponse({'error': error_msg}, status=response.status_code)

    except Exception as e:
        logger.error(f"Error kicking user: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='10/m', method='POST', block=True)
def api_mod_ban(request, guild_id):
    """POST /api/guild/<id>/mod/ban/ - Ban a user from the server."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_id = data.get('user_id')
    reason = data.get('reason', 'Banned via web dashboard')

    if not user_id:
        return JsonResponse({'error': 'user_id is required'}, status=400)

    # Get requester's Discord ID from session
    discord_user = request.session.get('discord_user')
    requester_id = discord_user.get('id') if discord_user else None

    try:
        # Call Discord bot API to ban user
        response = requests.post(
            f"{DISCORD_BOT_API_URL}/mod/ban",
            json={
                'guild_id': guild_id,
                'user_id': user_id,
                'requester_id': requester_id,
                'reason': reason,
            },
            headers={'Authorization': f'Bearer {DISCORD_BOT_API_TOKEN}'},
            timeout=10
        )

        if response.status_code == 200:
            return JsonResponse({'success': True, 'message': 'User banned successfully'})
        else:
            error_msg = response.json().get('error', 'Failed to ban user')
            return JsonResponse({'error': error_msg}, status=response.status_code)

    except Exception as e:
        logger.error(f"Error banning user: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_mod_settings_update(request, guild_id):
    """POST /api/guild/<id>/moderation/settings/ - Update moderation settings."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import Guild

        with get_db_session() as db:
            db_guild = db.get(Guild, int(guild_id))
            if not db_guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Update settings
            if 'mod_log_channel_id' in data:
                db_guild.mod_log_channel_id = int(data['mod_log_channel_id']) if data['mod_log_channel_id'] else None

            if 'jail_channel_id' in data:
                db_guild.jail_channel_id = int(data['jail_channel_id']) if data['jail_channel_id'] else None

            if 'jail_role_id' in data:
                db_guild.jail_role_id = int(data['jail_role_id']) if data['jail_role_id'] else None

            if 'muted_role_id' in data:
                db_guild.muted_role_id = int(data['muted_role_id']) if data['muted_role_id'] else None

            if data.get('mod_enabled') is False:
                db_guild.mod_log_channel_id = None
                db_guild.jail_channel_id = None
                db_guild.jail_role_id = None
                db_guild.muted_role_id = None
                db_guild.mod_enabled = False
            elif data.get('mod_enabled') is True:
                db_guild.mod_enabled = True

            db.commit()

            logger.info(f"Updated moderation settings for guild {guild_id}")
            return JsonResponse({'success': True, 'message': 'Settings updated successfully'})

    except Exception as e:
        logger.error(f"Error updating moderation settings: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_mod_unban(request, guild_id):
    """POST /api/guild/<id>/mod/unban/ - Unban a user from the server."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_id = data.get('user_id')
    reason = data.get('reason', 'Unbanned via web dashboard')

    if not user_id:
        return JsonResponse({'error': 'user_id is required'}, status=400)

    # Get requester's Discord ID from session
    discord_user = request.session.get('discord_user')
    requester_id = discord_user.get('id') if discord_user else None

    try:
        # Call Discord bot API to unban user
        response = requests.post(
            f"{DISCORD_BOT_API_URL}/mod/unban",
            json={
                'guild_id': guild_id,
                'user_id': user_id,
                'requester_id': requester_id,
                'reason': reason,
            },
            headers={'Authorization': f'Bearer {DISCORD_BOT_API_TOKEN}'},
            timeout=10
        )

        if response.status_code == 200:
            return JsonResponse({'success': True, 'message': 'User unbanned successfully'})
        else:
            error_msg = response.json().get('error', 'Failed to unban user')
            return JsonResponse({'error': error_msg}, status=response.status_code)

    except Exception as e:
        logger.error(f"Error unbanning user: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_mod_unmute(request, guild_id):
    """POST /api/guild/<id>/mod/unmute/ - Unmute a user."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_id = data.get('user_id')
    reason = data.get('reason', 'Unmuted via web dashboard')

    if not user_id:
        return JsonResponse({'error': 'user_id is required'}, status=400)

    # Get requester's Discord ID from session
    discord_user = request.session.get('discord_user')
    requester_id = discord_user.get('id') if discord_user else None

    try:
        # Call Discord bot API to unmute user
        response = requests.post(
            f"{DISCORD_BOT_API_URL}/mod/unmute",
            json={
                'guild_id': guild_id,
                'user_id': user_id,
                'requester_id': requester_id,
                'reason': reason,
            },
            headers={'Authorization': f'Bearer {DISCORD_BOT_API_TOKEN}'},
            timeout=10
        )

        if response.status_code == 200:
            return JsonResponse({'success': True, 'message': 'User unmuted successfully'})
        else:
            error_msg = response.json().get('error', 'Failed to unmute user')
            return JsonResponse({'error': error_msg}, status=response.status_code)

    except Exception as e:
        logger.error(f"Error unmuting user: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_mod_unjail(request, guild_id):
    """POST /api/guild/<id>/mod/unjail/ - Unjail a user."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_id = data.get('user_id')
    reason = data.get('reason', 'Unjailed via web dashboard')

    if not user_id:
        return JsonResponse({'error': 'user_id is required'}, status=400)

    # Get requester's Discord ID from session
    discord_user = request.session.get('discord_user')
    requester_id = discord_user.get('id') if discord_user else None

    try:
        # Call Discord bot API to unjail user
        response = requests.post(
            f"{DISCORD_BOT_API_URL}/mod/unjail",
            json={
                'guild_id': guild_id,
                'user_id': user_id,
                'requester_id': requester_id,
                'reason': reason,
            },
            headers={'Authorization': f'Bearer {DISCORD_BOT_API_TOKEN}'},
            timeout=10
        )

        if response.status_code == 200:
            return JsonResponse({'success': True, 'message': 'User unjailed successfully'})
        else:
            error_msg = response.json().get('error', 'Failed to unjail user')
            return JsonResponse({'error': error_msg}, status=response.status_code)

    except Exception as e:
        logger.error(f"Error unjailing user: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_warnings_list(request, guild_id):
    """GET /api/guild/<id>/warnings/ - Get warnings list with pagination."""
    try:
        from .db import get_db_session
        from .models import Warning

        page = int(request.GET.get('page', 1))
        per_page = min(int(request.GET.get('per_page', 25)), 100)
        offset = (page - 1) * per_page
        user_filter = request.GET.get('user', '')
        active_only = request.GET.get('active', 'false').lower() == 'true'

        with get_db_session() as db:
            query = db.query(Warning).filter(Warning.guild_id == int(guild_id))

            if user_filter:
                query = query.filter(Warning.user_id == int(user_filter))
            if active_only:
                query = query.filter(Warning.is_active == True)

            total = query.count()
            warnings = query.order_by(Warning.issued_at.desc()).offset(offset).limit(per_page).all()

            return JsonResponse({
                'success': True,
                'warnings': [
                    {
                        'id': w.id,
                        'user_id': str(w.user_id),
                        'warning_type': w.warning_type.value,
                        'reason': w.reason,
                        'severity': w.severity,
                        'issued_by_name': w.issued_by_name or 'AutoMod',
                        'issued_at': w.issued_at,
                        'is_active': w.is_active,
                        'pardoned': w.pardoned,
                        'action_taken': w.action_taken,
                    }
                    for w in warnings
                ],
                'total': total,
                'page': page,
                'per_page': per_page,
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# Templates Dashboard (Channel & Role Templates)

@discord_required
def guild_templates(request, guild_id):
    """Templates dashboard for channel and role templates."""
    from .module_utils import has_module_access, has_any_module_access

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    guild = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild:
        messages.error(request, "You don't have admin access to this server.")
        return redirect('questlog_dashboard')

    channel_templates = []
    role_templates = []
    is_premium = False
    guild_record = None
    total_templates = 0

    try:
        from .db import get_db_session
        from .models import ChannelTemplate, RoleTemplate, Guild as GuildModel

        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            is_premium = guild_record.is_premium() if guild_record else False

            # Get channel templates
            ch_templates = db.query(ChannelTemplate).filter(
                ChannelTemplate.guild_id == int(guild_id)
            ).order_by(ChannelTemplate.created_at.desc()).all()

            channel_templates = [
                {
                    'id': t.id,
                    'name': t.name,
                    'description': t.description,
                    'use_count': t.use_count,
                    'created_at': t.created_at,
                }
                for t in ch_templates
            ]

            # Get role templates
            r_templates = db.query(RoleTemplate).filter(
                RoleTemplate.guild_id == int(guild_id)
            ).order_by(RoleTemplate.created_at.desc()).all()

            role_templates = [
                {
                    'id': t.id,
                    'name': t.name,
                    'description': t.description,
                    'use_count': t.use_count,
                    'created_at': t.created_at,
                }
                for t in r_templates
            ]

            # Calculate total templates (combined channel + role)
            total_templates = len(ch_templates) + len(r_templates)

    except Exception as e:
        logger.warning(f"Could not fetch templates: {e}")

    # Check if guild has roles module access
    has_roles_module = has_module_access(guild_id, 'roles')
    has_any_module = has_any_module_access(guild_id)

    context = {
        'discord_user': discord_user,
        'guild': guild,
        'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
        'is_admin': True,
        'channel_templates': channel_templates,
        'role_templates': role_templates,
        'is_premium': is_premium,
        'has_roles_module': has_roles_module,
        'has_any_module': has_any_module,
        'active_page': 'templates',
        'guild_record': guild_record,
        'total_templates': total_templates,
    }
    return render(request, 'questlog/templates.html', context)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_channel_template_create(request, guild_id):
    """POST /api/guild/<id>/templates/channels/ - Create channel template."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import ChannelTemplate, RoleTemplate, Guild as GuildModel

        discord_user = request.session.get('discord_user', {})
        created_by = int(discord_user.get('id', 0))

        with get_db_session() as db:
            # Check tier limits
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Count existing templates (combined channel + role)
            channel_count = db.query(ChannelTemplate).filter_by(guild_id=int(guild_id)).count()
            role_count = db.query(RoleTemplate).filter_by(guild_id=int(guild_id)).count()
            total_templates = channel_count + role_count

            # Check limits (VIP and Complete = unlimited, Free = 5)
            if not guild_record.is_vip and guild_record.subscription_tier != 'complete':
                limit = 5
                if total_templates >= limit:
                    return JsonResponse({
                        'error': f'Template limit reached ({total_templates}/{limit} templates). Upgrade to Complete Suite for unlimited templates!'
                    }, status=403)

            template = ChannelTemplate(
                guild_id=int(guild_id),
                name=(data.get('name') or 'Untitled')[:100],
                description=(data.get('description') or '')[:500],
                template_data=json_lib.dumps(data.get('channels', [])),
                created_by=created_by,
            )
            db.add(template)
            db.flush()

            return JsonResponse({'success': True, 'id': template.id})

    except Exception as e:
        logger.error(f"Error creating channel template for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to create channel template'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_role_template_create(request, guild_id):
    """POST /api/guild/<id>/templates/roles/ - Create role template."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import RoleTemplate, ChannelTemplate, Guild as GuildModel

        discord_user = request.session.get('discord_user', {})
        created_by = int(discord_user.get('id', 0))

        with get_db_session() as db:
            # Check tier limits
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Count existing templates (combined channel + role)
            channel_count = db.query(ChannelTemplate).filter_by(guild_id=int(guild_id)).count()
            role_count = db.query(RoleTemplate).filter_by(guild_id=int(guild_id)).count()
            total_templates = channel_count + role_count

            # Check limits (VIP and Complete = unlimited, Free = 5)
            if not guild_record.is_vip and guild_record.subscription_tier != 'complete':
                limit = 5
                if total_templates >= limit:
                    return JsonResponse({
                        'error': f'Template limit reached ({total_templates}/{limit} templates). Upgrade to Complete Suite for unlimited templates!'
                    }, status=403)

            template = RoleTemplate(
                guild_id=int(guild_id),
                name=(data.get('name') or 'Untitled')[:100],
                description=(data.get('description') or '')[:500],
                template_data=json_lib.dumps(data.get('roles', [])),
                created_by=created_by,
            )
            db.add(template)
            db.flush()

            return JsonResponse({'success': True, 'id': template.id})

    except Exception as e:
        logger.error(f"Error creating role template for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to create role template'}, status=500)


@require_http_methods(["DELETE"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_template_delete(request, guild_id, template_type, template_id):
    """DELETE /api/guild/<id>/templates/<type>/<id>/ - Delete template."""
    try:
        from .db import get_db_session
        from .models import ChannelTemplate, RoleTemplate

        with get_db_session() as db:
            if template_type == 'channels':
                template = db.query(ChannelTemplate).filter(
                    ChannelTemplate.id == int(template_id),
                    ChannelTemplate.guild_id == int(guild_id)
                ).first()
            else:
                template = db.query(RoleTemplate).filter(
                    RoleTemplate.id == int(template_id),
                    RoleTemplate.guild_id == int(guild_id)
                ).first()

            if not template:
                return JsonResponse({'error': 'Template not found'}, status=404)

            db.delete(template)
            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_template_apply(request, guild_id, template_type, template_id):
    """POST /api/guild/<id>/templates/<type>/<id>/apply/ - Apply template."""
    try:
        from .db import get_db_session
        from .models import ChannelTemplate, RoleTemplate
        from .actions import queue_action, ActionType

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        with get_db_session() as db:
            if template_type == 'channels':
                template = db.query(ChannelTemplate).filter(
                    ChannelTemplate.id == int(template_id),
                    ChannelTemplate.guild_id == int(guild_id)
                ).first()
                action_type = ActionType.CHANNEL_CREATE
            else:
                template = db.query(RoleTemplate).filter(
                    RoleTemplate.id == int(template_id),
                    RoleTemplate.guild_id == int(guild_id)
                ).first()
                action_type = ActionType.ROLE_CREATE

            if not template:
                return JsonResponse({'error': 'Template not found'}, status=404)

            # Queue the action
            action_id = queue_action(
                guild_id=int(guild_id),
                action_type=action_type,
                payload={
                    'template_id': template.id,
                    'template_data': template.template_data,
                    'type': template_type,
                },
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name,
                source='website'
            )

            # Increment use count
            template.use_count += 1
            db.commit()  # Commit the usage count increment

            return JsonResponse({'success': True, 'action_id': action_id})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_template_detail(request, guild_id, template_type, template_id):
    """GET /api/guild/<id>/templates/<type>/<id>/ - Get template details."""
    try:
        from .db import get_db_session
        from .models import ChannelTemplate, RoleTemplate

        with get_db_session() as db:
            if template_type == 'channels':
                template = db.query(ChannelTemplate).filter(
                    ChannelTemplate.id == int(template_id),
                    ChannelTemplate.guild_id == int(guild_id)
                ).first()
            else:
                template = db.query(RoleTemplate).filter(
                    RoleTemplate.id == int(template_id),
                    RoleTemplate.guild_id == int(guild_id)
                ).first()

            if not template:
                return JsonResponse({'error': 'Template not found'}, status=404)

            return JsonResponse({
                'id': template.id,
                'name': template.name,
                'description': template.description,
                'template_data': template.template_data,
                'use_count': template.use_count,
                'created_at': template.created_at,
            })

    except Exception as e:
        logger.error(f"Error fetching template {template_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred.'}, status=500)


@require_http_methods(["PUT"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='PUT', block=True)
def api_template_update(request, guild_id, template_type, template_id):
    """PUT /api/guild/<id>/templates/<type>/<id>/ - Update template."""
    try:
        from .db import get_db_session
        from .models import ChannelTemplate, RoleTemplate
        import json as json_lib

        data = json_lib.loads(request.body)

        with get_db_session() as db:
            if template_type == 'channels':
                template = db.query(ChannelTemplate).filter(
                    ChannelTemplate.id == int(template_id),
                    ChannelTemplate.guild_id == int(guild_id)
                ).first()
            else:
                template = db.query(RoleTemplate).filter(
                    RoleTemplate.id == int(template_id),
                    RoleTemplate.guild_id == int(guild_id)
                ).first()

            if not template:
                return JsonResponse({'error': 'Template not found'}, status=404)

            # Update template fields
            template.name = (data.get('name') or 'Untitled')[:100]
            template.description = (data.get('description') or '')[:500]
            template.template_data = json_lib.dumps(
                data.get('channels' if template_type == 'channels' else 'roles', [])
            )

            db.commit()

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error(f"Error updating template {template_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred.'}, status=500)


# Wrapper functions for specific template types - Channels
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method=['PUT', 'DELETE'], block=True)
def api_channel_template_detail_update_delete(request, guild_id, template_id):
    """GET/PUT/DELETE /api/guild/<id>/templates/channels/<id>/ - Channel template operations."""
    if request.method == 'GET':
        return api_template_detail(request, guild_id, 'channels', template_id)
    elif request.method == 'PUT':
        return api_template_update(request, guild_id, 'channels', template_id)
    elif request.method == 'DELETE':
        return api_template_delete(request, guild_id, 'channels', template_id)
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_channel_template_apply(request, guild_id, template_id):
    """POST /api/guild/<id>/templates/channels/<id>/apply/ - Apply channel template."""
    return api_template_apply(request, guild_id, 'channels', template_id)


# Wrapper functions for specific template types - Roles
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method=['PUT', 'DELETE'], block=True)
def api_role_template_detail_update_delete(request, guild_id, template_id):
    """GET/PUT/DELETE /api/guild/<id>/templates/roles/<id>/ - Role template operations."""
    if request.method == 'GET':
        return api_template_detail(request, guild_id, 'roles', template_id)
    elif request.method == 'PUT':
        return api_template_update(request, guild_id, 'roles', template_id)
    elif request.method == 'DELETE':
        return api_template_delete(request, guild_id, 'roles', template_id)
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_role_template_apply(request, guild_id, template_id):
    """POST /api/guild/<id>/templates/roles/<id>/apply/ - Apply role template."""
    return api_template_apply(request, guild_id, 'roles', template_id)




# DISCOVERY / SELF-PROMO
@discord_required
def guild_discovery(request, guild_id):
    """Discovery/Self-Promo dashboard."""
    from .module_utils import has_module_access, has_any_module_access

    user_guilds = request.session.get('discord_admin_guilds', [])
    admin_guilds = user_guilds  # For sidebar navigation
    guild = next((g for g in user_guilds if str(g.get('id')) == str(guild_id)), None)

    if not guild:
        return redirect('questlog_dashboard')

    # Check admin permission
    permissions = int(guild.get('permissions', 0))
    is_admin = (permissions & 0x8) == 0x8 or (permissions & 0x20) == 0x20
    if not is_admin:
        return redirect('questlog_dashboard')

    try:
        from .db import get_db_session
        from .models import DiscoveryConfig, FeaturedPool, Guild, GuildMember, AnnouncedGame, GameSearchConfig
        import json as json_lib
        from datetime import datetime

        now = int(time.time())

        with get_db_session() as db:
            # Ensure guild exists first
            guild_db = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild_db:
                # Create guild record if it doesn't exist
                guild_db = Guild(
                    guild_id=int(guild_id),
                    guild_name=guild.get('name', 'Unknown'),
                    subscription_tier='free'
                )
                db.add(guild_db)
                db.flush()

            # Get or create discovery config
            discovery_config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if not discovery_config:
                discovery_config = DiscoveryConfig(guild_id=int(guild_id))
                db.add(discovery_config)
                db.flush()

            # Get pool entries and enrich with user data
            pool_entries_raw = db.query(FeaturedPool).filter(
                FeaturedPool.guild_id == int(guild_id),
                FeaturedPool.was_selected == False,
                FeaturedPool.expires_at > now
            ).order_by(FeaturedPool.entered_at.desc()).all()

            # Enrich pool entries with user info and extract URLs
            import re
            from datetime import datetime, timezone
            now_dt = datetime.fromtimestamp(now, tz=timezone.utc)
            pool_entries = []
            for entry in pool_entries_raw:
                # Get user info
                member = db.query(GuildMember).filter_by(
                    guild_id=int(guild_id),
                    user_id=entry.user_id
                ).first()

                # Extract URLs from content
                extracted_links = []
                if entry.content:
                    # Better regex to capture full URLs
                    url_pattern = r'https?://[^\s<>"\']+'
                    # url_pattern = r'https?://\S+'
                    urls = re.findall(url_pattern, entry.content)
                    for full_url in urls:
                        # Identify platform based on domain
                        platform_name = None
                        icon = None
                        color = None
                        url_lower = full_url.lower()

                        if 'twitch.tv' in url_lower:
                            platform_name = 'Twitch'
                            icon = 'fab fa-twitch'
                            color = 'purple'
                        elif 'youtube.com' in url_lower or 'youtu.be' in url_lower:
                            platform_name = 'YouTube'
                            icon = 'fab fa-youtube'
                            color = 'red'
                        elif 'twitter.com' in url_lower or 'x.com' in url_lower:
                            platform_name = 'Twitter'
                            icon = 'fab fa-twitter'
                            color = 'sky'
                        elif 'tiktok.com' in url_lower:
                            platform_name = 'TikTok'
                            icon = 'fab fa-tiktok'
                            color = 'black'
                        elif 'instagram.com' in url_lower:
                            platform_name = 'Instagram'
                            icon = 'fab fa-instagram'
                            color = 'pink'
                        elif 'bsky.app' in url_lower:
                            platform_name = 'bsky'
                            icon = 'fab fa-bluesky'
                            color = 'teal'
                        elif 'facebook.com' in url_lower or 'fb.com' in url_lower:
                            platform_name = 'Facebook'
                            icon = 'fab fa-facebook'
                            color = 'blue'
                        elif 'kick.com' in url_lower:
                            platform_name = 'Kick'
                            icon = 'fas fa-video'
                            color = 'green'
                        else:
                            platform_name = 'Link'
                            icon = 'fas fa-external-link-alt'
                            color = 'gray'

                        extracted_links.append({
                            'url': full_url,
                            'platform': platform_name,
                            'icon': icon,
                            'color': color
                        })

                # Create clean content without URLs
                clean_content = entry.content
                if clean_content and extracted_links:
                    for link in extracted_links:
                        clean_content = clean_content.replace(link['url'], '')
                    # Clean up extra whitespace
                    clean_content = ' '.join(clean_content.split())

                # Generate avatar URL
                if member and member.avatar_hash:
                    avatar_url = f"https://cdn.discordapp.com/avatars/{entry.user_id}/{member.avatar_hash}.png?size=128"
                else:
                    # Default Discord avatar (matches Discord's new avatar system)
                    default_num = (int(entry.user_id) >> 22) % 6
                    avatar_url = f"https://cdn.discordapp.com/embed/avatars/{default_num}.png"

                # Format timestamp for human readability
                entered_dt = datetime.fromtimestamp(entry.entered_at, tz=timezone.utc)
                time_diff = (now_dt - entered_dt).total_seconds()

                # Generate relative time string
                if time_diff < 60:
                    time_ago = "just now"
                elif time_diff < 3600:
                    mins = int(time_diff / 60)
                    time_ago = f"{mins} min{'s' if mins != 1 else ''} ago"
                elif time_diff < 86400:
                    hours = int(time_diff / 3600)
                    time_ago = f"{hours} hour{'s' if hours != 1 else ''} ago"
                elif time_diff < 604800:
                    days = int(time_diff / 86400)
                    time_ago = f"{days} day{'s' if days != 1 else ''} ago"
                else:
                    weeks = int(time_diff / 604800)
                    time_ago = f"{weeks} week{'s' if weeks != 1 else ''} ago"

                pool_entries.append({
                    'id': entry.id,
                    'user_id': entry.user_id,
                    'display_name': member.display_name or member.username or f'User {entry.user_id}' if member else f'User {entry.user_id}',
                    'platform': entry.platform,
                    'content': entry.content,
                    'clean_content': clean_content,
                    'link_url': entry.link_url,
                    'entered_at': entry.entered_at,
                    'entered_at_formatted': entered_dt.strftime('%b %d, %Y at %I:%M %p UTC'),
                    'time_ago': time_ago,
                    'extracted_links': extracted_links,
                    'avatar_url': avatar_url,
                })

            # Get recent features and enrich with user data
            recent_features_raw = db.query(FeaturedPool).filter(
                FeaturedPool.guild_id == int(guild_id),
                FeaturedPool.was_selected == True
            ).order_by(FeaturedPool.selected_at.desc()).limit(10).all()

            # Enrich recent features with user info and formatted dates
            recent_features = []
            for feature in recent_features_raw:
                # Get user info
                member = db.query(GuildMember).filter_by(
                    guild_id=int(guild_id),
                    user_id=feature.user_id
                ).first()

                # Format timestamp
                formatted_date = datetime.fromtimestamp(feature.selected_at).strftime('%b %d, %Y at %I:%M %p') if feature.selected_at else 'Unknown'

                recent_features.append({
                    'user_id': feature.user_id,
                    'username': member.display_name or member.username or f'User {feature.user_id}' if member else f'User {feature.user_id}',
                    'selected_at': feature.selected_at,
                    'formatted_date': formatted_date,
                    'content': feature.content,
                    'link_url': feature.link_url,
                })

            # Check premium
            guild_record = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            is_premium = guild_record.is_premium() if guild_record else False

            # Game Discovery Data - Comprehensive lists from IGDB
            available_genres = [
                "Pinball", "Adventure", "Indie", "Arcade", "Visual Novel",
                "Card & Board Game", "MOBA", "Point-and-click", "Fighting", "Shooter",
                "Music", "Platform", "Puzzle", "Racing", "Real Time Strategy (RTS)",
                "Role-playing (RPG)", "Simulator", "Sport", "Strategy",
                "Turn-based strategy (TBS)", "Tactical", "Hack and slash/Beat 'em up",
                "Quiz/Trivia"
            ]

            available_themes = [
                "Action", "Fantasy", "Science fiction", "Horror", "Thriller",
                "Survival", "Historical", "Stealth", "Comedy", "Business",
                "Drama", "Non-fiction", "Sandbox", "Educational", "Kids",
                "Open world", "Warfare", "Party", "4X (explore, expand, exploit, and exterminate)",
                "Erotic", "Mystery", "Romance"
            ]

            available_modes = [
                "Single player",
                "Multiplayer",
                "Co-operative",
                "Split screen",
                "Massively Multiplayer Online (MMO)",
                "Battle Royale",
            ]

            available_platforms = [
                "PC (Microsoft Windows)",
                "Mac",
                "Linux",
                "PlayStation 3",
                "PlayStation 4",
                "PlayStation 5",
                "Xbox 360",
                "Xbox One",
                "Xbox Series X|S",
                "Nintendo Switch",
                "Android",
                "iOS",
                "Web browser",
            ]

            # Parse selected filters from JSON
            selected_genres = json_lib.loads(discovery_config.game_genres) if discovery_config.game_genres else []
            selected_themes = json_lib.loads(discovery_config.game_themes) if discovery_config.game_themes else []
            selected_modes = json_lib.loads(discovery_config.game_modes) if discovery_config.game_modes else []
            selected_platforms = json_lib.loads(discovery_config.game_platforms) if discovery_config.game_platforms else []
            selected_tags = json_lib.loads(discovery_config.game_tags) if discovery_config.game_tags else []

            # Get recent game announcements (exclude private searches)
            recent_game_announcements_raw = db.query(AnnouncedGame).filter(
                AnnouncedGame.guild_id == int(guild_id),
                AnnouncedGame.game_name != '[PRIVATE]'  # Exclude private discoveries
            ).order_by(AnnouncedGame.announced_at.desc()).limit(10).all()

            recent_game_announcements = []
            for game in recent_game_announcements_raw:
                formatted_date = datetime.fromtimestamp(game.announced_at).strftime('%b %d, %Y') if game.announced_at else 'Unknown'
                genres_list = json_lib.loads(game.genres) if game.genres else []

                # Construct IGDB URL if slug exists
                igdb_url = f"https://www.igdb.com/games/{game.igdb_slug}" if game.igdb_slug else None

                recent_game_announcements.append({
                    'game_name': game.game_name,
                    'announced_at': game.announced_at,
                    'formatted_date': formatted_date,
                    'genres_list': genres_list,
                    'cover_url': game.cover_url,
                    'igdb_url': igdb_url,
                })

            # Count total announced games (exclude private searches)
            announced_games_count = db.query(AnnouncedGame).filter(
                AnnouncedGame.guild_id == int(guild_id),
                AnnouncedGame.game_name != '[PRIVATE]'  # Exclude private discoveries
            ).count()

            # Format last check time
            last_game_check = None
            if discovery_config.last_game_check_at:
                hours_ago = (now - discovery_config.last_game_check_at) // 3600
                last_game_check = f"{hours_ago}h ago" if hours_ago > 0 else "Just now"

            # Get game search configs
            search_configs_raw = db.query(GameSearchConfig).filter(
                GameSearchConfig.guild_id == int(guild_id)
            ).order_by(GameSearchConfig.created_at.desc()).all()

            # Process search configs for template
            search_configs = []
            for config in search_configs_raw:
                search_configs.append({
                    'id': config.id,
                    'name': config.name,
                    'enabled': config.enabled,
                    'days_ahead': config.days_ahead,
                    'genres_list': json_lib.loads(config.genres) if config.genres else [],
                    'themes_list': json_lib.loads(config.themes) if config.themes else [],
                    'modes_list': json_lib.loads(config.game_modes) if config.game_modes else [],
                    'platforms_list': json_lib.loads(config.platforms) if config.platforms else [],
                    'min_hype': config.min_hype,
                    'min_rating': config.min_rating,
                })

            # Check if guild has discovery module access
            has_discovery_module = has_module_access(guild_id, 'discovery')
            has_any_module = has_any_module_access(guild_id)

            # Get text channels and roles from Discord API for YouTube integration
            text_channels = []
            try:
                import requests
                bot_token = os.getenv('DISCORD_BOT_TOKEN', '')
                if bot_token:
                    headers = {'Authorization': f'Bot {bot_token}'}
                    channels_resp = requests.get(
                        f'https://discord.com/api/v10/guilds/{guild_id}/channels',
                        headers=headers,
                        timeout=5
                    )
                    if channels_resp.status_code == 200:
                        all_channels = channels_resp.json()
                        text_channels = [ch for ch in all_channels if ch.get('type') == 0]  # 0 = text channels
            except Exception as e:
                logger.warning(f"Failed to fetch channels for guild {guild_id}: {e}")

            context = {
                'guild': guild,
                'guild_record': guild_record,
                'discovery_config': discovery_config,
                'pool_entries': pool_entries,
                'pool_count': len(pool_entries),
                'recent_features': recent_features,
                'is_premium': is_premium,
                'is_admin': is_admin,
                # Game Discovery
                'available_genres': available_genres,
                'available_themes': available_themes,
                'available_modes': available_modes,
                'available_platforms': available_platforms,
                'selected_genres': selected_genres,
                'selected_themes': selected_themes,
                'selected_modes': selected_modes,
                'selected_platforms': selected_platforms,
                'selected_tags': selected_tags,
                'recent_game_announcements': recent_game_announcements,
                'announced_games_count': announced_games_count,
                'last_game_check': last_game_check,
                'search_configs': search_configs,
                'total_search_configs': len(search_configs),
                'has_discovery_module': has_discovery_module,
                'has_any_module': has_any_module,
                'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
                'active_page': 'discovery',
                'discovery_enabled': discovery_config.enabled if discovery_config else False,
                # YouTube Integration
                'text_channels': text_channels,
            }

            return render(request, 'questlog/discovery.html', context)

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        # Check if guild has discovery module access (for error case too)
        has_discovery_module = has_module_access(guild_id, 'discovery')
        has_any_module = has_any_module_access(guild_id)

        return render(request, 'questlog/discovery.html', {
            'guild': guild,
            'error': 'An internal error occurred. Please try again later.',

            'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
            'is_admin': is_admin,
            'has_discovery_module': has_discovery_module,
            'has_any_module': has_any_module,
            'active_page': 'discovery',
        })


@discord_required
def guild_discovery_network(request, guild_id):
    """Discovery Network dashboard - cross-server creator discovery."""
    from .module_utils import has_module_access, has_any_module_access

    # Serve Open Graph meta tags for social media crawlers (no auth required)
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_crawler = any(bot in user_agent for bot in [
        'facebookexternalhit', 'twitterbot', 'linkedinbot', 'discordbot',
        'slackbot', 'telegrambot', 'whatsapp', 'pinterest'
    ])

    if is_crawler:
        from django.http import HttpResponse
        og_html = f"""<!DOCTYPE html>
<html>
<head>
    <meta property="og:title" content="QuestLog Dashboard - Casual Heroes">
    <meta property="og:description" content="Discover gaming communities, find groups, and connect with players across the Casual Heroes network. Manage your Discord server with powerful moderation, leveling, and discovery tools.">
    <meta property="og:image" content="https://dashboard.casual-heroes.com/static/img/siteassets/homepage/CHLogoFinal.png">
    <meta property="og:url" content="https://dashboard.casual-heroes.com/questlog/guild/{guild_id}/discovery-network/">
    <meta property="og:type" content="website">
    <meta property="article:author" content="Ryven">
    <meta name="author" content="Ryven">
    <meta name="twitter:card" content="summary_large_image">
    <meta name="twitter:title" content="QuestLog Dashboard - Casual Heroes">
    <meta name="twitter:description" content="Discover gaming communities, find groups, and connect with players across the Casual Heroes network.">
    <meta name="twitter:image" content="https://dashboard.casual-heroes.com/static/img/siteassets/homepage/CHLogoFinal.png">
    <meta name="twitter:creator" content="@Ryven">
</head>
<body>
    <h1>QuestLog Dashboard - Casual Heroes</h1>
    <p>Visit <a href="https://dashboard.casual-heroes.com/">Casual Heroes</a> to explore gaming communities.</p>
</body>
</html>"""
        return HttpResponse(og_html)

    # Check if user is admin of this guild
    admin_guilds = request.session.get('discord_admin_guilds', [])
    guild = next((g for g in admin_guilds if str(g.get('id')) == str(guild_id)), None)

    is_admin = False
    if guild:
        permissions = int(guild.get('permissions', 0))
        is_admin = (permissions & 0x8) == 0x8 or (permissions & 0x20) == 0x20

    # If not admin, check if user is a member and guild is approved
    if not is_admin:
        all_guilds = request.session.get('discord_all_guilds', [])
        guild = next((g for g in all_guilds if str(g.get('id')) == str(guild_id)), None)

        if not guild:
            return redirect('questlog_dashboard')

        # Check if this guild is approved in Discovery Network
        from .db import get_db_session
        from .models import DiscoveryNetworkApplication

        with get_db_session() as db:
            application = db.query(DiscoveryNetworkApplication).filter_by(
                guild_id=int(guild_id),
                status='approved'
            ).first()

            if not application:
                # Not an admin and guild is not approved - redirect
                return redirect('questlog_dashboard')

    # Get guild owner ID from Discord
    guild_owner_id = guild.get('owner_id')
    user_id = request.session.get('discord_user', {}).get('id')
    is_owner = str(guild_owner_id) == str(user_id)

    # Check if user is bot owner (superadmin for Discovery Network)
    import os
    bot_owner_id = os.getenv('BOT_OWNER_ID')
    is_bot_owner = str(user_id) == str(bot_owner_id) if bot_owner_id else False

    # Check if user is a Discovery Approver (can manage Network COTW/COTM)
    approvers_env = os.getenv('DISCOVERY_APPROVERS', '')
    approved_ids = [id.strip() for id in approvers_env.split(',') if id.strip()]
    if bot_owner_id:
        approved_ids.append(bot_owner_id)
    is_discovery_approver = str(user_id) in approved_ids

    try:
        from .db import get_db_session
        from .models import Guild, DiscoveryNetworkApplication, DiscoveryNetworkBan

        with get_db_session() as db:
            # Get guild record
            guild_record = db.query(Guild).filter_by(guild_id=int(guild_id)).first()

            # Check SERVER's network status (per-guild, not per-user)
            server_network_status = None
            application_date = None
            denial_reason = None
            ban_reason = None
            ban_violation_type = None
            ban_appeal_allowed = False
            ban_appeal_submitted = False
            ban_appeal_reviewed = False
            ban_appeal_approved = False

            # Check if user (owner) is banned from adding ANY servers
            ban = db.query(DiscoveryNetworkBan).filter_by(user_id=int(user_id)).first()
            if ban:
                server_network_status = 'banned'
                ban_reason = ban.reason
                ban_violation_type = ban.violation_type
                ban_appeal_allowed = ban.appeal_allowed
                ban_appeal_submitted = ban.appeal_submitted
                ban_appeal_reviewed = ban.appeal_reviewed
                ban_appeal_approved = ban.appeal_approved
            else:
                # Check application status FOR THIS SPECIFIC SERVER/GUILD
                application = db.query(DiscoveryNetworkApplication).filter_by(
                    guild_id=int(guild_id)
                ).order_by(DiscoveryNetworkApplication.applied_at.desc()).first()

                if application:
                    server_network_status = application.status
                    application_date = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(application.applied_at))
                    if application.status == 'denied':
                        denial_reason = application.denial_reason

            # Check module access
            has_discovery_module = has_module_access(guild_id, 'discovery')
            has_any_module = has_any_module_access(guild_id)

            # Get user's Discovery Network preferences
            from .models import DiscoveryNetworkPreferences
            user_prefs = db.query(DiscoveryNetworkPreferences).filter_by(
                user_id=int(user_id)
            ).first()

            # Default to all enabled if no preferences exist
            enable_lfg = user_prefs.enable_lfg if user_prefs else True
            enable_games = user_prefs.enable_games if user_prefs else True
            # Only enable creators tab if guild is approved in the network
            enable_creators = (user_prefs.enable_creators if user_prefs else True) and server_network_status == 'approved'
            enable_directory = user_prefs.enable_directory if user_prefs else True

            # Check if user has a creator profile for this guild
            from .models import CreatorProfile
            user_creator_profile = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id),
                discord_id=int(user_id)
            ).first()

            has_creator_profile = user_creator_profile is not None
            profile_shared_to_network = user_creator_profile.share_to_network if user_creator_profile else False

            context = {
                'guild': guild,
                'guild_record': guild_record,
                'is_admin': is_admin,
                'is_owner': is_owner,
                'is_bot_owner': is_bot_owner,
                'is_discovery_approver': is_discovery_approver,
                'admin_guilds': admin_guilds,
                'member_guilds': get_member_guilds(request),
                'discord_user': request.session.get('discord_user', {}),
                'active_page': 'discovery_network',
                'has_discovery_module': has_discovery_module,
                'has_any_module': has_any_module,
                'user_network_status': server_network_status,  # Renamed for clarity - this is SERVER status
                'application_date': application_date,
                'denial_reason': denial_reason,
                'ban_reason': ban_reason,
                'ban_violation_type': ban_violation_type,
                'ban_appeal_allowed': ban_appeal_allowed,
                'ban_appeal_submitted': ban_appeal_submitted,
                'ban_appeal_reviewed': ban_appeal_reviewed,
                'ban_appeal_approved': ban_appeal_approved,
                'enable_lfg': enable_lfg,
                'enable_games': enable_games,
                'enable_creators': enable_creators,
                'enable_directory': enable_directory,
                'has_creator_profile': has_creator_profile,
                'profile_shared_to_network': profile_shared_to_network,
            }

            return render(request, 'questlog/discovery_network.html', context)

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return render(request, 'questlog/discovery_network.html', {
            'guild': guild,
            'error': 'An internal error occurred. Please try again later.',

            'admin_guilds': admin_guilds,
            'member_guilds': get_member_guilds(request),
            'is_admin': is_admin,
            'is_owner': False,
            'has_discovery_module': False,
            'has_any_module': False,
            'active_page': 'discovery_network',
        })


# Discovery Network API Endpoints

@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_network_servers(request):
    """Get list of servers in the Discovery Network - ONLY APPROVED members."""
    try:
        from .db import get_db_session
        from .models import Guild, DiscoveryNetworkApplication

        with get_db_session() as db:
            # ONLY get SERVERS (guilds) with APPROVED applications
            # Applications are per-server, not per-user
            approved_applications = db.query(DiscoveryNetworkApplication).filter_by(
                status='approved'
            ).all()

            # If no approved applications, return empty list
            if not approved_applications:
                return JsonResponse({
                    'success': True,
                    'servers': []
                })

            # Get approved guild IDs (server-specific approvals)
            approved_guild_ids = [app.guild_id for app in approved_applications if app.guild_id]

            # Get guilds that have been approved
            guilds = db.query(Guild).filter(
                Guild.guild_id.in_(approved_guild_ids)
            ).limit(100).all()

            servers = []
            for guild in guilds:
                # Get the server's application to get bio/description
                guild_app = next((app for app in approved_applications if app.guild_id == guild.guild_id), None)

                # Parse tags from JSON
                tags = []
                if guild_app and guild_app.tags:
                    try:
                        import json as json_lib
                        tags = json_lib.loads(guild_app.tags)
                    except:
                        tags = []  # Empty if parsing fails
                # If no tags, leave empty array instead of defaults

                # Check if join is allowed - handle NULL values properly
                # If allow_join is None (NULL in DB), default to False for safety
                allow_join = bool(guild_app.allow_join) if (guild_app and guild_app.allow_join is not None) else False

                # Get actual member count from guild
                member_count = guild.member_count if guild.member_count else 0

                # Calculate activity level based on recent messages
                # You can enhance this logic based on your needs
                activity_level = 'Low'  # Default
                if hasattr(guild, 'total_messages') and guild.total_messages:
                    if guild.total_messages > 1000:
                        activity_level = 'High'
                    elif guild.total_messages > 100:
                        activity_level = 'Medium'

                # Get primary game from guild settings if available
                primary_game = None
                if hasattr(guild, 'primary_game') and guild.primary_game:
                    primary_game = guild.primary_game

                # Construct Discord CDN URL for guild icon
                icon_url = None
                if hasattr(guild, 'guild_icon_hash') and guild.guild_icon_hash:
                    # Discord CDN format: https://cdn.discordapp.com/icons/{guild_id}/{icon_hash}.png
                    icon_url = f'https://cdn.discordapp.com/icons/{guild.guild_id}/{guild.guild_icon_hash}.png'

                # Get Discord invite URL (use invite_code from application if available)
                invite_url = None
                if guild_app and guild_app.invite_code and allow_join:
                    # Use invite code from the application (e.g., "abc123" -> "https://discord.gg/abc123")
                    invite_url = f'https://discord.gg/{guild_app.invite_code}'

                servers.append({
                    'id': str(guild.guild_id),
                    'name': guild.guild_name or 'QuestLog Community',
                    'description': guild_app.bio[:200] if guild_app and guild_app.bio else 'A QuestLog gaming community',
                    'icon_url': icon_url,  # Discord CDN URL or None
                    'member_count': member_count,  # From Guild model
                    'activity_level': activity_level,  # Calculated from guild activity
                    'primary_game': primary_game,  # From guild settings
                    'tags': tags,  # From application (empty if none selected)
                    'allow_join': allow_join,  # From application
                    'invite_url': invite_url,  # Discord invite URL (if available)
                })

            return JsonResponse({
                'success': True,
                'servers': servers
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_network_lfg(request):
    """Get cross-server LFG posts from Discovery Network guilds."""
    try:
        from .db import get_db_session
        from .models import LFGGroup, LFGGame, Guild, DiscoveryNetworkApplication
        from sqlalchemy import and_

        with get_db_session() as db:
            # Get all guilds in Discovery Network (approved applications)
            network_guilds = db.query(DiscoveryNetworkApplication.guild_id).filter(
                DiscoveryNetworkApplication.status == 'approved'
            ).all()
            network_guild_ids = [g[0] for g in network_guilds]

            if not network_guild_ids:
                return JsonResponse({
                    'success': True,
                    'posts': []
                })

            # Get active LFG groups from Discovery Network guilds (only shared groups)
            active_groups = db.query(
                LFGGroup, LFGGame, Guild
            ).join(
                LFGGame, LFGGroup.game_id == LFGGame.id
            ).join(
                Guild, LFGGroup.guild_id == Guild.guild_id
            ).filter(
                and_(
                    LFGGroup.guild_id.in_(network_guild_ids),
                    LFGGroup.is_active == True,
                    LFGGroup.is_full == False,
                    LFGGroup.shared_to_network == True  # Only show cross-server groups
                )
            ).order_by(
                LFGGroup.created_at.desc()
            ).limit(100).all()

            posts = []
            for lfg_group, game, guild in active_groups:
                # Parse custom_data for additional details
                custom_data = {}
                if lfg_group.custom_data:
                    try:
                        import json as json_lib
                        custom_data = json_lib.loads(lfg_group.custom_data)
                    except:
                        custom_data = {}

                # Determine platform, activity, skill level, player_role from custom_data
                platform = custom_data.get('platform', 'PC')
                # Activity is now stored as an array, but we return it as-is for frontend to handle
                activity = custom_data.get('activity', ['casual'])
                # Ensure activity is always an array for consistency
                if not isinstance(activity, list):
                    activity = [activity] if activity else ['casual']
                skill_level = custom_data.get('skill_level', 'any')
                voice_required = custom_data.get('voice_required', False)
                player_role = custom_data.get('player_role', '')

                # Calculate start time
                start_time = 'now'
                if lfg_group.scheduled_time:
                    now = int(time.time())
                    time_diff = lfg_group.scheduled_time - now
                    if time_diff > 7200:  # More than 2 hours
                        start_time = 'custom'
                    elif time_diff > 3600:  # More than 1 hour
                        start_time = '2hours'
                    elif time_diff > 0:
                        start_time = 'hour'

                # Get members for this group with avatar info
                from .models import LFGMember, GuildMember
                members = db.query(LFGMember).filter(
                    LFGMember.group_id == lfg_group.id,
                    LFGMember.left_at == None
                ).all()

                members_list = []
                for member in members:
                    # Try to get avatar from GuildMember table
                    guild_member = db.query(GuildMember).filter(
                        GuildMember.guild_id == lfg_group.guild_id,
                        GuildMember.user_id == member.user_id
                    ).first()

                    if guild_member and guild_member.avatar_hash:
                        # Construct Discord CDN URL with custom avatar
                        avatar_url = f"https://cdn.discordapp.com/avatars/{member.user_id}/{guild_member.avatar_hash}.png"
                    else:
                        # Use Discord's default avatar (calculated from user ID)
                        # New Discord system uses (user_id >> 22) % 6 for default avatar index
                        default_avatar_index = (int(member.user_id) >> 22) % 6
                        avatar_url = f"https://cdn.discordapp.com/embed/avatars/{default_avatar_index}.png"

                    member_data = {
                        'user_id': str(member.user_id),
                        'display_name': member.display_name or 'Unknown',
                        'is_creator': member.is_creator,
                        'is_co_leader': member.is_co_leader,
                        'avatar_url': avatar_url
                    }
                    # Add role info if available
                    if member.selections:
                        try:
                            import json as json_lib
                            selections = json_lib.loads(member.selections)
                            if selections.get('player_role'):
                                member_data['player_role'] = selections['player_role']
                        except:
                            pass
                    members_list.append(member_data)

                # Get creator's avatar
                creator_guild_member = db.query(GuildMember).filter(
                    GuildMember.guild_id == lfg_group.guild_id,
                    GuildMember.user_id == lfg_group.creator_id
                ).first()

                if creator_guild_member and creator_guild_member.avatar_hash:
                    # Use custom avatar if available
                    creator_avatar = f"https://cdn.discordapp.com/avatars/{lfg_group.creator_id}/{creator_guild_member.avatar_hash}.png"
                else:
                    # Use Discord's default avatar (calculated from user ID)
                    default_avatar_index = (int(lfg_group.creator_id) >> 22) % 6
                    creator_avatar = f"https://cdn.discordapp.com/embed/avatars/{default_avatar_index}.png"

                posts.append({
                    'id': str(lfg_group.id),
                    'game': game.game_name if game else 'Unknown',
                    'title': lfg_group.thread_name or f"LFG for {game.game_name if game else 'game'}",
                    'description': lfg_group.description or '',
                    'platform': platform,
                    'activity': activity,
                    'skill_level': skill_level,
                    'spots_needed': (lfg_group.max_group_size or game.max_group_size or 5) - lfg_group.member_count,
                    'voice_required': voice_required,
                    'start_time': start_time,
                    'scheduled_time': lfg_group.scheduled_time if lfg_group.scheduled_time else None,
                    'created_at': lfg_group.created_at,
                    'user_id': str(lfg_group.creator_id),
                    'username': lfg_group.creator_name or 'Unknown',
                    'user_avatar': creator_avatar,
                    'server_name': guild.guild_name or 'Unknown Server',
                    'guild_id': str(guild.guild_id),
                    'player_role': player_role,
                    'member_count': lfg_group.member_count,
                    'max_group_size': lfg_group.max_group_size or game.max_group_size or 5,
                    'is_raid': lfg_group.is_raid or False,
                    'tanks_needed': lfg_group.tanks_needed,
                    'healers_needed': lfg_group.healers_needed,
                    'dps_needed': lfg_group.dps_needed,
                    'members': members_list,
                    'cover_url': game.cover_url if game else None,
                    'igdb_id': game.igdb_id if game else None,
                    'igdb_slug': game.igdb_slug if game else None,
                    'event_duration': lfg_group.event_duration,
                    'activity_type': activity
                })

            return JsonResponse({
                'success': True,
                'posts': posts
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@require_http_methods(["POST"])
@ratelimit(key='user_or_ip', rate='20/h', method='POST', block=True)
def api_discovery_lfg_create(request):
    """Create a new Cross-Server LFG post in the Discovery Network."""
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import LFGGroup, LFGGame, Guild, DiscoveryNetworkApplication

        # Parse request body
        data = json_lib.loads(request.body)

        # Required fields
        guild_id = data.get('guild_id')
        game_name = data.get('game')
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()

        # Input length validation (prevent DoS)
        MAX_TITLE_LENGTH = 200
        MAX_DESCRIPTION_LENGTH = 2000
        MAX_GAME_NAME_LENGTH = 100

        if not guild_id or not game_name:
            return JsonResponse({
                'success': False,
                'error': 'Missing required fields: guild_id and game'
            }, status=400)

        if not title:
            return JsonResponse({
                'success': False,
                'error': 'Title is required'
            }, status=400)

        if len(title) > MAX_TITLE_LENGTH:
            return JsonResponse({
                'success': False,
                'error': f'Title must be {MAX_TITLE_LENGTH} characters or less'
            }, status=400)

        if len(description) > MAX_DESCRIPTION_LENGTH:
            return JsonResponse({
                'success': False,
                'error': f'Description must be {MAX_DESCRIPTION_LENGTH} characters or less'
            }, status=400)

        if len(game_name) > MAX_GAME_NAME_LENGTH:
            return JsonResponse({
                'success': False,
                'error': f'Game name must be {MAX_GAME_NAME_LENGTH} characters or less'
            }, status=400)

        with get_db_session() as db:
            # Verify guild is in Discovery Network
            network_app = db.query(DiscoveryNetworkApplication).filter(
                DiscoveryNetworkApplication.guild_id == int(guild_id),
                DiscoveryNetworkApplication.status == 'approved'
            ).first()

            if not network_app:
                return JsonResponse({
                    'success': False,
                    'error': 'Your guild is not in the Discovery Network'
                }, status=403)

            # Get guild to check tier limits
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({
                    'success': False,
                    'error': 'Guild not found'
                }, status=404)

            # Check Discovery Network LFG posting limit
            can_post, current_count, limit, reset_date = check_discovery_lfg_post_limit(db, guild_id, guild)
            if not can_post:
                return JsonResponse({
                    'success': False,
                    'error': f'Monthly LFG posting limit reached. You have posted {current_count}/{limit} LFG posts this month. Limit resets on {reset_date}. Upgrade to the LFG Module or Complete Suite for unlimited posting!',
                    'limit_reached': True,
                    'current_count': current_count,
                    'limit': limit,
                    'reset_date': reset_date
                }, status=429)

            # Get or create LFGGame
            game = db.query(LFGGame).filter(
                LFGGame.guild_id == int(guild_id),
                LFGGame.game_name == game_name
            ).first()

            if not game:
                # Try to find IGDB data from another server's game entry
                from sqlalchemy import func
                existing_game_with_igdb = db.query(LFGGame).filter(
                    func.lower(LFGGame.game_name) == game_name.lower(),
                    LFGGame.igdb_id != None,
                    LFGGame.cover_url != None
                ).first()

                # Create new game entry for this guild
                game = LFGGame(
                    guild_id=int(guild_id),
                    game_name=game_name,
                    game_short=game_name.lower().replace(' ', '_')[:50],
                    enabled=True,
                    max_group_size=5,  # Default
                    # Copy IGDB data if found
                    igdb_id=existing_game_with_igdb.igdb_id if existing_game_with_igdb else None,
                    igdb_slug=existing_game_with_igdb.igdb_slug if existing_game_with_igdb else None,
                    cover_url=existing_game_with_igdb.cover_url if existing_game_with_igdb else None
                )
                db.add(game)
                db.flush()
            elif not game.cover_url or not game.igdb_id:
                # Game exists but is missing IGDB data - try to enrich it
                from sqlalchemy import func
                existing_game_with_igdb = db.query(LFGGame).filter(
                    func.lower(LFGGame.game_name) == game_name.lower(),
                    LFGGame.igdb_id != None,
                    LFGGame.cover_url != None
                ).first()

                if existing_game_with_igdb:
                    # Update existing game with IGDB data
                    if not game.igdb_id:
                        game.igdb_id = existing_game_with_igdb.igdb_id
                    if not game.igdb_slug:
                        game.igdb_slug = existing_game_with_igdb.igdb_slug
                    if not game.cover_url:
                        game.cover_url = existing_game_with_igdb.cover_url
                    db.flush()

            # Build custom_data JSON with player role/class/spec
            # Activity can now be an array (multi-select up to 3)
            activity_value = data.get('activity', 'casual')
            # If activity is a list, keep it as is; otherwise wrap single value in list for consistency
            if isinstance(activity_value, list):
                activity = activity_value
            else:
                activity = [activity_value] if activity_value else ['casual']

            custom_data = {
                'platform': data.get('platform', 'PC'),
                'activity': activity,  # Now stored as array
                'skill_level': data.get('skill_level', 'any'),
                'voice_required': data.get('voice_required', False)
            }

            # Add player role/class/spec info
            if data.get('player_role'):
                custom_data['player_role'] = data.get('player_role')
            if data.get('player_class'):
                custom_data['player_class'] = data.get('player_class')
            if data.get('player_spec'):
                custom_data['player_spec'] = data.get('player_spec')

            # Calculate scheduled_time from scheduled_time input (datetime-local from form)
            scheduled_time = None
            scheduled_time_input = data.get('scheduled_time', '').strip()

            if scheduled_time_input:
                # Parse datetime-local format: "2025-12-23T15:30"
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(scheduled_time_input)
                    scheduled_time = int(dt.timestamp())
                except:
                    # If parsing fails, default to now
                    scheduled_time = int(time.time())
            else:
                # No scheduled time provided, default to "now"
                scheduled_time = int(time.time())

            # Get user info from session (same pattern as api_lfg_browser_create)
            discord_user = request.session.get('discord_user', {})
            user_id = discord_user.get('id')
            user_name = discord_user.get('username', 'Unknown')

            if not user_id:
                return JsonResponse({
                    'success': False,
                    'error': 'You must be logged in to create LFG posts'
                }, status=401)

            # Create LFG group (using a placeholder thread_id since this is web-based)
            # The bot will need to create actual Discord threads when users join
            lfg_group = LFGGroup(
                guild_id=int(guild_id),
                game_id=game.id,
                thread_id=0,  # Placeholder - bot will create thread on first join
                thread_name=title,
                creator_id=int(user_id),
                creator_name=user_name,
                scheduled_time=scheduled_time,
                description=description,
                custom_data=json_lib.dumps(custom_data),
                max_group_size=int(data.get('group_size', game.max_group_size or 5)),
                event_duration=int(data.get('event_duration')) if data.get('event_duration') else None,
                is_active=True,
                is_full=False,
                member_count=1,  # Creator counts as first member
                shared_to_network=True  # Discovery Network LFG posts are cross-server
            )

            db.add(lfg_group)
            db.flush()  # Get the group ID

            # Add creator as first member (leader)
            from .models import LFGMember

            # Build selections JSON for creator's class/spec/role
            creator_selections = {}
            if data.get('player_role'):
                creator_selections['player_role'] = data.get('player_role')

            creator_member = LFGMember(
                group_id=lfg_group.id,
                user_id=int(user_id),
                display_name=user_name,
                is_creator=True,
                is_co_leader=False,
                selections=json_lib.dumps(creator_selections) if creator_selections else None,
                joined_at=int(time.time())
            )

            db.add(creator_member)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'LFG post created successfully!',
                'post_id': lfg_group.id
            })

    except json_lib.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@require_http_methods(["POST"])
@ratelimit(key='user_or_ip', rate='30/h', method='POST', block=True)
def api_discovery_lfg_join(request, post_id):
    """POST /api/discovery/lfg/<post_id>/join - Join a Discovery Network LFG post."""
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import LFGGroup, LFGMember, Guild, DiscoveryNetworkApplication

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')
        username = discord_user.get('username', 'Unknown')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'You must be logged in to join LFG posts'
            }, status=401)

        # Parse request body for options (if provided)
        data = json_lib.loads(request.body) if request.body else {}
        options = data.get('options', {})

        with get_db_session() as db:
            # Get the LFG group
            group = db.query(LFGGroup).filter_by(id=int(post_id)).first()
            if not group:
                return JsonResponse({
                    'success': False,
                    'error': 'LFG post not found'
                }, status=404)

            if not group.is_active:
                return JsonResponse({
                    'success': False,
                    'error': 'This LFG post is no longer active'
                }, status=400)

            if group.is_full:
                return JsonResponse({
                    'success': False,
                    'error': 'This group is already full'
                }, status=400)

            # Check if user has any existing member record (active or left)
            existing = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id)
            ).first()

            # Extract rank if provided
            rank_value = options.get('rank')

            # Extract other selections (exclude 'rank' as it's stored separately)
            selections = {k: v for k, v in options.items() if k != 'rank'}
            selections_json = json_lib.dumps(selections) if selections else None

            if existing:
                # Check if they're currently in the group
                if existing.left_at is None:
                    return JsonResponse({
                        'success': False,
                        'error': 'You are already in this group'
                    }, status=400)

                # They left before, so rejoin them by updating their record
                existing.display_name = username
                existing.rank_value = rank_value
                existing.selections = selections_json
                existing.joined_at = int(time.time())
                existing.left_at = None  # Clear the left timestamp

                # Update group member count
                group.member_count += 1
                if group.member_count >= group.max_group_size:
                    group.is_full = True
            else:
                # Add new member
                new_member = LFGMember(
                    group_id=group.id,
                    user_id=int(user_id),
                    display_name=username,
                    is_creator=False,
                    rank_value=rank_value,
                    selections=selections_json,
                    joined_at=int(time.time())
                )
                db.add(new_member)

                # Update group member count
                group.member_count += 1
                if group.member_count >= group.max_group_size:
                    group.is_full = True

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Successfully joined the group!',
                'member_count': group.member_count
            })

    except json_lib.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@require_http_methods(["PATCH"])
@ratelimit(key='user_or_ip', rate='30/h', method='PATCH', block=True)
def api_discovery_lfg_update(request, post_id):
    """PATCH /api/discovery/lfg/<post_id>/update - Update a Discovery Network LFG post."""
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'You must be logged in'
            }, status=401)

        # Parse request body
        data = json_lib.loads(request.body)

        with get_db_session() as db:
            # Get the LFG group
            group = db.query(LFGGroup).filter_by(id=int(post_id)).first()
            if not group:
                return JsonResponse({
                    'success': False,
                    'error': 'LFG post not found'
                }, status=404)

            # Check if user is creator or co-leader
            is_creator = str(group.creator_id) == str(user_id)
            member = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id)
            ).filter(LFGMember.left_at == None).first()

            is_co_leader = member and member.is_co_leader if member else False

            if not is_creator and not is_co_leader:
                return JsonResponse({
                    'success': False,
                    'error': 'Only the creator or co-leaders can edit this post'
                }, status=403)

            # Update fields
            if 'title' in data:
                group.thread_name = data['title']
            if 'description' in data:
                group.description = data['description']
            if 'scheduled_time' in data:
                group.scheduled_time = data['scheduled_time']
            if 'event_duration' in data:
                group.event_duration = data['event_duration']
            if 'max_group_size' in data:
                group.max_group_size = data['max_group_size']
                # Update is_full status
                group.is_full = group.member_count >= group.max_group_size

            # Update custom_data (platform, activity, skill_level, voice_required)
            custom_data = json_lib.loads(group.custom_data) if group.custom_data else {}
            if 'platform' in data:
                custom_data['platform'] = data['platform']
            if 'activity_type' in data:
                custom_data['activity'] = data['activity_type']
            if 'skill_level' in data:
                custom_data['skill_level'] = data['skill_level']
            if 'voice_required' in data:
                custom_data['voice_required'] = data['voice_required']

            group.custom_data = json_lib.dumps(custom_data)

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'LFG post updated successfully'
            })

    except json_lib.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@require_http_methods(["PATCH"])
@ratelimit(key='user_or_ip', rate='60/h', method='PATCH', block=True)
def api_discovery_lfg_update_class(request, post_id):
    """PATCH /api/discovery/lfg/<post_id>/update-class - Update member's class/role."""
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'You must be logged in'
            }, status=401)

        # Parse request body
        data = json_lib.loads(request.body)
        player_role = data.get('player_role')

        if not player_role:
            return JsonResponse({
                'success': False,
                'error': 'Player role is required'
            }, status=400)

        with get_db_session() as db:
            # Get the LFG group
            group = db.query(LFGGroup).filter_by(id=int(post_id)).first()
            if not group:
                return JsonResponse({
                    'success': False,
                    'error': 'LFG post not found'
                }, status=404)

            # Get member record
            member = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id)
            ).filter(LFGMember.left_at == None).first()

            if not member:
                return JsonResponse({
                    'success': False,
                    'error': 'You are not a member of this group'
                }, status=403)

            # Update member's selections with player_role
            selections = json_lib.loads(member.selections) if member.selections else {}
            selections['player_role'] = player_role
            member.selections = json_lib.dumps(selections)

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Class updated successfully'
            })

    except json_lib.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@require_http_methods(["DELETE"])
@ratelimit(key='user_or_ip', rate='30/h', method='DELETE', block=True)
def api_discovery_lfg_delete(request, post_id):
    """DELETE /api/discovery/lfg/<post_id>/delete - Delete a Discovery Network LFG post."""
    try:
        from .db import get_db_session
        from .models import LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'You must be logged in'
            }, status=401)

        with get_db_session() as db:
            # Get the LFG group
            group = db.query(LFGGroup).filter_by(id=int(post_id)).first()
            if not group:
                return JsonResponse({
                    'success': False,
                    'error': 'LFG post not found'
                }, status=404)

            # Only creator can delete
            if str(group.creator_id) != str(user_id):
                return JsonResponse({
                    'success': False,
                    'error': 'Only the creator can delete this post'
                }, status=403)

            # Mark as inactive instead of deleting
            group.is_active = False
            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'LFG post deleted successfully'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_network_games(request):
    """Get games from LFG settings across all Discovery Network servers (for Server Directory filtering)."""
    try:
        from .db import get_db_session
        from .models import LFGGame, DiscoveryNetworkApplication

        with get_db_session() as db:
            # Get all guilds in Discovery Network (approved applications)
            network_guilds = db.query(DiscoveryNetworkApplication.guild_id).filter(
                DiscoveryNetworkApplication.status == 'approved'
            ).all()
            network_guild_ids = [g[0] for g in network_guilds]

            if not network_guild_ids:
                return JsonResponse({
                    'success': True,
                    'games': []
                })

            # Get all unique games from LFG settings in Discovery Network
            lfg_games = db.query(LFGGame).filter(
                LFGGame.guild_id.in_(network_guild_ids)
            ).order_by(LFGGame.game_name).all()

            # Build unique games list
            game_set = set()
            games_list = []

            for game in lfg_games:
                if game.game_name and game.game_name not in game_set:
                    game_set.add(game.game_name)
                    games_list.append({
                        'name': game.game_name,
                        'slug': game.game_name.lower().replace(' ', '-').replace(':', '').replace("'", '')
                    })

            # Sort alphabetically
            games_list.sort(key=lambda x: x['name'])

            return JsonResponse({
                'success': True,
                'games': games_list
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_game_templates(request):
    """Get ALL custom templates for a specific game from Discovery Network servers."""
    try:
        from .db import get_db_session
        from .models import LFGGame, DiscoveryNetworkApplication
        import json as json_lib

        game_name = request.GET.get('game')
        if not game_name:
            return JsonResponse({
                'success': False,
                'error': 'Game name required'
            }, status=400)

        with get_db_session() as db:
            from sqlalchemy import and_, or_
            import logging
            logger = logging.getLogger(__name__)

            # Get all guilds in Discovery Network
            network_guilds = db.query(DiscoveryNetworkApplication.guild_id).filter(
                DiscoveryNetworkApplication.status == 'approved'
            ).all()
            network_guild_ids = [g[0] for g in network_guilds]
            logger.info(f"Found {len(network_guild_ids)} guilds in Discovery Network")

            # Find all LFG game configurations for this game across Discovery Network
            # Check both exact match and partial match for game name
            lfg_games = db.query(LFGGame).filter(
                and_(
                    or_(
                        LFGGame.game_name.ilike(f'%{game_name}%'),
                        LFGGame.game_name == game_name
                    ),
                    LFGGame.guild_id.in_(network_guild_ids) if network_guild_ids else True,
                    LFGGame.custom_options.isnot(None)
                )
            ).all()
            logger.info(f"Found {len(lfg_games)} LFG games matching '{game_name}'")

            # Collect all unique templates
            templates = []
            seen_template_names = set()

            for lfg_game in lfg_games:
                if not lfg_game.custom_options:
                    logger.debug(f"Game {lfg_game.game_name} has no custom_options")
                    continue

                try:
                    custom_options = json_lib.loads(lfg_game.custom_options)
                    logger.debug(f"Parsed custom_options for {lfg_game.game_name}: {type(custom_options)}")

                    # Handle both array format and dict format
                    options_list = []
                    if isinstance(custom_options, list):
                        # Direct array format: [{name: "Class", choices: [...]}]
                        options_list = custom_options
                        logger.debug(f"Found array format with {len(custom_options)} options")
                    elif isinstance(custom_options, dict) and 'options' in custom_options:
                        # Dict format: {options: [{name: "Class", choices: [...]}]}
                        options_list = custom_options['options']
                        logger.debug(f"Found dict format with {len(options_list)} options")
                    else:
                        logger.debug(f"custom_options structure invalid: {custom_options}")
                        continue

                    # Process the options
                    for option in options_list:
                        if not isinstance(option, dict):
                            continue
                        template_name = option.get('name', '')
                        if template_name and template_name not in seen_template_names:
                            seen_template_names.add(template_name)
                            template_data = {
                                'name': template_name,
                                'choices': option.get('choices', [])
                            }
                            # Include depends_on if it exists (for Class + Spec systems)
                            if 'depends_on' in option:
                                template_data['depends_on'] = option['depends_on']
                            templates.append(template_data)
                            logger.info(f"Added template: {template_name} with {len(option.get('choices', []) if isinstance(option.get('choices'), list) else option.get('choices', {}))} choices")
                except Exception as e:
                    logger.error(f"Failed to parse custom_options for {lfg_game.game_name}: {e}")
                    continue

            return JsonResponse({
                'success': True,
                'templates': templates
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_game_roles(request):
    """DEPRECATED: Use api_discovery_game_templates instead. Kept for backwards compatibility."""
    return api_discovery_game_templates(request)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_network_lfg_games(request):
    """Get games from LFG settings across all Discovery Network servers."""
    try:
        from .db import get_db_session
        from .models import LFGGame, Guild, DiscoveryNetworkApplication
        from sqlalchemy import and_

        with get_db_session() as db:
            # Get all guilds in Discovery Network (approved applications)
            network_guilds = db.query(DiscoveryNetworkApplication.guild_id).filter(
                DiscoveryNetworkApplication.status == 'approved'
            ).all()
            network_guild_ids = [g[0] for g in network_guilds]

            if not network_guild_ids:
                return JsonResponse({
                    'success': True,
                    'games': []
                })

            # Get all unique games from LFG settings in Discovery Network
            lfg_games = db.query(LFGGame).filter(
                LFGGame.guild_id.in_(network_guild_ids)
            ).order_by(LFGGame.game_name).all()

            # Build unique games list
            games_set = set()
            games_list = []

            for game in lfg_games:
                if game.game_name and game.game_name not in games_set:
                    games_set.add(game.game_name)
                    games_list.append({
                        'name': game.game_name,
                        'slug': game.game_name.lower().replace(' ', '-').replace(':', '').replace("'", '')
                    })

            # Sort alphabetically
            games_list.sort(key=lambda x: x['name'])

            return JsonResponse({
                'success': True,
                'games': games_list
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_network_lfg_activities(request):
    """Get all unique activity options from LFG games across Discovery Network."""
    try:
        from .db import get_db_session
        from .models import LFGGame, DiscoveryNetworkApplication
        import json

        with get_db_session() as db:
            # Get all guilds in Discovery Network (approved applications)
            network_guilds = db.query(DiscoveryNetworkApplication.guild_id).filter(
                DiscoveryNetworkApplication.status == 'approved'
            ).all()
            network_guild_ids = [g[0] for g in network_guilds]

            if not network_guild_ids:
                return JsonResponse({
                    'success': True,
                    'activities': []
                })

            # Get all LFG games from Discovery Network
            lfg_games = db.query(LFGGame).filter(
                LFGGame.guild_id.in_(network_guild_ids)
            ).all()

            # Extract all unique activity values from custom_options
            activities_set = set()

            for game in lfg_games:
                if not game.custom_options:
                    continue

                try:
                    custom_options = json.loads(game.custom_options)
                    for option in custom_options:
                        # Look for "Activity" field
                        if option.get('name', '').lower() == 'activity':
                            choices = option.get('choices', [])
                            # Handle both array and object-based choices
                            if isinstance(choices, list):
                                for choice in choices:
                                    if choice and isinstance(choice, str):
                                        activities_set.add(choice.strip())
                            elif isinstance(choices, dict):
                                # For conditional dropdowns, get all values
                                for values in choices.values():
                                    if isinstance(values, list):
                                        for v in values:
                                            if v and isinstance(v, str):
                                                activities_set.add(v.strip())
                except:
                    continue

            # Convert to list and sort
            activities_list = sorted(list(activities_set))

            return JsonResponse({
                'success': True,
                'activities': activities_list
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_game_config(request):
    """Get custom_options configuration for a specific game across Discovery Network."""
    try:
        from .db import get_db_session
        from .models import LFGGame, DiscoveryNetworkApplication
        import json

        game_name = request.GET.get('game_name')

        if not game_name:
            return JsonResponse({
                'success': False,
                'error': 'game_name parameter is required'
            }, status=400)

        with get_db_session() as db:
            # Get all guilds in Discovery Network (approved applications)
            network_guilds = db.query(DiscoveryNetworkApplication.guild_id).filter(
                DiscoveryNetworkApplication.status == 'approved'
            ).all()
            network_guild_ids = [g[0] for g in network_guilds]

            if not network_guild_ids:
                return JsonResponse({
                    'success': True,
                    'custom_options': []
                })

            # Find this game in any Discovery Network server
            # Try exact match first, then case-insensitive
            lfg_game = db.query(LFGGame).filter(
                LFGGame.guild_id.in_(network_guild_ids),
                LFGGame.game_name == game_name
            ).first()

            if not lfg_game:
                # Try case-insensitive match
                lfg_game = db.query(LFGGame).filter(
                    LFGGame.guild_id.in_(network_guild_ids),
                    LFGGame.game_name.ilike(game_name)
                ).first()

            if not lfg_game or not lfg_game.custom_options:
                return JsonResponse({
                    'success': True,
                    'custom_options': []
                })

            # Parse and return custom_options
            try:
                custom_options = json.loads(lfg_game.custom_options)
                return JsonResponse({
                    'success': True,
                    'custom_options': custom_options
                })
            except:
                return JsonResponse({
                    'success': True,
                    'custom_options': []
                })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@api_owner_required  # SECURITY: Owner-only access (not admins)
@require_http_methods(["POST"])
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_network_apply(request, guild_id):
    """Submit application to join Discovery Network.

    Security: OWNER-ONLY - Only the server owner can apply to Discovery Network.
    This prevents non-owner admins from applying for guilds they don't own.
    """
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import DiscoveryNetworkApplication, DiscoveryNetworkBan

        data = json_lib.loads(request.body)
        user = request.session.get('discord_user', {})
        user_id = user.get('id')
        # guild_id now comes from URL parameter, validated by @api_auth_required

        # Note: user_id and guild_id are guaranteed by @api_auth_required decorator
        with get_db_session() as db:
            # Check if user (owner) is banned from adding ANY servers
            ban = db.query(DiscoveryNetworkBan).filter_by(user_id=int(user_id)).first()
            if ban:
                return JsonResponse({
                    'success': False,
                    'error': 'You are banned from the Discovery Network'
                }, status=403)

            # Check for existing pending application FOR THIS SPECIFIC SERVER
            existing = db.query(DiscoveryNetworkApplication).filter_by(
                guild_id=int(guild_id),
                status='pending'
            ).first()

            if existing:
                return JsonResponse({
                    'success': False,
                    'error': 'This server already has a pending application'
                }, status=400)

            # Serialize tags to JSON
            tags = data.get('tags', [])
            tags_json = json_lib.dumps(tags) if tags else None

            # Create new application for this specific server
            application = DiscoveryNetworkApplication(
                user_id=int(user_id),
                guild_id=int(guild_id),  # Server-specific application
                bio=data.get('bio', ''),
                twitch_url=data.get('twitch_url'),
                youtube_url=data.get('youtube_url'),
                twitter_url=data.get('twitter_url'),
                tiktok_url=data.get('tiktok_url'),
                instagram_url=data.get('instagram_url'),
                bsky_url=data.get('bsky_url'),
                username=user.get('username', 'Unknown'),
                display_name=user.get('global_name'),
                avatar_url=f"https://cdn.discordapp.com/avatars/{user_id}/{user.get('avatar')}.png" if user.get('avatar') else None,
                account_created_at=int(time.time()),  # Would need actual Discord account creation time
                guidelines_accepted=data.get('guidelines_accepted', False),
                tos_accepted=data.get('tos_accepted', False),
                content_policy_accepted=data.get('content_policy_accepted', False),
                tags=tags_json,  # JSON array of tags
                allow_join=data.get('allow_join', False),  # Allow others to join
                status='pending',
                applied_at=int(time.time()),
                updated_at=int(time.time())
            )

            db.add(application)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Application submitted successfully'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@api_auth_required  # SEC-005 fix: Validates guild admin access + enables CSRF
@require_http_methods(["GET", "POST"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_network_preferences(request, guild_id):
    """Get or update Discovery Network preferences for the current user.

    Security: SEC-005 fix - guild_id now comes from URL parameter and is validated by @api_auth_required.
    """
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import DiscoveryNetworkPreferences

        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'Not authenticated'
            }, status=401)

        with get_db_session() as db:
            # GET - Load preferences
            if request.method == 'GET':
                prefs = db.query(DiscoveryNetworkPreferences).filter_by(user_id=int(user_id)).first()

                # Also get allow_join and invite_code from the server's application (server setting, not user pref)
                from .models import DiscoveryNetworkApplication
                # guild_id comes from URL parameter, validated by @api_auth_required
                allow_join = False
                invite_code = None

                application = db.query(DiscoveryNetworkApplication).filter_by(
                    guild_id=int(guild_id),
                    status='approved'
                ).first()

                if application:
                    allow_join = application.allow_join
                    invite_code = application.invite_code

                if not prefs:
                    # Return default preferences
                    return JsonResponse({
                        'success': True,
                        'preferences': {
                            'enable_lfg': True,
                            'enable_games': True,
                            'enable_creators': True,
                            'enable_directory': True,
                            'preferred_games': [],
                            'preferred_tags': [],
                            'preferred_activities': [],
                            'preferred_skill_levels': [],
                            'preferred_size': '',
                            'lfg_filter_games': False,
                            'lfg_filter_activities': False,
                            'lfg_filter_skill_levels': False,
                            'lfg_show_now': True,
                            'lfg_hide_voice': False,
                            'notify_lfg': False,
                            'notify_servers': False,
                            'notify_digest': False,
                            'privacy_show_profile': True,
                            'privacy_show_server': True,
                            'privacy_allow_dms': True,
                            'allow_join': allow_join,
                            'invite_code': invite_code
                        }
                    })

                # Parse JSON fields
                preferred_games = []
                preferred_tags = []
                preferred_activities = []
                preferred_skill_levels = []
                try:
                    if prefs.preferred_games:
                        preferred_games = json_lib.loads(prefs.preferred_games)
                    if prefs.preferred_tags:
                        preferred_tags = json_lib.loads(prefs.preferred_tags)
                    if prefs.preferred_activities:
                        preferred_activities = json_lib.loads(prefs.preferred_activities)
                    if prefs.preferred_skill_levels:
                        preferred_skill_levels = json_lib.loads(prefs.preferred_skill_levels)
                except:
                    pass

                return JsonResponse({
                    'success': True,
                    'preferences': {
                        'enable_lfg': prefs.enable_lfg,
                        'enable_games': prefs.enable_games,
                        'enable_creators': prefs.enable_creators,
                        'enable_directory': prefs.enable_directory,
                        'preferred_games': preferred_games,
                        'preferred_tags': preferred_tags,
                        'preferred_activities': preferred_activities,
                        'preferred_skill_levels': preferred_skill_levels,
                        'preferred_size': prefs.preferred_size or '',
                        'lfg_filter_games': prefs.lfg_filter_games,
                        'lfg_filter_activities': prefs.lfg_filter_activities,
                        'lfg_filter_skill_levels': prefs.lfg_filter_skill_levels,
                        'lfg_show_now': prefs.lfg_show_now,
                        'lfg_hide_voice': prefs.lfg_hide_voice,
                        'notify_lfg': prefs.notify_lfg,
                        'notify_servers': prefs.notify_servers,
                        'notify_digest': prefs.notify_digest,
                        'privacy_show_profile': prefs.privacy_show_profile,
                        'privacy_show_server': prefs.privacy_show_server,
                        'privacy_allow_dms': prefs.privacy_allow_dms,
                        'allow_join': allow_join,
                        'invite_code': invite_code
                    }
                })

            # POST - Save preferences
            elif request.method == 'POST':
                data = json_lib.loads(request.body)

                # Get or create preferences record
                prefs = db.query(DiscoveryNetworkPreferences).filter_by(user_id=int(user_id)).first()
                if not prefs:
                    prefs = DiscoveryNetworkPreferences(user_id=int(user_id))
                    db.add(prefs)

                # Update boolean fields
                if 'enable_lfg' in data:
                    prefs.enable_lfg = bool(data['enable_lfg'])
                if 'enable_games' in data:
                    prefs.enable_games = bool(data['enable_games'])
                if 'enable_creators' in data:
                    prefs.enable_creators = bool(data['enable_creators'])
                if 'enable_directory' in data:
                    prefs.enable_directory = bool(data['enable_directory'])

                # Update preference arrays (stored as JSON)
                if 'preferred_games' in data:
                    prefs.preferred_games = json_lib.dumps(data['preferred_games']) if data['preferred_games'] else None
                if 'preferred_tags' in data:
                    prefs.preferred_tags = json_lib.dumps(data['preferred_tags']) if data['preferred_tags'] else None
                if 'preferred_activities' in data:
                    prefs.preferred_activities = json_lib.dumps(data['preferred_activities']) if data['preferred_activities'] else None
                if 'preferred_skill_levels' in data:
                    prefs.preferred_skill_levels = json_lib.dumps(data['preferred_skill_levels']) if data['preferred_skill_levels'] else None
                if 'preferred_size' in data:
                    prefs.preferred_size = data['preferred_size'] or None

                # Update LFG preferences
                if 'lfg_filter_games' in data:
                    prefs.lfg_filter_games = bool(data['lfg_filter_games'])
                if 'lfg_filter_activities' in data:
                    prefs.lfg_filter_activities = bool(data['lfg_filter_activities'])
                if 'lfg_filter_skill_levels' in data:
                    prefs.lfg_filter_skill_levels = bool(data['lfg_filter_skill_levels'])
                if 'lfg_show_now' in data:
                    prefs.lfg_show_now = bool(data['lfg_show_now'])
                if 'lfg_hide_voice' in data:
                    prefs.lfg_hide_voice = bool(data['lfg_hide_voice'])

                # Update notification preferences
                if 'notify_lfg' in data:
                    prefs.notify_lfg = bool(data['notify_lfg'])
                if 'notify_servers' in data:
                    prefs.notify_servers = bool(data['notify_servers'])
                if 'notify_digest' in data:
                    prefs.notify_digest = bool(data['notify_digest'])

                # Update privacy preferences
                if 'privacy_show_profile' in data:
                    prefs.privacy_show_profile = bool(data['privacy_show_profile'])
                if 'privacy_show_server' in data:
                    prefs.privacy_show_server = bool(data['privacy_show_server'])
                if 'privacy_allow_dms' in data:
                    prefs.privacy_allow_dms = bool(data['privacy_allow_dms'])

                # Update timestamp
                prefs.updated_at = int(time.time())

                # ALSO UPDATE ALLOW_JOIN AND INVITE_CODE IN THE APPLICATION TABLE (SERVER SETTINGS)
                # These are server settings, not user preferences, so they need to update the application
                # guild_id comes from URL parameter and user is already validated as admin by @api_auth_required
                if 'allow_join' in data or 'invite_code' in data:
                    from .models import DiscoveryNetworkApplication

                    # Find the server's application
                    application = db.query(DiscoveryNetworkApplication).filter_by(
                        guild_id=int(guild_id),
                        status='approved'
                    ).first()

                    if application:
                        if 'allow_join' in data:
                            application.allow_join = bool(data['allow_join'])
                        if 'invite_code' in data:
                            # Strip whitespace and set to None if empty
                            invite_code_value = str(data['invite_code']).strip() if data['invite_code'] else None
                            application.invite_code = invite_code_value if invite_code_value else None
                        application.updated_at = int(time.time())

                db.commit()

                return JsonResponse({
                    'success': True,
                    'message': 'Preferences saved successfully'
                })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@api_owner_required  # SECURITY: Owner-only access (not admins)
@require_http_methods(["POST"])
@ratelimit(key='user', rate='10/h', method='POST', block=True)
def api_discovery_network_leave(request, guild_id):
    """Leave the Discovery Network - keeps approval status so they can rejoin without reapplying.

    Security: OWNER-ONLY - Only the server owner can leave the Discovery Network.
    This prevents non-owner admins from leaving the network for guilds they don't own.
    """
    try:
        from .db import get_db_session
        from .models import DiscoveryNetworkApplication

        user = request.session.get('discord_user', {})
        user_id = user.get('id')
        # guild_id comes from URL parameter, validated by @api_auth_required

        logger.info(
            f"[SECURITY] Admin {user_id} leaving Discovery Network for guild {guild_id} | "
            f"IP: {request.META.get('REMOTE_ADDR')}"
        )

        with get_db_session() as db:
            # Find the server's application
            application = db.query(DiscoveryNetworkApplication).filter_by(
                guild_id=int(guild_id),
                status='approved'
            ).first()

            if not application:
                return JsonResponse({
                    'success': False,
                    'error': 'No approved application found for this server'
                }, status=404)

            # Don't delete the application - just mark it as inactive
            # This way they keep their approved status and can rejoin anytime
            # We'll add a new field to track this
            # For now, we'll change status to 'left' (we can add this status)
            application.status = 'left'
            application.updated_at = int(time.time())

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Successfully left the Discovery Network. You can rejoin anytime.'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@api_owner_required  # SECURITY: Owner-only access (not admins)
@require_http_methods(["POST"])
@ratelimit(key='user', rate='10/h', method='POST', block=True)
def api_discovery_network_rejoin(request, guild_id):
    """Rejoin the Discovery Network - restore approved status for servers who left within 90 days.

    Security: OWNER-ONLY - Only the server owner can rejoin the Discovery Network.
    This prevents non-owner admins from rejoining the network for guilds they don't own.
    """
    try:
        from .db import get_db_session
        from .models import DiscoveryNetworkApplication

        user = request.session.get('discord_user', {})
        user_id = user.get('id')
        # guild_id comes from URL parameter, validated by @api_auth_required

        with get_db_session() as db:
            # Find the server's application with 'left' status
            application = db.query(DiscoveryNetworkApplication).filter_by(
                guild_id=int(guild_id),
                status='left'
            ).first()

            if not application:
                return JsonResponse({
                    'success': False,
                    'error': 'No previous membership found. Please apply to join the network.'
                }, status=404)

            # Check if they left more than 90 days ago
            current_time = int(time.time())
            ninety_days_seconds = 90 * 24 * 60 * 60  # 90 days in seconds
            time_since_left = current_time - application.updated_at

            if time_since_left > ninety_days_seconds:
                # More than 90 days - they need to reapply
                return JsonResponse({
                    'success': False,
                    'error': 'reapply_required',
                    'message': 'It has been more than 90 days since you left the network. Please submit a new application.'
                }, status=400)

            # Restore to approved status (within 90 days)
            application.status = 'approved'
            application.updated_at = int(time.time())

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Welcome back to the Discovery Network!'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='30/m', method='GET', block=True)
def api_discovery_games_list(request):
    """
    GET /api/discovery/games-list - Get aggregated game list from Discovery Network.
    Aggregates games from announced_games and lfg_groups, showing counts and ratings.
    Query params: guild_id, search, genre, sort
    """
    try:
        from .db import get_db_session
        from .models import AnnouncedGame, LFGGroup, LFGGame, DiscoveryGameReview, Guild
        from sqlalchemy import func, case
        from collections import defaultdict

        guild_id = request.GET.get('guild_id')
        search_term = request.GET.get('search', '').strip().lower()
        genre_filter = request.GET.get('genre', '').strip().lower()
        sort_by = request.GET.get('sort', 'trending')  # trending, servers, recent, rating

        with get_db_session() as db:
            # Verify guild is in Discovery Network (approved application)
            from .models import DiscoveryNetworkApplication

            # Get all approved guilds (status='approved')
            approved_apps = db.query(DiscoveryNetworkApplication.guild_id).filter(
                DiscoveryNetworkApplication.status == 'approved'
            ).all()
            approved_guild_ids = [g[0] for g in approved_apps if g[0] is not None]

            # If checking specific guild, verify it's approved
            if guild_id:
                if int(guild_id) not in approved_guild_ids:
                    return JsonResponse({
                        'success': False,
                        'error': 'Guild is not a member of the Discovery Network. Please apply first!'
                    }, status=403)

            if not approved_guild_ids:
                return JsonResponse({
                    'success': True,
                    'games': []
                })

            # Aggregate game data from multiple sources
            games_data = defaultdict(lambda: {
                'name': '',
                'genre': None,
                'description': None,
                'release_date': None,
                'genres': None,
                'platforms': None,
                'server_count': 0,
                'server_guild_ids': set(),  # Track which servers announced this game
                'lfg_count': 0,
                'total_reviews': 0,
                'avg_rating': 0,
                'latest_timestamp': 0,
                'cover_url': None,
                'igdb_slug': None,
                'steam_id': None,
                'steam_url': None,
                'hypes': None,
                'shared_by_user_id': None,
                'shared_by_guild_id': None,
                'is_manual': False
            })

            # 1. Get games from announced_games (Found Games) - Fetch ALL instances to track guild IDs
            announced_games = db.query(
                AnnouncedGame.game_name,
                AnnouncedGame.guild_id,
                AnnouncedGame.genre,
                AnnouncedGame.description,
                AnnouncedGame.release_date,
                AnnouncedGame.genres,
                AnnouncedGame.platforms,
                AnnouncedGame.cover_url,
                AnnouncedGame.igdb_slug,
                AnnouncedGame.steam_id,
                AnnouncedGame.created_at
            ).filter(
                AnnouncedGame.guild_id.in_(approved_guild_ids)
            ).all()

            for game in announced_games:
                game_name_lower = game.game_name.lower()
                games_data[game_name_lower]['name'] = game.game_name  # Keep original casing
                games_data[game_name_lower]['genre'] = game.genre
                # Track which servers announced this game
                games_data[game_name_lower]['server_guild_ids'].add(game.guild_id)
                if game.created_at and game.created_at > games_data[game_name_lower]['latest_timestamp']:
                    games_data[game_name_lower]['latest_timestamp'] = game.created_at
                # Use first non-null values found
                if not games_data[game_name_lower]['cover_url'] and game.cover_url:
                    games_data[game_name_lower]['cover_url'] = game.cover_url
                if not games_data[game_name_lower]['igdb_slug'] and game.igdb_slug:
                    games_data[game_name_lower]['igdb_slug'] = game.igdb_slug
                if not games_data[game_name_lower]['steam_id'] and game.steam_id:
                    games_data[game_name_lower]['steam_id'] = game.steam_id
                if not games_data[game_name_lower]['description'] and game.description:
                    games_data[game_name_lower]['description'] = game.description
                if not games_data[game_name_lower]['release_date'] and game.release_date:
                    games_data[game_name_lower]['release_date'] = game.release_date
                if not games_data[game_name_lower]['genres'] and game.genres:
                    games_data[game_name_lower]['genres'] = game.genres
                if not games_data[game_name_lower]['platforms'] and game.platforms:
                    games_data[game_name_lower]['platforms'] = game.platforms

            # 2. Get games from found_games (has Steam URLs and rich IGDB data)
            # ONLY include games from public search configs (show_on_website=True)
            from .models import FoundGame, GameSearchConfig
            found_games = db.query(
                FoundGame.game_name,
                FoundGame.cover_url,
                FoundGame.igdb_slug,
                FoundGame.steam_url,
                FoundGame.release_date,
                FoundGame.summary,
                FoundGame.genres,
                FoundGame.platforms,
                FoundGame.hypes,
                FoundGame.rating,
                func.count(func.distinct(FoundGame.guild_id)).label('found_count')
            ).join(
                GameSearchConfig,
                FoundGame.search_config_id == GameSearchConfig.id
            ).filter(
                FoundGame.guild_id.in_(approved_guild_ids),
                GameSearchConfig.show_on_website == True  # ONLY public games
            ).group_by(
                FoundGame.game_name,
                FoundGame.cover_url,
                FoundGame.igdb_slug,
                FoundGame.steam_url,
                FoundGame.release_date,
                FoundGame.summary,
                FoundGame.genres,
                FoundGame.platforms,
                FoundGame.hypes,
                FoundGame.rating
            ).all()

            for game in found_games:
                game_name_lower = game.game_name.lower()
                if game_name_lower not in games_data:
                    games_data[game_name_lower]['name'] = game.game_name
                # Enrich with found_games data (prefer this over announced_games since it has Steam URLs)
                if not games_data[game_name_lower]['cover_url'] and game.cover_url:
                    games_data[game_name_lower]['cover_url'] = game.cover_url
                if not games_data[game_name_lower]['igdb_slug'] and game.igdb_slug:
                    games_data[game_name_lower]['igdb_slug'] = game.igdb_slug
                if game.steam_url:  # Steam URL from found_games
                    games_data[game_name_lower]['steam_url'] = game.steam_url
                if not games_data[game_name_lower]['release_date'] and game.release_date:
                    games_data[game_name_lower]['release_date'] = game.release_date
                if not games_data[game_name_lower]['description'] and game.summary:
                    games_data[game_name_lower]['description'] = game.summary
                if not games_data[game_name_lower]['genres'] and game.genres:
                    games_data[game_name_lower]['genres'] = game.genres
                if not games_data[game_name_lower]['platforms'] and game.platforms:
                    games_data[game_name_lower]['platforms'] = game.platforms
                if game.hypes:
                    games_data[game_name_lower]['hypes'] = game.hypes
                if game.rating and not games_data[game_name_lower]['avg_rating']:
                    games_data[game_name_lower]['avg_rating'] = game.rating

            # 3. Get active Cross-Server LFG count per game (only shared_to_network=True)
            active_lfgs = db.query(
                LFGGame.game_name,
                LFGGame.cover_url,
                LFGGame.igdb_slug,
                func.count(func.distinct(LFGGroup.id)).label('lfg_count')
            ).join(
                LFGGroup, LFGGroup.game_id == LFGGame.id
            ).filter(
                LFGGame.guild_id.in_(approved_guild_ids),
                LFGGroup.is_active == True,
                LFGGroup.shared_to_network == True  # Only count Cross-Server LFG groups
            ).group_by(
                LFGGame.game_name,
                LFGGame.cover_url,
                LFGGame.igdb_slug
            ).all()

            for game in active_lfgs:
                if game.game_name:
                    game_name_lower = game.game_name.lower()
                    if game_name_lower not in games_data:
                        games_data[game_name_lower]['name'] = game.game_name
                    games_data[game_name_lower]['lfg_count'] = game.lfg_count
                    # Use LFG game cover if we don't have one yet
                    if not games_data[game_name_lower]['cover_url'] and game.cover_url:
                        games_data[game_name_lower]['cover_url'] = game.cover_url
                    if not games_data[game_name_lower]['igdb_slug'] and game.igdb_slug:
                        games_data[game_name_lower]['igdb_slug'] = game.igdb_slug

            # 3. Get review stats
            review_stats = db.query(
                func.lower(DiscoveryGameReview.game_name).label('game_name_lower'),
                func.avg(DiscoveryGameReview.rating).label('avg_rating'),
                func.count(DiscoveryGameReview.id).label('review_count')
            ).filter(
                DiscoveryGameReview.is_flagged == False
            ).group_by(
                func.lower(DiscoveryGameReview.game_name)
            ).all()

            for stat in review_stats:
                game_name_lower = stat.game_name_lower
                if game_name_lower in games_data:
                    games_data[game_name_lower]['avg_rating'] = float(stat.avg_rating) if stat.avg_rating else 0
                    games_data[game_name_lower]['total_reviews'] = stat.review_count

            # 4. Get ONE example manual share per game (for attribution)
            manual_shares = db.query(
                func.lower(AnnouncedGame.game_name).label('game_name_lower'),
                AnnouncedGame.shared_by_user_id,
                AnnouncedGame.guild_id,
                AnnouncedGame.is_manual
            ).filter(
                AnnouncedGame.is_manual == True,
                AnnouncedGame.shared_by_user_id.isnot(None),
                AnnouncedGame.guild_id.in_(approved_guild_ids)
            ).distinct(
                func.lower(AnnouncedGame.game_name)
            ).all()

            for share in manual_shares:
                game_name_lower = share.game_name_lower
                if game_name_lower in games_data:
                    games_data[game_name_lower]['shared_by_user_id'] = share.shared_by_user_id
                    games_data[game_name_lower]['shared_by_guild_id'] = share.guild_id
                    games_data[game_name_lower]['is_manual'] = True

            # Convert to list and apply filters
            games_list = []
            for game_key, data in games_data.items():
                # Apply search filter
                if search_term and search_term not in game_key:
                    continue

                # Apply genre filter
                if genre_filter and (not data['genre'] or data['genre'].lower() != genre_filter):
                    continue

                # Build IGDB URL
                igdb_url = None
                if data['igdb_slug']:
                    igdb_url = f"https://www.igdb.com/games/{data['igdb_slug']}"

                # Use steam_url from found_games if available, otherwise construct from steam_id
                steam_url = data['steam_url']  # Prefer full URL from found_games
                if not steam_url and data['steam_id']:
                    steam_url = f"https://store.steampowered.com/app/{data['steam_id']}"

                # Parse JSON fields
                import json
                genres_list = None
                platforms_list = None
                if data['genres']:
                    try:
                        genres_list = json.loads(data['genres']) if isinstance(data['genres'], str) else data['genres']
                    except:
                        pass
                if data['platforms']:
                    try:
                        platforms_list = json.loads(data['platforms']) if isinstance(data['platforms'], str) else data['platforms']
                    except:
                        pass

                # Format release date
                release_date_str = None
                if data['release_date']:
                    try:
                        from datetime import datetime
                        release_date_str = datetime.fromtimestamp(data['release_date']).strftime('%b %d, %Y')
                    except:
                        pass

                # Use description or fallback to summary from IGDB if available
                summary = data['description']

                # Get user and guild info for manual shares
                shared_by_user = None
                shared_by_server = None
                if data['is_manual'] and data['shared_by_user_id'] and data['shared_by_guild_id']:
                    # Get Discord user info via API
                    try:
                        # Get member from cache (no Discord API call!)
                        from .discord_resources import get_guild_member

                        member_data = get_guild_member(str(data["shared_by_guild_id"]), str(data["shared_by_user_id"]))
                        if member_data:
                            # Get display name from cached data
                            shared_by_user = member_data.get('display_name') or member_data.get('username', 'Unknown User')
                    except Exception as e:
                        logger.warning(f"Failed to fetch Discord user {data['shared_by_user_id']} from cache: {e}")

                    # Get guild name
                    sharing_guild = db.query(Guild).filter_by(guild_id=data['shared_by_guild_id']).first()
                    if sharing_guild:
                        shared_by_server = sharing_guild.guild_name

                # Get list of server names that announced this game
                server_names = []
                if data['server_guild_ids']:
                    guild_lookup = db.query(Guild.guild_id, Guild.guild_name).filter(
                        Guild.guild_id.in_(list(data['server_guild_ids']))
                    ).all()
                    server_names = [name for _, name in guild_lookup if name]

                games_list.append({
                    'id': game_key,  # Use lowercase name as ID
                    'name': data['name'],
                    'genre': data['genre'],
                    'summary': summary,
                    'release_date': release_date_str,
                    'genres': genres_list,
                    'platforms': platforms_list,
                    'server_count': len(data['server_guild_ids']),  # Use actual count from set
                    'server_names': server_names,  # List of server names that announced this game
                    'lfg_count': data['lfg_count'],
                    'rating': round(data['avg_rating'], 1) if data['avg_rating'] > 0 else None,
                    'review_count': data['total_reviews'],
                    'latest_timestamp': data['latest_timestamp'],
                    'cover_url': data['cover_url'],
                    'igdb_url': igdb_url,
                    'steam_url': steam_url,
                    'hypes': data['hypes'],
                    'is_manual': data['is_manual'],
                    'shared_by_user': shared_by_user,
                    'shared_by_server': shared_by_server
                })

            # Apply sorting
            if sort_by == 'servers':
                games_list.sort(key=lambda x: x['server_count'], reverse=True)
            elif sort_by == 'recent':
                games_list.sort(key=lambda x: x['latest_timestamp'], reverse=True)
            elif sort_by == 'rating':
                games_list.sort(key=lambda x: (x['rating'] or 0, x['review_count']), reverse=True)
            else:  # trending (default)
                # Trending = combination of server count, LFG count, and recency
                now = time.time()
                for game in games_list:
                    # Calculate trending score
                    recency_bonus = 1.0
                    if game['latest_timestamp'] > 0:
                        days_old = (now - game['latest_timestamp']) / 86400
                        recency_bonus = max(0.1, 1.0 - (days_old / 30))  # Decay over 30 days

                    game['trending_score'] = (
                        game['server_count'] * 10 +
                        game['lfg_count'] * 5 +
                        (game['rating'] or 0) * 3
                    ) * recency_bonus

                games_list.sort(key=lambda x: x['trending_score'], reverse=True)

            return JsonResponse({
                'success': True,
                'games': games_list
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)

        # Log the full error for debugging
        logger.error(f"Error in api_discovery_games_list: {str(e)}")

        # Return user-friendly error message
        return JsonResponse({
            'success': False,
            'error': 'Unable to load games at this time. Please try again later.'
        }, status=500)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_game_share_limit(request):
    """
    GET /api/discovery/game-share-limit - Check remaining game share limit for guild.
    Query params: guild_id
    """
    try:
        from .db import get_db_session
        from .models import Guild

        guild_id = request.GET.get('guild_id')
        if not guild_id:
            return JsonResponse({
                'success': False,
                'error': 'Guild ID is required'
            }, status=400)

        with get_db_session() as db:
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({
                    'success': False,
                    'error': 'Guild not found'
                }, status=404)

            # Check limit
            can_share, current_count, limit, reset_date = check_discovery_game_share_limit(db, guild_id, guild)

            if limit is None:
                # Unlimited
                return JsonResponse({
                    'success': True,
                    'unlimited': True,
                    'current': current_count,
                    'limit': None,
                    'remaining': None,
                    'reset_date': None
                })
            else:
                return JsonResponse({
                    'success': True,
                    'unlimited': False,
                    'current': current_count,
                    'limit': limit,
                    'remaining': max(0, limit - current_count),
                    'reset_date': reset_date,
                    'can_share': can_share
                })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='10/m', method='GET', block=True)
def api_igdb_search(request):
    """
    GET /api/igdb/search - Search IGDB for games.
    Query params: query (search string)
    """
    try:
        from .utils.igdb import search_games
        import asyncio

        query = request.GET.get('query', '').strip()
        if not query or len(query) < 2:
            return JsonResponse({
                'success': False,
                'error': 'Query must be at least 2 characters'
            }, status=400)

        # Run async search in sync context
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            games = loop.run_until_complete(search_games(query, limit=10))
        finally:
            loop.close()

        # Convert to JSON-serializable format
        games_data = []
        for game in games:
            games_data.append({
                'id': game.id,
                'name': game.name,
                'slug': game.slug,
                'summary': game.summary,
                'cover_url': game.cover_url,
                'platforms': game.platforms,
                'release_year': game.release_year,
                'steam_id': game.steam_id
            })

        return JsonResponse({
            'success': True,
            'games': games_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Failed to search IGDB. Please try again.'
        }, status=500)


@discord_required
@require_http_methods(["POST"])
@ratelimit(key='user', rate='30/h', method='POST', block=True)
def api_discovery_share_game(request):
    """
    POST /api/discovery/share-game - Manually share a game to Discovery Network.
    Now uses IGDB data instead of manual entry.
    Body: guild_id, igdb_id, igdb_slug, game_name, cover_url, summary, release_year, platforms, recommendation
    """
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import AnnouncedGame, Guild
        from sqlalchemy import func
        import html

        data = json_lib.loads(request.body)
        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'Not authenticated'
            }, status=401)

        guild_id = data.get('guild_id')

        # IGDB data
        igdb_id = data.get('igdb_id')
        igdb_slug = data.get('igdb_slug', '').strip()
        game_name = data.get('game_name', '').strip()
        cover_url = data.get('cover_url', '').strip()
        summary = data.get('summary', '').strip()
        release_year = data.get('release_year')
        platforms = data.get('platforms', '[]')  # JSON string
        recommendation = data.get('recommendation', '').strip()
        steam_id = data.get('steam_id')  # Steam ID from IGDB

        # Validation
        if not guild_id or not game_name or not igdb_id:
            return JsonResponse({
                'success': False,
                'error': 'Guild ID, game name, and IGDB ID are required'
            }, status=400)

        if len(game_name) > 255:
            return JsonResponse({
                'success': False,
                'error': 'Game name must be 255 characters or less'
            }, status=400)

        if recommendation and len(recommendation) > 1000:
            return JsonResponse({
                'success': False,
                'error': 'Recommendation must be 1000 characters or less'
            }, status=400)

        # Sanitize inputs (prevent XSS)
        game_name = html.escape(game_name)
        recommendation = html.escape(recommendation) if recommendation else None

        with get_db_session() as db:
            # Verify guild exists and user has permission
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({
                    'success': False,
                    'error': 'Guild not found'
                }, status=404)

            # Check if guild is in Discovery Network (has approved application)
            from .models import DiscoveryNetworkApplication, DiscoveryNetworkPreferences
            network_app = db.query(DiscoveryNetworkApplication).filter(
                DiscoveryNetworkApplication.guild_id == int(guild_id),
                DiscoveryNetworkApplication.status == 'approved'
            ).first()

            if not network_app:
                return JsonResponse({
                    'success': False,
                    'error': 'Guild must be a member of the Discovery Network to share games'
                }, status=403)

            # Check if user is trying to share from a server that is NOT their main server (anti-abuse)
            user_prefs = db.query(DiscoveryNetworkPreferences).filter_by(user_id=int(user_id)).first()

            # If user has set a main server, they can ONLY share from that server
            if user_prefs and user_prefs.main_server_id:
                if user_prefs.main_server_id != int(guild_id):
                    return JsonResponse({
                        'success': False,
                        'error': 'You can only share games from your main server. This prevents multi-server promotion abuse. Change your main server in Profile Settings (30-day cooldown applies).'
                    }, status=403)

            # Check share limit
            can_share, current_count, limit, reset_date = check_discovery_game_share_limit(db, guild_id, guild)
            if not can_share:
                return JsonResponse({
                    'success': False,
                    'error': f'Monthly game sharing limit reached. You have shared {current_count}/{limit} games this month. Limit resets on {reset_date}. Upgrade to the Discovery Module or Complete Suite for unlimited sharing!',
                    'limit_reached': True,
                    'current_count': current_count,
                    'limit': limit,
                    'reset_date': reset_date
                }, status=429)

            # Check for duplicate - multiple checks needed:
            # 1. Check if THIS SERVER already has this game
            existing_in_server = db.query(AnnouncedGame).filter(
                AnnouncedGame.guild_id == int(guild_id)
            ).filter(
                (AnnouncedGame.igdb_id == igdb_id) |
                (func.lower(AnnouncedGame.game_name) == game_name.lower())
            ).first()

            if existing_in_server:
                # Determine if it was bot-discovered or manually shared
                source = "manually shared" if existing_in_server.is_manual else "automatically discovered by the bot"
                return JsonResponse({
                    'success': False,
                    'error': f'"{game_name}" was already {source} for your server',
                    'duplicate': True
                }, status=400)

            # 2. Check if THIS USER already shared this game from THIS SERVER (extra safety)
            user_already_shared = db.query(AnnouncedGame).filter(
                AnnouncedGame.guild_id == int(guild_id),
                AnnouncedGame.shared_by_user_id == int(user_id),
                AnnouncedGame.is_manual == True
            ).filter(
                (AnnouncedGame.igdb_id == igdb_id) |
                (func.lower(AnnouncedGame.game_name) == game_name.lower())
            ).first()

            if user_already_shared:
                return JsonResponse({
                    'success': False,
                    'error': f'You already shared "{game_name}" from this server',
                    'duplicate': True
                }, status=400)

            # Create the game entry with IGDB data
            current_time = int(time.time())
            new_game = AnnouncedGame(
                guild_id=int(guild_id),
                igdb_id=igdb_id,
                igdb_slug=igdb_slug,
                steam_id=steam_id,  # Steam ID from IGDB
                game_name=game_name,
                cover_url=cover_url or None,
                release_date=int(time.mktime(time.strptime(f"{release_year}-01-01", "%Y-%m-%d"))) if release_year else None,
                platforms=platforms,  # JSON string
                announced_at=current_time,  # Required field
                created_at=current_time,
                description=recommendation,  # User's recommendation goes in description
                is_manual=True,  # Flag to indicate manually shared
                shared_by_user_id=int(user_id)  # Track who shared it
            )

            db.add(new_game)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': f'Successfully shared "{game_name}" to the Discovery Network!',
                'game': {
                    'id': new_game.id,
                    'name': game_name,
                    'igdb_id': igdb_id,
                    'cover_url': cover_url
                }
            })

    except Exception as e:
        # Never expose raw database errors to users
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while sharing your game. Please try again or contact support if the issue persists.'
        }, status=500)


@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_user_main_server(request):
    """
    GET /api/discovery/user/main-server - Get user's main server settings.
    Returns main server ID, when it was set, and when it can be changed.
    """
    try:
        from .db import get_db_session
        from .models import DiscoveryNetworkPreferences, Guild

        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'Not authenticated'
            }, status=401)

        with get_db_session() as db:
            # Get user preferences
            prefs = db.query(DiscoveryNetworkPreferences).filter_by(user_id=int(user_id)).first()

            if not prefs or not prefs.main_server_id:
                return JsonResponse({
                    'success': True,
                    'main_server': None,
                    'can_change': True,
                    'message': 'No main server set'
                })

            # Get guild info
            guild = db.query(Guild).filter_by(guild_id=prefs.main_server_id).first()

            # Auto-select next available server if main server no longer exists
            if not guild:
                # Main server was deleted - reset to None and allow immediate change
                prefs.main_server_id = None
                prefs.main_server_set_at = None
                prefs.main_server_can_change_at = None
                db.commit()

                return JsonResponse({
                    'success': True,
                    'main_server': None,
                    'can_change': True,
                    'message': 'Your main server was removed. Please select a new one.',
                    'auto_reset': True
                })

            # Check if user can change (30 days since last change)
            current_time = int(time.time())
            can_change = not prefs.main_server_can_change_at or current_time >= prefs.main_server_can_change_at

            days_remaining = 0
            if prefs.main_server_can_change_at and current_time < prefs.main_server_can_change_at:
                days_remaining = (prefs.main_server_can_change_at - current_time) // 86400

            return JsonResponse({
                'success': True,
                'main_server': {
                    'id': prefs.main_server_id,
                    'name': guild.guild_name,
                    'set_at': prefs.main_server_set_at,
                    'can_change_at': prefs.main_server_can_change_at
                },
                'can_change': can_change,
                'days_remaining': days_remaining
            })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Failed to get main server settings'
        }, status=500)


@discord_required
@require_http_methods(["POST"])
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_user_set_main_server(request):
    """
    POST /api/discovery/user/set-main-server - Set or change user's main server.
    Body: guild_id
    Enforces 30-day cooldown between changes.
    """
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import DiscoveryNetworkPreferences, Guild

        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'Not authenticated'
            }, status=401)

        data = json_lib.loads(request.body)
        guild_id = data.get('guild_id')

        if not guild_id:
            return JsonResponse({
                'success': False,
                'error': 'Guild ID is required'
            }, status=400)

        with get_db_session() as db:
            # Verify guild exists
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({
                    'success': False,
                    'error': 'Guild not found'
                }, status=404)

            # Get or create user preferences
            prefs = db.query(DiscoveryNetworkPreferences).filter_by(user_id=int(user_id)).first()
            if not prefs:
                prefs = DiscoveryNetworkPreferences(user_id=int(user_id))
                db.add(prefs)

            current_time = int(time.time())

            # Check if user can change (30 days since last change)
            if prefs.main_server_can_change_at and current_time < prefs.main_server_can_change_at:
                days_remaining = (prefs.main_server_can_change_at - current_time) // 86400
                return JsonResponse({
                    'success': False,
                    'error': f'You can change your main server again in {days_remaining} days. This cooldown prevents abuse.',
                    'cooldown': True,
                    'days_remaining': days_remaining
                }, status=429)

            # Set the new main server
            prefs.main_server_id = int(guild_id)
            prefs.main_server_set_at = current_time
            # 30 days = 30 * 24 * 60 * 60 = 2592000 seconds
            prefs.main_server_can_change_at = current_time + 2592000
            prefs.updated_at = current_time

            db.commit()

            return JsonResponse({
                'success': True,
                'message': f'Main server set to "{guild.guild_name}". You can change it again in 30 days.',
                'main_server': {
                    'id': prefs.main_server_id,
                    'name': guild.guild_name,
                    'set_at': prefs.main_server_set_at,
                    'can_change_at': prefs.main_server_can_change_at
                }
            })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Failed to set main server'
        }, status=500)


@csrf_exempt
@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_creators_list(request):
    """
    GET /api/discovery/creators-list - Get featured creators from Discovery Network.
    TODO: This is a stub - Twitch/YouTube integration coming soon.
    """
    try:
        # For now, return empty list
        # TODO: Implement Twitch and YouTube discovery
        return JsonResponse({
            'success': True,
            'creators': [],
            'message': 'Creator discovery coming soon! Twitch and YouTube integration in progress.'
        })
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@api_auth_required
@require_http_methods(["POST"])
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_feature_creator(request):
    """
    POST /api/discovery/feature-creator - Feature a creator to Discovery Network.
    TODO: This is a stub - Twitch/YouTube integration coming soon.
    SECURITY: Requires authentication (admin-only via @api_auth_required).
    """
    try:
        return JsonResponse({
            'success': False,
            'error': 'Creator featuring coming soon! This feature requires Twitch/YouTube integration.'
        }, status=501)  # 501 Not Implemented
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discord_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_network_creators(request):
    """
    GET /api/discovery/network-creators - Get all network creators with COTW/COTM.
    Returns creators who have share_to_network=True.
    """
    try:
        from .db import get_db_session
        from .models import CreatorProfile, Guild
        import json as json_lib

        guild_id = request.GET.get('guild_id')
        if not guild_id:
            return JsonResponse({'success': False, 'error': 'guild_id required'}, status=400)

        with get_db_session() as session:
            # Get all creators who opted into network sharing
            creators_query = session.query(CreatorProfile).filter(
                CreatorProfile.share_to_network == True
            ).all()

            # Get Network COTW
            network_cotw = session.query(CreatorProfile).filter(
                CreatorProfile.is_current_network_cotw == True,
                CreatorProfile.share_to_network == True
            ).first()

            # Get Network COTM
            network_cotm = session.query(CreatorProfile).filter(
                CreatorProfile.is_current_network_cotm == True,
                CreatorProfile.share_to_network == True
            ).first()

            # Helper function to parse content categories
            def parse_categories(categories_str):
                """Parse content_categories from various formats to a list."""
                if not categories_str:
                    return []
                try:
                    if categories_str.startswith('['):
                        # Try JSON first
                        try:
                            return json.loads(categories_str)
                        except json.JSONDecodeError:
                            # Might be Python string representation - use ast.literal_eval
                            import ast
                            try:
                                return ast.literal_eval(categories_str)
                            except:
                                # Last resort - parse as comma-separated
                                return [cat.strip().strip("'\"") for cat in categories_str.strip('[]').split(',') if cat.strip()]
                    else:
                        return [cat.strip() for cat in categories_str.split(',') if cat.strip()]
                except:
                    return []

            # Helper function to get avatar URL and home server name
            def get_creator_extras(creator):
                try:
                    from .models import GuildMember

                    # Get home guild
                    guild = session.query(Guild).filter(Guild.guild_id == creator.guild_id).first()
                    home_server = guild.guild_name if guild else None

                    # Get member data for avatar and level
                    member_data = session.query(GuildMember).filter_by(
                        guild_id=creator.guild_id,
                        user_id=creator.discord_id
                    ).first()

                    if member_data:
                        username = member_data.username or f"user_{creator.discord_id}"
                        avatar_url = f"https://cdn.discordapp.com/avatars/{creator.discord_id}/{member_data.avatar_hash}.png" if member_data.avatar_hash else f"https://cdn.discordapp.com/embed/avatars/{creator.discord_id % 5}.png"
                        discord_level = member_data.level if member_data.level else 0
                        role_flair = member_data.flair if member_data.flair else None
                    else:
                        username = f"user_{creator.discord_id}"
                        avatar_url = f"https://cdn.discordapp.com/embed/avatars/{creator.discord_id % 5}.png"
                        discord_level = 0
                        role_flair = None

                    return {
                        'home_server': home_server,
                        'username': username,
                        'avatar_url': avatar_url,
                        'discord_level': discord_level,
                        'role_flair': role_flair
                    }
                except Exception as e:
                    logger.error(f"Error getting extras for creator {creator.id}", exc_info=True)
                    return {
                        'home_server': None,
                        'username': f"user_{creator.discord_id}",
                        'avatar_url': f"https://cdn.discordapp.com/embed/avatars/{creator.discord_id % 5}.png",
                        'discord_level': 0,
                        'role_flair': None
                    }

            # Build creators list
            creators_list = []
            for creator in creators_query:
                extras = get_creator_extras(creator)
                creators_list.append({
                    'id': creator.id,
                    'discord_id': str(creator.discord_id),
                    'guild_id': str(creator.guild_id),
                    'display_name': creator.display_name,
                    'bio': creator.bio,
                    'content_categories': parse_categories(creator.content_categories),
                    'twitter_handle': creator.twitter_handle,
                    'tiktok_handle': creator.tiktok_handle,
                    'instagram_handle': creator.instagram_handle,
                    'bluesky_handle': creator.bluesky_handle,
                    'twitch_handle': creator.twitch_handle,
                    'youtube_handle': creator.youtube_handle,
                    'youtube_url': creator.youtube_url,
                    'youtube_channel_id': creator.youtube_channel_id,
                    'stream_schedule': creator.stream_schedule,
                    'times_featured': creator.times_featured,
                    'is_current_cotw': creator.is_current_cotw,
                    'is_current_cotm': creator.is_current_cotm,
                    'is_current_network_cotw': creator.is_current_network_cotw,
                    'is_current_network_cotm': creator.is_current_network_cotm,
                    'home_server': extras['home_server'],
                    'username': extras['username'],
                    'avatar_url': extras['avatar_url'],
                    'discord_level': extras['discord_level'],
                    'role_flair': extras['role_flair'],
                })

            # Build Network COTW data
            network_cotw_data = None
            if network_cotw:
                extras = get_creator_extras(network_cotw)
                network_cotw_data = {
                    'id': network_cotw.id,
                    'discord_id': str(network_cotw.discord_id),
                    'display_name': network_cotw.display_name,
                    'bio': network_cotw.bio,
                    'content_categories': parse_categories(network_cotw.content_categories),
                    'twitter_handle': network_cotw.twitter_handle,
                    'tiktok_handle': network_cotw.tiktok_handle,
                    'instagram_handle': network_cotw.instagram_handle,
                    'bluesky_handle': network_cotw.bluesky_handle,
                    'twitch_handle': network_cotw.twitch_handle,
                    'youtube_handle': network_cotw.youtube_handle,
                    'youtube_url': network_cotw.youtube_url,
                    'youtube_channel_id': network_cotw.youtube_channel_id,
                    'home_server': extras['home_server'],
                    'username': extras['username'],
                    'avatar_url': extras['avatar_url'],
                }

            # Build Network COTM data
            network_cotm_data = None
            if network_cotm:
                extras = get_creator_extras(network_cotm)
                network_cotm_data = {
                    'id': network_cotm.id,
                    'discord_id': str(network_cotm.discord_id),
                    'display_name': network_cotm.display_name,
                    'bio': network_cotm.bio,
                    'content_categories': parse_categories(network_cotm.content_categories),
                    'twitter_handle': network_cotm.twitter_handle,
                    'tiktok_handle': network_cotm.tiktok_handle,
                    'instagram_handle': network_cotm.instagram_handle,
                    'bluesky_handle': network_cotm.bluesky_handle,
                    'twitch_handle': network_cotm.twitch_handle,
                    'youtube_handle': network_cotm.youtube_handle,
                    'youtube_url': network_cotm.youtube_url,
                    'youtube_channel_id': network_cotm.youtube_channel_id,
                    'home_server': extras['home_server'],
                    'username': extras['username'],
                    'avatar_url': extras['avatar_url'],
                }

            return JsonResponse({
                'success': True,
                'creators': creators_list,
                'network_cotw': network_cotw_data,
                'network_cotm': network_cotm_data,
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discord_required
@require_http_methods(["GET", "POST"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_game_reviews(request, game_id):
    """
    GET/POST /api/discovery/games/<game_id>/reviews
    Get all reviews for a game or submit a new review.
    """
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import DiscoveryGameReview, Guild
        from sqlalchemy import func
        import html

        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'Not authenticated'
            }, status=401)

        # Decode game_id (it's the lowercase game name)
        game_name = game_id.replace('_', ' ')

        if request.method == 'GET':
            # Get all reviews for this game
            with get_db_session() as db:
                reviews = db.query(DiscoveryGameReview).filter(
                    func.lower(DiscoveryGameReview.game_name) == game_name.lower(),
                    DiscoveryGameReview.is_flagged == False
                ).order_by(
                    DiscoveryGameReview.created_at.desc()
                ).all()

                reviews_data = []
                for review in reviews:
                    reviews_data.append({
                        'id': review.id,
                        'user_id': str(review.user_id),
                        'username': review.username,
                        'rating': review.rating,
                        'review_text': review.review_text,
                        'hours_played': review.hours_played,
                        'created_at': review.created_at,
                        'updated_at': review.updated_at
                    })

                # Calculate average rating
                avg_rating = db.query(func.avg(DiscoveryGameReview.rating)).filter(
                    func.lower(DiscoveryGameReview.game_name) == game_name.lower(),
                    DiscoveryGameReview.is_flagged == False
                ).scalar()

                return JsonResponse({
                    'success': True,
                    'reviews': reviews_data,
                    'average_rating': round(float(avg_rating), 1) if avg_rating else 0,
                    'total_reviews': len(reviews_data)
                })

        elif request.method == 'POST':
            # Submit a new review
            data = json_lib.loads(request.body)
            guild_id = data.get('guild_id')
            rating = data.get('rating')
            review_text = data.get('review_text', '').strip()
            hours_played = data.get('hours_played')

            # Validation
            if not rating or not isinstance(rating, int) or rating < 1 or rating > 5:
                return JsonResponse({
                    'success': False,
                    'error': 'Rating must be between 1 and 5 stars'
                }, status=400)

            if review_text and len(review_text) > 2000:
                return JsonResponse({
                    'success': False,
                    'error': 'Review text must be 2000 characters or less'
                }, status=400)

            # Sanitize inputs
            review_text = html.escape(review_text) if review_text else None

            with get_db_session() as db:
                # Get user's Discord username
                username = user.get('username', 'Unknown User')

                # Check if user already reviewed this game
                existing = db.query(DiscoveryGameReview).filter(
                    func.lower(DiscoveryGameReview.game_name) == game_name.lower(),
                    DiscoveryGameReview.user_id == int(user_id)
                ).first()

                if existing:
                    # Update existing review
                    existing.rating = rating
                    existing.review_text = review_text
                    existing.hours_played = hours_played
                    existing.updated_at = int(time.time())
                    db.commit()

                    return JsonResponse({
                        'success': True,
                        'message': 'Review updated successfully!',
                        'review': {
                            'id': existing.id,
                            'rating': rating,
                            'review_text': review_text,
                            'hours_played': hours_played
                        }
                    })
                else:
                    # Create new review
                    new_review = DiscoveryGameReview(
                        game_name=game_name,
                        user_id=int(user_id),
                        username=username,
                        guild_id=int(guild_id) if guild_id else None,
                        rating=rating,
                        review_text=review_text,
                        hours_played=hours_played,
                        created_at=int(time.time()),
                        updated_at=int(time.time())
                    )

                    db.add(new_review)
                    db.commit()

                    return JsonResponse({
                        'success': True,
                        'message': 'Review submitted successfully!',
                        'review': {
                            'id': new_review.id,
                            'rating': rating,
                            'review_text': review_text,
                            'hours_played': hours_played
                        }
                    })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discord_required
@require_http_methods(["GET", "POST"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_game_discussions(request, game_id):
    """
    GET/POST /api/discovery/games/<game_id>/discussions
    Get all discussions for a game or post a new comment.
    """
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import DiscoveryGameDiscussion
        from sqlalchemy import func
        import html

        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'Not authenticated'
            }, status=401)

        # Decode game_id
        game_name = game_id.replace('_', ' ')

        if request.method == 'GET':
            # Get all discussions for this game
            with get_db_session() as db:
                discussions = db.query(DiscoveryGameDiscussion).filter(
                    func.lower(DiscoveryGameDiscussion.game_name) == game_name.lower(),
                    DiscoveryGameDiscussion.is_deleted == False,
                    DiscoveryGameDiscussion.is_flagged == False
                ).order_by(
                    DiscoveryGameDiscussion.created_at.desc()
                ).all()

                discussions_data = []
                for disc in discussions:
                    discussions_data.append({
                        'id': disc.id,
                        'user_id': str(disc.user_id),
                        'username': disc.username,
                        'comment_text': disc.comment_text,
                        'parent_comment_id': disc.parent_comment_id,
                        'created_at': disc.created_at,
                        'updated_at': disc.updated_at,
                        'upvotes': disc.upvotes
                    })

                return JsonResponse({
                    'success': True,
                    'discussions': discussions_data,
                    'total_comments': len(discussions_data)
                })

        elif request.method == 'POST':
            # Post a new comment
            data = json_lib.loads(request.body)
            guild_id = data.get('guild_id')
            comment_text = data.get('comment_text', '').strip()
            parent_comment_id = data.get('parent_comment_id')

            # Validation
            if not comment_text:
                return JsonResponse({
                    'success': False,
                    'error': 'Comment text is required'
                }, status=400)

            if len(comment_text) > 2000:
                return JsonResponse({
                    'success': False,
                    'error': 'Comment must be 2000 characters or less'
                }, status=400)

            # Sanitize input
            comment_text = html.escape(comment_text)

            with get_db_session() as db:
                # Get user's Discord username
                username = user.get('username', 'Unknown User')

                # Create new comment
                new_comment = DiscoveryGameDiscussion(
                    game_name=game_name,
                    user_id=int(user_id),
                    username=username,
                    guild_id=int(guild_id) if guild_id else None,
                    comment_text=comment_text,
                    parent_comment_id=parent_comment_id,
                    created_at=int(time.time()),
                    updated_at=int(time.time())
                )

                db.add(new_comment)
                db.commit()

                return JsonResponse({
                    'success': True,
                    'message': 'Comment posted successfully!',
                    'comment': {
                        'id': new_comment.id,
                        'comment_text': comment_text,
                        'parent_comment_id': parent_comment_id,
                        'upvotes': 0
                    }
                })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discord_required
@require_http_methods(["POST"])
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_discussion_upvote(request, discussion_id):
    """
    POST /api/discovery/discussions/<discussion_id>/upvote
    Upvote a discussion comment.
    """
    try:
        from .db import get_db_session
        from .models import DiscoveryGameDiscussion

        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'Not authenticated'
            }, status=401)

        with get_db_session() as db:
            comment = db.query(DiscoveryGameDiscussion).filter_by(id=discussion_id).first()

            if not comment:
                return JsonResponse({
                    'success': False,
                    'error': 'Comment not found'
                }, status=404)

            # Increment upvote count
            comment.upvotes += 1
            db.commit()

            return JsonResponse({
                'success': True,
                'upvotes': comment.upvotes
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discovery_approvers_required
@ratelimit(key='user', rate='60/m', method='GET', block=True)
def api_discovery_network_admin_applications(request):
    """Get all applications for admin review (DISCOVERY APPROVERS ONLY)."""
    try:
        from .db import get_db_session
        from .models import DiscoveryNetworkApplication, DiscoveryNetworkBan, Guild

        # Authorization handled by @discovery_approvers_required decorator
        with get_db_session() as db:
            # Get all applications
            applications = db.query(DiscoveryNetworkApplication).order_by(
                DiscoveryNetworkApplication.applied_at.desc()
            ).all()

            # Get all bans
            bans = db.query(DiscoveryNetworkBan).all()

            # Format applications
            apps_data = []
            for app in applications:
                # Get guild info if guild_id exists
                guild_info = None
                if app.guild_id:
                    guild = db.query(Guild).filter_by(guild_id=app.guild_id).first()
                    if guild:
                        guild_info = {
                            'id': str(guild.guild_id),
                            'name': guild.guild_name or 'Unknown Server'
                        }

                apps_data.append({
                    'id': app.id,
                    'user_id': str(app.user_id),
                    'guild_id': str(app.guild_id) if app.guild_id else None,
                    'guild_info': guild_info,
                    'username': app.username,
                    'display_name': app.display_name,
                    'avatar_url': app.avatar_url,
                    'bio': app.bio,
                    'twitch_url': app.twitch_url,
                    'youtube_url': app.youtube_url,
                    'twitter_url': app.twitter_url,
                    'tiktok_url': app.tiktok_url,
                    'status': app.status,
                    'applied_at': app.applied_at,
                    'updated_at': app.updated_at,
                    'denial_reason': app.denial_reason,
                    'reviewed_by': str(app.reviewed_by) if app.reviewed_by else None
                })

            # Format bans
            bans_data = []
            for ban in bans:
                bans_data.append({
                    'id': ban.id,
                    'user_id': str(ban.user_id),
                    'reason': ban.reason,
                    'violation_type': ban.violation_type,
                    'banned_at': ban.banned_at,
                    'banned_by': str(ban.banned_by) if ban.banned_by else None,
                    'appeal_allowed': ban.appeal_allowed,
                    'appeal_submitted': ban.appeal_submitted
                })

            return JsonResponse({
                'success': True,
                'applications': apps_data,
                'bans': bans_data,
                'stats': {
                    'pending': len([a for a in applications if a.status == 'pending']),
                    'approved': len([a for a in applications if a.status == 'approved']),
                    'denied': len([a for a in applications if a.status == 'denied']),
                    'banned': len(bans)
                }
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discovery_approvers_required
@ratelimit(key='user', rate='20/m', method='POST', block=True)
def api_discovery_network_admin_approve(request, application_id):
    """Approve an application (DISCOVERY APPROVERS ONLY)."""
    try:
        from .db import get_db_session
        from .models import DiscoveryNetworkApplication

        # Authorization handled by @discovery_approvers_required decorator
        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        with get_db_session() as db:
            application = db.query(DiscoveryNetworkApplication).filter_by(id=application_id).first()

            if not application:
                return JsonResponse({
                    'success': False,
                    'error': 'Application not found'
                }, status=404)

            application.status = 'approved'
            application.reviewed_by = int(user_id)
            application.updated_at = int(time.time())

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Application approved'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discovery_approvers_required
@ratelimit(key='user', rate='20/m', method='POST', block=True)
def api_discovery_network_admin_deny(request, application_id):
    """Deny an application (DISCOVERY APPROVERS ONLY)."""
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import DiscoveryNetworkApplication

        # Authorization handled by @discovery_approvers_required decorator
        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        data = json_lib.loads(request.body)
        reason = data.get('reason', 'No reason provided')

        with get_db_session() as db:
            application = db.query(DiscoveryNetworkApplication).filter_by(id=application_id).first()

            if not application:
                return JsonResponse({
                    'success': False,
                    'error': 'Application not found'
                }, status=404)

            application.status = 'denied'
            application.denial_reason = reason
            application.reviewed_by = int(user_id)
            application.updated_at = int(time.time())

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Application denied'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discovery_approvers_required
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def api_discovery_network_admin_ban(request, application_id):
    """Ban a user from Discovery Network (DISCOVERY APPROVERS ONLY)."""
    try:
        import json as json_lib
        from .db import get_db_session
        from .models import DiscoveryNetworkApplication, DiscoveryNetworkBan

        # Authorization handled by @discovery_approvers_required decorator
        user = request.session.get('discord_user', {})
        user_id = user.get('id')

        data = json_lib.loads(request.body)
        reason = data.get('reason', 'No reason provided')
        violation_type = data.get('violation_type', 'other')

        with get_db_session() as db:
            application = db.query(DiscoveryNetworkApplication).filter_by(id=application_id).first()

            if not application:
                return JsonResponse({
                    'success': False,
                    'error': 'Application not found'
                }, status=404)

            # Check if already banned
            existing_ban = db.query(DiscoveryNetworkBan).filter_by(user_id=application.user_id).first()
            if existing_ban:
                return JsonResponse({
                    'success': False,
                    'error': 'User is already banned'
                }, status=400)

            # Create ban
            ban = DiscoveryNetworkBan(
                user_id=application.user_id,
                reason=reason,
                violation_type=violation_type,
                banned_by=int(user_id),
                banned_at=int(time.time()),
                appeal_allowed=True,
                appeal_submitted=False
            )

            # Update application status
            application.status = 'denied'
            application.denial_reason = f'Banned: {reason}'
            application.reviewed_by = int(user_id)
            application.updated_at = int(time.time())

            db.add(ban)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'User banned from Discovery Network'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred. Please try again later.'

        }, status=500)


@discord_required
def guild_found_games(request, guild_id):
    """GET /questlog/guild/<id>/found-games/ - View all games found in recent discovery checks."""
    from .module_utils import has_module_access, has_any_module_access
    import json as json_lib
    from datetime import datetime

    # Get guild from session - allow any guild member to view
    all_guilds = request.session.get('discord_all_guilds', [])
    admin_guilds = request.session.get('discord_admin_guilds', [])  # For sidebar navigation
    guild = get_guild_with_permissions(guild_id, admin_guilds, all_guilds)

    if not guild:
        messages.error(request, "You are not a member of this server.")
        return redirect('questlog_dashboard')

    try:
        from .db import get_db_session
        from .models import Guild, FoundGame, GameSearchConfig

        with get_db_session() as db:
            # Get guild record for sidebar navigation
            guild_record = db.query(Guild).filter_by(guild_id=int(guild_id)).first()

            # Get filters from query params
            check_id = request.GET.get('check_id')  # Optional: filter to specific check
            search_id = request.GET.get('search_id')  # Optional: filter by search config
            mode_filters = request.GET.getlist('mode')  # Optional: filter by one or more game modes
            min_hype_param = request.GET.get('min_hype')
            min_hype = int(min_hype_param) if min_hype_param and min_hype_param.isdigit() else None

            # Get available search configs (only public ones shown on website)
            # Get ALL search configs for this guild (both public and private)
            available_searches = db.query(GameSearchConfig).filter(
                GameSearchConfig.guild_id == int(guild_id)
            ).all()

            # Query found games (ALL games, both public and private searches)
            query = db.query(FoundGame).filter(
                FoundGame.guild_id == int(guild_id)
            )

            if check_id:
                query = query.filter(FoundGame.check_id == check_id)
            if search_id:
                query = query.filter(FoundGame.search_config_id == int(search_id))

            # Pull a reasonable window for filtering and option building
            base_games_raw = query.order_by(FoundGame.found_at.desc()).limit(300).all()

            # Build available game mode options from the base set
            available_modes = set()
            for game in base_games_raw:
                game_modes = json_lib.loads(game.game_modes) if game.game_modes else []
                for mode in game_modes:
                    available_modes.add(mode)

            # Apply filters in Python (game_modes stored as JSON text)
            filtered_games = []
            for game in base_games_raw:
                game_modes = json_lib.loads(game.game_modes) if game.game_modes else []
                # Mode filter: require at least one selected mode to be present
                if mode_filters:
                    if not any(mode in game_modes for mode in mode_filters):
                        continue
                # Hype filter
                if min_hype is not None:
                    if game.hypes is None or game.hypes < min_hype:
                        continue
                filtered_games.append((game, game_modes))

            # Order already by found_at desc; trim to 100 for display
            filtered_games = filtered_games[:100]

            # Process games for template
            found_games = []
            for game, game_modes in filtered_games:
                genres = json_lib.loads(game.genres) if game.genres else []
                themes = json_lib.loads(game.themes) if game.themes else []
                platforms = json_lib.loads(game.platforms) if game.platforms else []

                # Format release date
                release_date_str = None
                if game.release_date:
                    release_dt = datetime.fromtimestamp(game.release_date)
                    release_date_str = release_dt.strftime('%B %d, %Y')

                # Format found date
                found_dt = datetime.fromtimestamp(game.found_at)
                found_date_str = found_dt.strftime('%B %d, %Y %I:%M %p')

                found_games.append({
                    'id': game.id,
                    'igdb_id': game.igdb_id,
                    'igdb_slug': game.igdb_slug,
                    'name': game.game_name,
                    'summary': game.summary[:300] + '...' if game.summary and len(game.summary) > 300 else game.summary,
                    'release_date': release_date_str,
                    'genres': genres,
                    'themes': themes,
                    'game_modes': game_modes,
                    'platforms': platforms,
                    'cover_url': game.cover_url,
                    'igdb_url': game.igdb_url,
                    'steam_url': game.steam_url,
                    'hypes': game.hypes,
                    'rating': round(game.rating, 1) if game.rating else None,
                    'found_at': found_date_str,
                    'check_id': game.check_id,
                })

            # Get stats
            total_count = len(found_games)
            with_steam = sum(1 for g in found_games if g['steam_url'])

            # Format search configs for dropdown
            search_configs = [
                {
                    'id': s.id,
                    'name': s.name,
                }
                for s in available_searches
            ]

            # Check if user is admin
            is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)

            # Check if guild has discovery module access
            has_discovery_module = has_module_access(guild_id, 'discovery')
            has_any_module = has_any_module_access(guild_id)

            # Use guild from session (already set at top of function)
            context = {
                'guild': guild,
                'guild_record': guild_record,  # Add for sidebar navigation
                'found_games': found_games,
                'total_count': total_count,
                'with_steam': with_steam,
                'check_id': check_id,
                'search_id': int(search_id) if search_id else None,
                'search_configs': search_configs,
                'available_modes': sorted(available_modes),
                'selected_modes': mode_filters,
                'selected_min_hype': min_hype if min_hype is not None else '',
                'has_discovery_module': has_discovery_module,
                'has_any_module': has_any_module,
                'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
                'is_admin': is_admin,
                'active_page': 'found_games',
            }

            return render(request, 'questlog/found_games.html', context)

    except Exception as e:
        # Check if user is admin
        is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)

        # Check if guild has discovery module access (for error case too)
        has_discovery_module = has_module_access(guild_id, 'discovery')
        has_any_module = has_any_module_access(guild_id)

        # Get guild_record for sidebar navigation
        guild_record = None
        try:
            with get_db_session() as db:
                guild_record = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
        except:
            pass

        # Use guild from session (already set at top of function)
        return render(request, 'questlog/found_games.html', {
            'guild': guild,
            'guild_record': guild_record,  # Add for sidebar navigation
            'error': 'An internal error occurred. Please try again later.',

            'has_discovery_module': has_discovery_module,
            'has_any_module': has_any_module,
            'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
            'is_admin': is_admin,
            'active_page': 'found_games',
        })


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_config_update(request, guild_id):
    """POST /api/guild/<id>/discovery/config/update/ - Update discovery config."""
    import json as json_lib

    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import DiscoveryConfig

        with get_db_session() as db:
            config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                config = DiscoveryConfig(guild_id=int(guild_id))
                db.add(config)

            # Update fields
            if 'enabled' in data:
                config.enabled = bool(data['enabled'])
            if 'channel_enabled' in data:
                config.channel_enabled = bool(data['channel_enabled'])
            if 'forum_enabled' in data:
                config.forum_enabled = bool(data['forum_enabled'])
            if 'selfpromo_channel_id' in data:
                config.selfpromo_channel_id = int(data['selfpromo_channel_id']) if data['selfpromo_channel_id'] else None
            if 'selfpromo_quick_feature' in data:
                config.selfpromo_quick_feature = bool(data['selfpromo_quick_feature'])
            if 'feature_channel_id' in data:
                config.feature_channel_id = int(data['feature_channel_id']) if data['feature_channel_id'] else None
            if 'message_response_channel_id' in data:
                config.message_response_channel_id = int(data['message_response_channel_id']) if data['message_response_channel_id'] else None
            if 'reminder_schedule' in data:
                config.reminder_schedule = data['reminder_schedule']
            if 'intro_forum_channel_id' in data:
                config.intro_forum_channel_id = int(data['intro_forum_channel_id']) if data['intro_forum_channel_id'] else None
            if 'forum_feature_channel_id' in data:
                config.forum_feature_channel_id = int(data['forum_feature_channel_id']) if data['forum_feature_channel_id'] else None
            if 'test_channel_id' in data:
                config.test_channel_id = int(data['test_channel_id']) if data['test_channel_id'] else None
            if 'test_forum_id' in data:
                config.test_forum_id = int(data['test_forum_id']) if data['test_forum_id'] else None
            if 'cotw_enabled' in data:
                config.cotw_enabled = bool(data['cotw_enabled'])
            if 'cotw_channel_id' in data:
                config.cotw_channel_id = int(data['cotw_channel_id']) if data['cotw_channel_id'] else None
            if 'cotw_auto_rotate' in data:
                config.cotw_auto_rotate = bool(data['cotw_auto_rotate'])
            if 'cotw_rotation_day' in data:
                config.cotw_rotation_day = max(0, min(6, int(data['cotw_rotation_day'])))  # 0-6 for days of week
            if 'cotm_enabled' in data:
                config.cotm_enabled = bool(data['cotm_enabled'])
            if 'cotm_channel_id' in data:
                config.cotm_channel_id = int(data['cotm_channel_id']) if data['cotm_channel_id'] else None
            if 'cotm_auto_rotate' in data:
                config.cotm_auto_rotate = bool(data['cotm_auto_rotate'])
            if 'cotm_rotation_day' in data:
                config.cotm_rotation_day = max(1, min(28, int(data['cotm_rotation_day'])))  # 1-28 for days of month
            # Channel timing settings
            if 'channel_feature_interval_hours' in data:
                config.channel_feature_interval_hours = max(1, min(24, int(data['channel_feature_interval_hours'])))
            if 'channel_pool_entry_duration_hours' in data:
                config.channel_pool_entry_duration_hours = max(1, min(168, int(data['channel_pool_entry_duration_hours'])))
            if 'channel_entry_cooldown_hours' in data:
                config.channel_entry_cooldown_hours = max(1, min(168, int(data['channel_entry_cooldown_hours'])))
            if 'channel_feature_cooldown_hours' in data:
                config.channel_feature_cooldown_hours = max(0, min(168, int(data['channel_feature_cooldown_hours'])))

            # Forum timing settings
            if 'intro_scan_interval_hours' in data:
                config.intro_scan_interval_hours = max(1, min(24, int(data['intro_scan_interval_hours'])))
            if 'forum_resubmit_cooldown_days' in data:
                config.forum_resubmit_cooldown_days = max(1, min(365, int(data['forum_resubmit_cooldown_days'])))
            if 'forum_min_post_age_hours' in data:
                config.forum_min_post_age_hours = max(0, min(168, int(data['forum_min_post_age_hours'])))

            # Token costs (separate for channel and forum)
            if 'token_cost' in data:
                config.token_cost = max(0, int(data['token_cost']))
            if 'token_cost_forum' in data:
                config.token_cost_forum = max(0, int(data['token_cost_forum']))

            # Timing settings
            if 'feature_interval_hours' in data:
                config.feature_interval_hours = max(1, min(24, int(data['feature_interval_hours'])))
            if 'pool_entry_duration_hours' in data:
                config.pool_entry_duration_hours = max(1, min(168, int(data['pool_entry_duration_hours'])))
            if 'entry_cooldown_hours' in data:
                config.entry_cooldown_hours = max(0, min(168, int(data['entry_cooldown_hours'])))
            if 'feature_cooldown_hours' in data:
                config.feature_cooldown_hours = max(0, min(168, int(data['feature_cooldown_hours'])))

            # Messages and embeds
            if 'how_to_enter_response' in data:
                config.how_to_enter_response = data['how_to_enter_response'][:500]
            if 'post_response' in data:
                config.post_response = data['post_response'][:500]
            if 'feature_message' in data:
                config.feature_message = data['feature_message'][:500]
            if 'cooldown_message' in data:
                config.cooldown_message = data['cooldown_message'][:500]
            if 'use_embed' in data:
                config.use_embed = bool(data['use_embed'])
            if 'embed_color' in data:
                config.embed_color = int(data['embed_color'])

            # Other settings
            if 'require_tokens' in data:
                config.require_tokens = bool(data['require_tokens'])
            if 'require_tokens_forum' in data:
                config.require_tokens_forum = bool(data['require_tokens_forum'])
            if 'remove_after_feature' in data:
                config.remove_after_feature = bool(data['remove_after_feature'])

            # Network Creator Discovery settings
            if 'network_announcements_enabled' in data:
                config.network_announcements_enabled = bool(data['network_announcements_enabled'])
            if 'network_announcement_channel_id' in data:
                config.network_announcement_channel_id = int(data['network_announcement_channel_id']) if data['network_announcement_channel_id'] else None

            config.updated_at = int(time.time())

            db.commit()
            return JsonResponse({'success': True})

    except Exception as e:
        logger.error(f"Error updating discovery config for guild {guild_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to update discovery configuration'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_discovery_pool(request, guild_id):
    """GET /api/guild/<id>/discovery/pool/ - Get current pool entries."""
    try:
        from .db import get_db_session
        from .models import FeaturedPool

        now = int(time.time())

        with get_db_session() as db:
            entries = db.query(FeaturedPool).filter(
                FeaturedPool.guild_id == int(guild_id),
                FeaturedPool.was_selected == False,
                FeaturedPool.expires_at > now
            ).order_by(FeaturedPool.entered_at.desc()).all()

            pool_data = []
            for entry in entries:
                pool_data.append({
                    'id': entry.id,
                    'user_id': str(entry.user_id),
                    'content': entry.content,
                    'link_url': entry.link_url,
                    'platform': entry.platform,
                    'entered_at': entry.entered_at,
                    'expires_at': entry.expires_at,
                })

            return JsonResponse({'success': True, 'pool': pool_data, 'count': len(pool_data)})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_discovery_pool_remove(request, guild_id, entry_id):
    """DELETE /api/guild/<id>/discovery/pool/<entry_id>/ - Remove entry from pool."""
    try:
        from .db import get_db_session
        from .models import FeaturedPool

        with get_db_session() as db:
            entry = db.query(FeaturedPool).filter(
                FeaturedPool.id == int(entry_id),
                FeaturedPool.guild_id == int(guild_id)
            ).first()

            if not entry:
                return JsonResponse({'error': 'Entry not found'}, status=404)

            db.delete(entry)
            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_force_feature(request, guild_id):
    """POST /api/guild/<id>/discovery/feature/ - Force a feature selection now."""
    try:
        from .db import get_db_session
        from .models import DiscoveryConfig, FeaturedPool
        from .actions import queue_action, ActionType

        now = int(time.time())
        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        with get_db_session() as db:
            config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if not config or not config.enabled:
                return JsonResponse({'error': 'Discovery is disabled'}, status=400)

            # Check if there are entries
            entry_count = db.query(FeaturedPool).filter(
                FeaturedPool.guild_id == int(guild_id),
                FeaturedPool.was_selected == False,
                FeaturedPool.expires_at > now
            ).count()

            if entry_count == 0:
                return JsonResponse({'error': 'No entries in pool'}, status=400)

            # Queue the action for the bot to process
            action_id = queue_action(
                guild_id=int(guild_id),
                action_type=ActionType.FORCE_FEATURE,
                payload={
                    'triggered_by': triggered_by,
                },
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name,
                source='website'
            )

            return JsonResponse({'success': True, 'action_id': action_id, 'pool_count': entry_count})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_clear_featured(request, guild_id):
    """POST /api/guild/<id>/discovery/clear/ - Clear the currently featured person."""
    try:
        from .db import get_db_session
        from .models import DiscoveryConfig
        from .actions import queue_clear_featured

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        with get_db_session() as db:
            config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                return JsonResponse({'error': 'Discovery is not configured'}, status=400)

            # Queue the action for the bot to process
            action_id = queue_clear_featured(
                guild_id=int(guild_id),
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name
            )

            return JsonResponse({
                'success': True,
                'action_id': action_id,
                'message': 'Clear action queued - check Discord in a few seconds'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_test_channel_embed(request, guild_id):
    """POST /api/guild/<id>/discovery/test-channel-embed/ - Send test channel embed."""
    try:
        import json
        from .db import get_db_session
        from .models import DiscoveryConfig
        from .actions import queue_test_channel_embed

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        # Parse request body to get channel_id
        data = json.loads(request.body.decode('utf-8'))
        channel_id = data.get('channel_id')

        if not channel_id:
            return JsonResponse({'error': 'No channel selected. Please select a test channel first.'}, status=400)

        with get_db_session() as db:
            config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                return JsonResponse({'error': 'Discovery is not configured'}, status=400)

            # Queue the action for the bot to process with the selected channel
            action_id = queue_test_channel_embed(
                guild_id=int(guild_id),
                channel_id=int(channel_id),
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name
            )

            return JsonResponse({
                'success': True,
                'action_id': action_id,
                'message': 'Test channel embed queued - check Discord in a few seconds'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_discovery_test_forum_embed(request, guild_id):
    """POST /api/guild/<id>/discovery/test-forum-embed/ - Send test forum embed."""
    try:
        import json
        from .db import get_db_session
        from .models import DiscoveryConfig
        from .actions import queue_test_forum_embed

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        # Parse request body to get channel_id
        data = json.loads(request.body.decode('utf-8'))
        channel_id = data.get('channel_id')

        if not channel_id:
            return JsonResponse({'error': 'No channel selected. Please select a test forum channel first.'}, status=400)

        with get_db_session() as db:
            config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                return JsonResponse({'error': 'Discovery is not configured'}, status=400)

            # Queue the action for the bot to process with the selected channel
            action_id = queue_test_forum_embed(
                guild_id=int(guild_id),
                channel_id=int(channel_id),
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name
            )

            return JsonResponse({
                'success': True,
                'action_id': action_id,
                'message': 'Test forum embed queued - check Discord and website in a few seconds'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_game_discovery_config_update(request, guild_id):
    """POST /api/guild/<id>/discovery/game-config/update/ - Update game discovery config."""
    import json as json_lib

    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import DiscoveryConfig

        with get_db_session() as db:
            config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                config = DiscoveryConfig(guild_id=int(guild_id))
                db.add(config)

            # Update game discovery settings
            if 'game_discovery_enabled' in data:
                config.game_discovery_enabled = bool(data['game_discovery_enabled'])

            if 'public_game_channel_id' in data:
                channel_id = data['public_game_channel_id']
                config.public_game_channel_id = int(channel_id) if channel_id else None

            if 'private_game_channel_id' in data:
                channel_id = data['private_game_channel_id']
                config.private_game_channel_id = int(channel_id) if channel_id else None

            if 'public_game_ping_role_id' in data:
                role_id = data['public_game_ping_role_id']
                config.public_game_ping_role_id = int(role_id) if role_id else None

            if 'private_game_ping_role_id' in data:
                role_id = data['private_game_ping_role_id']
                config.private_game_ping_role_id = int(role_id) if role_id else None

            if 'game_check_interval_hours' in data:
                config.game_check_interval_hours = int(data['game_check_interval_hours'])

            if 'game_days_ahead' in data:
                config.game_days_ahead = int(data['game_days_ahead'])

            if 'game_days_behind' in data:
                config.game_days_behind = int(data['game_days_behind'])

            if 'game_min_hype' in data:
                hype_value = data['game_min_hype']
                config.game_min_hype = int(hype_value) if hype_value else None

            if 'game_min_rating' in data:
                rating_value = data['game_min_rating']
                config.game_min_rating = float(rating_value) if rating_value else None

            # Update filters (stored as JSON)
            if 'game_genres' in data:
                genres = data['game_genres']
                config.game_genres = json_lib.dumps(genres) if genres else None

            if 'game_themes' in data:
                themes = data['game_themes']
                config.game_themes = json_lib.dumps(themes) if themes else None

            if 'game_modes' in data:
                modes = data['game_modes']
                config.game_modes = json_lib.dumps(modes) if modes else None

            if 'game_platforms' in data:
                platforms = data['game_platforms']
                config.game_platforms = json_lib.dumps(platforms) if platforms else None

            if 'game_tags' in data:
                tags = data['game_tags']
                config.game_tags = json_lib.dumps(tags) if tags else None

            db.commit()
            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_game_discovery_check(request, guild_id):
    """POST /api/guild/<id>/discovery/check-games/ - Manually trigger game discovery check."""
    try:
        from .db import get_db_session
        from .models import DiscoveryConfig, AnnouncedGame
        from .actions import queue_action, ActionType
        import os

        discord_user = request.session.get('discord_user', {})
        triggered_by = int(discord_user.get('id', 0))
        triggered_by_name = discord_user.get('global_name', discord_user.get('username'))

        with get_db_session() as db:
            config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            if not config or not config.game_discovery_enabled:
                return JsonResponse({'error': 'Game discovery is not enabled'}, status=400)

            if not config.public_game_channel_id and not config.private_game_channel_id:
                return JsonResponse({'error': 'No game discovery channels configured'}, status=400)

            # Check if IGDB is configured
            if not (os.getenv('IGDB_CLIENT_ID') and os.getenv('IGDB_CLIENT_SECRET')):
                return JsonResponse({
                    'error': 'IGDB is not configured. Please add IGDB_CLIENT_ID and TWITCH_CLIENT_SECRET to your .env file.'
                }, status=400)

            # Queue the action for the bot to process
            action_id = queue_action(
                guild_id=int(guild_id),
                action_type=ActionType.CHECK_GAMES,
                payload={
                    'triggered_by': triggered_by,
                },
                triggered_by=triggered_by,
                triggered_by_name=triggered_by_name,
                source='website'
            )

            return JsonResponse({
                'success': True,
                'action_id': action_id,
                'message': 'Game check queued using IGDB - results will appear in Discord shortly'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def api_purge_announced_games(request, guild_id):
    """POST /api/guild/<id>/discovery/purge-announced-games/ - Clear all announced games for re-announcement."""
    try:
        from .db import get_db_session
        from .models import AnnouncedGame

        with get_db_session() as db:
            # Delete all announced games for this guild
            count = db.query(AnnouncedGame).filter_by(guild_id=int(guild_id)).delete()
            db.commit()

            return JsonResponse({
                'success': True,
                'count': count,
                'message': f'Purged {count} games from Discovery Network. These can now be re-announced.'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_game_search_configs_list(request, guild_id):
    """GET /api/guild/<id>/discovery/searches/ - List all game search configurations."""
    try:
        from .db import get_db_session
        from .models import GameSearchConfig
        import json as json_lib

        with get_db_session() as db:
            searches = db.query(GameSearchConfig).filter_by(guild_id=int(guild_id)).order_by(GameSearchConfig.name).all()

            result = []
            for search in searches:
                result.append({
                    'id': search.id,
                    'name': search.name,
                    'enabled': search.enabled,
                    'genres': json_lib.loads(search.genres) if search.genres else [],
                    'themes': json_lib.loads(search.themes) if search.themes else [],
                    'game_modes': json_lib.loads(search.game_modes) if search.game_modes else [],
                    'platforms': json_lib.loads(search.platforms) if search.platforms else [],
                    'min_hype': search.min_hype,
                    'min_rating': search.min_rating,
                    'days_ahead': search.days_ahead,
                    'show_on_website': search.show_on_website,
                    'auto_join_role_id': search.auto_join_role_id,
                    'created_at': search.created_at,
                    'updated_at': search.updated_at,
                })

            return JsonResponse({'success': True, 'searches': result})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_game_search_config_create(request, guild_id):
    """POST /api/guild/<id>/discovery/searches/ - Create a new game search configuration."""
    try:
        from .db import get_db_session
        from .models import GameSearchConfig, Guild as GuildModel, SubscriptionTier
        import json as json_lib

        data = json_lib.loads(request.body)

        # Check tier limits for search configurations
        with get_db_session() as db:
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Determine search limit based on tier
            # Check if guild has Discovery Module (grants unlimited searches)
            from .module_utils import has_module_access

            if has_module_access(guild_id, 'discovery') or guild_record.is_vip or guild_record.subscription_tier == SubscriptionTier.PREMIUM.value:
                max_searches = None  # Unlimited
            elif guild_record.subscription_tier == SubscriptionTier.PRO.value:
                max_searches = 5
            else:  # Free tier
                max_searches = 3

            # Count existing searches
            if max_searches is not None:
                existing_count = db.query(GameSearchConfig).filter_by(
                    guild_id=int(guild_id)
                ).count()

                if existing_count >= max_searches:
                    tier_name = 'Pro' if max_searches == 5 else 'Free'
                    return JsonResponse({
                        'error': f'{tier_name} tier limited to {max_searches} game discovery searches. You have {existing_count}/{max_searches}. Delete a search or upgrade to Premium for unlimited searches.',
                        'limit_exceeded': True,
                        'current_count': existing_count,
                        'max_allowed': max_searches
                    }, status=403)

        with get_db_session() as db:
            # Create new search config
            search = GameSearchConfig(
                guild_id=int(guild_id),
                name=data.get('name', 'New Search'),
                enabled=data.get('enabled', True),
                genres=json_lib.dumps(data.get('genres', [])) if data.get('genres') else None,
                themes=json_lib.dumps(data.get('themes', [])) if data.get('themes') else None,
                game_modes=json_lib.dumps(data.get('game_modes', [])) if data.get('game_modes') else None,
                platforms=json_lib.dumps(data.get('platforms', [])) if data.get('platforms') else None,
                min_hype=data.get('min_hype'),
                min_rating=data.get('min_rating'),
                days_ahead=data.get('days_ahead', 30),
                show_on_website=data.get('show_on_website', True),
                auto_join_role_id=data.get('auto_join_role_id'),
                created_at=int(time.time()),
                updated_at=int(time.time())
            )

            db.add(search)
            db.commit()

            return JsonResponse({
                'success': True,
                'search_id': search.id,
                'message': f"Search '{search.name}' created successfully"
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["PUT", "POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['PUT', 'POST'], block=True)
def api_game_search_config_update(request, guild_id, search_id):
    """PUT /api/guild/<id>/discovery/searches/<search_id>/ - Update a game search configuration."""
    try:
        from .db import get_db_session
        from .models import GameSearchConfig
        import json as json_lib

        data = json_lib.loads(request.body)

        with get_db_session() as db:
            search = db.query(GameSearchConfig).filter_by(
                id=int(search_id),
                guild_id=int(guild_id)
            ).first()

            if not search:
                return JsonResponse({'error': 'Search configuration not found'}, status=404)

            # Update fields
            if 'name' in data:
                search.name = data['name']
            if 'enabled' in data:
                search.enabled = data['enabled']
            if 'genres' in data:
                search.genres = json_lib.dumps(data['genres']) if data['genres'] else None
            if 'themes' in data:
                search.themes = json_lib.dumps(data['themes']) if data['themes'] else None
            if 'game_modes' in data:
                search.game_modes = json_lib.dumps(data['game_modes']) if data['game_modes'] else None
            if 'platforms' in data:
                search.platforms = json_lib.dumps(data['platforms']) if data['platforms'] else None
            if 'min_hype' in data:
                search.min_hype = data['min_hype']
            if 'min_rating' in data:
                search.min_rating = data['min_rating']
            if 'days_ahead' in data:
                search.days_ahead = data['days_ahead']
            if 'show_on_website' in data:
                search.show_on_website = data['show_on_website']
            if 'auto_join_role_id' in data:
                search.auto_join_role_id = data['auto_join_role_id']

            search.updated_at = int(time.time())
            db.commit()

            return JsonResponse({
                'success': True,
                'message': f"Search '{search.name}' updated successfully"
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE", "POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['DELETE', 'POST'], block=True)
def api_game_search_config_delete(request, guild_id, search_id):
    """DELETE /api/guild/<id>/discovery/searches/<search_id>/ - Delete a game search configuration."""
    try:
        from .db import get_db_session
        from .models import GameSearchConfig

        with get_db_session() as db:
            search = db.query(GameSearchConfig).filter_by(
                id=int(search_id),
                guild_id=int(guild_id)
            ).first()

            if not search:
                return JsonResponse({'error': 'Search configuration not found'}, status=404)

            search_name = search.name
            db.delete(search)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': f"Search '{search_name}' deleted successfully"
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_action_status(request, guild_id, action_id):
    """GET /api/guild/<id>/action/<action_id>/status/ - Get status of a queued action."""
    try:
        from .db import get_db_session
        from .models import PendingAction
        import json

        with get_db_session() as db:
            action = db.query(PendingAction).filter_by(
                id=int(action_id),
                guild_id=int(guild_id)
            ).first()

            if not action:
                return JsonResponse({'error': 'Action not found'}, status=404)

            # Parse result JSON if it exists
            result_data = None
            if action.result:
                try:
                    result_data = json.loads(action.result)
                except:
                    result_data = None

            return JsonResponse({
                'id': action.id,
                'status': action.status.value,
                'created_at': action.created_at,  # Already a Unix timestamp
                'completed_at': action.completed_at,  # Already a Unix timestamp
                'error_message': action.error_message,
                'result': result_data
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# LFG (Looking For Group) Dashboard

@discord_required
def guild_lfg(request, guild_id):
    """LFG dashboard."""
    from .module_utils import has_module_access, has_any_module_access
    import logging
    logger = logging.getLogger(__name__)

    user_guilds = request.session.get('discord_admin_guilds', [])
    logger.info(f"LFG: Looking for guild_id={guild_id}")
    logger.info(f"LFG: discord_admin_guilds has {len(user_guilds)} guilds")
    if user_guilds:
        logger.info(f"LFG: Guild IDs in session: {[str(g.get('id')) for g in user_guilds]}")

    guild = next((g for g in user_guilds if str(g.get('id')) == str(guild_id)), None)

    if not guild:
        logger.warning(f"LFG: Guild {guild_id} not found in session, redirecting to dashboard")
        return redirect('questlog_dashboard')

    # Check admin permission
    permissions = int(guild.get('permissions', 0))
    is_admin = (permissions & 0x8) == 0x8 or (permissions & 0x20) == 0x20

    if not is_admin:
        messages.error(request, 'You need Admin or Manage Server permission.')
        return redirect('questlog_dashboard')

    try:
        from .db import get_db_session
        from .models import Guild, LFGGame, LFGConfig

        with get_db_session() as db:
            guild_db = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            is_premium = guild_db.is_premium() if guild_db else False
            is_pro_plus = guild_db and (guild_db.subscription_tier in ['pro', 'premium'] or guild_db.is_vip)

            games = db.query(LFGGame).filter_by(guild_id=int(guild_id)).all()

            # Get LFG Browser notification config
            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if not lfg_config and is_pro_plus:
                # Create default config for Pro+ servers
                lfg_config = LFGConfig(guild_id=int(guild_id))
                db.add(lfg_config)
                db.commit()
            games_data = []
            for game in games:
                games_data.append({
                    'id': game.id,
                    'game_name': game.game_name,
                    'game_short': game.game_short,
                    'igdb_id': game.igdb_id,
                    'cover_url': game.cover_url,
                    'platforms': game.platforms,
                    'is_custom_game': game.is_custom_game,
                    'lfg_channel_id': game.lfg_channel_id,
                    'notify_role_id': game.notify_role_id,
                    'max_group_size': game.max_group_size,
                    'custom_options': json.loads(game.custom_options) if game.custom_options else None,
                    'require_rank': game.require_rank,
                    'rank_label': game.rank_label,
                    'enabled': game.enabled,
                    'current_player_count': game.current_player_count or 0,  # Live player count
                })

        # Get channels and roles from Discord API using bot token
        import os
        bot_token = os.getenv('DISCORD_BOT_TOKEN')

        channels = []
        roles = []

        if bot_token and bot_token != 'your_bot_token_here':
            try:
                # Fetch guild channels from cache (no Discord API call!)
                from .discord_resources import get_guild_channels, get_guild_roles

                all_channels = get_guild_channels(str(guild_id))
                # Filter for text channels (type 0) and forum channels (type 15)
                channels = [
                    {'id': ch['id'], 'name': ch['name']}
                    for ch in all_channels
                    if ch.get('type') in [0, 15]  # Text and Forum channels
                ]
                channels.sort(key=lambda x: x['name'])
                logger.info(f"Found {len(channels)} text/forum channels from cache")

                # Fetch guild roles from cache (no Discord API call!)
                all_roles = get_guild_roles(str(guild_id))
                # Sort by position
                roles = [
                    {'id': r['id'], 'name': r['name'], 'position': r.get('position', 0)}
                    for r in all_roles
                ]
                roles.sort(key=lambda x: x['position'], reverse=True)
                logger.info(f"Found {len(roles)} roles from cache")

            except Exception as e:
                logger.error(f"Error fetching Discord channels/roles from cache: {e}")
        else:
            logger.warning("DISCORD_BOT_TOKEN not configured")

        # Check if user is admin or LFG Manager (for audit log access)
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')
        has_lfg_manager_role = check_lfg_manager_role(guild_id, user_id) if user_id else False
        is_admin_or_manager = is_admin or has_lfg_manager_role

        # Check if guild has LFG module access
        has_lfg_module = has_module_access(guild_id, 'lfg')
        has_engagement_module = has_module_access(guild_id, 'engagement')
        has_any_module = has_any_module_access(guild_id)

        return render(request, 'questlog/lfg.html', {
            'guild': guild,
            'guild_record': guild_db,
            'games': games_data,
            'channels': channels,
            'roles': roles,
            'is_premium': is_premium,
            'is_pro_plus': is_pro_plus,
            'lfg_config': lfg_config,
            'is_admin': True,
            'is_admin_or_manager': is_admin_or_manager,
            'has_lfg_module': has_lfg_module,
            'has_engagement_module': has_engagement_module,
            'has_any_module': has_any_module,
            'admin_guilds': user_guilds,
            'active_page': 'lfg',
            'total_lfg_games': len(games_data),
        })

    except Exception as e:
        messages.error(request, f'Error loading LFG: {e}')
        return redirect('guild_dashboard', guild_id=guild_id)


def guild_lfg_browser(request, guild_id):
    """
    LFG Browser - Browse and create LFG groups (All tiers).

    This view renders the LFG Browser interface where users can:
    - Browse active LFG groups
    - Create new groups
    - Join/leave groups
    - Manage groups (if they have permissions)

    Permission Levels:
    - Regular users: Can create/join groups, manage their own groups
    - LFG Managers: Can edit/delete ANY group (requires "LFG Manager" Discord role with CREATE_EVENTS + MANAGE_EVENTS permissions)
    - Admins: Full management access to all groups
    """
    from .module_utils import has_module_access, has_any_module_access
    import logging
    logger = logging.getLogger(__name__)

    # Get user's guilds and authentication info from Discord OAuth session
    admin_guilds = request.session.get('discord_admin_guilds', [])
    all_guilds = request.session.get('discord_all_guilds', [])
    user_id = request.session.get('discord_user', {}).get('id')
    discord_user = request.session.get('discord_user', {})

    # Verify user is a member of this guild
    guild = next((g for g in all_guilds if str(g['id']) == str(guild_id)), None)
    is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)

    if not guild:
        messages.error(request, "You are not a member of this server.")
        return redirect('questlog_dashboard')

    # Check if user has "LFG Manager" role (requires both CREATE_EVENTS and MANAGE_EVENTS permissions)
    has_lfg_manager_role = check_lfg_manager_role(guild_id, user_id)

    # Users can manage ANY group if they're admin OR have the LFG Manager role
    # Otherwise, they can only manage their own groups (enforced in individual endpoints)
    can_manage_groups = is_admin or has_lfg_manager_role

    try:
        from .db import get_db_session
        from .models import Guild, LFGGame

        with get_db_session() as db:
            guild_db = db.query(Guild).filter_by(guild_id=int(guild_id)).first()

            if not guild_db:
                messages.error(request, 'Server not found in database.')
                return redirect('questlog_dashboard')

            # Get all enabled games for this guild
            games = db.query(LFGGame).filter_by(
                guild_id=int(guild_id),
                enabled=True
            ).all()

            games_data = []
            for game in games:
                # Parse custom_options safely - ensure it's always an array
                custom_options = None
                if game.custom_options:
                    try:
                        parsed = json.loads(game.custom_options)
                        # Ensure custom_options is a list and each option has a choices array
                        if isinstance(parsed, list):
                            custom_options = []
                            for opt in parsed:
                                if isinstance(opt, dict) and 'name' in opt:
                                    # Build option object
                                    option_obj = {'name': opt['name']}

                                    # Add depends_on if present
                                    if 'depends_on' in opt:
                                        option_obj['depends_on'] = opt['depends_on']

                                    # Add choices (can be array or object for conditional dropdowns)
                                    choices = opt.get('choices', [])
                                    # Don't validate type for conditional dropdowns - can be object or array
                                    option_obj['choices'] = choices

                                    custom_options.append(option_obj)
                    except (json.JSONDecodeError, TypeError, KeyError) as e:
                        logger.warning(f"Failed to parse custom_options for game {game.id}: {e}")
                        custom_options = None

                games_data.append({
                    'id': game.id,
                    'game_name': game.game_name,
                    'game_short': game.game_short,
                    'game_emoji': game.game_emoji,
                    'cover_url': game.cover_url,
                    'max_group_size': game.max_group_size,
                    'require_rank': game.require_rank,
                    'rank_label': game.rank_label,
                    'rank_min': game.rank_min,
                    'rank_max': game.rank_max,
                    'custom_options': custom_options,
                    'current_player_count': game.current_player_count or 0,
                })

        # Pass games_data directly to template - will use json_script tag for safe JSON injection prevention
        # Check if guild has LFG module access
        has_lfg_module = has_module_access(guild_id, 'lfg')
        has_engagement_module = has_module_access(guild_id, 'engagement')
        has_any_module = has_any_module_access(guild_id)

        return render(request, 'questlog/lfg_browser.html', {
            'guild': guild,
            'guild_record': guild_db,
            'games': games_data,
            'games_data': games_data,  # Pass for json_script tag
            'discord_user': discord_user,
            'is_admin': is_admin,
            'can_manage_groups': can_manage_groups,  # Pass can_manage_groups for audit logs button
            'has_lfg_module': has_lfg_module,
            'has_engagement_module': has_engagement_module,
            'has_any_module': has_any_module,
            'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
            'active_page': 'lfg_browser',
        })

    except Exception as e:
        logger.error(f"Error loading LFG browser: {e}", exc_info=True)
        messages.error(request, f'Error loading LFG browser: {e}')
        return redirect('guild_dashboard', guild_id=guild_id)


def guild_attendance(request, guild_id):
    """LFG Attendance tracking dashboard - Admin/Raid Leader only (Premium feature)."""
    import logging
    logger = logging.getLogger(__name__)

    # Get user's guilds from session
    admin_guilds = request.session.get('discord_admin_guilds', [])
    all_guilds = request.session.get('discord_all_guilds', [])
    user_id = request.session.get('discord_user', {}).get('id')

    # Check if user is admin OR member of this guild
    is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)
    guild = next((g for g in all_guilds if str(g['id']) == str(guild_id)), None)

    if not guild:
        messages.error(request, "You are not a member of this server.")
        return redirect('questlog_dashboard')

    # Check permissions
    permissions = int(guild.get('permissions', 0))
    has_create_events = (permissions & 0x100000000) == 0x100000000  # CREATE_EVENTS permission
    has_manage_events = (permissions & 0x200000000) == 0x200000000  # MANAGE_EVENTS permission
    has_manage_messages = (permissions & 0x2000) == 0x2000  # MANAGE_MESSAGES permission (moderator)

    # User has permission
    is_member = True

    try:
        from .db import get_db_session
        from .models import Guild, LFGConfig, LFGGroup, LFGMember, LFGAttendance, LFGMemberStats

        with get_db_session() as db:
            guild_db = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            is_premium = guild_db.is_premium() if guild_db else False

            # Check if guild has premium (required for attendance tracking)
            if not is_premium:
                messages.error(request, 'Attendance tracking requires Premium or VIP subscription.')
                return redirect('guild_dashboard', guild_id=guild_id)

            # Get or create LFG config
            config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if not config:
                config = LFGConfig(guild_id=int(guild_id))
                db.add(config)
                db.flush()

            # Check for fixed "Raid Leader" role via Discord API
            has_raid_leader_role = False
            if user_id:
                try:
                    # Get guild roles from cache (no Discord API call!)
                    from .discord_resources import get_guild_roles, get_guild_member

                    roles = get_guild_roles(str(guild_id))

                    raid_leader_role_id = None
                    for role in roles:
                        if role.get('name') == 'Raid Leader':
                            raid_leader_role_id = role.get('id')
                            break

                    # Check if user has the Raid Leader role (from cache, no API call!)
                    if raid_leader_role_id:
                        member_data = get_guild_member(str(guild_id), str(user_id))
                        if member_data:
                            member_roles = member_data.get('roles', [])
                            if str(raid_leader_role_id) in member_roles:
                                has_raid_leader_role = True
                except Exception as e:
                    logger.error(f"Error checking Raid Leader role from cache: {e}")

            # Permission logic: Admin OR Moderator OR (Raid Leader + Create Events + Manage Events)
            has_raid_leader_full_perms = has_raid_leader_role and has_create_events and has_manage_events

            # Deny access if not authorized
            if not (is_admin or has_manage_messages or has_raid_leader_full_perms):
                messages.error(request, "You need admin, moderator, or Raid Leader role (with Create Events and Manage Events permissions) to access attendance tracking.")
                return redirect('guild_dashboard', guild_id=guild_id)

            # Determine if user can manage attendance (for UI)
            can_manage_attendance = is_admin or has_manage_messages or has_raid_leader_full_perms

            # Fetch member display names from cache (no Discord API call!)
            from .discord_resources import get_guild_members
            import requests
            member_names_cache = {}

            try:
                guild_members = get_guild_members(str(guild_id))
                for m in guild_members:
                    user_id = int(m['id'])
                    display_name = m.get('display_name') or m.get('username', 'Unknown')
                    member_names_cache[user_id] = display_name
                logger.info(f"Loaded {len(member_names_cache)} member names from cache")
            except Exception as e:
                logger.error(f"Error fetching guild members from cache: {e}")

            # Get recent groups (last 30 days) with attendance tracking
            thirty_days_ago = int(time.time()) - (30 * 24 * 3600)

            # Only show groups that have attendance records (meaning attendance was ON when created)
            groups = db.query(LFGGroup).filter(
                LFGGroup.guild_id == int(guild_id),
                LFGGroup.created_at >= thirty_days_ago,
                LFGGroup.id.in_(
                    db.query(LFGAttendance.group_id).distinct()
                )
            ).order_by(LFGGroup.scheduled_time.desc()).limit(50).all()

            groups_data = []
            for group in groups:
                # Get game info
                from .models import LFGGame
                game = db.query(LFGGame).filter_by(id=group.game_id).first()
                game_name = game.game_name if game else "Unknown Game"

                # Get attendance records
                attendance_records = db.query(LFGAttendance).filter_by(group_id=group.id).all()
                members = db.query(LFGMember).filter_by(group_id=group.id).all()

                # Create a map of user_id -> member for quick lookup
                members_map = {m.user_id: m for m in members}

                attendance_data = []
                for att in attendance_records:
                    # Try Discord API cache first, then GuildMember table, then fallback to user ID
                    display_name = member_names_cache.get(att.user_id)

                    if not display_name:
                        # Check GuildMember table for synced Discord member data
                        from .models import GuildMember
                        guild_member = db.query(GuildMember).filter_by(
                            guild_id=int(guild_id),
                            user_id=att.user_id
                        ).first()
                        display_name = guild_member.display_name if guild_member else f"User {att.user_id}"

                    # Get member stats to check blacklist status
                    member_stats = db.query(LFGMemberStats).filter_by(
                        guild_id=int(guild_id),
                        user_id=att.user_id
                    ).first()

                    # Get member's class/role selections
                    member = members_map.get(att.user_id)
                    selections = {}
                    if member and member.selections:
                        import json
                        try:
                            selections = json.loads(member.selections)
                        except:
                            selections = {}
                    elif not member:
                        # Fallback: If no member record for THIS group (manually added attendance),
                        # try to find their most recent selections from ANY other group
                        import json
                        recent_member = db.query(LFGMember).filter_by(user_id=att.user_id).order_by(LFGMember.joined_at.desc()).first()
                        if recent_member and recent_member.selections:
                            try:
                                selections = json.loads(recent_member.selections)
                            except:
                                selections = {}

                    attendance_data.append({
                        'user_id': att.user_id,
                        'display_name': display_name,
                        'status': att.status.value if att.status else 'pending',
                        'confirmed_at': att.confirmed_at,
                        'showed_at': att.showed_at,
                        'is_blacklisted': member_stats.is_blacklisted if member_stats else False,
                        'selections': selections,  # Class/role selections
                    })

                groups_data.append({
                    'id': group.id,
                    'thread_name': group.thread_name,
                    'game_name': game_name,
                    'creator_name': group.creator_name,
                    'scheduled_time': group.scheduled_time,
                    'is_active': group.is_active,
                    'member_count': len(attendance_data),
                    'attendance': attendance_data,
                })

            # Get warning threshold
            warn_threshold = config.warn_at_reliability

            # Get top reliable members (above the at-risk range)
            top_members = db.query(LFGMemberStats).filter(
                LFGMemberStats.guild_id == int(guild_id),
                LFGMemberStats.total_signups >= 3,  # Must have attended at least 3 events
                LFGMemberStats.reliability_score >= warn_threshold + 15,  # Well above warning
                LFGMemberStats.is_blacklisted == False
            ).order_by(LFGMemberStats.reliability_score.desc()).limit(10).all()

            # Get members nearing warning threshold (within 15% above warning)
            at_risk_members = db.query(LFGMemberStats).filter(
                LFGMemberStats.guild_id == int(guild_id),
                LFGMemberStats.total_signups >= 3,
                LFGMemberStats.reliability_score >= warn_threshold,
                LFGMemberStats.reliability_score < warn_threshold + 15,
                LFGMemberStats.is_blacklisted == False
            ).order_by(LFGMemberStats.reliability_score.asc()).limit(10).all()

            # Get members below warning threshold (need attention)
            warned_members = db.query(LFGMemberStats).filter(
                LFGMemberStats.guild_id == int(guild_id),
                LFGMemberStats.total_signups >= 3,
                LFGMemberStats.reliability_score < warn_threshold,
                LFGMemberStats.is_blacklisted == False
            ).order_by(LFGMemberStats.reliability_score.asc()).limit(10).all()

            # Get blacklisted members (always show for admin visibility)
            blacklisted_members = db.query(LFGMemberStats).filter(
                LFGMemberStats.guild_id == int(guild_id),
                LFGMemberStats.is_blacklisted == True
            ).order_by(LFGMemberStats.blacklisted_at.desc()).all()

            # Build leaderboard using the member_names_cache we already fetched
            leaderboard = []
            for stats in top_members:
                # Try to get display name from Discord API cache first
                display_name = member_names_cache.get(stats.user_id)

                # Fall back to LFGMember table
                if not display_name:
                    member = db.query(LFGMember).filter_by(
                        user_id=stats.user_id
                    ).order_by(LFGMember.joined_at.desc()).first()
                    display_name = member.display_name if member else f"Unknown User"

                leaderboard.append({
                    'user_id': stats.user_id,
                    'display_name': display_name,
                    'reliability_score': stats.reliability_score if stats.reliability_score is not None else 100,
                    'total_showed': stats.total_showed or 0,
                    'total_no_shows': stats.total_no_shows or 0,
                    'total_late': stats.total_late or 0,
                    'total_cancelled': stats.total_cancelled or 0,
                    'total_pardoned': stats.total_pardoned or 0,
                    'total_signups': stats.total_signups or 0,
                    'is_blacklisted': stats.is_blacklisted,
                })

            # Process at-risk members
            at_risk_list = []
            for stats in at_risk_members:
                # Try to get display name from Discord API cache first
                display_name = member_names_cache.get(stats.user_id)

                # Fall back to LFGMember table
                if not display_name:
                    member = db.query(LFGMember).filter_by(
                        user_id=stats.user_id
                    ).order_by(LFGMember.joined_at.desc()).first()
                    display_name = member.display_name if member else f"Unknown User"

                at_risk_list.append({
                    'user_id': stats.user_id,
                    'display_name': display_name,
                    'reliability_score': stats.reliability_score if stats.reliability_score is not None else 100,
                    'total_showed': stats.total_showed or 0,
                    'total_no_shows': stats.total_no_shows or 0,
                    'total_late': stats.total_late or 0,
                    'total_cancelled': stats.total_cancelled or 0,
                    'total_pardoned': stats.total_pardoned or 0,
                    'total_signups': stats.total_signups or 0,
                })

            # Process warned members
            warned_list = []
            for stats in warned_members:
                # Try to get display name from Discord API cache first
                display_name = member_names_cache.get(stats.user_id)

                # Fall back to LFGMember table
                if not display_name:
                    member = db.query(LFGMember).filter_by(
                        user_id=stats.user_id
                    ).order_by(LFGMember.joined_at.desc()).first()
                    display_name = member.display_name if member else f"Unknown User"

                warned_list.append({
                    'user_id': stats.user_id,
                    'display_name': display_name,
                    'reliability_score': stats.reliability_score if stats.reliability_score is not None else 0,
                    'total_showed': stats.total_showed or 0,
                    'total_no_shows': stats.total_no_shows or 0,
                    'total_late': stats.total_late or 0,
                    'total_cancelled': stats.total_cancelled or 0,
                    'total_pardoned': stats.total_pardoned or 0,
                    'total_signups': stats.total_signups or 0,
                })

            # Process blacklisted members
            blacklisted_list = []
            for stats in blacklisted_members:
                # Try to get display name from Discord API cache first
                display_name = member_names_cache.get(stats.user_id)

                # Fall back to LFGMember table
                if not display_name:
                    member = db.query(LFGMember).filter_by(
                        user_id=stats.user_id
                    ).order_by(LFGMember.joined_at.desc()).first()
                    display_name = member.display_name if member else f"Unknown User"

                blacklisted_list.append({
                    'user_id': stats.user_id,
                    'display_name': display_name,
                    'reliability_score': stats.reliability_score if stats.reliability_score is not None else 0,
                    'total_showed': stats.total_showed or 0,
                    'total_no_shows': stats.total_no_shows or 0,
                    'total_late': stats.total_late or 0,
                    'total_cancelled': stats.total_cancelled or 0,
                    'total_pardoned': stats.total_pardoned or 0,
                    'total_signups': stats.total_signups or 0,
                    'blacklist_reason': stats.blacklist_reason or 'No reason provided',
                    'blacklisted_at': stats.blacklisted_at,
                })

            # Get unique game names for filter dropdown
            unique_games = sorted(list(set(g['game_name'] for g in groups_data)))

        return render(request, 'questlog/attendance.html', {
            'guild': guild,
            'guild_record': guild_db,
            'config': {
                'auto_noshow_hours': config.auto_noshow_hours,
                'warn_at_reliability': config.warn_at_reliability,
                'min_reliability_score': config.min_reliability_score,
                'auto_blacklist_noshows': config.auto_blacklist_noshows,
            },
            'groups': groups_data,
            'unique_games': unique_games,
            'leaderboard': leaderboard,
            'at_risk': at_risk_list,
            'warned': warned_list,
            'blacklisted': blacklisted_list,
            'is_premium': is_premium,
            'is_admin': is_admin,
            'is_member': is_member,
            'can_manage_attendance': can_manage_attendance,
            'admin_guilds': admin_guilds,
        'member_guilds': get_member_guilds(request),
            'active_page': 'attendance',
        })

    except Exception as e:
        logger.error(f"Error loading attendance: {e}", exc_info=True)
        messages.error(request, f'Error loading attendance: {e}')
        return redirect('guild_dashboard', guild_id=guild_id)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='GET', block=True)
def api_lfg_search(request, guild_id):
    """GET /api/guild/<id>/lfg/search/?q=query - Search IGDB for games."""
    query = request.GET.get('q', '')
    if not query or len(query) < 2:
        return JsonResponse({'error': 'Query too short', 'games': []})

    try:
        import asyncio
        from .utils import igdb

        # Run async search in sync context
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            games = loop.run_until_complete(igdb.search_games(query, limit=10))
        finally:
            loop.close()

        games_data = []
        for game in games:
            games_data.append({
                'id': game.id,
                'name': game.name,
                'slug': game.slug,
                'cover_url': game.cover_url,
                'platforms': ', '.join(game.platforms) if game.platforms else None,
                'release_year': game.release_year,
            })

        return JsonResponse({'success': True, 'games': games_data})

    except Exception as e:
        return JsonResponse({'error': 'Failed to load game data', 'games': []})


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_add(request, guild_id):
    """POST /api/guild/<id>/lfg/add/ - Add a game to LFG."""
    try:
        data = json_lib.loads(request.body)
    except json_lib.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    try:
        from .db import get_db_session
        from .models import Guild, LFGGame, SubscriptionTier

        with get_db_session() as db:
            # Check tier limits for LFG games
            guild_record = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild_record:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Count existing LFG games
            current_count = db.query(LFGGame).filter_by(guild_id=int(guild_id)).count()

            # Check tier limits
            if guild_record.is_vip or guild_record.subscription_tier == SubscriptionTier.PREMIUM.value:
                # Unlimited for Premium/VIP
                pass
            elif guild_record.subscription_tier == SubscriptionTier.PRO.value:
                if current_count >= 10:
                    return JsonResponse({
                        'error': f'Pro tier limit reached: You can create up to 10 LFG games. Currently have {current_count} games.',
                        'limit': 10,
                        'current': current_count
                    }, status=403)
            else:  # Free tier
                if current_count >= 3:
                    return JsonResponse({
                        'error': f'Free tier limit reached: You can create up to 3 LFG games. Currently have {current_count} games. Unlock the Events & Attendance module or Complete Suite for unlimited games.',
                        'limit': 3,
                        'current': current_count
                    }, status=403)

            is_custom = data.get('is_custom', False)

            short_code = data.get('short_code', '').upper().strip()
            if not short_code:
                return JsonResponse({'error': 'Short code required'}, status=400)

            # Check if short code already exists
            existing = db.query(LFGGame).filter(
                LFGGame.guild_id == int(guild_id),
                LFGGame.game_short.ilike(short_code)
            ).first()

            if existing:
                return JsonResponse({'error': f'Short code "{short_code}" already exists'}, status=400)

            # Create game
            game = LFGGame(
                guild_id=int(guild_id),
                game_name=data.get('game_name', 'Unknown Game'),
                game_short=short_code,
                igdb_id=int(data.get('igdb_id')) if data.get('igdb_id') else None,
                cover_url=data.get('cover_url'),
                platforms=data.get('platforms'),
                is_custom_game=is_custom,
                lfg_channel_id=int(data.get('channel_id')) if data.get('channel_id') else None,
                notify_role_id=int(data.get('notify_role_id')) if data.get('notify_role_id') else None,
                max_group_size=int(data.get('max_size', 4)),
            )
            db.add(game)

            return JsonResponse({'success': True, 'game_id': game.id})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_lfg_remove(request, guild_id, game_id):
    """DELETE /api/guild/<id>/lfg/<game_id>/ - Remove a game from LFG."""
    try:
        from .db import get_db_session
        from .models import LFGGame

        with get_db_session() as db:
            game = db.query(LFGGame).filter(
                LFGGame.id == int(game_id),
                LFGGame.guild_id == int(guild_id)
            ).first()

            if not game:
                return JsonResponse({'error': 'Game not found'}, status=404)

            db.delete(game)
            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET", "POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_game_update(request, guild_id, game_id):
    """GET/POST /api/guild/<id>/lfg/<game_id>/update/ - Get or update a game."""
    try:
        from .db import get_db_session
        from .models import LFGGame
        import json

        with get_db_session() as db:
            game = db.query(LFGGame).filter(
                LFGGame.id == int(game_id),
                LFGGame.guild_id == int(guild_id)
            ).first()

            if not game:
                return JsonResponse({'error': 'Game not found'}, status=404)

            if request.method == 'GET':
                return JsonResponse({
                    'id': game.id,
                    'game_name': game.game_name,
                    'game_short': game.game_short,
                    'game_emoji': game.game_emoji,
                    'igdb_id': game.igdb_id,
                    'cover_url': game.cover_url,
                    'platforms': game.platforms,
                    'is_custom_game': game.is_custom_game,
                    'lfg_channel_id': str(game.lfg_channel_id) if game.lfg_channel_id else None,
                    'notify_role_id': str(game.notify_role_id) if game.notify_role_id else None,
                    'custom_options': json.loads(game.custom_options) if game.custom_options else None,
                    'max_group_size': game.max_group_size,
                    'thread_auto_archive_hours': game.thread_auto_archive_hours,
                    'enabled': game.enabled,
                    'require_rank': game.require_rank,
                    'rank_label': game.rank_label,
                    'rank_min': game.rank_min,
                    'rank_max': game.rank_max,
                })

            # POST - update game
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON'}, status=400)

            if 'game_name' in data:
                game.game_name = data['game_name']
            if 'game_emoji' in data:
                game.game_emoji = data['game_emoji']
            if 'lfg_channel_id' in data:
                game.lfg_channel_id = int(data['lfg_channel_id']) if data['lfg_channel_id'] else None
            if 'notify_role_id' in data:
                game.notify_role_id = int(data['notify_role_id']) if data['notify_role_id'] else None
            if 'max_group_size' in data:
                game.max_group_size = int(data['max_group_size'])
            if 'thread_auto_archive_hours' in data:
                game.thread_auto_archive_hours = int(data['thread_auto_archive_hours'])
            if 'enabled' in data:
                game.enabled = bool(data['enabled'])
            if 'require_rank' in data:
                game.require_rank = bool(data['require_rank'])
            if 'rank_label' in data:
                game.rank_label = data['rank_label']
            if 'rank_min' in data:
                game.rank_min = int(data['rank_min'])
            if 'rank_max' in data:
                game.rank_max = int(data['rank_max'])
            if 'custom_options' in data:
                game.custom_options = json.dumps(data['custom_options']) if data['custom_options'] else None

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET", "POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_config(request, guild_id):
    """GET/POST /api/guild/<id>/lfg/config/ - Get or update LFG config (Premium)."""
    try:
        from .db import get_db_session
        from .models import LFGConfig, Guild
        import json

        with get_db_session() as db:
            # Check premium
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            is_premium = guild.is_premium() if guild else False

            config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()

            if request.method == 'GET':
                if not config:
                    return JsonResponse({
                        'is_premium': is_premium,
                        'attendance_tracking_enabled': False,
                        'auto_noshow_hours': 1,
                        'require_confirmation': False,
                        'min_reliability_score': 0,
                        'warn_at_reliability': 50,
                        'auto_blacklist_noshows': 0,
                        'notify_on_noshow': False,
                        'notify_channel_id': None,
                    })

                return JsonResponse({
                    'is_premium': is_premium,
                    'attendance_tracking_enabled': config.attendance_tracking_enabled,
                    'auto_noshow_hours': config.auto_noshow_hours,
                    'require_confirmation': config.require_confirmation,
                    'min_reliability_score': config.min_reliability_score,
                    'warn_at_reliability': config.warn_at_reliability,
                    'auto_blacklist_noshows': config.auto_blacklist_noshows,
                    'notify_on_noshow': config.notify_on_noshow,
                    'notify_channel_id': str(config.notify_channel_id) if config.notify_channel_id else None,
                })

            # POST - update config (requires premium)
            if not is_premium:
                return JsonResponse({'error': 'Premium required'}, status=403)

            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON'}, status=400)

            if not config:
                config = LFGConfig(guild_id=int(guild_id))
                db.add(config)

            if 'attendance_tracking_enabled' in data:
                config.attendance_tracking_enabled = bool(data['attendance_tracking_enabled'])
            if 'auto_noshow_hours' in data:
                config.auto_noshow_hours = max(0, int(data['auto_noshow_hours']))
            if 'require_confirmation' in data:
                config.require_confirmation = bool(data['require_confirmation'])
            if 'min_reliability_score' in data:
                config.min_reliability_score = max(0, min(100, int(data['min_reliability_score'])))
            if 'warn_at_reliability' in data:
                config.warn_at_reliability = max(0, min(100, int(data['warn_at_reliability'])))
            if 'auto_blacklist_noshows' in data:
                config.auto_blacklist_noshows = max(0, int(data['auto_blacklist_noshows']))
            if 'notify_on_noshow' in data:
                config.notify_on_noshow = bool(data['notify_on_noshow'])
            if 'notify_channel_id' in data:
                config.notify_channel_id = int(data['notify_channel_id']) if data['notify_channel_id'] else None

            config.updated_at = int(time.time())

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_lfg_stats(request, guild_id):
    """GET /api/guild/<id>/lfg/stats/ - Get member stats/leaderboard (Premium)."""
    try:
        from .db import get_db_session
        from .models import LFGMemberStats, Guild, GuildMember

        with get_db_session() as db:
            # Check premium
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild or not guild.is_premium():
                return JsonResponse({'error': 'Premium required', 'members': []}, status=403)

            # Get query params
            sort = request.GET.get('sort', 'reliability')  # reliability, active, flaky
            limit = min(50, int(request.GET.get('limit', 20)))

            query = db.query(LFGMemberStats).filter(
                LFGMemberStats.guild_id == int(guild_id)
            )

            if sort == 'reliability':
                query = query.order_by(LFGMemberStats.reliability_score.desc())
            elif sort == 'active':
                query = query.order_by(LFGMemberStats.total_signups.desc())
            elif sort == 'flaky':
                query = query.order_by(LFGMemberStats.total_no_shows.desc())

            members = query.limit(limit).all()

            # Enrich with usernames
            enriched_members = []
            for m in members:
                # Get user info
                member = db.query(GuildMember).filter_by(
                    guild_id=int(guild_id),
                    user_id=m.user_id
                ).first()

                username = member.display_name or member.username or f'User {m.user_id}' if member else f'User {m.user_id}'

                enriched_members.append({
                    'user_id': str(m.user_id),
                    'display_name': username,
                    'total_signups': m.total_signups,
                    'total_showed': m.total_showed,
                    'total_no_shows': m.total_no_shows,
                    'total_cancelled': m.total_cancelled,
                    'total_late': m.total_late,
                    'reliability_score': m.reliability_score,
                    'current_show_streak': m.current_show_streak,
                    'best_show_streak': m.best_show_streak,
                    'is_blacklisted': m.is_blacklisted,
                    'blacklist_reason': m.blacklist_reason,
                })

            return JsonResponse({'success': True, 'stats': enriched_members})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_lfg_blacklist(request, guild_id):
    """GET /api/guild/<id>/lfg/blacklist/ - Get blacklisted members (Premium)."""
    try:
        from .db import get_db_session
        from .models import LFGMemberStats, Guild, GuildMember
        from datetime import datetime

        with get_db_session() as db:
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild or not guild.is_premium():
                return JsonResponse({'error': 'Premium required', 'members': []}, status=403)

            members = db.query(LFGMemberStats).filter(
                LFGMemberStats.guild_id == int(guild_id),
                LFGMemberStats.is_blacklisted == True
            ).all()

            # Enrich with usernames and format dates
            enriched_members = []
            for m in members:
                # Get user info
                member = db.query(GuildMember).filter_by(
                    guild_id=int(guild_id),
                    user_id=m.user_id
                ).first()

                username = member.display_name or member.username or f'User {m.user_id}' if member else f'User {m.user_id}'
                formatted_date = datetime.fromtimestamp(m.blacklisted_at).strftime('%b %d, %Y at %I:%M %p') if m.blacklisted_at else 'Unknown'

                enriched_members.append({
                    'user_id': str(m.user_id),
                    'username': username,
                    'reliability_score': m.reliability_score,
                    'total_no_shows': m.total_no_shows,
                    'blacklist_reason': m.blacklist_reason,
                    'blacklisted_at': m.blacklisted_at,
                    'formatted_date': formatted_date,
                    'blacklisted_by': str(m.blacklisted_by) if m.blacklisted_by else None,
                })

            return JsonResponse({'members': enriched_members})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_blacklist_update(request, guild_id, user_id):
    """POST /api/guild/<id>/lfg/blacklist/<user_id>/ - Update blacklist status (Premium)."""
    try:
        from .db import get_db_session
        from .models import LFGMemberStats, Guild
        import json

        with get_db_session() as db:
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild or not guild.is_premium():
                return JsonResponse({'error': 'Premium required'}, status=403)

            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON'}, status=400)
            action = data.get('action', 'add')  # 'add' or 'remove'
            reason = data.get('reason', '')

            stats = db.query(LFGMemberStats).filter_by(
                guild_id=int(guild_id), user_id=int(user_id)
            ).first()

            if not stats:
                stats = LFGMemberStats(guild_id=int(guild_id), user_id=int(user_id))
                db.add(stats)

            now = int(time.time())

            if action == 'add':
                stats.is_blacklisted = True
                stats.blacklisted_at = now
                stats.blacklist_reason = reason or 'Blacklisted via dashboard'
            else:
                stats.is_blacklisted = False
                stats.blacklisted_at = None
                stats.blacklisted_by = None
                stats.blacklist_reason = None

            return JsonResponse({'success': True})

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_lfg_groups(request, guild_id):
    """GET /api/guild/<id>/lfg/groups/ - List active LFG groups."""
    try:
        from .db import get_db_session
        from .models import LFGGroup, LFGGame

        with get_db_session() as db:
            # Get groups from last 7 days
            week_ago = int(time.time()) - (7 * 24 * 3600)

            groups = db.query(LFGGroup).filter(
                LFGGroup.guild_id == int(guild_id),
                LFGGroup.created_at >= week_ago
            ).order_by(LFGGroup.created_at.desc()).limit(50).all()

            return JsonResponse({
                'groups': [{
                    'id': g.id,
                    'game_id': g.game_id,
                    'thread_id': str(g.thread_id) if g.thread_id else None,
                    'thread_name': g.thread_name,
                    'creator_id': str(g.creator_id),
                    'creator_name': g.creator_name,
                    'scheduled_time': g.scheduled_time,
                    'status': g.status,
                    'member_count': g.member_count,
                    'created_at': g.created_at,
                } for g in groups]
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@require_subscription_tier('pro', 'premium')
@validate_json_schema({
    "type": "object",
    "properties": {
        "group_id": {"type": "integer", "minimum": 1},
        "user_id": {"type": "integer", "minimum": 1},
        "status": {"type": "string", "enum": ["showed", "no_show", "late", "cancelled", "confirmed", "pardoned"]}
    },
    "required": ["group_id", "user_id", "status"]
})
def api_lfg_attendance_update(request, guild_id):
    """POST /api/guild/<id>/lfg/attendance/update/ - Update attendance status (Pro/Premium only)."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGConfig, LFGGroup, LFGMember, LFGAttendance, LFGMemberStats
        import logging
        logger = logging.getLogger(__name__)

        # Get validated data
        data = request.validated_data
        group_id = data['group_id']
        user_id = data['user_id']
        status = data['status']

        with get_db_session() as db:
            # Tier check handled by @require_subscription_tier decorator
            # Check if attendance tracking is enabled
            config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if not config or not config.attendance_tracking_enabled:
                return JsonResponse({'error': 'Attendance tracking not enabled'}, status=403)

            # Verify group belongs to this guild
            group = db.query(LFGGroup).filter_by(
                id=int(group_id),
                guild_id=int(guild_id)
            ).first()

            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            # Check permissions - reuse same logic as guild_attendance view
            # Get user info from session
            session_user_id = request.session.get('discord_user', {}).get('id')
            admin_guilds = request.session.get('discord_admin_guilds', [])
            all_guilds = request.session.get('discord_all_guilds', [])

            is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)
            user_guild = next((g for g in all_guilds if str(g['id']) == str(guild_id)), None)

            # Check permissions
            has_permission = False
            if is_admin:
                has_permission = True
            elif user_guild:
                permissions = int(user_guild.get('permissions', 0))
                has_manage_messages = (permissions & 0x2000) == 0x2000
                has_create_events = (permissions & 0x100000000) == 0x100000000
                has_manage_events = (permissions & 0x200000000) == 0x200000000

                if has_manage_messages:
                    has_permission = True
                else:
                    # Check for Raid Leader role
                    try:
                        # Get guild roles from cache (no Discord API call!)
                        from .discord_resources import get_guild_roles, get_guild_member

                        if session_user_id:
                            roles = get_guild_roles(str(guild_id))

                            raid_leader_role_id = None
                            for role in roles:
                                if role.get('name') == 'Raid Leader':
                                    raid_leader_role_id = role.get('id')
                                    break

                            # Check if user has the Raid Leader role (from cache, no API call!)
                            if raid_leader_role_id:
                                member_data = get_guild_member(str(guild_id), str(session_user_id))
                                if member_data:
                                    member_roles = member_data.get('roles', [])
                                    if str(raid_leader_role_id) in member_roles:
                                        # Check if they have required permissions
                                        if has_create_events and has_manage_events:
                                            has_permission = True
                    except Exception as e:
                        logger.error(f"Error checking Raid Leader permissions from cache: {e}")

            if not has_permission:
                return JsonResponse({'error': 'Insufficient permissions'}, status=403)

            # Check if user is blacklisted (prevents joining new groups)
            user_stats = db.query(LFGMemberStats).filter_by(
                guild_id=int(guild_id),
                user_id=int(user_id)
            ).first()

            # Block blacklisted users from joining NEW groups
            # Check if they're trying to join a group they're not already in
            existing_attendance = db.query(LFGAttendance).filter_by(
                group_id=int(group_id),
                user_id=int(user_id)
            ).first()

            if user_stats and user_stats.is_blacklisted and not existing_attendance:
                return JsonResponse({
                    'error': 'User is blacklisted and cannot join new groups',
                    'is_blacklisted': True,
                    'blacklist_reason': user_stats.blacklist_reason
                }, status=403)

            # Update or create attendance record
            now = int(time.time())
            attendance = existing_attendance

            if attendance:
                # Update existing record
                attendance.status = status
                attendance.updated_at = now

                # Update timestamps based on status
                if status == 'showed':
                    attendance.showed_at = now
                elif status == 'no_show':
                    attendance.no_show_at = now
                elif status == 'late':
                    attendance.late_at = now
                elif status == 'cancelled':
                    attendance.cancelled_at = now
            else:
                # Create new record
                attendance = LFGAttendance(
                    group_id=int(group_id),
                    user_id=int(user_id),
                    status=status,
                    joined_at=now,
                    updated_at=now
                )

                if status == 'showed':
                    attendance.showed_at = now
                elif status == 'no_show':
                    attendance.no_show_at = now
                elif status == 'late':
                    attendance.late_at = now
                elif status == 'cancelled':
                    attendance.cancelled_at = now

                db.add(attendance)

            db.flush()  # Ensure attendance record is saved before recalculating stats

            # Get or create stats record FIRST (need to check pardon timestamp)
            stats = db.query(LFGMemberStats).filter_by(
                guild_id=int(guild_id),
                user_id=int(user_id)
            ).first()

            if not stats:
                stats = LFGMemberStats(
                    guild_id=int(guild_id),
                    user_id=int(user_id)
                )
                db.add(stats)

            # Get attendance records for recalculation
            # If user has been globally pardoned, ONLY count records AFTER the pardon timestamp
            # Historical records from before pardon are kept for admin reference but don't affect calculations
            all_attendance_query = db.query(LFGAttendance).filter_by(
                user_id=int(user_id)
            ).join(LFGGroup).filter(
                LFGGroup.guild_id == int(guild_id)
            )

            # If globally pardoned, exclude records from before the pardon
            if stats.blacklist_pardoned and stats.blacklist_pardoned_at:
                # Only count records created/updated AFTER the pardon timestamp
                # Use the latest timestamp available (updated_at, joined_at, or created_at)
                all_attendance_query = all_attendance_query.filter(
                    or_(
                        LFGAttendance.updated_at >= stats.blacklist_pardoned_at,
                        and_(
                            LFGAttendance.updated_at == None,
                            LFGAttendance.joined_at >= stats.blacklist_pardoned_at
                        )
                    )
                )

            all_attendance = all_attendance_query.all()

            # Recalculate from attendance records (excluding historical data if pardoned)
            # Global Pardon = fresh start, only NEW records count toward stats
            total_signups = len(all_attendance)
            total_showed = sum(1 for a in all_attendance if a.status == 'showed')
            total_no_shows = sum(1 for a in all_attendance if a.status == 'no_show')
            total_late = sum(1 for a in all_attendance if a.status == 'late')
            total_cancelled = sum(1 for a in all_attendance if a.status == 'cancelled')
            total_pardoned = sum(1 for a in all_attendance if a.status == 'pardoned')

            stats.total_signups = total_signups
            stats.total_showed = total_showed
            stats.total_no_shows = total_no_shows
            stats.total_late = total_late
            stats.total_cancelled = total_cancelled
            stats.total_pardoned = total_pardoned

            # Recalculate reliability score (0-100)
            # Pardoned counts as showed (100%), late counts 80%, no-show counts 0%, cancelled excluded
            total_counted = total_showed + total_pardoned + total_no_shows + total_late
            if total_counted > 0:
                # Pardoned = 100% (valid excuse, treat as showed), showed = 100%, late = 80%, no-show = 0%
                score = ((total_showed + total_pardoned) * 100 + (total_late * 80)) / total_counted
                stats.reliability_score = int(score)
            else:
                stats.reliability_score = 100

            # Update timestamps
            stats.updated_at = now
            if all_attendance:
                # Get all non-None timestamps
                timestamps = [a.updated_at or a.joined_at or a.created_at for a in all_attendance if (a.updated_at or a.joined_at or a.created_at)]
                if timestamps:
                    stats.last_event = max(timestamps)

                # Get first join timestamp
                join_times = [a.joined_at for a in all_attendance if a.joined_at]
                if join_times:
                    stats.first_event = min(join_times)
                elif timestamps:
                    stats.first_event = min(timestamps)

            # Check auto-blacklist threshold (always check, even for pardoned users)
            # Global Pardon is a fresh start, not permanent immunity - they can be blacklisted again
            if config.auto_blacklist_noshows > 0:
                if stats.total_no_shows >= config.auto_blacklist_noshows and not stats.is_blacklisted:
                    # Add to blacklist when threshold is reached
                    stats.is_blacklisted = True
                    stats.blacklisted_at = now
                    stats.blacklist_reason = f"Auto-blacklisted: {stats.total_no_shows} no-shows"
                elif stats.total_no_shows < config.auto_blacklist_noshows and stats.is_blacklisted and stats.blacklist_reason and "Auto-blacklisted" in stats.blacklist_reason:
                    # Remove from blacklist if no-show count drops below threshold (e.g., pardoned)
                    # Only remove if they were auto-blacklisted (not manually blacklisted)
                    stats.is_blacklisted = False
                    stats.blacklisted_at = None
                    stats.blacklist_reason = None

            db.commit()

            return JsonResponse({
                'success': True,
                'attendance': {
                    'status': status,
                    'reliability_score': stats.reliability_score
                },
                'stats': {
                    'total_showed': stats.total_showed,
                    'total_no_shows': stats.total_no_shows,
                    'total_late': stats.total_late,
                    'total_cancelled': stats.total_cancelled,
                    'total_pardoned': stats.total_pardoned,
                    'reliability_score': stats.reliability_score,
                    'is_blacklisted': stats.is_blacklisted or False
                }
            })

    except Exception as e:
        logger.error(f"Error updating attendance: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_blacklist_toggle(request, guild_id):
    """POST /api/guild/<id>/lfg/blacklist/toggle/ - Manually toggle blacklist status for a member."""
    import json
    try:
        from .db import get_db_session
        from .models import Guild, LFGConfig, LFGMemberStats

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        user_id = data.get('user_id')
        reason = data.get('reason', '')

        if not user_id:
            return JsonResponse({'error': 'Missing user_id'}, status=400)

        with get_db_session() as db:
            # Check permissions (admin, moderator, or raid leader)
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get LFG config
            config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if not config or not config.attendance_tracking_enabled:
                return JsonResponse({'error': 'Attendance tracking not enabled'}, status=403)

            # Get or create member stats
            stats = db.query(LFGMemberStats).filter_by(
                guild_id=int(guild_id),
                user_id=int(user_id)
            ).first()

            if not stats:
                return JsonResponse({'error': 'Member not found in LFG system'}, status=404)

            # Toggle blacklist status
            now = int(time.time())
            if stats.is_blacklisted:
                # Unblacklist
                stats.is_blacklisted = False
                stats.blacklisted_at = None
                stats.blacklist_reason = None
                action = 'unblacklisted'
            else:
                # Blacklist
                stats.is_blacklisted = True
                stats.blacklisted_at = now
                stats.blacklist_reason = f"Manually blacklisted by admin{': ' + reason if reason else ''}"
                action = 'blacklisted'

            db.commit()

            return JsonResponse({
                'success': True,
                'action': action,
                'is_blacklisted': stats.is_blacklisted,
                'blacklist_reason': stats.blacklist_reason
            })

    except Exception as e:
        logger.error(f"Error toggling blacklist: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_pardon(request, guild_id):
    """POST /api/guild/<id>/lfg/pardon/ - Grant permanent pardon to a member, removing them from blacklist and preventing auto-blacklist."""
    import json
    try:
        from .db import get_db_session
        from .models import Guild, LFGConfig, LFGMemberStats

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        user_id = data.get('user_id')

        if not user_id:
            return JsonResponse({'error': 'Missing user_id'}, status=400)

        with get_db_session() as db:
            # Check permissions (admin, moderator, or raid leader)
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get LFG config
            config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if not config or not config.attendance_tracking_enabled:
                return JsonResponse({'error': 'Attendance tracking not enabled'}, status=403)

            # Get or create member stats
            stats = db.query(LFGMemberStats).filter_by(
                guild_id=int(guild_id),
                user_id=int(user_id)
            ).first()

            if not stats:
                return JsonResponse({'error': 'Member not found in LFG system'}, status=404)

            # Grant pardon - RESET ALL counters for complete fresh start
            # Historical LFGAttendance records are kept for admin viewing but excluded from calculations
            now = int(time.time())
            stats.blacklist_pardoned = True
            stats.blacklist_pardoned_at = now

            # Reset ALL stats to 0 (complete fresh start)
            stats.total_showed = 0
            stats.total_no_shows = 0
            stats.total_cancelled = 0
            stats.total_late = 0
            stats.total_pardoned = 0
            stats.total_signups = 0
            stats.current_noshow_streak = 0
            stats.current_show_streak = 0
            stats.best_show_streak = 0
            stats.reliability_score = 100  # Reset to perfect score

            # If currently blacklisted, remove them from blacklist
            was_blacklisted = stats.is_blacklisted
            if stats.is_blacklisted:
                stats.is_blacklisted = False
                stats.blacklisted_at = None
                stats.blacklist_reason = None

            db.commit()

            return JsonResponse({
                'success': True,
                'pardoned': True,
                'was_blacklisted': was_blacklisted,
                'message': 'Member has been pardoned and will not be auto-blacklisted in the future'
            })

    except Exception as e:
        logger.error(f"Error pardoning member: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_auth_required
@require_subscription_tier('pro', 'premium')
@ratelimit(key='user_or_ip', rate='10/h', method='GET', block=True)
def api_lfg_attendance_export(request, guild_id):
    """GET /api/guild/<id>/lfg/attendance/export/ - Export attendance data as CSV (Pro/Premium only, 10 req/hour)."""
    import csv
    import io
    from django.http import HttpResponse

    try:
        from .db import get_db_session
        from .models import LFGConfig, LFGGroup, LFGAttendance, LFGMemberStats

        with get_db_session() as db:
            # Tier check handled by @require_subscription_tier decorator
            # Check if attendance tracking is enabled
            config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if not config or not config.attendance_tracking_enabled:
                return JsonResponse({'error': 'Attendance tracking not enabled'}, status=403)

            # Fetch member display names from Discord API (using bot token)
            import requests
            member_names_cache = {}

            # Fetch member names from cache (no Discord API call!)
            from .discord_resources import get_guild_members

            try:
                guild_members = get_guild_members(str(guild_id))
                for m in guild_members:
                    user_id = int(m['id'])
                    display_name = m.get('display_name') or m.get('username', 'Unknown')
                    member_names_cache[user_id] = display_name
                logger.info(f"Loaded {len(member_names_cache)} member names from cache for export")
            except Exception as e:
                logger.error(f"Error fetching guild members from cache for export: {e}")

            # Get all groups with attendance from last 90 days
            ninety_days_ago = int(time.time()) - (90 * 24 * 3600)

            groups = db.query(LFGGroup).filter(
                LFGGroup.guild_id == int(guild_id),
                LFGGroup.created_at >= ninety_days_ago
            ).order_by(LFGGroup.scheduled_time.desc()).all()

            # Create CSV
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                'Group Name',
                'Scheduled Date',
                'Member Name',
                'User ID',
                'Status',
                'Joined At',
                'Confirmed At',
                'Showed At',
                'Cancelled At',
                'Late At',
                'No Show At',
                'Reliability Score'
            ])

            # Write data
            for group in groups:
                attendance_records = db.query(LFGAttendance).filter_by(
                    group_id=group.id
                ).all()

                for att in attendance_records:
                    # Get member stats
                    stats = db.query(LFGMemberStats).filter_by(
                        guild_id=int(guild_id),
                        user_id=att.user_id
                    ).first()

                    # Get display name from Discord API cache first, then fallback to GuildMember table
                    display_name = member_names_cache.get(att.user_id)

                    if not display_name:
                        from .models import GuildMember
                        member = db.query(GuildMember).filter_by(
                            guild_id=int(guild_id),
                            user_id=att.user_id
                        ).first()
                        display_name = member.display_name if member else f"User {att.user_id}"

                    # Format timestamps
                    from datetime import datetime
                    scheduled_date = datetime.fromtimestamp(group.scheduled_time).strftime('%Y-%m-%d %H:%M') if group.scheduled_time else 'N/A'
                    joined_at = datetime.fromtimestamp(att.joined_at).strftime('%Y-%m-%d %H:%M') if att.joined_at else ''
                    confirmed_at = datetime.fromtimestamp(att.confirmed_at).strftime('%Y-%m-%d %H:%M') if att.confirmed_at else ''
                    showed_at = datetime.fromtimestamp(att.showed_at).strftime('%Y-%m-%d %H:%M') if att.showed_at else ''
                    cancelled_at = datetime.fromtimestamp(att.cancelled_at).strftime('%Y-%m-%d %H:%M') if att.cancelled_at else ''
                    late_at = datetime.fromtimestamp(att.late_at).strftime('%Y-%m-%d %H:%M') if att.late_at else ''
                    no_show_at = datetime.fromtimestamp(att.no_show_at).strftime('%Y-%m-%d %H:%M') if att.no_show_at else ''

                    writer.writerow([
                        group.thread_name or f"Group {group.id}",
                        scheduled_date,
                        display_name,
                        att.user_id,
                        att.status.value if hasattr(att.status, 'value') else att.status,
                        joined_at,
                        confirmed_at,
                        showed_at,
                        cancelled_at,
                        late_at,
                        no_show_at,
                        stats.reliability_score if stats else 'N/A'
                    ])

            # Return CSV
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="attendance_guild_{guild_id}.csv"'
            return response

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting attendance: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_lfg_user_history(request, guild_id, user_id):
    """GET /api/guild/<id>/lfg/attendance/user/<user_id>/ - Get complete attendance history for a user (Premium)."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGConfig, LFGGroup, LFGAttendance, LFGMemberStats, GuildMember
        from datetime import datetime

        with get_db_session() as db:
            # Check if guild is premium
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild or not guild.is_premium():
                return JsonResponse({'error': 'Premium required'}, status=403)

            # Check if attendance tracking is enabled
            config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if not config or not config.attendance_tracking_enabled:
                return JsonResponse({'error': 'Attendance tracking not enabled'}, status=403)

            # Get user stats
            stats = db.query(LFGMemberStats).filter_by(
                guild_id=int(guild_id),
                user_id=int(user_id)
            ).first()

            # Get user display name
            member = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=int(user_id)
            ).first()
            display_name = member.display_name if member else f"User {user_id}"

            # Get ALL attendance records for this user (including historical pre-pardon records)
            attendance_records = db.query(LFGAttendance).filter_by(
                user_id=int(user_id)
            ).join(LFGGroup).filter(
                LFGGroup.guild_id == int(guild_id)
            ).order_by(LFGGroup.scheduled_time.desc()).all()

            # Format attendance history
            history = []
            from .models import LFGGame
            for att in attendance_records:
                group = db.query(LFGGroup).filter_by(id=att.group_id).first()
                if not group:
                    continue

                # Get game name from LFGGame table
                game = db.query(LFGGame).filter_by(id=group.game_id).first()
                game_name = game.game_name if game else "Unknown Game"

                # Determine if this record is from before pardon
                is_pre_pardon = False
                if stats and stats.blacklist_pardoned and stats.blacklist_pardoned_at:
                    record_time = att.updated_at or att.joined_at or 0
                    if record_time < stats.blacklist_pardoned_at:
                        is_pre_pardon = True

                history.append({
                    'group_id': group.id,
                    'group_name': group.thread_name or f"Group {group.id}",
                    'game_name': game_name,
                    'scheduled_time': group.scheduled_time,
                    'scheduled_date': datetime.fromtimestamp(group.scheduled_time).strftime('%Y-%m-%d %H:%M') if group.scheduled_time else 'N/A',
                    'status': att.status,
                    'joined_at': att.joined_at,
                    'updated_at': att.updated_at,
                    'is_pre_pardon': is_pre_pardon,  # Flag to show in UI
                })

            return JsonResponse({
                'success': True,
                'user_id': user_id,
                'display_name': display_name,
                'stats': {
                    'total_signups': stats.total_signups if stats else 0,
                    'total_showed': stats.total_showed if stats else 0,
                    'total_no_shows': stats.total_no_shows if stats else 0,
                    'total_late': stats.total_late if stats else 0,
                    'total_cancelled': stats.total_cancelled if stats else 0,
                    'total_pardoned': stats.total_pardoned if stats else 0,
                    'reliability_score': stats.reliability_score if stats else 100,
                    'is_blacklisted': stats.is_blacklisted if stats else False,
                    'blacklist_reason': stats.blacklist_reason if stats else None,
                    'blacklist_pardoned': stats.blacklist_pardoned if stats else False,
                    'blacklist_pardoned_at': stats.blacklist_pardoned_at if stats else None,
                },
                'history': history,
                'total_events': len(history)
            })

    except Exception as e:
        logger.error(f"Error fetching user history: {e}")
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# ====== LFG Browser API Endpoints (All Tiers - Game limits enforced) ======

@require_http_methods(["GET"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_lfg_manager_check(request, guild_id):
    """
    GET /api/guild/<id>/lfg/manager-check/ - Check if user has LFG Manager role (real-time).

    This endpoint is polled every 30 seconds by the frontend to detect permission changes
    in real-time without requiring a page refresh. This ensures that:
    - When a user is granted the "LFG Manager" role, Edit/Delete buttons appear immediately
    - When the role is revoked, the buttons disappear immediately
    - Role permission changes (adding/removing CREATE_EVENTS or MANAGE_EVENTS) are detected

    Returns:
        JSON with is_admin, has_lfg_manager_role, and can_manage flags
    """
    try:
        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        # Check if user is admin
        admin_guilds = request.session.get('discord_admin_guilds', [])
        is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)

        # Check for "LFG Manager" role with required permissions
        has_lfg_manager_role = check_lfg_manager_role(guild_id, user_id)

        can_manage = is_admin or has_lfg_manager_role

        return JsonResponse({
            'success': True,
            'is_admin': is_admin,
            'has_lfg_manager_role': has_lfg_manager_role,
            'can_manage': can_manage
        })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in LFG Manager check: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred'}, status=500)


@api_member_auth_required
@require_http_methods(["GET"])
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_lfg_browser_groups(request, guild_id):
    """GET /api/guild/<id>/lfg/browser/groups/ - List active LFG groups with game info (Pro/Premium/VIP)."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGGame, LFGMember

        with get_db_session() as db:
            # Get guild (LFG Browser now available to all tiers)
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get filter parameters
            game_id = request.GET.get('game_id')
            status_filter = request.GET.get('status', 'active')  # active, full, all

            # Build query - only show server-only groups (not Discovery Network)
            query = db.query(LFGGroup).filter(
                LFGGroup.guild_id == int(guild_id),
                LFGGroup.is_active == True,
                LFGGroup.shared_to_network == False  # Find Groups = server-only
            )

            # Filter by game
            if game_id:
                query = query.filter(LFGGroup.game_id == int(game_id))

            # Filter by status
            if status_filter == 'active':
                query = query.filter(LFGGroup.is_full == False)
            elif status_filter == 'full':
                query = query.filter(LFGGroup.is_full == True)

            # Get groups ordered by scheduled time (upcoming first)
            current_time = int(time.time())
            groups = query.order_by(LFGGroup.scheduled_time.asc()).all()

            # PERFORMANCE: Batch load all games and members to avoid N+1 queries
            group_ids = [g.id for g in groups]
            game_ids = list(set([g.game_id for g in groups]))

            # Single query for all games
            games_dict = {}
            if game_ids:
                games = db.query(LFGGame).filter(LFGGame.id.in_(game_ids)).all()
                games_dict = {g.id: g for g in games}

            # Single query for all members
            members_by_group = {}
            if group_ids:
                all_members = db.query(LFGMember).filter(
                    LFGMember.group_id.in_(group_ids),
                    LFGMember.left_at == None
                ).all()

                for member in all_members:
                    if member.group_id not in members_by_group:
                        members_by_group[member.group_id] = []
                    members_by_group[member.group_id].append(member)

            # Single query for all attendance records (if attendance tracking enabled and Pro/Premium/VIP)
            from .models import LFGConfig, LFGAttendance, AttendanceStatus
            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            attendance_by_group = {}
            # Attendance tracking only for Pro/Premium/VIP
            is_attendance_tier = guild.subscription_tier in ['pro', 'Pro', 'premium', 'Premium'] or guild.is_vip
            if lfg_config and lfg_config.attendance_tracking_enabled and is_attendance_tier and group_ids:
                all_attendance = db.query(LFGAttendance).filter(
                    LFGAttendance.group_id.in_(group_ids),
                    LFGAttendance.status == AttendanceStatus.CONFIRMED
                ).all()

                for attendance in all_attendance:
                    if attendance.group_id not in attendance_by_group:
                        attendance_by_group[attendance.group_id] = []
                    attendance_by_group[attendance.group_id].append(attendance)

            groups_data = []
            for group in groups:
                # Get game from pre-loaded dict
                game = games_dict.get(group.game_id)

                # Get members from pre-loaded dict
                members = members_by_group.get(group.id, [])

                # Populate display names if null (for legacy records)
                from .models import GuildMember as GuildMemberModel
                members_data = []
                for m in members:
                    display_name = m.display_name
                    if not display_name:
                        # Look up display name from GuildMember table
                        guild_member = db.query(GuildMemberModel).filter_by(
                            guild_id=int(guild_id),
                            user_id=m.user_id
                        ).first()
                        display_name = guild_member.display_name if guild_member else f'User {m.user_id}'
                        # Update the LFGMember record with the display name
                        m.display_name = display_name

                    members_data.append({
                        'user_id': str(m.user_id),
                        'display_name': display_name,
                        'rank_value': m.rank_value,
                        'is_creator': m.is_creator,
                        'is_co_leader': m.is_co_leader,
                        'selections': json_lib.loads(m.selections) if m.selections else {},
                        'joined_at': m.joined_at,
                    })

                # Commit any display name updates
                db.commit()

                # Determine if group is in the past, present, or future
                time_status = 'upcoming'
                if group.scheduled_time:
                    if group.scheduled_time < current_time - (2 * 3600):  # More than 2 hours ago
                        time_status = 'past'
                    elif group.scheduled_time < current_time + (1800):  # Within 30 minutes
                        time_status = 'active_now'

                # Get confirmed attendance for this group
                confirmed_attendance = attendance_by_group.get(group.id, [])
                confirmed_ids = [a.user_id for a in confirmed_attendance]

                groups_data.append({
                    'id': group.id,
                    'game_id': group.game_id,
                    'game_name': game.game_name if game else 'Unknown',
                    'game_short': game.game_short if game else '',
                    'game_emoji': game.game_emoji if game else '',
                    'cover_url': game.cover_url if game else None,
                    'thread_id': str(group.thread_id) if group.thread_id else None,
                    'thread_name': group.thread_name,
                    'creator_id': str(group.creator_id),
                    'creator_name': group.creator_name,
                    'scheduled_time': group.scheduled_time,
                    'description': group.description,
                    'max_group_size': group.max_group_size or (game.max_group_size if game else 4),
                    'is_raid': group.is_raid or False,
                    'tanks_needed': group.tanks_needed,
                    'healers_needed': group.healers_needed,
                    'dps_needed': group.dps_needed,
                    'is_active': group.is_active,
                    'is_full': group.is_full,
                    'member_count': group.member_count,
                    'members': members_data,
                    'created_at': group.created_at,
                    'time_status': time_status,
                    'confirmed_attendance_count': len(confirmed_ids),
                    'confirmed_attendance_ids': [str(uid) for uid in confirmed_ids],
                    'attendance_tracking_enabled': (lfg_config.attendance_tracking_enabled and is_attendance_tier) if lfg_config else False,
                })

            response = JsonResponse({
                'success': True,
                'groups': groups_data,
                'count': len(groups_data)
            })
            # No caching - always fetch fresh data for instant updates
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'
            return response

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error fetching LFG browser groups: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_browser_create(request, guild_id):
    """POST /api/guild/<id>/lfg/browser/create/ - Create LFG group from web (Pro/Premium/VIP)."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGGame, LFGMember
        import logging
        logger = logging.getLogger(__name__)

        data = json.loads(request.body)
        game_id = data.get('game_id')
        title = data.get('title', '').strip()
        scheduled_time = data.get('scheduled_time')
        event_duration = data.get('event_duration')
        ping_role_id = data.get('ping_role_id')
        description = data.get('description', '')
        max_group_size = data.get('max_group_size')
        co_leader_ids = data.get('co_leader_ids', [])  # List of user IDs
        creator_options = data.get('creator_options', {})  # Creator's game-specific selections
        create_discord_thread = data.get('create_discord_thread', False)  # Whether to create Discord thread

        # Raid composition fields (unified LFG/LFR system)
        is_raid = data.get('is_raid', False)
        tanks_needed = data.get('tanks_needed')
        healers_needed = data.get('healers_needed')
        dps_needed = data.get('dps_needed')

        if not game_id:
            return JsonResponse({'error': 'Game ID required'}, status=400)

        if not title:
            return JsonResponse({'error': 'Group title required'}, status=400)

        if len(title) > 150:
            return JsonResponse({'error': 'Title must be 150 characters or less'}, status=400)

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')
        username = discord_user.get('username', 'Unknown')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        with get_db_session() as db:
            # Get guild (LFG Browser available to all tiers)
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get game
            game = db.query(LFGGame).filter_by(id=int(game_id), guild_id=int(guild_id)).first()
            if not game:
                return JsonResponse({'error': 'Game not found'}, status=404)

            # Create group (note: thread_id will be 0 for web-created groups until bot creates thread)
            new_group = LFGGroup(
                guild_id=int(guild_id),
                game_id=game.id,
                thread_id=0,  # Will be updated by bot when thread is created
                thread_name=title,
                ping_role_id=int(ping_role_id) if ping_role_id else None,
                creator_id=int(user_id),
                creator_name=username,
                scheduled_time=scheduled_time,
                event_duration=event_duration,
                description=description,
                max_group_size=max_group_size or game.max_group_size,
                is_raid=is_raid,
                tanks_needed=tanks_needed,
                healers_needed=healers_needed,
                dps_needed=dps_needed,
                is_active=True,
                is_full=False,
                member_count=1,
                shared_to_network=False,  # Find Groups are server-only, not cross-server
                created_at=int(time.time())
            )
            db.add(new_group)
            db.flush()  # Get the ID

            # Add creator as first member
            # Extract rank and selections from creator_options
            rank_value = creator_options.get('rank') if creator_options else None
            selections_dict = {k: v for k, v in creator_options.items() if k != 'rank'} if creator_options else {}

            creator_member = LFGMember(
                group_id=new_group.id,
                user_id=int(user_id),
                display_name=username,
                rank_value=rank_value,
                selections=json_lib.dumps(selections_dict) if selections_dict else None,
                is_creator=True,
                joined_at=int(time.time())
            )
            db.add(creator_member)

            # Add co-leaders as members
            from .models import GuildMember as GuildMemberModel
            co_leaders_added = []
            for co_leader_id in co_leader_ids:
                # Skip if co-leader is the creator themselves
                if str(co_leader_id) == str(user_id):
                    continue

                # Fetch display name from GuildMember table
                guild_member = db.query(GuildMemberModel).filter_by(
                    guild_id=int(guild_id),
                    user_id=int(co_leader_id)
                ).first()
                co_leader_display_name = guild_member.display_name if guild_member else f'User {co_leader_id}'

                co_leader_member = LFGMember(
                    group_id=new_group.id,
                    user_id=int(co_leader_id),
                    display_name=co_leader_display_name,
                    is_co_leader=True,
                    joined_at=int(time.time())
                )
                db.add(co_leader_member)
                new_group.member_count += 1
                co_leaders_added.append(int(co_leader_id))

            # Log audit entry for group creation
            log_lfg_audit(
                db=db,
                guild_id=guild_id,
                group_id=new_group.id,
                action='create',
                actor_id=user_id,
                actor_name=username,
                group_name=title,
                game_name=game.game_name
            )

            # Auto-confirm attendance for creator and co-leaders if attendance tracking is enabled (Pro/Premium/VIP only)
            from .models import LFGConfig
            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            is_attendance_tier = guild.subscription_tier in ['pro', 'Pro', 'premium', 'Premium'] or guild.is_vip
            if lfg_config and lfg_config.attendance_tracking_enabled and is_attendance_tier:
                from .models import LFGAttendance, AttendanceStatus

                now = int(time.time())

                # Auto-confirm creator
                creator_attendance = LFGAttendance(
                    group_id=new_group.id,
                    user_id=int(user_id),
                    status=AttendanceStatus.CONFIRMED,
                    confirmed_at=now,
                    joined_at=now
                )
                db.add(creator_attendance)

                # Auto-confirm co-leaders
                for co_leader_id in co_leaders_added:
                    co_leader_attendance = LFGAttendance(
                        group_id=new_group.id,
                        user_id=co_leader_id,
                        status=AttendanceStatus.CONFIRMED,
                        confirmed_at=now,
                        joined_at=now
                    )
                    db.add(co_leader_attendance)

            db.commit()

            # Queue Discord thread creation if user requested it (independent of lfg_config)
            if create_discord_thread and game.lfg_channel_id:
                from .models import PendingAction, ActionType, ActionStatus
                import time as time_lib

                action = PendingAction(
                    guild_id=int(guild_id),
                    action_type=ActionType.LFG_THREAD_CREATE,
                    payload=json.dumps({
                        'group_id': new_group.id,
                        'channel_id': str(game.lfg_channel_id)
                    }),
                    status=ActionStatus.PENDING,
                    priority=1,
                    created_at=int(time_lib.time())
                )
                db.add(action)
                db.commit()

            # Send webhook notifications if lfg_config exists and webhooks are enabled
            if not lfg_config:
                lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()

            if lfg_config:
                # Build notification embed with full group details
                scheduled_text = f"<t:{scheduled_time}:F>" if scheduled_time else "Now / Flexible"

                # Build members list with roles
                # Extract creator's role from selections if available
                creator_role = 'Leader'
                if creator_options:
                    for key, value in creator_options.items():
                        if key != 'rank' and value:
                            creator_role = value
                            break
                members_text = f"<@{user_id}> - {creator_role}"
                member_count = 1

                # Add co-leaders to members list
                for co_leader_id in co_leaders_added:
                    members_text += f"\n<@{co_leader_id}> - Co-Leader"
                    member_count += 1

                # Build confirmed attendance list (for premium guilds with attendance enabled)
                confirmed_attendance = ""
                if hasattr(lfg_config, 'attendance_tracking_enabled') and lfg_config.attendance_tracking_enabled:
                    # For now, creator is auto-confirmed
                    confirmed_attendance = f"<@{user_id}>"
                    for co_leader_id in co_leaders_added:
                        confirmed_attendance += f" <@{co_leader_id}>"

                # Format event duration
                duration_text = 'Not set'
                if new_group.event_duration:
                    duration_text = f"{new_group.event_duration} hour{'s' if new_group.event_duration != 1 else ''}"

                embed_data = {
                    'title': f"{game.game_emoji or '🎮'} {title}",
                    'description': f"Group by <@{user_id}>",
                    'color': 0x5865F2,  # Discord blurple
                    'fields': [
                        {'name': 'Scheduled Time', 'value': scheduled_text, 'inline': True},
                        {'name': 'Event Duration ⏱️', 'value': duration_text, 'inline': True},
                        {'name': 'Activity 🎮', 'value': 'No one currently playing', 'inline': True},
                        {'name': f'Members ({member_count}/{new_group.max_group_size})', 'value': members_text, 'inline': False},
                    ],
                    'footer': {'text': f'Group ID: {new_group.id}'},
                    'group_id': new_group.id,  # For bot action processing
                }

                if description:
                    embed_data['fields'].insert(3, {'name': '📝 Description', 'value': description[:1024], 'inline': False})

                # Add confirmed attendance if tracking is enabled
                if confirmed_attendance:
                    confirm_count = len(confirmed_attendance.split())
                    embed_data['fields'].append({
                        'name': f'📋 Confirmed Attendance ({confirm_count}/{member_count})',
                        'value': confirmed_attendance,
                        'inline': False
                    })

                # Send webhook notification if configured
                if lfg_config.webhook_url and lfg_config.notify_on_group_create:
                    send_lfg_webhook_notification(lfg_config.webhook_url, embed_data)

            # Success message varies based on whether Discord thread was requested
            success_message = 'LFG group created!'
            if create_discord_thread:
                success_message = 'LFG group created! Bot will create a Discord thread shortly.'

            return JsonResponse({
                'success': True,
                'message': success_message,
                'group': {
                    'id': new_group.id,
                    'game_name': game.game_name,
                    'scheduled_time': new_group.scheduled_time,
                    'co_leaders': co_leaders_added,
                }
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error creating LFG group: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_browser_join(request, guild_id, group_id):
    """POST /api/guild/<id>/lfg/browser/<group_id>/join/ - Join LFG group (Pro/Premium/VIP)."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')
        username = discord_user.get('username', 'Unknown')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        # Parse request body for options
        data = json.loads(request.body) if request.body else {}
        options = data.get('options', {})

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get group
            group = db.query(LFGGroup).filter_by(id=int(group_id), guild_id=int(guild_id)).first()
            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            if not group.is_active:
                return JsonResponse({'error': 'Group is no longer active'}, status=400)

            if group.is_full:
                return JsonResponse({'error': 'Group is full'}, status=400)

            # Check if user has any existing member record (active or left)
            existing = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id)
            ).first()

            # Extract rank if provided
            rank_value = options.get('rank')

            # Extract other selections (exclude 'rank' as it's stored separately)
            selections = {k: v for k, v in options.items() if k != 'rank'}
            selections_json = json.dumps(selections) if selections else None

            if existing:
                # Check if they're currently in the group
                if existing.left_at is None:
                    return JsonResponse({'error': 'You are already in this group'}, status=400)

                # They left before, so rejoin them by updating their record
                existing.display_name = username
                existing.rank_value = rank_value
                existing.selections = selections_json
                existing.joined_at = int(time.time())
                existing.left_at = None  # Clear the left timestamp

                # Update group member count
                group.member_count += 1
                if group.member_count >= group.max_group_size:
                    group.is_full = True
            else:
                # Add new member
                new_member = LFGMember(
                    group_id=group.id,
                    user_id=int(user_id),
                    display_name=username,
                    is_creator=False,
                    rank_value=rank_value,
                    selections=selections_json,
                    joined_at=int(time.time())
                )
                db.add(new_member)

                # Update group member count
                group.member_count += 1
                if group.member_count >= group.max_group_size:
                    group.is_full = True

            # Get game info for notifications (before commit)
            from .models import LFGGame
            game = db.query(LFGGame).filter_by(id=group.game_id).first()

            # Auto-confirm attendance if attendance tracking is enabled (Pro/Premium/VIP only)
            from .models import LFGConfig
            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            is_attendance_tier = guild.subscription_tier in ['pro', 'Pro', 'premium', 'Premium'] or guild.is_vip
            if lfg_config and lfg_config.attendance_tracking_enabled and is_attendance_tier:
                from .models import LFGAttendance, AttendanceStatus

                # Check if user already has an attendance record
                attendance = db.query(LFGAttendance).filter_by(
                    group_id=group.id,
                    user_id=int(user_id)
                ).first()

                now = int(time.time())
                if not attendance:
                    # Create new attendance record with CONFIRMED status
                    attendance = LFGAttendance(
                        group_id=group.id,
                        user_id=int(user_id),
                        status=AttendanceStatus.CONFIRMED,
                        confirmed_at=now,
                        joined_at=now
                    )
                    db.add(attendance)
                else:
                    # Update existing record to CONFIRMED
                    attendance.status = AttendanceStatus.CONFIRMED
                    attendance.confirmed_at = now
                    if not attendance.joined_at:
                        attendance.joined_at = now

            db.commit()

            # Queue bot action to update thread embed and add user to thread
            from .models import PendingAction, ActionType, ActionStatus
            import logging
            logger = logging.getLogger(__name__)

            if lfg_config and group.thread_id:
                import time as time_lib

                action = PendingAction(
                    guild_id=int(guild_id),
                    action_type=ActionType.LFG_THREAD_UPDATE,
                    payload=json.dumps({
                        'group_id': group.id,
                        'add_user_to_thread': int(user_id),  # Add this user to the thread
                    }),
                    status=ActionStatus.PENDING,
                    priority=2,
                    created_at=int(time_lib.time())
                )
                db.add(action)
                db.commit()
                logger.info(f"Queued LFG thread update action for join: user {user_id} to group {group.id}")

            return JsonResponse({
                'success': True,
                'message': 'Successfully joined the group!',
                'member_count': group.member_count
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error joining LFG group: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE"])
@api_member_auth_required
@ratelimit(key='user', rate='10/m', method='DELETE', block=True)
def api_lfg_browser_leave(request, guild_id, group_id):
    """DELETE /api/guild/<id>/lfg/browser/<group_id>/leave/ - Leave LFG group (Pro/Premium/VIP)."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get group
            group = db.query(LFGGroup).filter_by(id=int(group_id), guild_id=int(guild_id)).first()
            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            # Find membership
            member = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id)
            ).filter(LFGMember.left_at == None).first()

            if not member:
                return JsonResponse({'error': 'You are not in this group'}, status=400)

            if member.is_creator:
                return JsonResponse({'error': 'Group creator cannot leave. Delete the group instead.'}, status=400)

            # Mark as left
            member.left_at = int(time.time())

            # Update group member count
            group.member_count = max(1, group.member_count - 1)
            group.is_full = False

            # Get game info for notifications (before commit)
            from .models import LFGGame
            game = db.query(LFGGame).filter_by(id=group.game_id).first()

            db.commit()

            # Queue bot action to update thread embed (user will be removed from embed automatically)
            from .models import LFGConfig, PendingAction, ActionType, ActionStatus
            import logging
            logger = logging.getLogger(__name__)

            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if lfg_config and group.thread_id:
                import time as time_lib

                action = PendingAction(
                    guild_id=int(guild_id),
                    action_type=ActionType.LFG_THREAD_UPDATE,
                    payload=json.dumps({
                        'group_id': group.id,
                    }),
                    status=ActionStatus.PENDING,
                    priority=2,
                    created_at=int(time_lib.time())
                )
                db.add(action)
                db.commit()
                logger.info(f"Queued LFG thread update action for leave: user {user_id} from group {group.id}")

            return JsonResponse({
                'success': True,
                'message': 'Successfully left the group',
                'member_count': group.member_count
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error leaving LFG group: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_browser_remove_member(request, guild_id, group_id):
    """POST /api/guild/<id>/lfg/browser/<group_id>/remove-member/ - Remove a member from LFG group (Creator/Co-Leader only)."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        data = json.loads(request.body)
        target_user_id = data.get('user_id')

        if not target_user_id:
            return JsonResponse({'error': 'user_id is required'}, status=400)

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get group
            group = db.query(LFGGroup).filter_by(id=int(group_id), guild_id=int(guild_id)).first()
            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            # Check permissions: only creator or co-leaders can remove members
            requester_member = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id)
            ).filter(LFGMember.left_at == None).first()

            is_creator = requester_member and requester_member.is_creator
            is_co_leader = requester_member and requester_member.is_co_leader

            if not (is_creator or is_co_leader):
                return JsonResponse({'error': 'Permission denied. Only the creator or co-leaders can remove members.'}, status=403)

            # Find target member
            target_member = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(target_user_id)
            ).filter(LFGMember.left_at == None).first()

            if not target_member:
                return JsonResponse({'error': 'Member not found in group'}, status=404)

            # Cannot remove the creator
            if target_member.is_creator:
                return JsonResponse({'error': 'Cannot remove the group creator'}, status=400)

            # Mark as left
            target_member.left_at = int(time.time())
            target_display_name = target_member.display_name or f"User {target_user_id}"

            # Update group member count
            group.member_count = max(1, group.member_count - 1)
            group.is_full = False

            db.commit()

            # Queue bot action to remove user from thread and update embed
            from .models import LFGConfig, PendingAction, ActionType, ActionStatus
            import logging
            logger = logging.getLogger(__name__)

            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if lfg_config and group.thread_id:
                import time as time_lib

                action = PendingAction(
                    guild_id=int(guild_id),
                    action_type=ActionType.LFG_THREAD_UPDATE,
                    payload=json.dumps({
                        'group_id': group.id,
                        'remove_users_from_thread': [int(target_user_id)],  # Remove from thread
                    }),
                    status=ActionStatus.PENDING,
                    priority=2,
                    created_at=int(time_lib.time())
                )
                db.add(action)
                db.commit()
                logger.info(f"Queued LFG thread update action to remove user {target_user_id} from group {group.id}")

            return JsonResponse({
                'success': True,
                'message': f'Removed {target_display_name} from the group',
                'member_count': group.member_count
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error removing member from LFG group: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_browser_join_thread(request, guild_id, group_id):
    """POST /api/guild/<id>/lfg/browser/<group_id>/join-thread/ - Add user to Discord thread."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, PendingAction, ActionType, ActionStatus
        import logging
        logger = logging.getLogger(__name__)

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get group
            group = db.query(LFGGroup).filter_by(id=int(group_id), guild_id=int(guild_id)).first()
            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            if not group.thread_id:
                return JsonResponse({'error': 'This group does not have a Discord thread'}, status=400)

            # Queue bot action to add user to thread
            action = PendingAction(
                guild_id=int(guild_id),
                action_type=ActionType.LFG_THREAD_UPDATE,
                payload=json.dumps({
                    'group_id': group.id,
                    'add_user_to_thread': int(user_id),
                }),
                status=ActionStatus.PENDING,
                priority=2,
                created_at=int(time.time())
            )
            db.add(action)
            db.commit()

            logger.info(f"Queued action to add user {user_id} to thread {group.thread_id}")

            return JsonResponse({
                'success': True,
                'message': 'You will be added to the Discord thread shortly'
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error joining thread: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["DELETE"])
@api_member_auth_required
@ratelimit(key='user', rate='10/m', method='DELETE', block=True)
def api_lfg_browser_delete(request, guild_id, group_id):
    """
    DELETE /api/guild/<id>/lfg/browser/<group_id>/delete/ - Delete LFG group.

    Permission hierarchy (any of these grants delete access):
    1. Group creator (always can delete their own group)
    2. Co-leaders (designated by creator)
    3. LFG Managers (users with "LFG Manager" Discord role with proper permissions)
    4. Server admins

    Returns:
        200: Group deleted successfully
        403: Permission denied
        404: Group not found
    """
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        # Check if user is server admin
        admin_guilds = request.session.get('admin_guilds', [])
        is_admin = str(guild_id) in [str(g['id']) for g in admin_guilds]

        # Check if user has "LFG Manager" role (users with this role can manage ANY group)
        has_lfg_manager_role = check_lfg_manager_role(guild_id, user_id)

        can_manage = is_admin or has_lfg_manager_role

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get group
            group = db.query(LFGGroup).filter_by(id=int(group_id), guild_id=int(guild_id)).first()
            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            # Check permissions: creator, co-leader, or admin
            is_creator = group.creator_id == int(user_id)

            # Check if user is a co-leader
            is_co_leader = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id),
                is_co_leader=True
            ).filter(LFGMember.left_at == None).first() is not None

            if not (is_creator or is_co_leader or can_manage):
                return JsonResponse({'error': 'Permission denied. Only the creator, co-leaders, admins, or users with the LFG Manager role can delete this group.'}, status=403)

            # Capture group info before deletion for audit log
            from .models import LFGGame, LFGConfig, PendingAction, ActionType, ActionStatus
            import logging
            logger = logging.getLogger(__name__)

            game = db.query(LFGGame).filter_by(id=group.game_id).first()
            group_name_snapshot = group.thread_name
            game_name_snapshot = game.game_name if game else 'Unknown'
            game_emoji = game.game_emoji if game else '🎮'
            thread_id_snapshot = group.thread_id

            # Log audit entry for group deletion
            log_lfg_audit(
                db=db,
                guild_id=guild_id,
                group_id=group.id,
                action='delete',
                actor_id=user_id,
                actor_name=discord_user.get('username', 'Unknown'),
                group_name=group_name_snapshot,
                game_name=game_name_snapshot
            )

            # Queue bot action to delete thread and send cancellation notice
            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if lfg_config and lfg_config.browser_notify_channel_id and lfg_config.notify_on_group_delete:
                import time as time_lib

                action = PendingAction(
                    guild_id=int(guild_id),
                    action_type=ActionType.LFG_THREAD_DELETE,
                    payload=json.dumps({
                        'group_id': group.id,
                        'thread_id': str(thread_id_snapshot) if thread_id_snapshot else None,
                        'thread_name': group_name_snapshot,
                        'game_name': game_name_snapshot,
                        'game_emoji': game_emoji,
                        'deleted_by_id': user_id,
                        'channel_id': str(lfg_config.browser_notify_channel_id)
                    }),
                    status=ActionStatus.PENDING,
                    priority=1,  # High priority
                    created_at=int(time_lib.time())
                )
                db.add(action)
                db.commit()  # Commit action BEFORE deleting group so bot can process it
                logger.info(f"Queued LFG thread deletion action for group {group.id}")

            # Delete all members first
            db.query(LFGMember).filter_by(group_id=group.id).delete()

            # Delete the group
            db.delete(group)
            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Group deleted successfully'
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error deleting LFG group: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["POST"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='10/m', method='POST', block=True)
def api_lfg_convert_to_thread(request, guild_id, group_id):
    """
    POST /api/guild/<id>/lfg/browser/<group_id>/convert-to-thread/ - Convert web-only LFG group to Discord thread.

    Converts a browser-only LFG group into a Discord thread, allowing members to join from Discord.
    Thread will be created in the game's configured LFG channel and will stay active for 7 days.

    Permission Hierarchy:
    - Server admins: Full control of all groups
    - Group creator: Can convert their own group
    - Co-leaders: Can convert appointed groups

    Security:
    - Rate limited to 10 conversions per minute to prevent spam
    - Validates group has no existing thread
    - Requires game's LFG channel to be configured

    Returns:
        200: Thread creation queued successfully
        400: Group already has a thread or LFG channel not configured
        403: Permission denied (not admin, creator, or co-leader)
        404: Group or game not found
    """
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGConfig, LFGGame, LFGMember, PendingAction, ActionType, ActionStatus
        import logging
        import time as time_lib
        logger = logging.getLogger(__name__)

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')
        username = discord_user.get('username', 'Unknown')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get group
            group = db.query(LFGGroup).filter_by(id=int(group_id), guild_id=int(guild_id)).first()
            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            # Permission hierarchy:
            # 1. Server admins (full control of all groups)
            # 2. Group creator (control their own group)
            # 3. Co-leaders (control appointed groups)

            # Check if user is server admin
            admin_guilds = request.session.get('discord_admin_guilds', [])
            logger.info(f"DEBUG: admin_guilds from session: {admin_guilds}")
            logger.info(f"DEBUG: guild_id checking: {guild_id} (type: {type(guild_id)})")

            is_admin = str(guild_id) in [str(g['id']) for g in admin_guilds]
            logger.info(f"DEBUG: is_admin result: {is_admin}")

            # Check if user is the group creator
            is_creator = group.creator_id == int(user_id)
            logger.info(f"DEBUG: is_creator: {is_creator} (group.creator_id={group.creator_id}, user_id={user_id})")

            # Check if user is a co-leader
            co_leader_member = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id),
                is_co_leader=True
            ).filter(LFGMember.left_at == None).first()
            is_co_leader = co_leader_member is not None
            logger.info(f"DEBUG: is_co_leader: {is_co_leader}")

            # Admin, creator, or co-leader can convert
            if not (is_admin or is_creator or is_co_leader):
                return JsonResponse({
                    'success': False,
                    'error': 'Only server admins, the group creator, or co-leaders can convert this group to a Discord thread'
                }, status=403)

            # Check if group already has a thread
            if group.thread_id and group.thread_id > 0:
                return JsonResponse({
                    'success': False,
                    'error': 'This group already has a Discord thread'
                }, status=400)

            # Get game to check if LFG channel is configured
            game = db.query(LFGGame).filter_by(id=group.game_id).first()
            if not game:
                return JsonResponse({
                    'success': False,
                    'error': 'Game not found'
                }, status=404)

            if not game.lfg_channel_id:
                return JsonResponse({
                    'success': False,
                    'error': f'LFG channel not configured for {game.game_name}. Please ask an admin to set it up in the LFG settings.'
                }, status=400)

            # Get LFG config for notification settings
            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()

            # Queue bot action to create the Discord thread
            action = PendingAction(
                guild_id=int(guild_id),
                action_type=ActionType.LFG_THREAD_CREATE,
                payload=json.dumps({
                    'group_id': group.id,
                    'channel_id': str(game.lfg_channel_id)  # Use game's LFG channel, not notify channel
                }),
                status=ActionStatus.PENDING,
                priority=1,  # High priority
                created_at=int(time_lib.time())
            )
            db.add(action)

            # Log audit entry
            log_lfg_audit(
                db=db,
                guild_id=guild_id,
                group_id=group.id,
                action='convert_to_thread',
                actor_id=user_id,
                actor_name=username,
                group_name=group.thread_name,
                game_name=game.game_name if game else 'Unknown'
            )

            db.commit()
            logger.info(f"User {username} ({user_id}) converted LFG group {group.id} to thread")

            return JsonResponse({
                'success': True,
                'message': 'Discord thread creation queued! The bot will create the thread shortly.'
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error converting LFG group to thread: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["PUT", "PATCH"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method=['PUT', 'PATCH'], block=True)
def api_lfg_browser_update(request, guild_id, group_id):
    """
    PUT/PATCH /api/guild/<id>/lfg/browser/<group_id>/update/ - Update LFG group.

    Permission hierarchy (any of these grants edit access):
    1. Group creator (always can edit their own group)
    2. Co-leaders (designated by creator)
    3. LFG Managers (users with "LFG Manager" Discord role with proper permissions)
    4. Server admins

    Updatable fields:
    - description: Group description/notes
    - scheduled_time: Unix timestamp of when group is scheduled
    - max_group_size: Maximum number of members
    - co_leader_ids: List of user IDs to designate as co-leaders

    Returns:
        200: Group updated successfully with updated group data
        403: Permission denied
        404: Group not found
    """
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        # Check if user is server admin
        admin_guilds = request.session.get('admin_guilds', [])
        is_admin = str(guild_id) in [str(g['id']) for g in admin_guilds]

        # Check if user has "LFG Manager" role (users with this role can manage ANY group)
        has_lfg_manager_role = check_lfg_manager_role(guild_id, user_id)

        can_manage = is_admin or has_lfg_manager_role

        data = json.loads(request.body)
        description = data.get('description')
        scheduled_time = data.get('scheduled_time')
        max_group_size = data.get('max_group_size')
        co_leader_ids = data.get('co_leader_ids')

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get group
            group = db.query(LFGGroup).filter_by(id=int(group_id), guild_id=int(guild_id)).first()
            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            # Check permissions: creator, co-leader, or admin
            is_creator = group.creator_id == int(user_id)

            # Check if user is a co-leader
            is_co_leader = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id),
                is_co_leader=True
            ).filter(LFGMember.left_at == None).first() is not None

            if not (is_creator or is_co_leader or can_manage):
                return JsonResponse({'error': 'Permission denied. Only the creator, co-leaders, admins, or users with the LFG Manager role can update this group.'}, status=403)

            # Get game info for audit log
            from .models import LFGGame
            game = db.query(LFGGame).filter_by(id=group.game_id).first()
            group_name_snapshot = group.thread_name
            game_name_snapshot = game.game_name if game else 'Unknown'
            actor_name = discord_user.get('username', 'Unknown')

            # Update fields if provided and log changes
            if description is not None and group.description != description:
                log_lfg_audit(
                    db=db, guild_id=guild_id, group_id=group.id, action='update',
                    actor_id=user_id, actor_name=actor_name,
                    field_changed='description',
                    old_value=group.description, new_value=description,
                    group_name=group_name_snapshot, game_name=game_name_snapshot
                )
                group.description = description

            if scheduled_time is not None and group.scheduled_time != scheduled_time:
                log_lfg_audit(
                    db=db, guild_id=guild_id, group_id=group.id, action='update',
                    actor_id=user_id, actor_name=actor_name,
                    field_changed='scheduled_time',
                    old_value=str(group.scheduled_time), new_value=str(scheduled_time),
                    group_name=group_name_snapshot, game_name=game_name_snapshot
                )
                group.scheduled_time = scheduled_time

            if max_group_size is not None and group.max_group_size != max_group_size:
                log_lfg_audit(
                    db=db, guild_id=guild_id, group_id=group.id, action='update',
                    actor_id=user_id, actor_name=actor_name,
                    field_changed='max_group_size',
                    old_value=str(group.max_group_size), new_value=str(max_group_size),
                    group_name=group_name_snapshot, game_name=game_name_snapshot
                )
                group.max_group_size = max_group_size
                # Update is_full status based on new size
                group.is_full = group.member_count >= max_group_size

            # Track removed co-leaders for thread management
            removed_co_leader_ids = []

            # Update co-leaders if provided (only allow creator to modify co-leaders)
            if co_leader_ids is not None and is_creator:
                # Get existing co-leaders before removing them
                existing_co_leaders = db.query(LFGMember).filter_by(
                    group_id=group.id,
                    is_co_leader=True
                ).filter(LFGMember.left_at == None).all()

                # Track which co-leaders are being removed
                existing_co_leader_ids = [co.user_id for co in existing_co_leaders]
                new_co_leader_ids = [int(cid) for cid in co_leader_ids]
                removed_co_leader_ids = [cid for cid in existing_co_leader_ids if cid not in new_co_leader_ids and cid != group.creator_id]

                for co_leader in existing_co_leaders:
                    # Remove co-leader status
                    db.delete(co_leader)
                    group.member_count = max(1, group.member_count - 1)

                # Add new co-leaders
                for co_leader_id in co_leader_ids:
                    # Skip if co-leader is the creator themselves
                    if str(co_leader_id) == str(group.creator_id):
                        continue

                    # Check if they're already a regular member
                    existing_member = db.query(LFGMember).filter_by(
                        group_id=group.id,
                        user_id=int(co_leader_id)
                    ).filter(LFGMember.left_at == None).first()

                    if existing_member:
                        # Promote to co-leader
                        existing_member.is_co_leader = True
                    else:
                        # Fetch display name from GuildMember table
                        from .models import GuildMember as GuildMemberModel
                        guild_member = db.query(GuildMemberModel).filter_by(
                            guild_id=int(guild_id),
                            user_id=int(co_leader_id)
                        ).first()
                        co_leader_display_name = guild_member.display_name if guild_member else f'User {co_leader_id}'

                        # Add as new co-leader member
                        new_co_leader = LFGMember(
                            group_id=group.id,
                            user_id=int(co_leader_id),
                            display_name=co_leader_display_name,
                            is_co_leader=True,
                            joined_at=int(time.time())
                        )
                        db.add(new_co_leader)
                        group.member_count += 1

            # Update game-specific options if provided
            leader_options = data.get('leader_options')
            if leader_options:
                # Find the current user's member record
                current_user_member = db.query(LFGMember).filter_by(
                    group_id=group.id,
                    user_id=int(user_id)
                ).filter(LFGMember.left_at == None).first()

                if current_user_member:
                    # Extract rank and selections
                    rank_value = leader_options.get('rank') if leader_options else None
                    selections_dict = {k: v for k, v in leader_options.items() if k != 'rank'} if leader_options else {}

                    # Separate Activity (group-level) from personal options (Class, Spec, etc.)
                    activity_value = selections_dict.pop('Activity', None)

                    # Update current user's personal selections (Class, Spec, etc.)
                    current_user_member.rank_value = rank_value
                    current_user_member.selections = json_lib.dumps(selections_dict) if selections_dict else None

                    # Update Activity on the leader's record (group-level setting)
                    # This allows co-leaders to modify the group's activities
                    if activity_value is not None:
                        leader_member = db.query(LFGMember).filter_by(
                            group_id=group.id,
                            is_creator=True
                        ).filter(LFGMember.left_at == None).first()

                        if leader_member:
                            # Get leader's current selections
                            leader_selections = json_lib.loads(leader_member.selections) if leader_member.selections else {}
                            # Update Activity
                            leader_selections['Activity'] = activity_value
                            # Save back to leader
                            leader_member.selections = json_lib.dumps(leader_selections)

            db.commit()

            # Queue bot action to update the thread embed
            from .models import LFGConfig, PendingAction, ActionType, ActionStatus
            import logging
            logger = logging.getLogger(__name__)

            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if lfg_config and lfg_config.browser_notify_channel_id and lfg_config.notify_on_group_update:
                import time as time_lib

                action = PendingAction(
                    guild_id=int(guild_id),
                    action_type=ActionType.LFG_THREAD_UPDATE,
                    payload=json.dumps({
                        'group_id': group.id,
                        'remove_users_from_thread': removed_co_leader_ids,  # Remove ex-co-leaders from thread
                    }),
                    status=ActionStatus.PENDING,
                    priority=2,  # Medium priority
                    created_at=int(time_lib.time())
                )
                db.add(action)
                logger.info(f"Queued LFG thread update action for group {group.id}, removing {len(removed_co_leader_ids)} users from thread")

            return JsonResponse({
                'success': True,
                'message': 'Group updated successfully',
                'group': {
                    'id': group.id,
                    'description': group.description,
                    'scheduled_time': group.scheduled_time,
                    'max_group_size': group.max_group_size,
                    'member_count': group.member_count,
                }
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error updating LFG group: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_lfg_browser_audit_logs(request, guild_id):
    """
    GET /api/guild/<id>/lfg/browser/audit-logs/ - Get audit logs for LFG groups.

    Query parameters:
    - group_id (optional): Filter logs for specific group
    - limit (optional): Number of logs to return (default 50, max 200)
    - offset (optional): Pagination offset

    Access: Admins and LFG Managers only

    Returns:
        JSON with audit log entries including actor, action, changes, and timestamps
    """
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroupAuditLog

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        # Check if user is admin or LFG Manager
        admin_guilds = request.session.get('discord_admin_guilds', [])
        is_admin = any(str(g['id']) == str(guild_id) for g in admin_guilds)
        has_lfg_manager_role = check_lfg_manager_role(guild_id, user_id)

        if not (is_admin or has_lfg_manager_role):
            return JsonResponse({'error': 'Permission denied. Only admins and LFG Managers can view audit logs.'}, status=403)

        # Get query parameters
        group_id_filter = request.GET.get('group_id')
        limit = min(int(request.GET.get('limit', 50)), 200)
        offset = int(request.GET.get('offset', 0))

        with get_db_session() as db:
            # Check if guild exists
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Build query
            query = db.query(LFGGroupAuditLog).filter_by(guild_id=int(guild_id))

            if group_id_filter:
                query = query.filter_by(group_id=int(group_id_filter))

            # Order by most recent first
            query = query.order_by(LFGGroupAuditLog.created_at.desc())

            # Get total count
            total_count = query.count()

            # Apply pagination
            logs = query.offset(offset).limit(limit).all()

            # Format logs for response
            logs_data = []
            for log in logs:
                from datetime import datetime
                logs_data.append({
                    'id': log.id,
                    'group_id': log.group_id,
                    'action': log.action,
                    'actor_id': log.actor_id,
                    'actor_name': log.actor_name,
                    'field_changed': log.field_changed,
                    'old_value': log.old_value,
                    'new_value': log.new_value,
                    'group_name': log.group_name,
                    'game_name': log.game_name,
                    'created_at': log.created_at,
                    'created_at_formatted': datetime.fromtimestamp(log.created_at).strftime('%Y-%m-%d %H:%M:%S') if log.created_at else None
                })

            return JsonResponse({
                'success': True,
                'logs': logs_data,
                'total_count': total_count,
                'limit': limit,
                'offset': offset
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error fetching audit logs: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["PATCH"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='PATCH', block=True)
def api_lfg_browser_update_class(request, guild_id, group_id):
    """PATCH /api/guild/<id>/lfg/browser/<group_id>/update-class/ - Update member's class/role selections."""
    try:
        from .db import get_db_session
        from .models import Guild, LFGGroup, LFGMember

        # Get Discord user from session
        discord_user = request.session.get('discord_user', {})
        user_id = discord_user.get('id')

        if not user_id:
            return JsonResponse({'error': 'Not authenticated'}, status=401)

        data = json.loads(request.body)
        options = data.get('options', {})

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get group
            group = db.query(LFGGroup).filter_by(id=int(group_id), guild_id=int(guild_id)).first()
            if not group:
                return JsonResponse({'error': 'Group not found'}, status=404)

            # Find member
            member = db.query(LFGMember).filter_by(
                group_id=group.id,
                user_id=int(user_id)
            ).filter(LFGMember.left_at == None).first()

            if not member:
                return JsonResponse({'error': 'You are not in this group'}, status=400)

            # Extract rank and selections
            rank_value = options.get('rank')
            selections = {k: v for k, v in options.items() if k != 'rank'}
            selections_json = json_lib.dumps(selections) if selections else None

            # Update member's selections
            member.rank_value = rank_value
            member.selections = selections_json

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Class updated successfully'
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error updating member class: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


@require_http_methods(["GET", "POST"])
@discord_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_lfg_browser_notifications(request, guild_id):
    """
    GET/POST /api/guild/<id>/lfg/browser-notifications/ - Get or save LFG Browser notification settings.

    Pro/Premium/VIP only feature.

    GET Returns:
        200: Current notification settings
        403: Not Pro/Premium/VIP or not admin

    POST Payload:
        browser_notify_channel_id: Channel ID for announcements (optional)
        notify_on_group_create: Announce new groups (bool)
        notify_on_group_update: Announce group updates (bool)
        notify_on_group_delete: Announce group deletions (bool)
        notify_on_member_join: Announce member joins (bool)
        notify_on_member_leave: Announce member leaves (bool)
        dm_members_on_update: DM members on group update (bool)
        dm_members_on_delete: DM members on group deletion (bool)
        webhook_url: Optional webhook URL for custom integrations

    POST Returns:
        200: Settings saved successfully
        403: Not Pro/Premium/VIP or not admin
    """
    try:
        from .db import get_db_session
        from .models import Guild, LFGConfig

        # Verify admin permission
        admin_guilds = request.session.get('discord_admin_guilds', [])
        is_admin = str(guild_id) in [str(g['id']) for g in admin_guilds]

        if not is_admin:
            return JsonResponse({'error': 'Admin permission required'}, status=403)

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # GET: Return current settings
            if request.method == 'GET':
                lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
                if not lfg_config:
                    # Return default settings
                    return JsonResponse({
                        'browser_notify_channel_id': None,
                        'notify_on_group_create': True,
                        'notify_on_group_update': False,
                        'notify_on_group_delete': False,
                        'notify_on_member_join': False,
                        'notify_on_member_leave': False,
                        'webhook_url': None
                    })

                return JsonResponse({
                    'browser_notify_channel_id': str(lfg_config.browser_notify_channel_id) if lfg_config.browser_notify_channel_id else None,
                    'notify_on_group_create': lfg_config.notify_on_group_create,
                    'notify_on_group_update': lfg_config.notify_on_group_update,
                    'notify_on_group_delete': lfg_config.notify_on_group_delete,
                    'notify_on_member_join': lfg_config.notify_on_member_join,
                    'notify_on_member_leave': lfg_config.notify_on_member_leave,
                    'webhook_url': lfg_config.webhook_url or ''
                })

        # POST: Save settings
        data = json.loads(request.body)

        with get_db_session() as db:
            # Get guild
            guild = db.query(Guild).filter_by(guild_id=int(guild_id)).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            # Get or create LFG config
            lfg_config = db.query(LFGConfig).filter_by(guild_id=int(guild_id)).first()
            if not lfg_config:
                lfg_config = LFGConfig(guild_id=int(guild_id))
                db.add(lfg_config)

            # Update notification settings
            lfg_config.browser_notify_channel_id = int(data.get('browser_notify_channel_id')) if data.get('browser_notify_channel_id') else None
            lfg_config.notify_on_group_create = data.get('notify_on_group_create', True)
            lfg_config.notify_on_group_update = data.get('notify_on_group_update', False)
            lfg_config.notify_on_group_delete = data.get('notify_on_group_delete', False)
            lfg_config.notify_on_member_join = data.get('notify_on_member_join', False)
            lfg_config.notify_on_member_leave = data.get('notify_on_member_leave', False)
            lfg_config.dm_members_on_update = data.get('dm_members_on_update', True)
            lfg_config.dm_members_on_delete = data.get('dm_members_on_delete', True)
            lfg_config.webhook_url = data.get('webhook_url')

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Browser notification settings saved successfully'
            })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error saving LFG Browser notification settings: {e}", exc_info=True)
        return JsonResponse({'error': 'An internal error occurred. Please try again later.'}, status=500)


# Bot Installation Callback
def bot_install_callback(request):
    """
    Handle bot installation callback from Discord.
    After user authorizes the bot, Discord redirects here.
    """
    # Get parameters from Discord
    guild_id = request.GET.get('guild_id')
    permissions = request.GET.get('permissions')
    error = request.GET.get('error')
    error_description = request.GET.get('error_description')

    # Handle authorization errors
    if error:
        messages.error(request, f'Bot authorization failed: {error_description or error}')
        return redirect('home')

    # Check if guild_id was provided
    if not guild_id:
        messages.warning(request, 'No server selected. Please try adding the bot again.')
        return redirect('home')

    try:
        # SECURITY FIX: Don't leak guild existence or premium status
        # Display generic success message without querying database
        context = {
            'guild_id': guild_id,
            'guild_name': 'Your Server',  # Generic - don't leak actual name
            'bot_name': 'Wardenbot',
            'permissions': permissions,
            'is_premium': False,  # Generic - don't leak premium status
        }

        messages.success(request, f'Wardenbot has been successfully added to your server!')
        return render(request, 'questlog/bot_install_success.html', context)

    except Exception as e:
        logger.error(f"Error in bot install callback: {e}", exc_info=True)
        messages.error(request, 'An error occurred. The bot should still be installed.')
        return redirect('home')


# ============================================================================
# FEATURED CREATORS - Hall of Fame
# ============================================================================

def guild_featured_creators(request, guild_id):
    """
    Featured Creators hall of fame page - PUBLIC (no login required).

    Anyone can view featured creators to discover content creators.
    Admin features (Discovery settings) still require authentication.
    """
    from .module_utils import has_module_access, has_any_module_access
    import logging
    logger = logging.getLogger(__name__)

    # Check if user is authenticated (optional for public view)
    all_guilds = request.session.get('discord_all_guilds', [])
    admin_guilds = request.session.get('discord_admin_guilds', [])
    is_authenticated = bool(request.session.get('discord_user'))

    # Try to get guild from session if authenticated
    guild_from_session = next((g for g in all_guilds if str(g.get('id')) == str(guild_id)), None)

    # Check if user is admin (only if authenticated)
    is_admin = is_authenticated and any(str(g.get('id')) == str(guild_id) for g in admin_guilds)
    show_admin_view = is_admin  # For template clarity

    try:
        from .db import get_db_session
        from .models import Guild as DBGuild, DiscoveryConfig, CreatorProfile, GuildMember

        with get_db_session() as db:
            guild_db = db.query(DBGuild).filter_by(guild_id=int(guild_id)).first()

            if not guild_db:
                logger.warning(f"Featured Creators: Guild {guild_id} not found in database")
                messages.error(request, "Guild not found.")
                return redirect('home')

            # Build guild object for template (from DB or session)
            if guild_from_session:
                # Use session data if available (has icon, name, etc.)
                guild = guild_from_session
            else:
                # Build minimal guild object from database for public view
                guild = {
                    'id': str(guild_db.guild_id),
                    'name': guild_db.guild_name or f"Guild {guild_db.guild_id}",
                    'icon': None,  # We don't store icon in DB
                }

            # Get Discovery config to check if enabled
            discovery_config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()
            discovery_enabled = discovery_config.enabled if discovery_config else False
            selfpromo_channel_id = discovery_config.selfpromo_channel_id if discovery_config else None

            # Get CreatorProfile creators for this guild
            import json
            profile_creators = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id)
            ).order_by(CreatorProfile.updated_at.desc()).all()

            # Process CreatorProfile creators
            creators_data = []
            for profile in profile_creators:
                # Get Discord user info from GuildMember
                member_data = db.query(GuildMember).filter_by(
                    guild_id=int(guild_id),
                    user_id=profile.discord_id
                ).first()

                # Get user info - use member data if available, otherwise use defaults
                if member_data:
                    username = member_data.username or f"User_{profile.discord_id}"
                    avatar_url = f"https://cdn.discordapp.com/avatars/{profile.discord_id}/{member_data.avatar_hash}.png" if member_data.avatar_hash else f"https://cdn.discordapp.com/embed/avatars/{profile.discord_id % 5}.png"
                    member_level = member_data.level if member_data else 0
                    member_flair = member_data.flair if member_data else None
                else:
                    username = f"User_{profile.discord_id}"
                    avatar_url = f"https://cdn.discordapp.com/embed/avatars/{profile.discord_id % 5}.png"
                    member_level = 0
                    member_flair = None

                # SECURITY: Sanitize bio HTML to prevent XSS
                import bleach
                import html
                clean_bio = ""
                if profile.bio:
                    # Escape HTML first
                    bio_text = html.escape(profile.bio)
                    # Convert line breaks to <br>
                    bio_text = bio_text.replace('\n', '<br>')
                    # Sanitize
                    ALLOWED_TAGS = ['br']
                    clean_bio = bleach.clean(bio_text, tags=ALLOWED_TAGS, strip=True)

                # Build social media links from handles
                # IMPORTANT: Twitch and YouTube MUST be first
                extracted_links = []
                shown_platforms = set()

                # Twitch FIRST (connected via OAuth)
                if profile.twitch_handle or profile.twitch_user_id:
                    handle = profile.twitch_handle or profile.twitch_user_id
                    extracted_links.append({
                        'url': f"https://twitch.tv/{handle}",
                        'platform': 'Twitch',
                        'icon': 'fab fa-twitch',
                        'color': 'purple'
                    })
                    shown_platforms.add('twitch')

                # YouTube SECOND (connected via OAuth)
                if profile.youtube_handle:
                    # YouTube handles can be either @username or channel ID
                    if profile.youtube_handle.startswith('@'):
                        extracted_links.append({
                            'url': f"https://youtube.com/{profile.youtube_handle}",
                            'platform': 'YouTube',
                            'icon': 'fab fa-youtube',
                            'color': 'red'
                        })
                    else:
                        # Assume it's a channel ID or custom URL
                        extracted_links.append({
                            'url': f"https://youtube.com/@{profile.youtube_handle}",
                            'platform': 'YouTube',
                            'icon': 'fab fa-youtube',
                            'color': 'red'
                        })
                    shown_platforms.add('youtube')
                elif profile.youtube_channel_id:
                    # Fallback to channel ID if no handle
                    extracted_links.append({
                        'url': f"https://youtube.com/channel/{profile.youtube_channel_id}",
                        'platform': 'YouTube',
                        'icon': 'fab fa-youtube',
                        'color': 'red'
                    })
                    shown_platforms.add('youtube')

                # Other social platforms
                if profile.twitter_handle:
                    extracted_links.append({
                        'url': f"https://twitter.com/{profile.twitter_handle}",
                        'platform': 'Twitter',
                        'icon': 'fab fa-twitter',
                        'color': 'sky'
                    })
                    shown_platforms.add('twitter')

                if profile.tiktok_handle:
                    extracted_links.append({
                        'url': f"https://tiktok.com/@{profile.tiktok_handle}",
                        'platform': 'TikTok',
                        'icon': 'fab fa-tiktok',
                        'color': 'gray'
                    })
                    shown_platforms.add('tiktok')

                if profile.instagram_handle:
                    extracted_links.append({
                        'url': f"https://instagram.com/{profile.instagram_handle}",
                        'platform': 'Instagram',
                        'icon': 'fab fa-instagram',
                        'color': 'pink'
                    })
                    shown_platforms.add('instagram')

                if profile.bluesky_handle:
                    extracted_links.append({
                        'url': f"https://bsky.app/profile/{profile.bluesky_handle}",
                        'platform': 'Bluesky',
                        'icon': 'fas fa-cloud',
                        'color': 'teal'
                    })
                    shown_platforms.add('bsky')

                # Parse content categories if present
                content_categories = []
                if profile.content_categories:
                    try:
                        # Could be JSON array or comma-separated
                        if profile.content_categories.startswith('['):
                            # Try JSON first
                            try:
                                content_categories = json.loads(profile.content_categories)
                            except json.JSONDecodeError:
                                # Might be Python string representation - use ast.literal_eval
                                import ast
                                try:
                                    content_categories = ast.literal_eval(profile.content_categories)
                                except:
                                    # Last resort - parse as comma-separated
                                    content_categories = [cat.strip().strip("'\"") for cat in profile.content_categories.strip('[]').split(',') if cat.strip()]
                        else:
                            content_categories = [cat.strip() for cat in profile.content_categories.split(',') if cat.strip()]
                    except:
                        pass

                creators_data.append({
                    'user_id': profile.discord_id,
                    'username': username,
                    'display_name': profile.display_name,
                    'avatar_url': avatar_url,
                    'bio': profile.bio,
                    'clean_bio': clean_bio,
                    'extracted_bio_links': extracted_links,
                    'times_featured': profile.times_featured,
                    'first_featured_at': profile.created_at,
                    'last_featured_at': profile.updated_at,
                    'level': member_level,
                    'flair': member_flair,
                    'show_twitch': False,  # Profile system doesn't have direct URLs
                    'show_youtube': False,
                    'show_twitter': False,
                    'show_tiktok': False,
                    'show_instagram': False,
                    'show_bsky': False,
                    'twitch_url': None,
                    'youtube_url': None,
                    'twitter_url': None,
                    'tiktok_url': None,
                    'instagram_url': None,
                    'bsky_url': None,
                    'discord_connections': {},
                    'forum_thread_id': None,  # Profile creators don't have forum threads
                    'source': 'profile',  # Source type for template
                    'content_categories': content_categories,
                    'stream_schedule': profile.stream_schedule,
                    'is_current_cotw': profile.is_current_cotw,
                    'is_current_cotm': profile.is_current_cotm,
                })

        # Check if guild has discovery module access
        has_discovery_module = has_module_access(guild_id, 'discovery')
        has_any_module = has_any_module_access(guild_id)

        # Check if current user already has a creator profile
        user_has_profile = False
        if is_authenticated:
            discord_user = request.session.get('discord_user')
            user_id = int(discord_user['id'])
            existing_profile = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id),
                discord_id=user_id
            ).first()
            user_has_profile = existing_profile is not None

        # Get featured creators (COTW and COTM only - no random)
        cotw_creator = None
        cotm_creator = None

        # Get Creator of the Week and Creator of the Month (admin-selected)
        for creator in creators_data:
            if creator.get('is_current_cotw'):
                cotw_creator = creator
            if creator.get('is_current_cotm'):
                cotm_creator = creator

        return render(request, 'questlog/featured_creators.html', {
            'guild': guild,
            'guild_record': guild_db,
            'creators': creators_data,
            'total_creators': len(creators_data),
            'is_admin': is_admin,
            'show_admin_view': show_admin_view,
            'discovery_enabled': discovery_enabled,
            'selfpromo_channel_id': selfpromo_channel_id,
            'has_discovery_module': has_discovery_module,
            'has_any_module': has_any_module,
            'admin_guilds': admin_guilds,
            'member_guilds': get_member_guilds(request),
            'active_page': 'featured_creators',
            'user_has_profile': user_has_profile,
            'is_authenticated': is_authenticated,
            'cotw_creator': cotw_creator,
            'cotm_creator': cotm_creator,
        })

    except Exception as e:
        logger.error(f"Error loading featured creators: {e}", exc_info=True)
        messages.error(request, f'Error loading featured creators: {e}')
        return redirect('home')

def guild_cotw(request, guild_id):
    """Guild-specific Creator of the Week page (placeholder)."""
    # Redirect to global COTW page for now
    return redirect('creator_of_the_week')

def guild_cotm(request, guild_id):
    """Guild-specific Creator of the Month page (placeholder)."""
    # Redirect to global COTM page for now
    return redirect('creator_of_the_month')


def robots_txt(request):
    """Serve robots.txt with proper headers for social media and search engine crawlers."""
    from django.http import HttpResponse

    robots_content = """# Casual Heroes - robots.txt

# Allow social media crawlers (Discord, LinkedIn, Twitter, Facebook, etc.)
User-agent: Twitterbot
Allow: /

User-agent: facebookexternalhit
Allow: /

User-agent: LinkedInBot
Allow: /

User-agent: Discordbot
Allow: /

User-agent: Slackbot
Allow: /

User-agent: TelegramBot
Allow: /

# Allow search engines
User-agent: Googlebot
Allow: /

User-agent: Bingbot
Allow: /

# Block AI scrapers
User-agent: GPTBot
Disallow: /

User-agent: ChatGPT-User
Disallow: /

User-agent: CCBot
Disallow: /

User-agent: Google-Extended
Disallow: /

User-agent: anthropic-ai
Disallow: /

User-agent: Claude-Web
Disallow: /

User-agent: cohere-ai
Disallow: /

# Default: Allow everything else except API endpoints
User-agent: *
Allow: /
Disallow: /api/admin/
Disallow: /admin/

Sitemap: https://dashboard.casual-heroes.com/sitemap.xml
"""

    return HttpResponse(robots_content, content_type='text/plain')


# ============================================================================
# CREATOR DISCOVERY SYSTEM - Phase 1
# ============================================================================

@discord_required
def creator_profile_register(request, guild_id):
    """
    Creator self-registration form.

    FREE TIER: Anyone can create a creator profile.
    Users fill out bio, social links, and streaming info.
    """
    import logging
    import json
    import bleach
    from .db import get_db_session
    from .models import Guild as DBGuild, GuildMember, CreatorProfile
    from .module_utils import has_module_access

    logger = logging.getLogger(__name__)

    # Security: Get authenticated user
    discord_user = request.session.get('discord_user')
    if not discord_user:
        messages.error(request, "You must be logged in to register as a creator.")
        return redirect('login')

    user_id = int(discord_user['id'])

    # Get guild data
    all_guilds = request.session.get('discord_all_guilds', [])
    admin_guilds = request.session.get('discord_admin_guilds', [])
    guild = get_guild_with_permissions(guild_id, admin_guilds, all_guilds)

    if not guild:
        messages.error(request, "Guild not found.")
        return redirect('home')

    is_admin = any(str(g.get('id')) == str(guild_id) for g in admin_guilds)

    try:
        with get_db_session() as db:
            guild_db = db.query(DBGuild).filter_by(guild_id=int(guild_id)).first()
            if not guild_db:
                messages.error(request, "Guild not found in database.")
                return redirect('home')

            # Check if user is a member of this guild
            member = db.query(GuildMember).filter_by(
                guild_id=int(guild_id),
                user_id=user_id
            ).first()

            if not member:
                logger.warning(f"User {user_id} is not a member of guild {guild_id}")
                messages.error(request, "You must be a member of this server to register as a creator.")
                return redirect('guild_dashboard', guild_id=guild_id)

            # Check if profile already exists
            existing_profile = db.query(CreatorProfile).filter_by(
                discord_id=user_id,
                guild_id=int(guild_id)
            ).first()

            if request.method == 'POST':
                # Security: CSRF protection (Django middleware handles this)
                # Input validation and sanitization
                logger.info(f"CreatorProfile POST request: user_id={user_id}, guild_id={guild_id}")

                display_name = request.POST.get('display_name', '').strip()
                bio = request.POST.get('bio', '').strip()
                content_categories = request.POST.get('content_categories', '').strip()

                # Social media handles (sanitize input)
                twitter_handle = request.POST.get('twitter_handle', '').strip()
                tiktok_handle = request.POST.get('tiktok_handle', '').strip()
                instagram_handle = request.POST.get('instagram_handle', '').strip()
                bluesky_handle = request.POST.get('bluesky_handle', '').strip()
                twitch_handle = request.POST.get('twitch_handle', '').strip()
                youtube_handle = request.POST.get('youtube_handle', '').strip()

                stream_schedule = request.POST.get('stream_schedule', '').strip()

                # Discovery Network opt-in
                share_to_network = request.POST.get('share_to_network') in ['true', 'on']

                # Validation
                if not display_name:
                    messages.error(request, "Display name is required.")
                    return redirect('creator_profile_register', guild_id=guild_id)

                if len(display_name) > 100:
                    messages.error(request, "Display name must be 100 characters or less.")
                    return redirect('creator_profile_register', guild_id=guild_id)

                if bio and len(bio) > 1000:
                    messages.error(request, "Bio must be 1000 characters or less.")
                    return redirect('creator_profile_register', guild_id=guild_id)

                # Sanitize HTML in bio (prevent XSS)
                ALLOWED_TAGS = []  # No HTML allowed in bio input
                ALLOWED_ATTRS = {}
                bio = bleach.clean(bio, tags=ALLOWED_TAGS, attributes=ALLOWED_ATTRS, strip=True)

                # Sanitize handles
                import re
                def sanitize_handle(handle):
                    """Sanitize regular handles (alphanumeric, underscore, hyphen only)."""
                    if not handle:
                        return None
                    # Remove @ symbol if present
                    handle = handle.lstrip('@')
                    # Allow only alphanumeric, underscore, hyphen
                    return re.sub(r'[^a-zA-Z0-9_-]', '', handle)[:100]

                def sanitize_bluesky_handle(handle):
                    """Sanitize Bluesky handles - allows dots for domain format."""
                    if not handle:
                        return None
                    # Remove @ symbol if present
                    handle = handle.lstrip('@')
                    # Allow alphanumeric, underscore, hyphen, and dots (for domain format)
                    return re.sub(r'[^a-zA-Z0-9_.-]', '', handle)[:100]

                twitter_handle = sanitize_handle(twitter_handle)
                tiktok_handle = sanitize_handle(tiktok_handle)
                instagram_handle = sanitize_handle(instagram_handle)
                bluesky_handle = sanitize_bluesky_handle(bluesky_handle)
                twitch_handle = sanitize_handle(twitch_handle)
                youtube_handle = sanitize_handle(youtube_handle)

                # Parse content categories (JSON array from custom_select)
                categories_list = []
                if content_categories:
                    # DEBUG: Log raw POST data
                    logger.info(f"[CATEGORY DEBUG] Raw content_categories from POST: {repr(content_categories)}")

                    try:
                        # Try parsing as JSON first (from custom_select multiselect)
                        if content_categories.startswith('['):
                            categories_list = json.loads(content_categories)
                            logger.info(f"[CATEGORY DEBUG] After JSON parse: {repr(categories_list)}")
                        else:
                            # Fallback to comma-separated for backwards compatibility
                            categories_list = [cat.strip() for cat in content_categories.split(',') if cat.strip()]
                            logger.info(f"[CATEGORY DEBUG] After CSV parse: {repr(categories_list)}")
                    except json.JSONDecodeError as e:
                        # If JSON parsing fails, treat as comma-separated
                        logger.warning(f"[CATEGORY DEBUG] JSON parse failed: {e}, falling back to CSV")
                        categories_list = [cat.strip() for cat in content_categories.split(',') if cat.strip()]

                    # Filter out empty strings, bracket-only strings, and limit to 10 categories
                    categories_list = [
                        cat for cat in categories_list
                        if cat and cat.strip() and cat.strip() not in ['[', ']', '[]', '[[', ']]']
                    ][:10]
                    logger.info(f"[CATEGORY DEBUG] After filtering: {repr(categories_list)}")

                categories_json = json.dumps(categories_list) if categories_list else None
                logger.info(f"[CATEGORY DEBUG] Final JSON to save: {repr(categories_json)}")

                # Create or update profile
                if existing_profile:
                    # Update existing
                    existing_profile.display_name = display_name
                    existing_profile.bio = bio
                    existing_profile.content_categories = categories_json
                    existing_profile.twitter_handle = twitter_handle
                    existing_profile.tiktok_handle = tiktok_handle
                    existing_profile.instagram_handle = instagram_handle
                    existing_profile.bluesky_handle = bluesky_handle
                    existing_profile.twitch_handle = twitch_handle

                    # Only update YouTube handle if not OAuth-connected
                    if not existing_profile.youtube_channel_id:
                        existing_profile.youtube_handle = youtube_handle

                    existing_profile.stream_schedule = stream_schedule
                    existing_profile.share_to_network = share_to_network
                    existing_profile.updated_at = int(time.time())

                    db.commit()
                    messages.success(request, "Your creator profile has been updated!")
                    logger.info(f"Creator profile updated: user_id={user_id}, guild_id={guild_id}, share_to_network={share_to_network}")
                else:
                    # Create new
                    try:
                        new_profile = CreatorProfile(
                            discord_id=user_id,
                            guild_id=int(guild_id),
                            display_name=display_name,
                            bio=bio,
                            content_categories=categories_json,
                            twitter_handle=twitter_handle,
                            tiktok_handle=tiktok_handle,
                            instagram_handle=instagram_handle,
                            bluesky_handle=bluesky_handle,
                            twitch_handle=twitch_handle,
                            youtube_handle=youtube_handle,
                            stream_schedule=stream_schedule,
                            times_featured=0,
                            is_current_cotw=False,
                            is_current_cotm=False,
                            share_to_network=share_to_network,
                            created_at=int(time.time()),
                            updated_at=int(time.time())
                        )
                        db.add(new_profile)
                        db.commit()
                        messages.success(request, "Your creator profile has been created!")
                        logger.info(f"Creator profile created: user_id={user_id}, guild_id={guild_id}, profile_id={new_profile.id}, share_to_network={share_to_network}")
                    except Exception as save_error:
                        logger.error(f"Failed to save creator profile: {save_error}", exc_info=True)
                        db.rollback()
                        messages.error(request, f"Error saving profile: {str(save_error)}")
                        return redirect('creator_profile_register', guild_id=guild_id)

                return redirect('guild_featured_creators', guild_id=guild_id)

            # GET request - show form
            profile_data = None
            if existing_profile:
                # Refresh YouTube/Twitch stats on page load
                stats_updated = False

                # Refresh YouTube stats if connected
                if existing_profile.youtube_channel_id and existing_profile.youtube_access_token:
                    try:
                        from app.services.youtube_service import YouTubeService, YouTubeAPIError
                        from app.utils.encryption import decrypt_token, encrypt_token

                        yt_service = YouTubeService()
                        access_token = decrypt_token(existing_profile.youtube_access_token)

                        # Try to get channel info
                        try:
                            channel_info = yt_service.get_channel_info(access_token)
                            existing_profile.youtube_subscriber_count = channel_info.get('subscriber_count', 0)
                            existing_profile.youtube_last_synced = int(time.time())
                            stats_updated = True
                            logger.debug(f"YouTube stats refreshed for creator {existing_profile.id}: {channel_info.get('subscriber_count')} subs")
                        except YouTubeAPIError as e:
                            # Token might be expired, try to refresh
                            if existing_profile.youtube_refresh_token:
                                try:
                                    refresh_token = decrypt_token(existing_profile.youtube_refresh_token)
                                    new_tokens = yt_service.refresh_access_token(refresh_token)
                                    existing_profile.youtube_access_token = encrypt_token(new_tokens['access_token'])
                                    existing_profile.youtube_token_expires = int(time.time()) + new_tokens.get('expires_in', 3600)

                                    # Retry with new token
                                    channel_info = yt_service.get_channel_info(new_tokens['access_token'])
                                    existing_profile.youtube_subscriber_count = channel_info.get('subscriber_count', 0)
                                    existing_profile.youtube_last_synced = int(time.time())
                                    stats_updated = True
                                    logger.debug(f"YouTube stats refreshed after token refresh for creator {existing_profile.id}")
                                except Exception as refresh_err:
                                    error_str = str(refresh_err).lower()
                                    # Check if token was revoked/expired - clear connection so user can reconnect
                                    # Catch: invalid_grant, revoked, expired, 400 Bad Request, 401 Unauthorized
                                    if any(x in error_str for x in ['invalid_grant', 'revoked', 'expired', '400', '401', 'bad request', 'unauthorized']):
                                        logger.warning(f"YouTube token revoked for creator {existing_profile.id}, clearing connection")
                                        existing_profile.youtube_access_token = None
                                        existing_profile.youtube_refresh_token = None
                                        existing_profile.youtube_token_expires = None
                                        existing_profile.youtube_channel_id = None  # Clear so UI shows "Not connected"
                                        stats_updated = True  # Need to commit this change
                                    else:
                                        logger.warning(f"YouTube token refresh failed for creator {existing_profile.id}: {refresh_err}")
                    except Exception as yt_err:
                        logger.warning(f"Failed to refresh YouTube stats for creator {existing_profile.id}: {yt_err}")

                # Refresh Twitch stats if connected
                if existing_profile.twitch_user_id and existing_profile.twitch_access_token:
                    try:
                        from app.services.twitch_service import TwitchService, TwitchAPIError
                        from app.utils.encryption import decrypt_token, encrypt_token

                        twitch_service = TwitchService()
                        access_token = decrypt_token(existing_profile.twitch_access_token)

                        # Try to get channel info
                        try:
                            channel_info = twitch_service.get_channel_info(access_token, existing_profile.twitch_user_id)
                            existing_profile.twitch_follower_count = channel_info.get('follower_count', 0)
                            existing_profile.twitch_last_synced = int(time.time())
                            stats_updated = True
                            logger.debug(f"Twitch stats refreshed for creator {existing_profile.id}: {channel_info.get('follower_count')} followers")
                        except TwitchAPIError as e:
                            # Token might be expired, try to refresh
                            if existing_profile.twitch_refresh_token:
                                try:
                                    refresh_token = decrypt_token(existing_profile.twitch_refresh_token)
                                    new_tokens = twitch_service.refresh_access_token(refresh_token)
                                    existing_profile.twitch_access_token = encrypt_token(new_tokens['access_token'])
                                    if new_tokens.get('refresh_token'):
                                        existing_profile.twitch_refresh_token = encrypt_token(new_tokens['refresh_token'])
                                    existing_profile.twitch_token_expires = int(time.time()) + new_tokens.get('expires_in', 3600)

                                    # Retry with new token
                                    channel_info = twitch_service.get_channel_info(new_tokens['access_token'], existing_profile.twitch_user_id)
                                    existing_profile.twitch_follower_count = channel_info.get('follower_count', 0)
                                    existing_profile.twitch_last_synced = int(time.time())
                                    stats_updated = True
                                    logger.debug(f"Twitch stats refreshed after token refresh for creator {existing_profile.id}")
                                except Exception as refresh_err:
                                    error_str = str(refresh_err).lower()
                                    # Check if token was revoked/expired - clear connection so user can reconnect
                                    if 'invalid' in error_str or 'revoked' in error_str or 'expired' in error_str or 'unauthorized' in error_str:
                                        logger.warning(f"Twitch token revoked for creator {existing_profile.id}, clearing connection")
                                        existing_profile.twitch_access_token = None
                                        existing_profile.twitch_refresh_token = None
                                        existing_profile.twitch_token_expires = None
                                        existing_profile.twitch_user_id = None  # Clear so UI shows "Not connected"
                                        stats_updated = True  # Need to commit this change
                                    else:
                                        logger.warning(f"Twitch token refresh failed for creator {existing_profile.id}: {refresh_err}")
                    except Exception as tw_err:
                        logger.warning(f"Failed to refresh Twitch stats for creator {existing_profile.id}: {tw_err}")

                # Commit if any stats were updated
                if stats_updated:
                    try:
                        db.commit()
                    except Exception as commit_err:
                        logger.warning(f"Failed to commit stats update: {commit_err}")
                        db.rollback()

                # Keep categories as stored JSON string - template will parse it
                profile_data = {
                    'display_name': existing_profile.display_name,
                    'bio': existing_profile.bio,
                    'content_categories': existing_profile.content_categories or '[]',
                    'twitter_handle': existing_profile.twitter_handle or '',
                    'tiktok_handle': existing_profile.tiktok_handle or '',
                    'instagram_handle': existing_profile.instagram_handle or '',
                    'bluesky_handle': existing_profile.bluesky_handle or '',
                    'twitch_handle': existing_profile.twitch_handle or '',
                    'youtube_handle': existing_profile.youtube_handle or '',
                    'stream_schedule': existing_profile.stream_schedule or '',
                    'share_to_network': existing_profile.share_to_network,
                    'youtube_connected': bool(existing_profile.youtube_channel_id),
                    'twitch_connected': bool(existing_profile.twitch_user_id),
                }

            # Check if guild has Discovery module (for showing OAuth link options later)
            has_discovery_module = has_module_access(guild_id, 'discovery')

            # Check if guild is approved in Discovery Network
            from .models import DiscoveryNetworkApplication
            guild_network_status = None
            network_application = db.query(DiscoveryNetworkApplication).filter_by(
                guild_id=int(guild_id)
            ).order_by(DiscoveryNetworkApplication.applied_at.desc()).first()

            if network_application:
                guild_network_status = network_application.status

            return render(request, 'questlog/creator_profile_register.html', {
                'guild': guild,
                'guild_record': guild_db,
                'is_admin': is_admin,
                'profile': profile_data,
                'existing_profile': existing_profile,
                'has_discovery_module': has_discovery_module,
                'guild_network_status': guild_network_status,
                'member_guilds': get_member_guilds(request),
                'admin_guilds': admin_guilds,
                'active_page': 'featured_creators',
            })

    except Exception as e:
        logger.error(f"Error in creator profile registration: {e}", exc_info=True)
        messages.error(request, f"An error occurred: {e}")
        return redirect('guild_dashboard', guild_id=guild_id)


@discord_required
def creator_profile_delete(request, guild_id):
    """
    Delete creator profile.
    Security: Users can only delete their own profiles.
    """
    import logging
    from .db import get_db_session
    from .models import CreatorProfile

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    discord_user = request.session.get('discord_user')
    if not discord_user:
        return JsonResponse({'error': 'Not authenticated'}, status=401)

    user_id = int(discord_user['id'])

    try:
        with get_db_session() as db:
            profile = db.query(CreatorProfile).filter_by(
                discord_id=user_id,
                guild_id=int(guild_id)
            ).first()

            if not profile:
                return JsonResponse({'error': 'Profile not found'}, status=404)

            # Security: Verify ownership
            if profile.discord_id != user_id:
                logger.warning(f"Unauthorized delete attempt: user {user_id} tried to delete profile of user {profile.discord_id}")
                return JsonResponse({'error': 'Unauthorized'}, status=403)

            db.delete(profile)
            db.commit()

            logger.info(f"Creator profile deleted: user_id={user_id}, guild_id={guild_id}")
            return JsonResponse({'success': True, 'message': 'Profile deleted successfully'})

    except Exception as e:
        logger.error(f"Error deleting creator profile: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to delete creator profile'}, status=500)


@discord_required
def set_creator_of_week(request, guild_id):
    """
    Set a creator as Creator of the Week.
    Security: Only admins can set COTW.
    """
    import logging
    import json
    from .db import get_db_session
    from .models import CreatorProfile, DiscoveryConfig

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    # Check if user is admin
    admin_guilds = request.session.get('discord_admin_guilds', [])
    is_admin = any(str(g.get('id')) == str(guild_id) for g in admin_guilds)

    if not is_admin:
        return JsonResponse({'error': 'Admin access required'}, status=403)

    try:
        data = json.loads(request.body)
        target_user_id = int(data.get('user_id'))

        with get_db_session() as db:
            # Get the target profile first
            new_cotw = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id),
                discord_id=target_user_id
            ).first()

            if not new_cotw:
                return JsonResponse({'error': 'Creator profile not found'}, status=404)

            # Get discovery config for channel info
            discovery_config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()

            # If this person is currently COTM, remove that status first and delete the COTM announcement
            if new_cotw.is_current_cotm:
                new_cotw.is_current_cotm = False
                new_cotw.cotm_last_featured = int(time.time())

                # Delete COTM Discord announcement
                if discovery_config and discovery_config.cotm_last_message_id and discovery_config.cotm_channel_id:
                    try:
                        import requests
                        import os
                        bot_api_url = "http://localhost:8001/api/delete-message"
                        api_token = os.getenv('DISCORD_BOT_API_TOKEN')
                        headers = {'Authorization': f'Bearer {api_token}'} if api_token else {}
                        payload = {
                            'guild_id': int(guild_id),
                            'channel_id': discovery_config.cotm_channel_id,
                            'message_id': discovery_config.cotm_last_message_id
                        }
                        response = requests.post(bot_api_url, json=payload, headers=headers, timeout=5)
                        if response.status_code == 200:
                            discovery_config.cotm_last_message_id = None
                            logger.info(f"Deleted COTM announcement when setting user as COTW")
                    except Exception as bot_error:
                        logger.error(f"Error deleting COTM announcement: {bot_error}", exc_info=True)

            # Remove current COTW
            current_cotw = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id),
                is_current_cotw=True
            ).all()

            for profile in current_cotw:
                profile.is_current_cotw = False
                profile.cotw_last_featured = int(time.time())

            # Set new COTW
            new_cotw.is_current_cotw = True
            new_cotw.times_featured += 1

            db.commit()

            # Trigger Discord bot to delete old announcement and post new one
            if discovery_config and discovery_config.cotw_enabled and discovery_config.cotw_channel_id:
                try:
                    import requests
                    import os
                    bot_api_url = "http://localhost:8001/api/announce-cotw"
                    api_token = os.getenv('DISCORD_BOT_API_TOKEN')
                    headers = {'Authorization': f'Bearer {api_token}'} if api_token else {}
                    payload = {
                        'guild_id': int(guild_id),
                        'user_id': target_user_id,
                        'channel_id': discovery_config.cotw_channel_id,
                        'old_message_id': discovery_config.cotw_last_message_id
                    }
                    response = requests.post(bot_api_url, json=payload, headers=headers, timeout=5)
                    if response.status_code == 200:
                        logger.info(f"Discord bot notified of COTW change for guild {guild_id}")
                    else:
                        logger.warning(f"Failed to notify Discord bot of COTW change: {response.text}")
                except Exception as bot_error:
                    logger.error(f"Error notifying Discord bot of COTW change: {bot_error}", exc_info=True)

            logger.info(f"COTW set: guild_id={guild_id}, user_id={target_user_id}")
            return JsonResponse({'success': True, 'message': 'Creator of the Week set successfully'})

    except Exception as e:
        logger.error(f"Error setting COTW: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to set Creator of the Week'}, status=500)


@discord_required
def clear_creator_of_week(request, guild_id):
    """
    Clear the current Creator of the Week.
    Security: Only admins can clear COTW.
    """
    import logging
    import json
    from .db import get_db_session
    from .models import CreatorProfile, DiscoveryConfig

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    # Check if user is admin
    admin_guilds = request.session.get('discord_admin_guilds', [])
    is_admin = any(str(g.get('id')) == str(guild_id) for g in admin_guilds)

    if not is_admin:
        return JsonResponse({'error': 'Admin access required'}, status=403)

    try:
        with get_db_session() as db:
            # Remove current COTW
            current_cotw = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id),
                is_current_cotw=True
            ).all()

            for profile in current_cotw:
                profile.is_current_cotw = False
                profile.cotw_last_featured = int(time.time())

            # Get discovery config to delete announcement
            discovery_config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()

            db.commit()

            # Delete Discord announcement if it exists
            if discovery_config and discovery_config.cotw_last_message_id and discovery_config.cotw_channel_id:
                try:
                    import requests
                    import os
                    bot_api_url = "http://localhost:8001/api/delete-message"
                    api_token = os.getenv('DISCORD_BOT_API_TOKEN')
                    headers = {'Authorization': f'Bearer {api_token}'} if api_token else {}
                    payload = {
                        'guild_id': int(guild_id),
                        'channel_id': discovery_config.cotw_channel_id,
                        'message_id': discovery_config.cotw_last_message_id
                    }
                    response = requests.post(bot_api_url, json=payload, headers=headers, timeout=5)
                    if response.status_code == 200:
                        logger.info(f"Deleted COTW announcement message for guild {guild_id}")
                        # Clear the message ID from database
                        discovery_config.cotw_last_message_id = None
                        db.commit()
                    else:
                        logger.warning(f"Failed to delete COTW announcement: {response.text}")
                except Exception as bot_error:
                    logger.error(f"Error deleting COTW announcement: {bot_error}", exc_info=True)

            logger.info(f"COTW cleared for guild {guild_id}")
            return JsonResponse({'success': True, 'message': 'Creator of the Week cleared'})

    except Exception as e:
        logger.error(f"Error clearing COTW: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to clear Creator of the Week'}, status=500)


@discord_required
def clear_creator_of_month(request, guild_id):
    """
    Clear the current Creator of the Month.
    Security: Only admins can clear COTM.
    """
    import logging
    import json
    from .db import get_db_session
    from .models import CreatorProfile, DiscoveryConfig

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    # Check if user is admin
    admin_guilds = request.session.get('discord_admin_guilds', [])
    is_admin = any(str(g.get('id')) == str(guild_id) for g in admin_guilds)

    if not is_admin:
        return JsonResponse({'error': 'Admin access required'}, status=403)

    try:
        with get_db_session() as db:
            # Remove current COTM
            current_cotm = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id),
                is_current_cotm=True
            ).all()

            for profile in current_cotm:
                profile.is_current_cotm = False
                profile.cotm_last_featured = int(time.time())

            # Get discovery config to delete announcement
            discovery_config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()

            db.commit()

            # Delete Discord announcement if it exists
            if discovery_config and discovery_config.cotm_last_message_id and discovery_config.cotm_channel_id:
                try:
                    import requests
                    import os
                    bot_api_url = "http://localhost:8001/api/delete-message"
                    api_token = os.getenv('DISCORD_BOT_API_TOKEN')
                    headers = {'Authorization': f'Bearer {api_token}'} if api_token else {}
                    payload = {
                        'guild_id': int(guild_id),
                        'channel_id': discovery_config.cotm_channel_id,
                        'message_id': discovery_config.cotm_last_message_id
                    }
                    response = requests.post(bot_api_url, json=payload, headers=headers, timeout=5)
                    if response.status_code == 200:
                        logger.info(f"Deleted COTM announcement message for guild {guild_id}")
                        # Clear the message ID from database
                        discovery_config.cotm_last_message_id = None
                        db.commit()
                    else:
                        logger.warning(f"Failed to delete COTM announcement: {response.text}")
                except Exception as bot_error:
                    logger.error(f"Error deleting COTM announcement: {bot_error}", exc_info=True)

            logger.info(f"COTM cleared for guild {guild_id}")
            return JsonResponse({'success': True, 'message': 'Creator of the Month cleared'})

    except Exception as e:
        logger.error(f"Error clearing COTM: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to clear Creator of the Month'}, status=500)


@discord_required
def set_creator_of_month(request, guild_id):
    """
    Set a creator as Creator of the Month.
    Security: Only admins can set COTM.
    """
    import logging
    import json
    from .db import get_db_session
    from .models import CreatorProfile, DiscoveryConfig

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    # Check if user is admin
    admin_guilds = request.session.get('discord_admin_guilds', [])
    is_admin = any(str(g.get('id')) == str(guild_id) for g in admin_guilds)

    if not is_admin:
        return JsonResponse({'error': 'Admin access required'}, status=403)

    try:
        data = json.loads(request.body)
        target_user_id = int(data.get('user_id'))

        with get_db_session() as db:
            # Get the target profile first
            new_cotm = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id),
                discord_id=target_user_id
            ).first()

            if not new_cotm:
                return JsonResponse({'error': 'Creator profile not found'}, status=404)

            # Get discovery config for channel info
            discovery_config = db.query(DiscoveryConfig).filter_by(guild_id=int(guild_id)).first()

            # If this person is currently COTW, remove that status first and delete the COTW announcement
            if new_cotm.is_current_cotw:
                new_cotm.is_current_cotw = False
                new_cotm.cotw_last_featured = int(time.time())

                # Delete COTW Discord announcement
                if discovery_config and discovery_config.cotw_last_message_id and discovery_config.cotw_channel_id:
                    try:
                        import requests
                        import os
                        bot_api_url = "http://localhost:8001/api/delete-message"
                        api_token = os.getenv('DISCORD_BOT_API_TOKEN')
                        headers = {'Authorization': f'Bearer {api_token}'} if api_token else {}
                        payload = {
                            'guild_id': int(guild_id),
                            'channel_id': discovery_config.cotw_channel_id,
                            'message_id': discovery_config.cotw_last_message_id
                        }
                        response = requests.post(bot_api_url, json=payload, headers=headers, timeout=5)
                        if response.status_code == 200:
                            discovery_config.cotw_last_message_id = None
                            logger.info(f"Deleted COTW announcement when setting user as COTM")
                    except Exception as bot_error:
                        logger.error(f"Error deleting COTW announcement: {bot_error}", exc_info=True)

            # Remove current COTM
            current_cotm = db.query(CreatorProfile).filter_by(
                guild_id=int(guild_id),
                is_current_cotm=True
            ).all()

            for profile in current_cotm:
                profile.is_current_cotm = False
                profile.cotm_last_featured = int(time.time())

            # Set new COTM
            new_cotm.is_current_cotm = True
            new_cotm.times_featured += 1

            db.commit()

            # Trigger Discord bot to delete old announcement and post new one
            if discovery_config and discovery_config.cotm_enabled and discovery_config.cotm_channel_id:
                try:
                    import requests
                    import os
                    bot_api_url = "http://localhost:8001/api/announce-cotm"
                    api_token = os.getenv('DISCORD_BOT_API_TOKEN')
                    headers = {'Authorization': f'Bearer {api_token}'} if api_token else {}
                    payload = {
                        'guild_id': int(guild_id),
                        'user_id': target_user_id,
                        'channel_id': discovery_config.cotm_channel_id,
                        'old_message_id': discovery_config.cotm_last_message_id
                    }
                    response = requests.post(bot_api_url, json=payload, headers=headers, timeout=5)
                    if response.status_code == 200:
                        logger.info(f"Discord bot notified of COTM change for guild {guild_id}")
                    else:
                        logger.warning(f"Failed to notify Discord bot of COTM change: {response.text}")
                except Exception as bot_error:
                    logger.error(f"Error notifying Discord bot of COTM change: {bot_error}", exc_info=True)

            logger.info(f"COTM set: guild_id={guild_id}, user_id={target_user_id}")
            return JsonResponse({'success': True, 'message': 'Creator of the Month set successfully'})

    except Exception as e:
        logger.error(f"Error setting COTM: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to set Creator of the Month'}, status=500)


# ==============================================================================
# Network Creator of the Week/Month Management (DISCOVERY_APPROVERS ONLY)
# ==============================================================================

@discovery_approvers_required
@discord_required
def set_network_creator_of_week(request):
    """
    Set a creator as Network Creator of the Week (cross-server).
    Security: Only DISCOVERY_APPROVERS can set Network COTW.
    """
    import logging
    import json
    from .db import get_db_session
    from .models import CreatorProfile

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        profile_id = int(data.get('profile_id'))

        with get_db_session() as db:
            # Get the target profile
            new_cotw = db.query(CreatorProfile).filter_by(id=profile_id).first()

            if not new_cotw:
                return JsonResponse({'error': 'Creator profile not found'}, status=404)

            # Verify creator is shared to network
            if not new_cotw.share_to_network:
                return JsonResponse({'error': 'Creator is not shared to Discovery Network'}, status=400)

            # If this person is currently Network COTM, remove that status first
            if new_cotw.is_current_network_cotm:
                new_cotw.is_current_network_cotm = False
                new_cotw.network_cotm_last_featured = int(time.time())

            # Remove current Network COTW
            current_cotw = db.query(CreatorProfile).filter_by(
                is_current_network_cotw=True
            ).all()

            for profile in current_cotw:
                profile.is_current_network_cotw = False
                profile.network_cotw_last_featured = int(time.time())

            # Set new Network COTW
            new_cotw.is_current_network_cotw = True
            new_cotw.network_cotw_last_featured = int(time.time())

            db.commit()

            # Trigger Discord bot to post Network COTW announcements to all opted-in servers
            try:
                import requests
                import os
                bot_api_url = "http://localhost:8001/api/announce-network-cotw"
                api_token = os.getenv('DISCORD_BOT_API_TOKEN')
                headers = {'Authorization': f'Bearer {api_token}'} if api_token else {}
                payload = {
                    'profile_id': profile_id
                }
                response = requests.post(bot_api_url, json=payload, headers=headers, timeout=10)
                if response.status_code == 200:
                    logger.info(f"Discord bot notified of Network COTW change: profile_id={profile_id}")
                else:
                    logger.warning(f"Failed to notify Discord bot of Network COTW change: {response.text}")
            except Exception as bot_error:
                logger.error(f"Error notifying Discord bot of Network COTW change: {bot_error}", exc_info=True)

            logger.info(f"Network COTW set by DISCOVERY_APPROVER: profile_id={profile_id}, user_id={new_cotw.discord_id}")
            return JsonResponse({'success': True, 'message': 'Network Creator of the Week set successfully'})

    except Exception as e:
        logger.error(f"Error setting Network COTW: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to set Network Creator of the Week'}, status=500)


@discovery_approvers_required
@discord_required
def set_network_creator_of_month(request):
    """
    Set a creator as Network Creator of the Month (cross-server).
    Security: Only DISCOVERY_APPROVERS can set Network COTM.
    """
    import logging
    import json
    from .db import get_db_session
    from .models import CreatorProfile

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        profile_id = int(data.get('profile_id'))

        with get_db_session() as db:
            # Get the target profile
            new_cotm = db.query(CreatorProfile).filter_by(id=profile_id).first()

            if not new_cotm:
                return JsonResponse({'error': 'Creator profile not found'}, status=404)

            # Verify creator is shared to network
            if not new_cotm.share_to_network:
                return JsonResponse({'error': 'Creator is not shared to Discovery Network'}, status=400)

            # If this person is currently Network COTW, remove that status first
            if new_cotm.is_current_network_cotw:
                new_cotm.is_current_network_cotw = False
                new_cotm.network_cotw_last_featured = int(time.time())

            # Remove current Network COTM
            current_cotm = db.query(CreatorProfile).filter_by(
                is_current_network_cotm=True
            ).all()

            for profile in current_cotm:
                profile.is_current_network_cotm = False
                profile.network_cotm_last_featured = int(time.time())

            # Set new Network COTM
            new_cotm.is_current_network_cotm = True
            new_cotm.network_cotm_last_featured = int(time.time())

            db.commit()

            # Trigger Discord bot to post Network COTM announcements to all opted-in servers
            try:
                import requests
                import os
                bot_api_url = "http://localhost:8001/api/announce-network-cotm"
                api_token = os.getenv('DISCORD_BOT_API_TOKEN')
                headers = {'Authorization': f'Bearer {api_token}'} if api_token else {}
                payload = {
                    'profile_id': profile_id
                }
                response = requests.post(bot_api_url, json=payload, headers=headers, timeout=10)
                if response.status_code == 200:
                    logger.info(f"Discord bot notified of Network COTM change: profile_id={profile_id}")
                else:
                    logger.warning(f"Failed to notify Discord bot of Network COTM change: {response.text}")
            except Exception as bot_error:
                logger.error(f"Error notifying Discord bot of Network COTM change: {bot_error}", exc_info=True)

            logger.info(f"Network COTM set by DISCOVERY_APPROVER: profile_id={profile_id}, user_id={new_cotm.discord_id}")
            return JsonResponse({'success': True, 'message': 'Network Creator of the Month set successfully'})

    except Exception as e:
        logger.error(f"Error setting Network COTM: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to set Network Creator of the Month'}, status=500)


@discovery_approvers_required
@discord_required
def clear_network_creator_of_week(request):
    """
    Clear the current Network Creator of the Week.
    Security: Only DISCOVERY_APPROVERS can clear Network COTW.
    """
    import logging
    from .db import get_db_session
    from .models import CreatorProfile

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        with get_db_session() as db:
            # Remove current Network COTW
            current_cotw = db.query(CreatorProfile).filter_by(
                is_current_network_cotw=True
            ).all()

            for profile in current_cotw:
                profile.is_current_network_cotw = False
                profile.network_cotw_last_featured = int(time.time())

            db.commit()

            logger.info(f"Network COTW cleared by DISCOVERY_APPROVER")
            return JsonResponse({'success': True, 'message': 'Network Creator of the Week cleared'})

    except Exception as e:
        logger.error(f"Error clearing Network COTW: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to clear Network Creator of the Week'}, status=500)


@discovery_approvers_required
@discord_required
def clear_network_creator_of_month(request):
    """
    Clear the current Network Creator of the Month.
    Security: Only DISCOVERY_APPROVERS can clear Network COTM.
    """
    import logging
    from .db import get_db_session
    from .models import CreatorProfile

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        with get_db_session() as db:
            # Remove current Network COTM
            current_cotm = db.query(CreatorProfile).filter_by(
                is_current_network_cotm=True
            ).all()

            for profile in current_cotm:
                profile.is_current_network_cotm = False
                profile.network_cotm_last_featured = int(time.time())

            db.commit()

            logger.info(f"Network COTM cleared by DISCOVERY_APPROVER")
            return JsonResponse({'success': True, 'message': 'Network Creator of the Month cleared'})

    except Exception as e:
        logger.error(f"Error clearing Network COTM: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to clear Network Creator of the Month'}, status=500)


# ==============================================================================
# CSP Violation Reporting Endpoint
# ==============================================================================

@csrf_exempt
@require_http_methods(["POST"])
def csp_violation_report(request):
    """
    Endpoint to receive CSP violation reports.
    Logs violations for security monitoring and analysis.
    """
    import json as json_lib
    try:
        # Parse CSP violation report
        report = json_lib.loads(request.body.decode('utf-8'))

        # Extract key information
        violated_directive = report.get('csp-report', {}).get('violated-directive', 'unknown')
        blocked_uri = report.get('csp-report', {}).get('blocked-uri', 'unknown')
        document_uri = report.get('csp-report', {}).get('document-uri', 'unknown')
        source_file = report.get('csp-report', {}).get('source-file', 'unknown')
        line_number = report.get('csp-report', {}).get('line-number', 'unknown')

        # Log the violation
        logger.warning(
            f"CSP VIOLATION: directive={violated_directive}, "
            f"blocked_uri={blocked_uri}, document={document_uri}, "
            f"source={source_file}:{line_number}"
        )

        return JsonResponse({'status': 'ok'}, status=204)

    except Exception as e:
        logger.error(f"Error processing CSP violation report: {e}", exc_info=True)
        return JsonResponse({'status': 'error'}, status=400)


# ============================================================================
# YouTube OAuth & Streaming Integration
# ============================================================================

@discord_required
@require_http_methods(["GET"])
def youtube_oauth_initiate(request, guild_id):
    """
    Initiate YouTube OAuth flow for connecting a creator's YouTube channel.

    This endpoint generates a secure OAuth URL and redirects the creator to
    Google's consent screen to grant access to their YouTube channel data.
    """
    import secrets
    from app.services.youtube_service import YouTubeService

    discord_user = request.session.get('discord_user', {})
    user_id = discord_user.get('id')

    if not user_id:
        messages.error(request, "You must be logged in to connect YouTube.")
        return redirect('questlog_dashboard')

    try:
        # Generate secure state token for CSRF protection
        state = secrets.token_urlsafe(32)

        # Store state in session with metadata
        request.session[f'youtube_oauth_state_{state}'] = {
            'guild_id': guild_id,
            'user_id': user_id,
            'timestamp': int(time.time()),
        }
        request.session.modified = True

        # Generate authorization URL
        youtube_service = YouTubeService()
        auth_url = youtube_service.get_authorization_url(state)

        logger.info(f"YouTube OAuth initiated for user {user_id} in guild {guild_id}")
        return redirect(auth_url)

    except Exception as e:
        logger.error(f"YouTube OAuth initiation error: {e}", exc_info=True)
        messages.error(request, "Failed to initiate YouTube connection. Please try again.")
        return redirect('creator_profile_register', guild_id=guild_id)


@require_http_methods(["GET"])
def youtube_oauth_callback(request):
    """
    Handle OAuth callback from Google after user grants/denies access.

    This endpoint:
    1. Validates the state token (CSRF protection)
    2. Exchanges authorization code for access & refresh tokens
    3. Fetches channel information from YouTube API
    4. Creates/updates CreatorProfile with YouTube credentials
    5. Redirects back to creator profile page
    """
    from app.services.youtube_service import YouTubeService, YouTubeAPIError
    from app.db import get_db_session
    from app.models import CreatorProfile, GuildMember, Guild, ApprovedStreamer
    import json

    # Get OAuth response parameters
    code = request.GET.get('code')
    state = request.GET.get('state')
    error = request.GET.get('error')

    # Check if user denied access
    if error:
        logger.warning(f"YouTube OAuth denied: {error}")
        messages.error(request, "YouTube connection was cancelled.")
        return redirect('questlog_dashboard')

    if not code or not state:
        logger.error("YouTube OAuth callback missing code or state")
        messages.error(request, "Invalid YouTube OAuth response.")
        return redirect('questlog_dashboard')

    # Validate state token
    session_key = f'youtube_oauth_state_{state}'
    state_data = request.session.get(session_key)

    if not state_data:
        logger.error(f"YouTube OAuth invalid state token: {state}")
        messages.error(request, "Invalid or expired OAuth session. Please try again.")
        return redirect('questlog_dashboard')

    # Clean up state from session
    del request.session[session_key]
    request.session.modified = True

    # Extract metadata
    guild_id = state_data.get('guild_id')
    user_id = state_data.get('user_id')

    # Check for expired state (30 minute timeout)
    state_age = int(time.time()) - state_data.get('timestamp', 0)
    if state_age > 1800:  # 30 minutes
        logger.warning(f"YouTube OAuth expired state token: {state_age}s old")
        messages.error(request, "OAuth session expired. Please try again.")
        return redirect('creator_profile_register', guild_id=guild_id)

    try:
        youtube_service = YouTubeService()

        # Exchange code for tokens
        tokens = youtube_service.exchange_code_for_tokens(code)
        access_token = tokens.get('access_token')
        refresh_token = tokens.get('refresh_token')
        expires_in = tokens.get('expires_in', 3600)

        if not access_token:
            raise YouTubeAPIError("No access token received")

        # Calculate token expiration timestamp
        token_expires = int(time.time()) + expires_in

        # Get channel information
        channel_info = youtube_service.get_channel_info(access_token)

        # Update database
        with get_db_session() as db:
            # Find or create creator profile
            member = db.query(GuildMember).filter(
                GuildMember.guild_id == int(guild_id),
                GuildMember.user_id == int(user_id)
            ).first()

            if not member:
                raise Exception("Guild member not found")

            creator_profile = db.query(CreatorProfile).filter(
                CreatorProfile.guild_id == int(guild_id),
                CreatorProfile.discord_id == int(user_id)
            ).first()

            if not creator_profile:
                # Create new creator profile
                # Build avatar URL from avatar_hash
                avatar_url = None
                if member.avatar_hash:
                    avatar_url = f"https://cdn.discordapp.com/avatars/{user_id}/{member.avatar_hash}.png"

                creator_profile = CreatorProfile(
                    guild_id=int(guild_id),
                    discord_id=int(user_id),
                    display_name=member.display_name or member.username,
                    avatar_url=avatar_url,
                    created_at=int(time.time()),
                    updated_at=int(time.time()),
                )
                db.add(creator_profile)
                db.flush()  # Get ID

            # Update YouTube fields with encrypted tokens
            from app.utils.encryption import encrypt_token

            creator_profile.youtube_channel_id = channel_info['id']
            creator_profile.youtube_access_token = encrypt_token(access_token)
            creator_profile.youtube_refresh_token = encrypt_token(refresh_token) if refresh_token else None
            creator_profile.youtube_token_expires = token_expires
            creator_profile.youtube_subscriber_count = channel_info.get('subscriber_count')
            creator_profile.youtube_video_count = channel_info.get('video_count')
            creator_profile.youtube_last_synced = int(time.time())

            # Store YouTube handle from custom_url or title
            if channel_info.get('custom_url'):
                # Remove @ if it's already there, we'll add it in display
                youtube_handle = channel_info['custom_url'].lstrip('@')
                creator_profile.youtube_handle = youtube_handle
                creator_profile.youtube_url = f"https://youtube.com/@{youtube_handle}"
            else:
                # Fallback to channel title as handle
                creator_profile.youtube_handle = channel_info.get('title', '').replace(' ', '')
                creator_profile.youtube_url = f"https://youtube.com/channel/{channel_info['id']}"

            creator_profile.updated_at = int(time.time())

            db.commit()

            # Auto-approve guild owner as streamer
            guild_record = db.query(Guild).filter(Guild.guild_id == int(guild_id)).first()
            if guild_record and guild_record.owner_id == int(user_id):
                # Check if already approved
                existing_approval = db.query(ApprovedStreamer).filter(
                    ApprovedStreamer.guild_id == int(guild_id),
                    ApprovedStreamer.creator_profile_id == creator_profile.id,
                    ApprovedStreamer.revoked == False
                ).first()

                if not existing_approval:
                    # Auto-approve the owner
                    approval = ApprovedStreamer(
                        guild_id=int(guild_id),
                        creator_profile_id=creator_profile.id,
                        approved_by_user_id=int(user_id),  # Self-approved
                        approved_at=int(time.time()),
                        revoked=False
                    )
                    db.add(approval)
                    db.commit()
                    logger.info(f"Auto-approved guild owner {user_id} as streamer in guild {guild_id}")

            logger.info(
                f"YouTube connected: user={user_id}, guild={guild_id}, "
                f"channel={channel_info['id']} ({channel_info['title']})"
            )

            # Return a page that notifies parent window
            from django.http import HttpResponse
            return HttpResponse("""
                <!DOCTYPE html>
                <html>
                <head>
                    <title>YouTube Connected</title>
                    <style>
                        body {{
                            font-family: system-ui, -apple-system, sans-serif;
                            display: flex;
                            align-items: center;
                            justify-content: center;
                            min-height: 100vh;
                            margin: 0;
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: white;
                        }}
                        .container {{
                            text-align: center;
                            padding: 2rem;
                            background: rgba(255, 255, 255, 0.1);
                            border-radius: 1rem;
                            backdrop-filter: blur(10px);
                        }}
                        .icon {{
                            font-size: 4rem;
                            margin-bottom: 1rem;
                        }}
                        h1 {{
                            margin: 0 0 1rem 0;
                        }}
                        p {{
                            margin: 0.5rem 0;
                            opacity: 0.9;
                        }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="icon">✓</div>
                        <h1>YouTube Connected!</h1>
                        <p>Your YouTube channel has been successfully connected.</p>
                        <p style="margin-top: 1.5rem; font-size: 0.9rem;">You can safely close this tab now.</p>
                    </div>
                    <script>
                        // Notify parent window that OAuth is complete
                        if (window.opener) {{
                            window.opener.postMessage('youtube_oauth_complete', '*');
                            // Close immediately after sending message
                            setTimeout(() => {{
                                window.close();
                                // If close didn't work, try again
                                setTimeout(() => window.close(), 500);
                            }}, 500);
                        }} else {{
                            // No opener, close after showing message
                            setTimeout(() => window.close(), 2000);
                        }}
                    </script>
                </body>
                </html>
            """.format(guild_id))

    except YouTubeAPIError as e:
        logger.error(f"YouTube API error during OAuth: {e}", exc_info=True)
        messages.error(request, f"Failed to connect YouTube: {e}")
        return redirect('creator_profile_register', guild_id=guild_id)

    except Exception as e:
        logger.error(f"YouTube OAuth callback error: {e}", exc_info=True)
        messages.error(request, "An error occurred while connecting YouTube. Please try again.")
        return redirect('creator_profile_register', guild_id=guild_id)


@discord_required
@require_http_methods(["POST"])
def youtube_disconnect(request, guild_id):
    """
    Disconnect YouTube from creator profile.

    Removes all YouTube OAuth tokens and credentials from the database.
    Does NOT revoke the token with Google (user must do that manually).
    """
    from app.db import get_db_session
    from app.models import CreatorProfile

    discord_user = request.session.get('discord_user', {})
    user_id = discord_user.get('id')

    if not user_id:
        return JsonResponse({'success': False, 'error': 'Not logged in'}, status=401)

    try:
        with get_db_session() as db:
            creator_profile = db.query(CreatorProfile).filter(
                CreatorProfile.guild_id == int(guild_id),
                CreatorProfile.discord_id == int(user_id)
            ).first()

            if not creator_profile:
                return JsonResponse({'success': False, 'error': 'Creator profile not found'}, status=404)

            # Clear YouTube fields
            creator_profile.youtube_channel_id = None
            creator_profile.youtube_handle = None
            creator_profile.youtube_url = None
            creator_profile.youtube_access_token = None
            creator_profile.youtube_refresh_token = None
            creator_profile.youtube_token_expires = None
            creator_profile.youtube_subscriber_count = None
            creator_profile.youtube_video_count = None
            creator_profile.youtube_last_synced = None
            creator_profile.is_live_youtube = False
            creator_profile.updated_at = int(time.time())

            db.commit()

            logger.info(f"YouTube disconnected: user={user_id}, guild={guild_id}")

            return JsonResponse({
                'success': True,
                'message': 'YouTube channel disconnected successfully'
            })

    except Exception as e:
        logger.error(f"YouTube disconnect error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to disconnect YouTube'
        }, status=500)


# ============================================================================
# Twitch OAuth Integration
# ============================================================================

@discord_required
@require_http_methods(["GET"])
def twitch_oauth_initiate(request, guild_id):
    """
    Initiate Twitch OAuth flow for creator profile.
    Security: Uses state token for CSRF protection, validates user is creator owner.
    """
    from app.services.twitch_service import TwitchService
    import secrets

    discord_user = request.session.get('discord_user', {})
    user_id = discord_user.get('id')

    if not user_id:
        return JsonResponse({'error': 'Not authenticated'}, status=401)

    try:
        # Generate secure state token
        state = secrets.token_urlsafe(32)

        # Store state in session with metadata
        request.session['twitch_oauth_state'] = {
            'token': state,
            'guild_id': guild_id,
            'user_id': user_id,
            'created_at': int(time.time())
        }

        # Generate authorization URL
        twitch_service = TwitchService()
        auth_url = twitch_service.get_authorization_url(state)

        return redirect(auth_url)

    except Exception as e:
        logger.error(f"Error initiating Twitch OAuth: {e}")
        return JsonResponse({'error': 'Failed to initiate Twitch OAuth'}, status=500)


@require_http_methods(["GET"])
def twitch_oauth_callback(request):
    """
    Handle Twitch OAuth callback.
    Security:
    - Validates state token (30 min timeout)
    - Encrypts tokens before storage
    - Only updates creator profile owned by user
    - Auto-approves guild owners as streamers
    """
    from app.db import get_db_session
    from app.models import CreatorProfile, ApprovedStreamer, Guild
    from app.services.twitch_service import TwitchService, TwitchAPIError
    from app.utils.encryption import encrypt_token
    import secrets

    code = request.GET.get('code')
    state = request.GET.get('state')

    if not code or not state:
        return HttpResponse("""
            <html><body>
                <h1>Error</h1>
                <p>Missing authorization code or state</p>
                <script>
                    if (window.opener) {
                        window.opener.postMessage({type: 'twitch_oauth_error', error: 'Missing parameters'}, '*');
                        window.close();
                    }
                </script>
            </body></html>
        """, status=400)

    # Validate state token
    session_state = request.session.get('twitch_oauth_state')
    if not session_state:
        return HttpResponse("""
            <html><body>
                <h1>Error</h1>
                <p>Invalid or expired state token</p>
                <script>
                    if (window.opener) {
                        window.opener.postMessage({type: 'twitch_oauth_error', error: 'Invalid state'}, '*');
                        window.close();
                    }
                </script>
            </body></html>
        """, status=400)

    # Check state matches and is not expired (30 minutes)
    if session_state['token'] != state:
        return HttpResponse("""
            <html><body>
                <h1>Error</h1>
                <p>State token mismatch</p>
                <script>
                    if (window.opener) {
                        window.opener.postMessage({type: 'twitch_oauth_error', error: 'State mismatch'}, '*');
                        window.close();
                    }
                </script>
            </body></html>
        """, status=400)

    if int(time.time()) - session_state['created_at'] > 1800:  # 30 minutes
        return HttpResponse("""
            <html><body>
                <h1>Error</h1>
                <p>State token expired. Please try again.</p>
                <script>
                    if (window.opener) {
                        window.opener.postMessage({type: 'twitch_oauth_error', error: 'State expired'}, '*');
                        window.close();
                    }
                </script>
            </body></html>
        """, status=400)

    guild_id = session_state['guild_id']
    user_id = session_state['user_id']

    try:
        twitch_service = TwitchService()

        # Exchange code for tokens
        token_data = twitch_service.exchange_code_for_token(code)
        access_token = token_data['access_token']
        refresh_token = token_data['refresh_token']
        expires_in = token_data['expires_in']

        # Get user info
        user_info = twitch_service.get_user_info(access_token)
        twitch_user_id = user_info['id']
        twitch_login = user_info['login']
        display_name = user_info['display_name']
        profile_image = user_info['profile_image_url']

        # Get channel stats
        channel_info = twitch_service.get_channel_info(access_token, twitch_user_id)
        follower_count = channel_info['follower_count']

        # Encrypt tokens before storage
        encrypted_access_token = encrypt_token(access_token)
        encrypted_refresh_token = encrypt_token(refresh_token)

        # Calculate token expiry timestamp
        token_expires = int(time.time()) + expires_in

        with get_db_session() as db:
            from app.models import GuildMember

            # Find creator profile for this user
            creator = db.query(CreatorProfile).filter(
                CreatorProfile.discord_id == user_id
            ).first()

            if not creator:
                # Auto-create creator profile from Discord info
                member = db.query(GuildMember).filter(
                    GuildMember.guild_id == int(guild_id),
                    GuildMember.user_id == int(user_id)
                ).first()

                if not member:
                    raise Exception("Guild member not found")

                # Build avatar URL from avatar_hash
                avatar_url = profile_image  # Use Twitch profile image as fallback
                if member.avatar_hash:
                    avatar_url = f"https://cdn.discordapp.com/avatars/{user_id}/{member.avatar_hash}.png"

                creator = CreatorProfile(
                    guild_id=int(guild_id),
                    discord_id=int(user_id),
                    display_name=member.display_name or member.username or display_name,
                    avatar_url=avatar_url,
                    created_at=int(time.time()),
                    updated_at=int(time.time()),
                )
                db.add(creator)
                db.flush()  # Get ID

            # Update creator profile with Twitch OAuth data
            creator.twitch_user_id = twitch_user_id
            creator.twitch_access_token = encrypted_access_token
            creator.twitch_refresh_token = encrypted_refresh_token
            creator.twitch_token_expires = token_expires
            creator.twitch_follower_count = follower_count
            creator.twitch_last_synced = int(time.time())

            # Always update twitch_handle with the login name from API
            creator.twitch_handle = twitch_login

            # Update avatar if not set
            if not creator.avatar_url and profile_image:
                creator.avatar_url = profile_image

            db.commit()

            # Auto-approve guild owner as streamer
            guild = db.query(Guild).filter(Guild.guild_id == guild_id).first()
            if guild and str(guild.owner_id) == str(user_id):
                # Check if already approved
                existing_approval = db.query(ApprovedStreamer).filter(
                    ApprovedStreamer.guild_id == guild_id,
                    ApprovedStreamer.creator_profile_id == creator.id,
                    ApprovedStreamer.revoked == False
                ).first()

                if not existing_approval:
                    approval = ApprovedStreamer(
                        guild_id=guild_id,
                        creator_profile_id=creator.id,
                        approved_by_user_id=user_id,
                        approved_at=int(time.time()),
                        revoked=False
                    )
                    db.add(approval)
                    db.commit()

        # Clear state from session
        del request.session['twitch_oauth_state']

        # Return success page that notifies parent window
        return HttpResponse(f"""
            <html>
            <head>
                <title>Twitch Connected</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        text-align: center;
                        padding: 50px;
                        background: linear-gradient(135deg, #6441a5 0%, #9147ff 100%);
                        color: white;
                    }}
                    .success-box {{
                        background: rgba(255, 255, 255, 0.1);
                        border-radius: 10px;
                        padding: 30px;
                        max-width: 500px;
                        margin: 0 auto;
                        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                    }}
                    h1 {{
                        margin-bottom: 20px;
                    }}
                </style>
            </head>
            <body>
                <div class="success-box">
                    <h1>✅ Twitch Connected!</h1>
                    <p>Your Twitch account <strong>{display_name}</strong> has been successfully connected.</p>
                    <p>This window will close automatically...</p>
                </div>
                <script>
                    // Send message to parent window
                    if (window.opener) {{
                        window.opener.postMessage({{
                            type: 'twitch_oauth_complete',
                            data: {{
                                twitch_login: '{twitch_login}',
                                display_name: '{display_name}',
                                follower_count: {follower_count}
                            }}
                        }}, '*');
                        // Close immediately after sending message
                        setTimeout(() => {{
                            window.close();
                            // If close didn't work, try again
                            setTimeout(() => window.close(), 500);
                        }}, 500);
                    }} else {{
                        // No opener, close after showing message
                        setTimeout(() => window.close(), 2000);
                    }}
                </script>
            </body>
            </html>
        """)

    except TwitchAPIError as e:
        logger.error(f"Twitch API error during OAuth: {e}")
        return HttpResponse(f"""
            <html><body>
                <h1>Error</h1>
                <p>Twitch API error: {e}</p>
                <script>
                    if (window.opener) {{
                        window.opener.postMessage({{type: 'twitch_oauth_error', error: '{e}'}}, '*');
                        window.close();
                    }}
                </script>
            </body></html>
        """, status=500)

    except Exception as e:
        logger.error(f"Error in Twitch OAuth callback: {e}")
        return HttpResponse(f"""
            <html><body>
                <h1>Error</h1>
                <p>Failed to complete Twitch authentication</p>
                <script>
                    if (window.opener) {{
                        window.opener.postMessage({{type: 'twitch_oauth_error', error: 'Server error'}}, '*');
                        window.close();
                    }}
                </script>
            </body></html>
        """, status=500)


@discord_required
@require_http_methods(["POST"])
def twitch_disconnect(request, guild_id):
    """
    Disconnect Twitch from creator profile.
    Security: Only creator owner can disconnect, CSRF protected.
    Note: Does NOT revoke token with Twitch (user must do manually in Twitch settings).
    """
    from app.db import get_db_session
    from app.models import CreatorProfile

    discord_user = request.session.get('discord_user', {})
    user_id = discord_user.get('id')

    if not user_id:
        return JsonResponse({'error': 'Not authenticated'}, status=401)

    try:
        with get_db_session() as db:
            # Find creator profile for this user
            creator = db.query(CreatorProfile).filter(
                CreatorProfile.discord_id == user_id
            ).first()

            if not creator:
                return JsonResponse({
                    'success': False,
                    'error': 'Creator profile not found'
                }, status=404)

            # Clear all Twitch OAuth fields
            creator.twitch_user_id = None
            creator.twitch_access_token = None
            creator.twitch_refresh_token = None
            creator.twitch_token_expires = None
            creator.twitch_follower_count = None
            creator.twitch_last_synced = None
            creator.is_live_twitch = False
            creator.current_stream_title = None
            creator.current_stream_game = None
            creator.current_stream_started_at = None
            creator.current_stream_thumbnail = None
            creator.current_stream_viewer_count = None

            db.commit()

            return JsonResponse({
                'success': True,
                'message': 'Twitch disconnected successfully'
            })

    except Exception as e:
        logger.error(f"Error disconnecting Twitch: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to disconnect Twitch'
        }, status=500)


# ============================================================================
# Streaming Notifications Admin API
# ============================================================================

@discord_required
@require_http_methods(["GET", "POST"])
def streaming_notifications_config(request, guild_id):
    """
    Get or update streaming notification configuration for a guild.

    GET: Returns current configuration (requires guild membership)
    POST: Updates configuration (admin only)

    Security: IDOR fix - GET requests now require guild membership verification.
    This prevents users from reading configuration for guilds they don't belong to.
    """
    from app.db import get_db_session
    from app.models import StreamingNotificationsConfig

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])
    user_guilds = request.session.get('discord_guilds', [])

    # SECURITY FIX: Require admin access for both GET and POST
    # Streaming config contains sensitive channel/role IDs that should only be visible to admins
    guild_check = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild_check:
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)

    try:
        with get_db_session() as db:
            # Get or create config
            config = db.query(StreamingNotificationsConfig).filter(
                StreamingNotificationsConfig.guild_id == int(guild_id)
            ).first()

            if request.method == 'GET':
                # Return current config
                if not config:
                    return JsonResponse({
                        'success': True,
                        'config': {
                            'enabled': False,
                            'notification_channel_id': None,
                            'ping_role_id': None,
                            'minimum_level_required': 10,
                            'notification_title': '🔴 {creator} is now LIVE!',
                            'notification_message': 'Check out the stream!',
                            'embed_color': '#FF0000',
                        }
                    })

                return JsonResponse({
                    'success': True,
                    'config': {
                        'enabled': config.enabled,
                        'notification_channel_id': str(config.notification_channel_id) if config.notification_channel_id else None,
                        'ping_role_id': str(config.ping_role_id) if config.ping_role_id else None,
                        'minimum_level_required': config.minimum_level_required,
                        'notification_title': config.notification_title or '🔴 {creator} is now LIVE!',
                        'notification_message': config.notification_message or 'Check out the stream!',
                        'embed_color': config.embed_color or '#FF0000',
                    }
                })

            # POST - Update config
            # Check for Discovery module access (required to save streaming config)
            from app.models import Guild as GuildModel, GuildModule
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                has_discovery_module = db.query(GuildModule).filter_by(
                    guild_id=int(guild_id),
                    module_name='discovery',
                    enabled=True
                ).first() is not None

                has_access = (
                    guild_record.is_vip or
                    guild_record.subscription_tier == 'complete' or
                    guild_record.billing_cycle == 'lifetime' or
                    has_discovery_module
                )

                if not has_access:
                    return JsonResponse({
                        'success': False,
                        'error': 'Discovery module or Complete tier required for streaming notifications'
                    }, status=403)

            data = json.loads(request.body)

            if not config:
                config = StreamingNotificationsConfig(
                    guild_id=int(guild_id),
                    enabled=False,
                    minimum_level_required=10,
                    created_at=int(time.time()),
                    updated_at=int(time.time()),
                )
                db.add(config)

            # Update fields
            if 'enabled' in data:
                config.enabled = bool(data['enabled'])

            if 'notification_channel_id' in data:
                channel_id = data['notification_channel_id']
                config.notification_channel_id = int(channel_id) if channel_id else None

            if 'ping_role_id' in data:
                role_id = data['ping_role_id']
                config.ping_role_id = int(role_id) if role_id else None

            if 'minimum_level_required' in data:
                config.minimum_level_required = int(data['minimum_level_required'])

            if 'notification_title' in data:
                config.notification_title = data['notification_title'] or '🔴 {creator} is now LIVE!'

            if 'notification_message' in data:
                config.notification_message = data['notification_message'] or 'Check out the stream!'

            if 'embed_color' in data:
                config.embed_color = data['embed_color'] or '#FF0000'

            config.updated_at = int(time.time())
            db.commit()

            logger.info(f"Streaming notifications config updated for guild {guild_id}")

            return JsonResponse({
                'success': True,
                'message': 'Configuration updated successfully'
            })

    except Exception as e:
        logger.error(f"Streaming notifications config error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to update configuration'
        }, status=500)


@discord_required
@require_http_methods(["GET"])
def approved_streamers_list(request, guild_id):
    """
    Get list of approved streamers for a guild.

    Returns creator profiles that are approved for notifications.
    """
    from app.db import get_db_session
    from app.models import ApprovedStreamer, CreatorProfile, GuildMember

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Check admin access
    guild_check = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild_check:
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)

    try:
        with get_db_session() as db:
            # Get approved streamers with creator info and guild member data
            approvals = db.query(ApprovedStreamer, CreatorProfile, GuildMember).join(
                CreatorProfile,
                ApprovedStreamer.creator_profile_id == CreatorProfile.id
            ).outerjoin(
                GuildMember,
                (GuildMember.guild_id == ApprovedStreamer.guild_id) &
                (GuildMember.user_id == CreatorProfile.discord_id)
            ).filter(
                ApprovedStreamer.guild_id == int(guild_id),
                ApprovedStreamer.revoked == False
            ).all()

            streamers = []
            for approval, creator, member in approvals:
                # Construct avatar URL from hash
                avatar_url = None
                if member and member.avatar_hash:
                    avatar_url = f"https://cdn.discordapp.com/avatars/{creator.discord_id}/{member.avatar_hash}.png"

                streamers.append({
                    'creator_profile_id': creator.id,
                    'user_id': str(creator.discord_id),
                    'display_name': creator.display_name,
                    'avatar_url': avatar_url,
                    'youtube_connected': bool(creator.youtube_channel_id),
                    'youtube_handle': creator.youtube_handle,
                    'twitch_connected': bool(creator.twitch_user_id),
                    'twitch_handle': creator.twitch_handle,
                    'is_live_youtube': creator.is_live_youtube,
                    'is_live_twitch': creator.is_live_twitch,
                    'approved_at': approval.approved_at,
                    'approved_by': str(approval.approved_by_user_id),
                })

            return JsonResponse({
                'success': True,
                'approved_streamers': streamers  # Frontend expects 'approved_streamers' key
            })

    except Exception as e:
        logger.error(f"Approved streamers list error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to get approved streamers'
        }, status=500)


@discord_required
@require_http_methods(["GET"])
def all_creators_with_streaming(request, guild_id):
    """
    Get all creators in a guild that have YouTube or Twitch connected.
    Used by admins to see who can be approved for streaming notifications.

    Query params:
    - platform: 'youtube', 'twitch', or 'all' (default: 'all')
    """
    from app.db import get_db_session
    from app.models import CreatorProfile, ApprovedStreamer, GuildMember

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Check admin access
    guild_check = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild_check:
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)

    platform = request.GET.get('platform', 'all').lower()

    try:
        with get_db_session() as db:
            # Base query for creators in this guild
            query = db.query(CreatorProfile).filter(
                CreatorProfile.guild_id == int(guild_id)
            )

            # Filter by platform
            if platform == 'youtube':
                query = query.filter(CreatorProfile.youtube_channel_id.isnot(None))
            elif platform == 'twitch':
                query = query.filter(CreatorProfile.twitch_user_id.isnot(None))
            else:
                # 'all' - must have at least one platform connected
                query = query.filter(
                    (CreatorProfile.youtube_channel_id.isnot(None)) |
                    (CreatorProfile.twitch_user_id.isnot(None))
                )

            creators = query.all()

            # Get list of already approved creator IDs for this guild
            approved_ids = set(
                row[0] for row in db.query(ApprovedStreamer.creator_profile_id).filter(
                    ApprovedStreamer.guild_id == int(guild_id),
                    ApprovedStreamer.revoked == False
                ).all()
            )

            result = []
            for creator in creators:
                # Get guild member for avatar
                member = db.query(GuildMember).filter(
                    GuildMember.guild_id == int(guild_id),
                    GuildMember.user_id == creator.discord_id
                ).first()

                avatar_url = None
                if member and member.avatar_hash:
                    avatar_url = f"https://cdn.discordapp.com/avatars/{creator.discord_id}/{member.avatar_hash}.png"
                elif creator.avatar_url:
                    avatar_url = creator.avatar_url

                result.append({
                    'creator_profile_id': creator.id,
                    'user_id': str(creator.discord_id),
                    'display_name': creator.display_name,
                    'avatar_url': avatar_url,
                    'youtube_connected': bool(creator.youtube_channel_id),
                    'youtube_handle': creator.youtube_handle,
                    'youtube_subscriber_count': creator.youtube_subscriber_count,
                    'twitch_connected': bool(creator.twitch_user_id),
                    'twitch_handle': creator.twitch_handle,
                    'twitch_follower_count': creator.twitch_follower_count,
                    'is_live_youtube': creator.is_live_youtube,
                    'is_live_twitch': creator.is_live_twitch,
                    'is_approved': creator.id in approved_ids,
                })

            return JsonResponse({
                'success': True,
                'creators': result
            })

    except Exception as e:
        logger.error(f"All creators with streaming error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to get creators'
        }, status=500)


@discord_required
@require_http_methods(["POST"])
def approve_streamer(request, guild_id):
    """
    Approve a creator for stream notifications.

    Requires admin access.
    """
    from app.db import get_db_session
    from app.models import ApprovedStreamer, CreatorProfile

    discord_user = request.session.get('discord_user', {})
    user_id = discord_user.get('id')
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Check admin access
    guild_check = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild_check:
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)

    try:
        data = json.loads(request.body)
        creator_profile_id = data.get('creator_profile_id')

        if not creator_profile_id:
            return JsonResponse({'success': False, 'error': 'Creator profile ID required'}, status=400)

        with get_db_session() as db:
            # Check for Discovery module access
            from app.models import Guild as GuildModel, GuildModule
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                has_discovery_module = db.query(GuildModule).filter_by(
                    guild_id=int(guild_id),
                    module_name='discovery',
                    enabled=True
                ).first() is not None

                has_access = (
                    guild_record.is_vip or
                    guild_record.subscription_tier == 'complete' or
                    guild_record.billing_cycle == 'lifetime' or
                    has_discovery_module
                )

                if not has_access:
                    return JsonResponse({
                        'success': False,
                        'error': 'Discovery module or Complete tier required'
                    }, status=403)

            # Verify creator exists and belongs to this guild
            creator = db.query(CreatorProfile).filter(
                CreatorProfile.id == int(creator_profile_id),
                CreatorProfile.guild_id == int(guild_id)
            ).first()

            if not creator:
                return JsonResponse({'success': False, 'error': 'Creator not found'}, status=404)

            # Check if already approved
            existing = db.query(ApprovedStreamer).filter(
                ApprovedStreamer.guild_id == int(guild_id),
                ApprovedStreamer.creator_profile_id == int(creator_profile_id)
            ).first()

            if existing:
                if existing.revoked:
                    # Un-revoke
                    existing.revoked = False
                    existing.revoked_by_user_id = None
                    existing.revoked_at = None
                    message = 'Streamer approval restored'
                else:
                    return JsonResponse({'success': False, 'error': 'Streamer already approved'}, status=400)
            else:
                # Create new approval
                approval = ApprovedStreamer(
                    guild_id=int(guild_id),
                    creator_profile_id=int(creator_profile_id),
                    approved_by_user_id=int(user_id),
                    approved_at=int(time.time()),
                    revoked=False
                )
                db.add(approval)
                message = 'Streamer approved for notifications'

            db.commit()

            logger.info(f"Streamer {creator_profile_id} approved in guild {guild_id} by {user_id}")

            return JsonResponse({
                'success': True,
                'message': message
            })

    except Exception as e:
        logger.error(f"Approve streamer error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to approve streamer'
        }, status=500)


@discord_required
@require_http_methods(["POST"])
def revoke_streamer(request, guild_id):
    """
    Revoke a creator's stream notification approval.

    Requires admin access.
    """
    from app.db import get_db_session
    from app.models import ApprovedStreamer

    discord_user = request.session.get('discord_user', {})
    user_id = discord_user.get('id')
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Check admin access
    guild_check = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild_check:
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)

    try:
        data = json.loads(request.body)
        creator_profile_id = data.get('creator_profile_id')

        if not creator_profile_id:
            return JsonResponse({'success': False, 'error': 'Creator profile ID required'}, status=400)

        with get_db_session() as db:
            # Check for Discovery module access
            from app.models import Guild as GuildModel, GuildModule
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                has_discovery_module = db.query(GuildModule).filter_by(
                    guild_id=int(guild_id),
                    module_name='discovery',
                    enabled=True
                ).first() is not None

                has_access = (
                    guild_record.is_vip or
                    guild_record.subscription_tier == 'complete' or
                    guild_record.billing_cycle == 'lifetime' or
                    has_discovery_module
                )

                if not has_access:
                    return JsonResponse({
                        'success': False,
                        'error': 'Discovery module or Complete tier required'
                    }, status=403)

            approval = db.query(ApprovedStreamer).filter(
                ApprovedStreamer.guild_id == int(guild_id),
                ApprovedStreamer.creator_profile_id == int(creator_profile_id)
            ).first()

            if not approval:
                return JsonResponse({'success': False, 'error': 'Approval not found'}, status=404)

            if approval.revoked:
                return JsonResponse({'success': False, 'error': 'Already revoked'}, status=400)

            # Revoke approval
            approval.revoked = True
            approval.revoked_by_user_id = int(user_id)
            approval.revoked_at = int(time.time())

            db.commit()

            logger.info(f"Streamer {creator_profile_id} revoked in guild {guild_id} by {user_id}")

            return JsonResponse({
                'success': True,
                'message': 'Streamer approval revoked'
            })

    except Exception as e:
        logger.error(f"Revoke streamer error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to revoke streamer'
        }, status=500)


@discord_required
@require_http_methods(["POST"])
def test_streaming_notification(request, guild_id):
    """
    Send a test streaming notification to the configured channel.

    Requires admin access.
    """
    from app.db import get_db_session
    from app.models import StreamingNotificationsConfig
    import os
    import requests

    discord_user = request.session.get('discord_user', {})
    admin_guilds = request.session.get('discord_admin_guilds', [])

    # Check admin access
    guild_check = next((g for g in admin_guilds if g['id'] == guild_id), None)
    if not guild_check:
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)

    try:
        # Parse request body for platform parameter
        import json
        try:
            body = json.loads(request.body.decode('utf-8'))
            platform = body.get('platform', 'youtube').lower()
        except:
            platform = 'youtube'

        with get_db_session() as db:
            # Check for Discovery module access
            from app.models import Guild as GuildModel, GuildModule
            guild_record = db.query(GuildModel).filter_by(guild_id=int(guild_id)).first()
            if guild_record:
                has_discovery_module = db.query(GuildModule).filter_by(
                    guild_id=int(guild_id),
                    module_name='discovery',
                    enabled=True
                ).first() is not None

                has_access = (
                    guild_record.is_vip or
                    guild_record.subscription_tier == 'complete' or
                    guild_record.billing_cycle == 'lifetime' or
                    has_discovery_module
                )

                if not has_access:
                    return JsonResponse({
                        'success': False,
                        'error': 'Discovery module or Complete tier required'
                    }, status=403)

            config = db.query(StreamingNotificationsConfig).filter(
                StreamingNotificationsConfig.guild_id == int(guild_id)
            ).first()

            if not config or not config.enabled:
                return JsonResponse({
                    'success': False,
                    'error': 'Streaming notifications not enabled'
                }, status=400)

            if not config.notification_channel_id:
                return JsonResponse({
                    'success': False,
                    'error': 'No notification channel configured'
                }, status=400)

            # Get bot token
            bot_token = os.getenv('DISCORD_BOT_TOKEN', '')
            if not bot_token:
                return JsonResponse({
                    'success': False,
                    'error': 'Bot token not configured'
                }, status=500)

            # Get custom notification settings
            notification_title = config.notification_title or '🔴 {creator} is now LIVE!'
            notification_message = config.notification_message or 'Check out the stream!'
            embed_color_hex = config.embed_color or '#FF0000'

            # Replace {creator} placeholder with test value
            notification_title = notification_title.replace('{creator}', 'Test Creator')
            notification_message = notification_message.replace('{creator}', 'Test Creator')

            # Convert hex color to decimal
            embed_color_decimal = int(embed_color_hex.lstrip('#'), 16)

            # Platform-specific footer branding
            if platform == 'twitch':
                footer_text = 'Twitch'
                footer_icon = 'https://static-cdn.jtvnw.net/jtv_user_pictures/8a6381c7-d0c0-4576-b179-38bd5ce1d6af-profile_image-70x70.png'
            else:  # Default to YouTube
                footer_text = 'YouTube'
                footer_icon = 'https://www.youtube.com/s/desktop/f506bd45/img/favicon_32.png'

            # Build test embed
            embed = {
                'title': notification_title,
                'description': notification_message,
                'color': embed_color_decimal,
                'fields': [
                    {
                        'name': 'Playing',
                        'value': 'Test Game/Category',
                        'inline': True
                    },
                    {
                        'name': 'Viewers',
                        'value': '0',
                        'inline': True
                    }
                ],
                'footer': {
                    'text': footer_text,
                    'icon_url': footer_icon
                },
                'thumbnail': {
                    'url': 'https://cdn.casual-heroes.com/static/img/logo.png'
                }
            }

            # Build message payload
            payload = {
                'embeds': [embed]
            }

            # Add role ping if configured
            if config.ping_role_id:
                role_content = f'<@&{config.ping_role_id}> - Test stream notification!'
                payload['content'] = role_content

            # Send via Discord API
            response = requests.post(
                f'https://discord.com/api/v10/channels/{config.notification_channel_id}/messages',
                headers={
                    'Authorization': f'Bot {bot_token}',
                    'Content-Type': 'application/json'
                },
                json=payload,
                timeout=10
            )

            if response.status_code == 200 or response.status_code == 204:
                logger.info(f"Test notification sent to guild {guild_id} channel {config.notification_channel_id}")
                return JsonResponse({
                    'success': True,
                    'message': 'Test notification sent successfully!'
                })
            else:
                logger.error(f"Discord API error: {response.status_code} - {response.text}")
                return JsonResponse({
                    'success': False,
                    'error': f'Failed to send notification (Discord API error: {response.status_code})'
                }, status=500)

    except Exception as e:
        logger.error(f"Test notification error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to send test notification'
        }, status=500)


# ============================================================================
# RAID MANAGEMENT API ENDPOINTS
# ============================================================================

@require_http_methods(["POST"])
@api_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_raid_create(request, guild_id):
    """
    POST /api/guild/<guild_id>/raids/create/
    Create a new raid event.

    FREE: up to 3 active raids
    LFG Module: unlimited active raids
    Complete Edition: unlimited active raids
    """
    from .db import get_db_session
    from .models import RaidScheduleEvent, Guild
    from .module_utils import has_module_access
    import json as json_lib

    try:
        # Parse request body
        try:
            data = json_lib.loads(request.body)
        except json_lib.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        # Validate required fields
        required_fields = ['title', 'game', 'scheduled_at']
        for field in required_fields:
            if field not in data:
                return JsonResponse({'error': f'Missing required field: {field}'}, status=400)

        title = (data.get('title') or '').strip()
        description = (data.get('description') or '').strip()
        game = (data.get('game') or '').strip()
        raid_type = (data.get('raid_type') or '').strip()
        scheduled_at = data.get('scheduled_at')
        duration_minutes = data.get('duration_minutes', 120)
        tanks_needed = data.get('tanks_needed', 2)
        healers_needed = data.get('healers_needed', 2)
        dps_needed = data.get('dps_needed', 6)

        # Validate input lengths
        if len(title) < 3 or len(title) > 150:
            return JsonResponse({'error': 'Title must be between 3 and 150 characters'}, status=400)

        if description and len(description) > 2000:
            return JsonResponse({'error': 'Description must be less than 2000 characters'}, status=400)

        if len(game) < 2 or len(game) > 100:
            return JsonResponse({'error': 'Game must be between 2 and 100 characters'}, status=400)

        if raid_type and len(raid_type) > 100:
            return JsonResponse({'error': 'Raid type must be less than 100 characters'}, status=400)

        # Validate numeric fields
        try:
            scheduled_at = int(scheduled_at)
            duration_minutes = int(duration_minutes)
            tanks_needed = int(tanks_needed)
            healers_needed = int(healers_needed)
            dps_needed = int(dps_needed)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid numeric field'}, status=400)

        # Validate timestamp is in the future
        if scheduled_at < int(time.time()):
            return JsonResponse({'error': 'Cannot schedule raid in the past'}, status=400)

        # Validate composition numbers
        if tanks_needed < 0 or tanks_needed > 50:
            return JsonResponse({'error': 'Tanks needed must be between 0 and 50'}, status=400)

        if healers_needed < 0 or healers_needed > 50:
            return JsonResponse({'error': 'Healers needed must be between 0 and 50'}, status=400)

        if dps_needed < 0 or dps_needed > 200:
            return JsonResponse({'error': 'DPS needed must be between 0 and 200'}, status=400)

        if duration_minutes < 15 or duration_minutes > 1440:
            return JsonResponse({'error': 'Duration must be between 15 and 1440 minutes'}, status=400)

        # Check active raid limit based on tier
        with get_db_session() as db:
            guild = db.query(Guild).filter_by(guild_id=guild_id).first()
            if not guild:
                return JsonResponse({'error': 'Guild not found'}, status=404)

            active_raids = db.query(RaidScheduleEvent).filter_by(
                guild_id=guild_id,
                status='scheduled'
            ).count()

            # FREE: 3 raids, LFG Module: unlimited, Complete Edition: unlimited
            has_lfg = has_module_access(guild_id, 'lfg')
            has_complete = guild.subscription_tier == 'complete'

            if has_complete or has_lfg:
                max_raids = None  # Unlimited
            else:
                max_raids = 3  # FREE tier

            if max_raids and active_raids >= max_raids:
                return JsonResponse({
                    'error': f'Maximum {max_raids} active raids reached. Upgrade to LFG Module for unlimited raids.',
                    'upgrade_required': True,
                    'current_tier': 'free',
                    'required_tier': 'lfg'
                }, status=403)

            # Get creator info from session
            discord_user = request.session.get('discord_user', {})
            creator_id = int(discord_user.get('id', 0))
            creator_name = discord_user.get('username', 'Unknown')

            if not creator_id:
                return JsonResponse({'error': 'Not authenticated'}, status=401)

            # Create the raid event
            current_time = int(time.time())
            raid = RaidScheduleEvent(
                guild_id=guild_id,
                title=title,
                description=description or None,
                game=game,
                raid_type=raid_type or None,
                scheduled_at=scheduled_at,
                duration_minutes=duration_minutes,
                tanks_needed=tanks_needed,
                healers_needed=healers_needed,
                dps_needed=dps_needed,
                creator_id=creator_id,
                creator_name=creator_name,
                status='scheduled',
                created_at=current_time,
                updated_at=current_time
            )

            db.add(raid)
            db.commit()
            db.refresh(raid)

            logger.info(f"[RAID] Created raid {raid.id} for guild {guild_id} by user {creator_id}")

            return JsonResponse({
                'success': True,
                'raid': {
                    'id': raid.id,
                    'title': raid.title,
                    'description': raid.description,
                    'game': raid.game,
                    'raid_type': raid.raid_type,
                    'scheduled_at': raid.scheduled_at,
                    'duration_minutes': raid.duration_minutes,
                    'tanks_needed': raid.tanks_needed,
                    'healers_needed': raid.healers_needed,
                    'dps_needed': raid.dps_needed,
                    'status': raid.status,
                    'created_at': raid.created_at
                }
            }, status=201)

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'error': 'An internal error occurred. Please try again later.'
        }, status=500)


@require_http_methods(["GET"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_raid_list(request, guild_id):
    """
    GET /api/guild/<guild_id>/raids/
    List all raids for a guild with optional filters.

    Query params:
    - status: filter by status (scheduled, completed, cancelled)
    - game: filter by game
    - limit: max results (default 50, max 100)
    - offset: pagination offset
    """
    from .db import get_db_session
    from .models import RaidScheduleEvent, RaidSignup

    try:
        # Parse query parameters
        status_filter = request.GET.get('status', 'scheduled')
        game_filter = request.GET.get('game')
        limit = min(int(request.GET.get('limit', 50)), 100)
        offset = int(request.GET.get('offset', 0))

        with get_db_session() as db:
            # Build query
            query = db.query(RaidScheduleEvent).filter_by(guild_id=int(guild_id))

            if status_filter:
                query = query.filter_by(status=status_filter)

            if game_filter:
                query = query.filter_by(game=game_filter)

            # Order by scheduled time
            query = query.order_by(RaidScheduleEvent.scheduled_at.asc())

            # Apply pagination
            total_count = query.count()
            raids = query.limit(limit).offset(offset).all()

            # Build response with signup counts
            raid_list = []
            for raid in raids:
                # Count signups by role
                signups = db.query(RaidSignup).filter_by(
                    raid_id=raid.id,
                    status='confirmed'
                ).all()

                tanks_count = sum(1 for s in signups if s.role == 'tank')
                healers_count = sum(1 for s in signups if s.role == 'healer')
                dps_count = sum(1 for s in signups if s.role == 'dps')

                raid_list.append({
                    'id': raid.id,
                    'title': raid.title,
                    'description': raid.description,
                    'game': raid.game,
                    'raid_type': raid.raid_type,
                    'scheduled_at': raid.scheduled_at,
                    'duration_minutes': raid.duration_minutes,
                    'composition': {
                        'tanks': {'current': tanks_count, 'needed': raid.tanks_needed},
                        'healers': {'current': healers_count, 'needed': raid.healers_needed},
                        'dps': {'current': dps_count, 'needed': raid.dps_needed}
                    },
                    'status': raid.status,
                    'created_at': raid.created_at
                })

            return JsonResponse({
                'success': True,
                'raids': raid_list,
                'pagination': {
                    'total': total_count,
                    'limit': limit,
                    'offset': offset,
                    'has_more': (offset + limit) < total_count
                }
            })

    except ValueError:
        return JsonResponse({'error': 'Invalid pagination parameters'}, status=400)
    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'error': 'An internal error occurred. Please try again later.'
        }, status=500)


@require_http_methods(["GET"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
def api_raid_detail(request, guild_id, raid_id):
    """
    GET /api/guild/<guild_id>/raids/<raid_id>/
    Get detailed information about a specific raid including all signups.
    """
    from .db import get_db_session
    from .models import RaidScheduleEvent, RaidSignup

    try:
        with get_db_session() as db:
            # Get raid
            raid = db.query(RaidScheduleEvent).filter_by(
                id=raid_id,
                guild_id=guild_id
            ).first()

            if not raid:
                return JsonResponse({'error': 'Raid not found'}, status=404)

            # Get all signups
            signups = db.query(RaidSignup).filter_by(raid_id=raid.id).all()

            # Organize signups by role and status
            signups_by_role = {
                'tank': [],
                'healer': [],
                'dps': []
            }

            bench_signups = []

            for signup in signups:
                signup_data = {
                    'id': signup.id,
                    'user_id': signup.user_id,
                    'username': signup.username,
                    'display_name': signup.display_name,
                    'avatar_url': signup.avatar_url,
                    'role': signup.role,
                    'class_spec': signup.class_spec,
                    'status': signup.status,
                    'notes': signup.notes,
                    'signed_up_at': signup.signed_up_at
                }

                if signup.status == 'bench':
                    bench_signups.append(signup_data)
                elif signup.status in ['confirmed', 'tentative', 'late']:
                    if signup.role in signups_by_role:
                        signups_by_role[signup.role].append(signup_data)

            return JsonResponse({
                'success': True,
                'raid': {
                    'id': raid.id,
                    'title': raid.title,
                    'description': raid.description,
                    'game': raid.game,
                    'raid_type': raid.raid_type,
                    'scheduled_at': raid.scheduled_at,
                    'duration_minutes': raid.duration_minutes,
                    'composition': {
                        'tanks': {'current': len(signups_by_role['tank']), 'needed': raid.tanks_needed},
                        'healers': {'current': len(signups_by_role['healer']), 'needed': raid.healers_needed},
                        'dps': {'current': len(signups_by_role['dps']), 'needed': raid.dps_needed}
                    },
                    'status': raid.status,
                    'thread_id': raid.thread_id,
                    'creator_id': raid.creator_id,
                    'created_at': raid.created_at,
                    'updated_at': raid.updated_at
                },
                'signups': {
                    'tanks': signups_by_role['tank'],
                    'healers': signups_by_role['healer'],
                    'dps': signups_by_role['dps'],
                    'bench': bench_signups
                }
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'error': 'An internal error occurred. Please try again later.'
        }, status=500)


@require_http_methods(["POST"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='POST', block=True)
def api_raid_signup(request, guild_id, raid_id):
    """
    POST /api/guild/<guild_id>/raids/<raid_id>/signup/
    Sign up for a raid event.

    Body: {
        "role": "tank|healer|dps",
        "class_spec": "Protection Warrior" (optional),
        "notes": "Can swap to heals if needed" (optional)
    }
    """
    from .db import get_db_session
    from .models import RaidScheduleEvent, RaidSignup
    import json as json_lib

    try:
        # Parse request body
        try:
            data = json_lib.loads(request.body)
        except json_lib.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        role = data.get('role', '').strip().lower()
        class_spec = data.get('class_spec', '').strip()
        notes = data.get('notes', '').strip()

        # Validate role
        if role not in ['tank', 'healer', 'dps']:
            return JsonResponse({'error': 'Role must be tank, healer, or dps'}, status=400)

        # Validate input lengths
        if class_spec and len(class_spec) > 100:
            return JsonResponse({'error': 'Class/spec must be less than 100 characters'}, status=400)

        if notes and len(notes) > 500:
            return JsonResponse({'error': 'Notes must be less than 500 characters'}, status=400)

        with get_db_session() as db:
            # Verify raid exists
            raid = db.query(RaidScheduleEvent).filter_by(
                id=raid_id,
                guild_id=guild_id
            ).first()

            if not raid:
                return JsonResponse({'error': 'Raid not found'}, status=404)

            if raid.status != 'scheduled':
                return JsonResponse({'error': 'Cannot sign up for a raid that is not scheduled'}, status=400)

            # Check if user already signed up
            existing_signup = db.query(RaidSignup).filter_by(
                raid_id=raid_id,
                user_id=request.user_id
            ).first()

            if existing_signup:
                return JsonResponse({'error': 'You are already signed up for this raid'}, status=400)

            # Get user info from Discord cache
            user_info = request.session.get('discord_user', {})
            username = user_info.get('username', 'Unknown')
            display_name = user_info.get('global_name') or user_info.get('username', 'Unknown')
            avatar_hash = user_info.get('avatar', '')
            avatar_url = f"https://cdn.discordapp.com/avatars/{request.user_id}/{avatar_hash}.png" if avatar_hash else None

            # Create signup
            current_time = int(time.time())
            signup = RaidSignup(
                raid_id=raid_id,
                guild_id=guild_id,
                user_id=request.user_id,
                username=username,
                display_name=display_name,
                avatar_url=avatar_url,
                role=role,
                class_spec=class_spec or None,
                status='confirmed',
                notes=notes or None,
                signed_up_at=current_time,
                updated_at=current_time
            )

            db.add(signup)
            db.commit()
            db.refresh(signup)

            logger.info(f"[RAID] User {request.user_id} signed up for raid {raid_id} as {role}")

            return JsonResponse({
                'success': True,
                'signup': {
                    'id': signup.id,
                    'role': signup.role,
                    'class_spec': signup.class_spec,
                    'status': signup.status,
                    'signed_up_at': signup.signed_up_at
                }
            }, status=201)

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'error': 'An internal error occurred. Please try again later.'
        }, status=500)


@require_http_methods(["DELETE"])
@api_member_auth_required
@ratelimit(key='user_or_ip', rate='30/m', method='DELETE', block=True)
def api_raid_leave(request, guild_id, raid_id):
    """
    DELETE /api/guild/<guild_id>/raids/<raid_id>/signup/
    Leave a raid (remove your signup).
    """
    from .db import get_db_session
    from .models import RaidScheduleEvent, RaidSignup

    try:
        with get_db_session() as db:
            # Verify raid exists
            raid = db.query(RaidScheduleEvent).filter_by(
                id=raid_id,
                guild_id=guild_id
            ).first()

            if not raid:
                return JsonResponse({'error': 'Raid not found'}, status=404)

            # Find user's signup
            signup = db.query(RaidSignup).filter_by(
                raid_id=raid_id,
                user_id=request.user_id
            ).first()

            if not signup:
                return JsonResponse({'error': 'You are not signed up for this raid'}, status=404)

            # Delete the signup
            db.delete(signup)
            db.commit()

            logger.info(f"[RAID] User {request.user_id} left raid {raid_id}")

            return JsonResponse({
                'success': True,
                'message': 'Successfully left the raid'
            })

    except Exception as e:
        logger.error('API error occurred', exc_info=True)
        return JsonResponse({
            'error': 'An internal error occurred. Please try again later.'
        }, status=500)



# ============================================================================
# RAID MANAGEMENT PAGE VIEW
# ============================================================================

